"""
K-IFRS 1116 리스 완전성(Completeness) 검토 스크립트

실행 방법:
  [모드 1] 기중비용원장 직접 분석 (input_data/ 폴더 사용)
      python lease_filter.py

  [모드 2] Playwright 상세검색 결과 파일 연계
      python lease_filter.py --company dae_il
      python lease_filter.py --company dae_il --no-filter
"""

import argparse
import glob
import io
import os
import re
import sys

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── 리스 탐지 키워드 (상세검색 필터용) ───────────────────────────────────────
LEASE_KEYWORDS = [
    '호', '빌딩', '건물', '임차', '차량', '렌탈', '렌트', '리스',
    '사무실', '복합기', '정수기', '냉난방', '에어컨', '주차', '창고',
]

# ── K-IFRS 1116 자동 판정 기준 ─────────────────────────────────────────────

# 적요 기준 리스 인식 필요 키워드 (명확한 것만)
_LEASE_O_DESC = [
    '임차료', '임대료', '렌탈료', '리스료',
    '임차', '임대',
    '렌트카', '렌터카', '사무실', '창고', '공장', '호실',
]
# 거래처명 기준 명시적 렌탈/리스 업체 (단순 '리스' 포함 제외 — 오탐 방지)
_LEASE_O_VENDOR_EXACT = [
    'SK매직', '코웨이', '청호나이스', '현대렌탈', 'KT렌탈',
    'AJ렌터카', 'SK렌터카', '롯데렌터카', '쏘카', '오릭스', '한국리스',
]
# 거래처명 키워드 (단어 전체 포함이어야 유효 — '렌탈', '렌터카' 포함 업체)
_LEASE_O_VENDOR_KW = ['렌탈', '렌터카', '임대']

# 적요 기준 비리스 키워드 (리스 판정보다 먼저 체크)
_LEASE_X_DESC = [
    '유류', '주유', '기름', '경유', '휘발유', '통행료', '하이패스', '주차',
    '컨설팅', '강연', '사례금', '강사료', '자문', '채용', '헤드헌터',
    '특허', '상표', '출원', '인증심사', '심사비',
    '연구용역', '연구비', '과제비', '용역비', '보험료',
    '사이닝', '보너스', '법인카드', '로드쇼', '간담회',
    '중개수수료', '중계수수료', '중개비',
    '송금수수료', '해외송금',
    '정기구독', '구독료',
    'Consulting', 'Services', 'Flight', 'GMP', 'PKG', 'MDR',
]
# 저가자산 면제 검토 키워드 (K-IFRS 1116.5(b), USD 5,000 이하)
_LOW_VALUE_DESC = ['정수기', '공기청정기', '복합기', '프린터', '냉온수기', '냉정수기', '에어컨', '제빙기']
# 저가자산 면제 금액 기준 proxy (USD 5,000 × 1,400 = 7,000,000원)
_LOW_VALUE_AMT = 7_000_000


def classify_lease(row: dict) -> tuple:
    """
    K-IFRS 1116 기준으로 리스 인식 여부를 자동 판정.
    Returns: (리스인식여부, 면제검토, 판단근거)
      리스인식여부 : 'O' = 리스 인식 필요 / 'X' = 비리스 / '?' = 추가 검토
      면제검토     : '단기' / '저가' / '단기·저가' / ''
      판단근거     : 판정 이유 문자열
    """
    account = str(row.get('계정과목', ''))
    desc    = str(row.get('대표_적요', ''))
    vendor  = str(row.get('거래처', ''))
    amount  = float(row.get('연간_총발생액', 0) or 0)
    count   = int(row.get('거래_발생건수', 0) or 0)

    def _exemption(amt, cnt):
        parts = []
        if cnt < 12:
            parts.append('단기')
        if amt <= _LOW_VALUE_AMT or any(kw in desc for kw in _LOW_VALUE_DESC):
            parts.append('저가')
        return '·'.join(parts)

    # ── 0. 비리스 적요 선검사 (임차료 계정은 예외)
    if '임차료' not in account:
        for kw in _LEASE_X_DESC:
            if kw in desc or kw.lower() in desc.lower():
                return 'X', '', f'적요 [{kw}]'

    # ── 1. 임차료 계정 → O (K-IFRS 1116 적용 대상)
    if '임차료' in account:
        return 'O', _exemption(amount, count), f'{account} 계정'

    # ── 2. 적요에 리스 키워드 → O
    for kw in _LEASE_O_DESC:
        if kw in desc:
            return 'O', _exemption(amount, count), f'적요 [{kw}]'

    # ── 3. 거래처가 명시적 렌탈/리스 업체 → O
    for kw in _LEASE_O_VENDOR_EXACT:
        if kw in vendor:
            return 'O', _exemption(amount, count), f'거래처 [{kw}]'
    for kw in _LEASE_O_VENDOR_KW:
        if kw in vendor:
            return 'O', _exemption(amount, count), f'거래처명 [{kw}]'

    # ── 4. 차량유지비 + 렌트 → O
    if '차량유지비' in account and ('렌트' in desc or '렌탈' in desc):
        return 'O', _exemption(amount, count), '차량 렌트'

    # ── 5. 수수료 계정 → X
    if '수수료' in account:
        return 'X', '', f'{account} — 리스 키워드 없음'

    # ── 6. 차량유지비 → X
    if '차량유지비' in account:
        return 'X', '', '차량유지비 — 소비성'

    return '?', '', '추가 검토 필요'

# ── 경로 설정 ─────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(BASE_DIR)   # audit-automation/
INPUT_DIR   = os.path.join(BASE_DIR, 'input_data')
OUTPUT_DIR  = os.path.join(BASE_DIR, 'output')

KEYWORD_PATTERN = '|'.join(re.escape(kw) for kw in LEASE_KEYWORDS)
REQUIRED_COLS   = {'일자', '계정과목', '적요', '거래처', '차변'}

# 계정과목 코드 정규화 패턴: '205_이자비용(93100)' → '이자비용'
_ACCOUNT_RE = re.compile(r'^\d+_(.+?)\s*\(\d+\)\s*$')


# ── 공통 유틸 ─────────────────────────────────────────────────────────────────
def _clean_account(name: str) -> str:
    m = _ACCOUNT_RE.match(str(name).strip())
    return m.group(1).strip() if m else str(name).strip()


def _top_remarks(series: pd.Series, n: int = 2) -> str:
    cleaned = series[series.str.strip() != '']
    if cleaned.empty:
        return ''
    return ' / '.join(cleaned.value_counts().head(n).index.tolist())


def _normalize_desc_key(desc: str) -> str:
    """적요를 집계 키용으로 정규화 — 월/차수 등 변동 요소만 제거하여 동일 자산을 하나의 그룹으로."""
    s = str(desc).strip()
    s = re.sub(r'\d{4}년\s*', '', s)            # 연도 제거
    s = re.sub(r'\d{1,2}월분?', '', s)           # 1월, 2월분 등 제거
    s = re.sub(r'\d+\s*(차|회)차?\s*', '', s)    # 1차, 2회 등 차수 제거
    s = re.sub(r'\(\s*\)', '', s)               # 빈 괄호 제거
    s = re.sub(r'\s+', ' ', s).strip()
    return s if s else '(적요없음)'


# ── 모드 1: input_data/ 원장 로드 ────────────────────────────────────────────
def load_ledger() -> pd.DataFrame:
    files = glob.glob(os.path.join(INPUT_DIR, '*.xlsx'))
    files += glob.glob(os.path.join(INPUT_DIR, '*.xls'))
    if not files:
        raise FileNotFoundError(
            f"[오류] input_data 폴더에 엑셀 파일이 없습니다.\n  경로: {INPUT_DIR}"
        )
    frames = []
    for f in sorted(files):
        df = pd.read_excel(f)
        missing = REQUIRED_COLS - set(df.columns)
        if missing:
            raise ValueError(f"[오류] 필수 컬럼 누락 - {os.path.basename(f)}: {missing}")
        frames.append(df)
        print(f"  로드: {os.path.basename(f)}  ({len(df):,}행)")
    return pd.concat(frames, ignore_index=True)


# ── 모드 2: 회사 results/ 상세검색 결과 로드 ──────────────────────────────────
def _normalize_result_sheet(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """상세검색 결과 컬럼 → 표준 컬럼 정규화."""
    df = df.copy()

    # 컬럼명 공백 정규화: '적   요   란' → '적요란', '차    변' → '차변' 등
    df.columns = [re.sub(r'\s+', '', str(c)) for c in df.columns]

    # 날짜 → 일자
    if '날짜' in df.columns:
        df = df.rename(columns={'날짜': '일자'})

    # 적요란 → 적요 (실제 설명은 적요란에 있음)
    if '적요란' in df.columns:
        raw_desc = df['적요란'].fillna('')
        df['적요'] = raw_desc.astype(str).str.strip()

    # 거래처명 → 거래처 (detail_search_extractor 출력 형식 대응)
    if '거래처명' in df.columns and '거래처' not in df.columns:
        df = df.rename(columns={'거래처명': '거래처'})

    # 계정과목 코드 정규화
    if '계정과목' in df.columns:
        df['계정과목'] = df['계정과목'].astype(str).apply(_clean_account)

    # 계정과목이 모두 비어있으면 시트명으로 보완
    if '계정과목' not in df.columns or df['계정과목'].replace('', pd.NA).isna().all():
        df['계정과목'] = sheet_name

    # 전기이월 행 제거 (집계 왜곡 방지)
    if '적요' in df.columns:
        df['적요'] = df['적요'].fillna('').astype(str)
        df = df[~df['적요'].str.contains('전기이월', na=False)]

    return df


def load_from_result(company: str, base: str = None) -> pd.DataFrame:
    """회사 results/ 폴더에서 리스 완전성 결과 파일을 찾아 전체 시트 병합."""
    if base:
        results_dir = os.path.join(PROJECT_ROOT, base, company, 'results')
    else:
        results_dir = os.path.join(PROJECT_ROOT, company, 'results')
    if not os.path.isdir(results_dir):
        raise FileNotFoundError(f"[오류] results 폴더가 없습니다: {results_dir}")

    candidates = []
    for pat in ['*리스*완전성*.xlsx', '*리스완전성*.xlsx', '*리스거래*.xlsx']:
        candidates += glob.glob(os.path.join(results_dir, pat))
    candidates = sorted(set(f for f in candidates if not os.path.basename(f).startswith('~$')))

    if not candidates:
        raise FileNotFoundError(
            f"[오류] 리스 완전성 결과 파일을 찾을 수 없습니다.\n"
            f"  탐색 경로: {results_dir}\n"
            f"  Playwright 상세검색 시나리오를 먼저 실행해 주세요."
        )

    target = candidates[-1]   # 가장 최신 파일
    print(f"  대상 파일: {os.path.basename(target)}")

    xl = pd.ExcelFile(target)
    print(f"  시트 목록: {xl.sheet_names}")

    frames = []
    for sheet in xl.sheet_names:
        raw = pd.read_excel(target, sheet_name=sheet)
        normalized = _normalize_result_sheet(raw, sheet)
        frames.append(normalized)
        print(f"  시트 '{sheet}': {len(normalized):,}행")

    merged = pd.concat(frames, ignore_index=True)

    # 필수 컬럼 누락 체크
    missing = REQUIRED_COLS - set(merged.columns)
    if missing:
        raise ValueError(
            f"[오류] 정규화 후에도 필수 컬럼 누락: {missing}\n"
            f"  실제 컬럼: {list(merged.columns)}"
        )
    return merged


# ── 공통 파이프라인 ───────────────────────────────────────────────────────────
def preprocess(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df['적요']   = df['적요'].fillna('').astype(str).str.strip()
    df['거래처']  = df['거래처'].fillna('(미기재)').astype(str).str.strip()
    df['차변']   = pd.to_numeric(df['차변'], errors='coerce').fillna(0)
    return df


def filter_lease_rows(df: pd.DataFrame) -> pd.DataFrame:
    return df[df['적요'].str.contains(
        KEYWORD_PATTERN, flags=re.IGNORECASE, regex=True, na=False
    )].copy()


def aggregate(df: pd.DataFrame) -> pd.DataFrame:
    # 적요 정규화 키: 월/차수 제거 후 동일 자산은 같은 그룹, 다른 자산(차량번호 등)은 다른 그룹
    df = df.copy()
    df['_desc_key'] = df['적요'].apply(_normalize_desc_key)

    result = (
        df.groupby(['거래처', '계정과목', '_desc_key'], as_index=False, sort=False)
        .agg(
            연간_총발생액=('차변', 'sum'),
            거래_발생건수=('차변', 'count'),
            대표_적요=('적요', _top_remarks),
        )
        .drop(columns=['_desc_key'])
        .sort_values('연간_총발생액', ascending=False)
        .reset_index(drop=True)
    )

    classifications = result.to_dict('records')
    result['리스인식여부(O/X)'] = [classify_lease(r)[0] for r in classifications]
    result['면제검토']          = [classify_lease(r)[1] for r in classifications]
    result['판단근거']          = [classify_lease(r)[2] for r in classifications]
    result['비고'] = ''
    return result


def save_excel(df: pd.DataFrame, output_path: str) -> None:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    col_labels = {
        '연간_총발생액': '연간 총 발생액',
        '거래_발생건수': '거래 발생건수',
        '대표_적요':    '대표 적요',
    }
    df = df.rename(columns=col_labels)

    # 판정 결과별 색상
    _FILL = {
        'O': PatternFill('solid', fgColor='C6EFCE'),   # 녹색 — 리스 인식 필요
        'X': PatternFill('solid', fgColor='FFCCCC'),   # 빨강 — 비리스
        '?': PatternFill('solid', fgColor='FFEB9C'),   # 노랑 — 추가 검토
    }
    _FONT = {
        'O': Font(bold=True, color='276221'),
        'X': Font(bold=False, color='9C0006'),
        '?': Font(bold=False, color='7D4E00'),
    }

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='리스후보목록')
        ws = writer.sheets['리스후보목록']

        judgment_col = df.columns.get_loc('리스인식여부(O/X)') + 1

        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            j_cell = row_cells[judgment_col - 1]
            val = str(j_cell.value or '').strip()
            if val in _FILL:
                for cell in row_cells:
                    cell.fill = _FILL[val]
                j_cell.font = _FONT[val]

        # 천단위 콤마 서식
        amt_col = df.columns.get_loc('연간 총 발생액') + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=amt_col, max_col=amt_col):
            for cell in row:
                cell.number_format = '#,##0'

        # 헤더 스타일
        hdr_fill = PatternFill('solid', fgColor='D9E1F2')
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 18

        # 컬럼 너비 자동 조정
        col_widths = {'거래처': 28, '계정과목': 18, '연간 총 발생액': 16,
                      '거래 발생건수': 12, '대표 적요': 45,
                      '리스인식여부(O/X)': 14, '면제검토': 10, '판단근거': 30, '비고': 20}
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

    # 판정 요약 출력
    counts = df['리스인식여부(O/X)'].value_counts()
    o_amt  = df.loc[df['리스인식여부(O/X)'] == 'O', '연간 총 발생액'].sum()
    print(f"\n저장 완료: {output_path}")
    print(f"  전체 후보: {len(df):,}건 / 연간 총 발생액: {df['연간 총 발생액'].sum():,.0f}원")
    print(f"  ● O (리스 인식 필요): {counts.get('O', 0):,}건 / {o_amt:,.0f}원")
    print(f"  ● X (비리스):         {counts.get('X', 0):,}건")
    print(f"  ● ? (추가 검토):      {counts.get('?', 0):,}건")
    if '면제검토' in df.columns:
        exempt = df[df['면제검토'] != '']['면제검토'].value_counts()
        if not exempt.empty:
            print(f"  ※ 면제 검토 대상: {exempt.to_dict()}")


# ── CLI ───────────────────────────────────────────────────────────────────────
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description='K-IFRS 1116 리스 완전성 검토',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            '예시:\n'
            '  python lease_filter.py                     # input_data/ 원장 분석\n'
            '  python lease_filter.py --company dae_il    # Playwright 결과 연계\n'
            '  python lease_filter.py --company dae_il --no-filter  # 키워드 필터 생략\n'
        )
    )
    p.add_argument('--company', '-c', metavar='COMPANY',
                   help='회사 폴더명 (예: dae_il). 지정 시 results/*리스*완전성*.xlsx 자동 탐색')
    p.add_argument('--base', metavar='BASE',
                   help='루트 하위 기준 폴더 (예: --base journal_analyzer)')
    p.add_argument('--no-filter', action='store_true',
                   help='키워드 필터링 생략 (계정과목이 이미 리스 특정 계정인 경우)')
    p.add_argument('--output', '-o', metavar='FILE',
                   help='출력 파일 전체 경로 (기본: output/{company}_리스완전성검토_후보목록.xlsx)')
    return p.parse_args()


# ── main ──────────────────────────────────────────────────────────────────────
def main() -> None:
    args = parse_args()

    print('=' * 60)
    print('  K-IFRS 1116 리스 완전성 검토 - 후보 추출')
    if args.company:
        print(f'  모드: 상세검색 결과 연계  (회사: {args.company})')
    else:
        print('  모드: 기중비용원장 직접 분석')
    print('=' * 60)

    # ── 1. 데이터 로드
    print('\n[1/3] 데이터 로드')
    if args.company:
        raw = load_from_result(args.company, base=args.base)
        output_file = args.output if args.output else os.path.join(OUTPUT_DIR, f'{args.company}_리스완전성검토_후보목록.xlsx')
    else:
        raw = load_ledger()
        output_file = args.output if args.output else os.path.join(OUTPUT_DIR, '리스완전성검토_후보목록.xlsx')
    print(f'  전체 행 수: {len(raw):,}')

    # ── 2. 전처리 + 필터링
    print('\n[2/3] 전처리 및 필터링')
    df = preprocess(raw)

    if args.no_filter:
        filtered = df[df['차변'] > 0].copy()   # 차변 발생 건만
        print(f'  키워드 필터 생략 (--no-filter) → 차변 발생 건: {len(filtered):,}행')
    else:
        print(f'  적용 키워드: {LEASE_KEYWORDS}')
        filtered = filter_lease_rows(df)
        ratio = len(filtered) / len(df) * 100 if len(df) else 0
        print(f'  필터링 후: {len(filtered):,}행  (전체 대비 {ratio:.1f}%)')

    if filtered.empty:
        print('\n  [안내] 해당하는 거래 내역이 없습니다.')
        return

    # ── 3. 집계 + 저장
    print('\n[3/3] 거래처·계정과목별 집계 및 저장')
    result = aggregate(filtered)
    save_excel(result, output_file)

    print("\n완료. 결과 파일에서 '리스인식여부(O/X)' 컬럼을 체크해 주세요.")
    print('=' * 60)


if __name__ == '__main__':
    main()
