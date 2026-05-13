#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
K-IFRS 1116 리스 회계처리 자동화 스크립트

Usage:
    python lease_schedule.py

입력: input_data/lease_{회사명}_information_{연도}.xlsx  (header=1)
출력: output/lease_schedule_{회사명}_{연도}.xlsx
  - Sheet 1: 계약별 요약
  - Sheet 2: 월별 상각 스케줄

주요 컬럼:
    리스계약번호, 리스개시일, 리스종료일, 최종 산정리스기간(월수),
    정기리스료, 지급주기(월/분기/년), 지급시점(기초/기말),
    적용할인율(연 이자율), 보증잔존가치/매수선택권,
    선급리스료, 수취한 리스인센티브, 리스개설직접원가, 복구충당부채,
    지급보증금(명목가액)
"""

import glob
import os
import re
import shutil
import sys

import numpy_financial as npf
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BASE_DIR)   # audit-automation 루트
INPUT_DIR   = os.path.join(BASE_DIR, 'input_data')
OUTPUT_DIR  = os.path.join(BASE_DIR, 'output')

MONEY_FMT  = '#,##0'
RATE_FMT   = '0.00%'


# ── 이자율 변환 ───────────────────────────────────────────────────────────────

def _months_per_period(freq: str) -> int:
    freq = str(freq).strip()
    if freq in ('월', '월별', 'M', '매월'):                return 1
    if freq in ('분기', '분기별', 'Q', '매분기'):           return 3
    if freq in ('반기', '반기별', 'H', '매반기', '6개월'):  return 6
    if freq in ('년', '연', '연간', '년간', 'Y', '매년'):   return 12
    raise ValueError(f'지원하지 않는 지급주기: {freq}')


def _period_rate(annual_rate: float, freq: str) -> float:
    """연 이자율 → 지급주기 이자율 (유효이자율 방식)"""
    m = _months_per_period(freq)
    return (1 + annual_rate) ** (m / 12) - 1


def _monthly_rate(annual_rate: float) -> float:
    return (1 + annual_rate) ** (1 / 12) - 1


def _safe_float(v, default: float = 0.0) -> float:
    """NaN/None/공백 → default"""
    try:
        f = float(v)
        return default if (f != f) else f   # NaN check
    except (TypeError, ValueError):
        return default


def _col(row: dict, *candidates) -> object:
    """컬럼명 후보 중 값이 있는 첫 번째 반환 (괄호 제거 부분 일치 폴백)"""
    for key in candidates:
        if key in row:
            v = row[key]
            if v is not None and str(v).strip() not in ('', 'nan', 'NaN', 'None'):
                return v
    # 폴백: 괄호·특수문자 제거 후 부분 일치
    def _strip(s):
        return re.sub(r'[\(\)（）/\s]', '', str(s))
    for key in candidates:
        k_stripped = _strip(key)
        for col in row:
            if k_stripped in _strip(col) or _strip(col) in k_stripped:
                v = row[col]
                if v is not None and str(v).strip() not in ('', 'nan', 'NaN', 'None'):
                    return v
    return None


# ── 스케줄 생성 ───────────────────────────────────────────────────────────────

def build_schedule(c: dict) -> tuple:
    """
    단일 계약의 월별 상각 스케줄 생성.
    Returns (schedule_df, initial_liability, initial_rou, deposit, deposit_pv, deposit_discount)
    """
    start    = pd.Timestamp(c['리스개시일']).date()
    n_months = int(_safe_float(_col(c, '최종 산정리스기간(월수)'), 1))
    payment  = _safe_float(_col(c, '정기리스료'))
    freq     = str(_col(c, '지급주기(월/분기/년)', '지급주기(월/분기/반기/년)', '지급주기') or '월').strip()
    timing   = str(_col(c, '지급시점(기초/기말)', '지급시점') or '기말').strip()
    annual_r = _safe_float(_col(c, '적용할인율(연 이자율)', '적용할인율', '이자율'))
    if annual_r > 1:        # % 단위 입력 (예: 5.0 → 0.05)
        annual_r /= 100

    residual    = _safe_float(_col(c, '보증잔존가치/매수선택권', '보증잔존가치/매수선택권금액', '잔존가치'))
    prepaid     = _safe_float(_col(c, '선급리스료'))
    incentive   = _safe_float(_col(c, '수취한 리스인센티브', '리스인센티브'))
    direct_cost = _safe_float(_col(c, '리스개설직접원가', '직접원가'))
    restoration = _safe_float(_col(c, '복구충당부채', '복구충당부채(현재가치)', '복구충당'))
    deposit     = _safe_float(_col(c, '지급보증금(명목가액)', '지급보증금', '보증금'))

    mpp    = _months_per_period(freq)      # 지급주기(개월)
    n_prd  = n_months // mpp              # 총 지급 횟수
    r_prd  = _period_rate(annual_r, freq) # 기간 이자율
    r_mo   = _monthly_rate(annual_r)      # 월 이자율
    when   = 1 if timing == '기초' else 0  # numpy_financial 규칙

    # ── 최초 리스부채 현재가치 (numpy_financial.pv)
    # npf.pv: pmt 음수(유출) 입력 시 양수(자산/부채 가치)를 반환
    initial_liability = float(npf.pv(
        rate=r_prd, nper=n_prd, pmt=-payment, fv=-residual, when=when
    ))

    # ── 보증금 현재가치(PV) 및 현재가치할인차금
    # 할인 기간: n_months(월), 할인율: r_mo(월 단위)
    # 보증금 회수일 = 리스종료일 가정
    if deposit > 0 and r_mo > 0:
        deposit_pv = deposit / (1 + r_mo) ** n_months
    else:
        deposit_pv = deposit  # 할인율 0이면 PV = 명목가액
    deposit_discount = deposit - deposit_pv   # 현재가치할인차금

    # ── 사용권자산 최초 원가 (보증금 현재가치할인차금 가산)
    # 보증금 할인차금은 선급리스료 성격으로 ROU에 포함
    initial_rou = max(0.0, initial_liability + prepaid - incentive + direct_cost + restoration + deposit_discount)

    # ── 지급 월 집합 (1-based 월 인덱스)
    pay_months: set = set()
    if when == 1:   # 기초: 1, mpp+1, 2*mpp+1, ...
        for i in range(n_prd):
            pay_months.add(i * mpp + 1)
    else:           # 기말: mpp, 2*mpp, ...
        for i in range(1, n_prd + 1):
            pay_months.add(i * mpp)

    monthly_dep = initial_rou / n_months if n_months > 0 else 0.0   # 정액법 월 감가상각비

    rows = []
    liab    = initial_liability
    rou     = initial_rou
    dep_bal = deposit_pv   # 보증금 당월 기초PV

    for mi in range(1, n_months + 1):
        dt       = start + relativedelta(months=mi - 1)
        ym       = dt.strftime('%Y-%m')
        is_pay   = mi in pay_months
        is_last  = mi == n_months

        op_liab = liab
        op_rou  = rou
        dep_op  = dep_bal

        # 현금지급액
        cash = payment if is_pay else 0.0
        if is_last and residual > 0:
            cash += residual

        # 이자비용
        # 기초 지급 월: 지급 후 잔액에 이자 적용 (당월 지급액은 원금에서 선차감)
        if when == 1 and is_pay:
            interest = max(0.0, op_liab - payment) * r_mo
        else:
            interest = op_liab * r_mo

        # 기말 부채 잔액: 기초 + 이자 - 지급액 (음수 방지)
        cl_liab = max(0.0, op_liab + interest - cash)
        cl_rou  = max(0.0, op_rou - monthly_dep)

        # 보증금 상각 (유효이자율법)
        # 마지막 달은 단수 차이를 이자수익에서 가감하여 기말PV = 명목가액으로 맞춤
        if deposit > 0:
            if is_last:
                dep_int = deposit - dep_op
            else:
                dep_int = dep_op * r_mo
            dep_cl = dep_op + dep_int
        else:
            dep_int = None
            dep_cl  = None

        rows.append({
            '연월':               ym,
            '기초부채잔액':        op_liab,
            '이자비용':           interest,
            '현금지급액':         cash,
            '원금상환액':         cash - interest,   # 음수=이자만 발생(비지급월)
            '기말부채잔액':        cl_liab,
            '기초자산잔액':        op_rou,
            '감가상각비':         monthly_dep,
            '기말자산잔액':        cl_rou,
            '유동성대체대상액':     None,
            '비유동성리스부채잔액':  None,
            '보증금_기초PV':       dep_op  if deposit > 0 else None,
            '보증금_이자수익':      dep_int if deposit > 0 else None,
            '보증금_기말PV':       dep_cl  if deposit > 0 else None,
        })

        liab    = cl_liab
        rou     = cl_rou
        dep_bal = dep_cl if deposit > 0 else 0.0

    df = pd.DataFrame(rows)

    # ── 12월 말 유동성 대체 계산
    # 당해 12월 잔액 - 다음 해 12월 잔액 = 내년도 원금상환 예정액(유동성 대체액)
    dec_idx = df[df['연월'].str.endswith('-12')].index.tolist()
    for k, idx in enumerate(dec_idx):
        cur_bal  = df.loc[idx, '기말부채잔액']
        next_bal = df.loc[dec_idx[k + 1], '기말부채잔액'] if k + 1 < len(dec_idx) else 0.0
        df.loc[idx, '유동성대체대상액']    = cur_bal - next_bal
        df.loc[idx, '비유동성리스부채잔액'] = next_bal

    return df, initial_liability, initial_rou, deposit, deposit_pv, deposit_discount


# ── 연간 집계 ─────────────────────────────────────────────────────────────────

def _annual_summary(sched: pd.DataFrame) -> pd.DataFrame:
    """월별 스케줄 → 연간 집계 (리스부채 + 사용권자산)"""
    rows = []
    for yr, grp in sched.groupby(sched['연월'].str[:4], sort=True):
        dec = grp[grp['연월'].str.endswith('-12')]

        def _dec(col):
            if dec.empty or col not in dec.columns:
                return None
            v = dec.iloc[0][col]
            return None if (v is None or (isinstance(v, float) and v != v)) else float(v)

        rows.append({
            '연도':            yr,
            '기초리스부채잔액':  float(grp.iloc[0]['기초부채잔액']),
            '이자비용':         float(grp['이자비용'].sum()),
            '현금지급액':       float(grp['현금지급액'].sum()),
            '원금상환액':       float(grp['원금상환액'].sum()),
            '기말리스부채잔액':  float(grp.iloc[-1]['기말부채잔액']),
            '유동성리스부채':   _dec('유동성대체대상액'),
            '비유동성리스부채': _dec('비유동성리스부채잔액'),
            '기초사용권자산':   float(grp.iloc[0]['기초자산잔액']),
            '감가상각비':       float(grp['감가상각비'].sum()),
            '기말사용권자산':   float(grp.iloc[-1]['기말자산잔액']),
        })
    return pd.DataFrame(rows)


# ── Excel 저장 ────────────────────────────────────────────────────────────────

_HDR_FILL   = PatternFill('solid', fgColor='1F497D')
_HDR_FONT   = Font(bold=True, color='FFFFFF', size=10)
_DEC_FILL   = PatternFill('solid', fgColor='FFF2CC')
_CUR_FILL   = PatternFill('solid', fgColor='E2EFDA')
_INFO_FILL  = PatternFill('solid', fgColor='EBF3FF')
_TITLE_FILL = PatternFill('solid', fgColor='D6E4F0')
_SUM_FILL   = PatternFill('solid', fgColor='F2F2F2')
_DEP_FILL   = PatternFill('solid', fgColor='EDE7F6')   # 보증금 열 구분색
_THIN       = Side(style='thin', color='BFBFBF')
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _set_header(ws):
    for cell in ws[1]:
        cell.fill      = _HDR_FILL
        cell.font      = _HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30


def _money_cols(ws, df: pd.DataFrame, exclude=(), min_row=2):
    for ci, col in enumerate(df.columns, 1):
        if col in exclude:
            continue
        for cell in ws.iter_rows(min_row=min_row, max_row=ws.max_row,
                                  min_col=ci, max_col=ci):
            for c in cell:
                if isinstance(c.value, (int, float)):
                    c.number_format = MONEY_FMT


def _col_width(ws, widths: dict, default=14):
    for ci, col_dim in ws.column_dimensions.items():
        pass
    for ci, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        hdr = str(col[0].value or '')
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(hdr, default)


def _write_contract_sheet(ws, cid: str, info: dict, annual: pd.DataFrame, sched: pd.DataFrame):
    """단일 계약 시트: 계약정보 → 리스부채 연간요약 → 사용권자산 연간요약 → 월별상세"""

    CTR = Alignment(horizontal='center', vertical='center')
    RGT = Alignment(horizontal='right',  vertical='center')
    LFT = Alignment(horizontal='left',   vertical='center')

    deposit          = info.get('deposit', 0.0) or 0.0
    deposit_pv       = info.get('deposit_pv', 0.0) or 0.0
    deposit_discount = info.get('deposit_discount', 0.0) or 0.0
    has_deposit      = deposit > 0

    def _bc(r, c):
        cell = ws.cell(row=r, column=c)
        cell.border = _BORDER
        return cell

    def _hdr(r, c, text, fill=None):
        cell = _bc(r, c)
        cell.value = text
        cell.fill  = fill if fill else _HDR_FILL
        cell.font  = _HDR_FONT
        cell.alignment = CTR

    def _money(r, c, val):
        cell = _bc(r, c)
        if val is not None and not (isinstance(val, float) and val != val):
            cell.value = round(float(val))
            cell.number_format = MONEY_FMT
        cell.alignment = RGT

    def _title(r, text, n=9):
        cell = ws.cell(row=r, column=1)
        cell.value = text
        cell.font  = Font(bold=True, size=11, color='1F3864')
        cell.fill  = _TITLE_FILL
        cell.alignment = LFT
        if n > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
        ws.row_dimensions[r].height = 22

    cur = 1

    # ── 1. 계약 정보 ──────────────────────────────────────────────────────────
    desc = str(info.get('desc', '')).strip()
    title_cols = 11 if has_deposit else 9
    _title(cur, f'▶  리스계약 {cid}  /  {desc}' if desc else f'▶  리스계약 {cid}', title_cols)
    cur += 1

    info_labels = ['개시일', '종료일', '기간(월)', '정기리스료', '지급주기', '지급시점', '할인율(연)']
    info_values = [
        info.get('start'), info.get('end'), info.get('n_months'),
        info.get('payment'), info.get('freq'), info.get('timing'), info.get('rate'),
    ]
    if has_deposit:
        info_labels += ['지급보증금', '보증금PV', '현가할인차금']
        info_values += [deposit, deposit_pv, deposit_discount]

    for i, (lbl, val) in enumerate(zip(info_labels, info_values), 1):
        lc = ws.cell(row=cur, column=i)
        lc.value = lbl; lc.font = Font(bold=True); lc.fill = _INFO_FILL
        lc.alignment = CTR; lc.border = _BORDER
        vc = ws.cell(row=cur + 1, column=i)
        vc.value = val; vc.alignment = CTR; vc.border = _BORDER
        if lbl == '할인율(연)':
            vc.number_format = RATE_FMT
        elif lbl in ('정기리스료', '지급보증금', '보증금PV', '현가할인차금'):
            vc.number_format = MONEY_FMT
        elif lbl in ('개시일', '종료일'):
            vc.number_format = 'YYYY-MM-DD'
    cur += 3   # label + value + blank

    # ── 2. 리스부채 연간 요약 ─────────────────────────────────────────────────
    L_HDRS = ['연도', '기초리스부채잔액', '이자비용', '현금지급액', '원금상환액',
              '기말리스부채잔액', '유동성리스부채', '비유동성리스부채']
    L_DATA = L_HDRS[1:]
    L_SUM  = {'이자비용', '현금지급액', '원금상환액'}

    _title(cur, '■  리스부채 상각 연간 요약', len(L_HDRS))
    cur += 1
    for ci, h in enumerate(L_HDRS, 1):
        _hdr(cur, ci, h)
    cur += 1

    for _, ar in annual.iterrows():
        c0 = _bc(cur, 1); c0.value = ar['연도']; c0.alignment = CTR
        for ci, col in enumerate(L_DATA, 2):
            _money(cur, ci, ar.get(col))
        cur += 1

    c0 = _bc(cur, 1); c0.value = '합  계'; c0.font = Font(bold=True)
    c0.fill = _SUM_FILL; c0.alignment = CTR
    for ci, col in enumerate(L_DATA, 2):
        cell = _bc(cur, ci); cell.fill = _SUM_FILL; cell.alignment = RGT
        if col in L_SUM:
            cell.value = round(annual[col].sum()); cell.number_format = MONEY_FMT
    cur += 2

    # ── 3. 사용권자산 연간 요약 ───────────────────────────────────────────────
    # 취득원가(고정) + 상각누계액 계산
    rou_init = float(annual.iloc[0]['기초사용권자산']) if not annual.empty else 0.0
    ar_rou   = annual.copy()
    ar_rou['취득원가']  = rou_init
    ar_rou['상각누계액'] = rou_init - ar_rou['기말사용권자산']

    R_HDRS = ['연도', '취득원가', '감가상각비(해당연도)', '상각누계액', '기말잔액']
    R_DATA = ['취득원가', '감가상각비', '상각누계액', '기말사용권자산']

    _title(cur, '■  사용권자산 감가상각 연간 요약', len(L_HDRS))
    cur += 1
    for ci, h in enumerate(R_HDRS, 1):
        _hdr(cur, ci, h)
    cur += 1

    for _, ar in ar_rou.iterrows():
        c0 = _bc(cur, 1); c0.value = ar['연도']; c0.alignment = CTR
        for ci, col in enumerate(R_DATA, 2):
            _money(cur, ci, ar.get(col))
        cur += 1

    # 합계 행: 감가상각비(해당연도)만 합산, 나머지는 최종연도 값
    c0 = _bc(cur, 1); c0.value = '합  계'; c0.font = Font(bold=True)
    c0.fill = _SUM_FILL; c0.alignment = CTR
    for ci, col in enumerate(R_DATA, 2):
        cell = _bc(cur, ci); cell.fill = _SUM_FILL; cell.alignment = RGT
        if col == '감가상각비':
            cell.value = round(ar_rou['감가상각비'].sum()); cell.number_format = MONEY_FMT
        elif col in ('상각누계액', '기말사용권자산'):
            cell.value = round(ar_rou.iloc[-1][col]); cell.number_format = MONEY_FMT
    cur += 2

    # ── 4. 월별 상세 스케줄 ───────────────────────────────────────────────────
    M_COLS = ['연월', '기초부채잔액', '이자비용', '현금지급액', '원금상환액',
              '기말부채잔액', '기초자산잔액', '감가상각비', '기말자산잔액',
              '유동성대체대상액', '비유동성리스부채잔액']
    DEP_COLS = ['보증금_기초PV', '보증금_이자수익', '보증금_기말PV']

    all_cols    = M_COLS + (DEP_COLS if has_deposit else [])
    n_title_col = len(all_cols)

    _title(cur, '■  월별 상각 스케줄 (상세)', n_title_col)
    cur += 1
    mhdr_row = cur

    # 기본 리스 컬럼 헤더
    for ci, h in enumerate(M_COLS, 1):
        _hdr(cur, ci, h)

    # 보증금 컬럼 헤더 (보라색 구분)
    if has_deposit:
        dep_hdr_fill = PatternFill('solid', fgColor='7E57C2')
        dep_hdr_font = Font(bold=True, color='FFFFFF', size=10)
        for ci, h in enumerate(DEP_COLS, len(M_COLS) + 1):
            cell = _bc(cur, ci)
            cell.value = h
            cell.fill  = dep_hdr_fill
            cell.font  = dep_hdr_font
            cell.alignment = CTR
    cur += 1

    for _, mr in sched.iterrows():
        ym     = str(mr.get('연월', ''))
        is_dec = ym.endswith('-12')
        for ci, col in enumerate(all_cols, 1):
            cell = _bc(cur, ci)
            val  = mr.get(col)
            is_dep_col = col in DEP_COLS
            if col == '연월':
                cell.value = ym; cell.alignment = CTR
            else:
                if val is not None and not (isinstance(val, float) and val != val):
                    cell.value = round(float(val))
                    cell.number_format = MONEY_FMT
                cell.alignment = RGT
            if is_dec:
                cell.fill = _DEP_FILL if is_dep_col else _DEC_FILL
                if col in ('유동성대체대상액', '비유동성리스부채잔액') and cell.value:
                    cell.fill = _CUR_FILL
                    cell.font = Font(bold=True)
            elif is_dep_col:
                # 보증금 열은 항상 연보라 배경으로 구분
                cell.fill = PatternFill('solid', fgColor='F3E5F5')
        cur += 1

    # ── 열 너비 ───────────────────────────────────────────────────────────────
    base_widths = [10, 16, 14, 14, 14, 16, 14, 14, 16, 18, 20]
    dep_widths  = [16, 16, 16]
    col_widths  = base_widths + (dep_widths if has_deposit else [])
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = f'B{mhdr_row + 1}'


def save_results(summary_df: pd.DataFrame, contracts: list, output_path: str):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # ── Sheet 1: 계약별 요약 ─────────────────────────────────────────────
        summary_df.to_excel(writer, sheet_name='계약별 요약', index=False)
        ws1 = writer.sheets['계약별 요약']
        _set_header(ws1)
        _money_cols(ws1, summary_df, exclude=('리스계약번호', '리스개시일', '리스종료일',
                                               '리스기간(월)', '지급주기', '지급시점', '할인율(연)'))
        rate_ci = summary_df.columns.get_loc('할인율(연)') + 1 if '할인율(연)' in summary_df.columns else None
        if rate_ci:
            for cell in ws1.iter_rows(min_row=2, max_row=ws1.max_row,
                                       min_col=rate_ci, max_col=rate_ci):
                for c in cell:
                    c.number_format = RATE_FMT
        s_widths = {
            '리스계약번호': 18, '리스개시일': 12, '리스종료일': 12,
            '리스기간(월)': 10, '정기리스료': 14, '지급주기': 8, '지급시점': 8, '할인율(연)': 10,
            '최초 리스부채': 16, '사용권자산(최초)': 16,
            '지급보증금': 14, '보증금PV': 14, '보증금할인차금': 14,
        }
        _col_width(ws1, s_widths, default=18)

        # ── 계약별 개별 시트 ──────────────────────────────────────────────────
        for c in contracts:
            desc  = str(c['info'].get('desc', '')).strip()
            raw   = f"{c['cid']}_{desc}" if desc else f"계약_{c['cid']}"
            sname = raw[:31]
            ws    = writer.book.create_sheet(title=sname)
            _write_contract_sheet(ws, c['cid'], c['info'], c['annual'], c['sched'])

    print(f'  → {os.path.basename(output_path)} 저장 완료')


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    pattern = os.path.join(INPUT_DIR, 'lease_*_information_*.xlsx')
    files   = sorted(f for f in glob.glob(pattern)
                     if not os.path.basename(f).startswith('~$'))

    if not files:
        print(f'[안내] 처리할 파일 없음 — 경로: {INPUT_DIR}')
        print(f'       파일명 형식: lease_{{회사명}}_information_{{연도}}.xlsx')
        return

    for fpath in files:
        fname = os.path.basename(fpath)
        # 연도: 4자리 숫자(2025) 또는 fy##(fy25 → 2025) 모두 허용
        m = re.match(r'^lease_(.+)_information_(\d{4}|fy\d{2})\.xlsx$', fname, re.IGNORECASE)
        if not m:
            print(f'[건너뜀] 파일명 형식 불일치: {fname}')
            continue

        company  = m.group(1)
        raw_year = m.group(2)
        year     = raw_year if raw_year.isdigit() else f'20{raw_year[2:]}'  # fy25 → 2025
        print(f'\n{"="*60}')
        print(f'  처리: {fname}  (회사: {company}, 연도: {year})')
        print(f'{"="*60}')

        try:
            df_in = pd.read_excel(fpath, header=1)
        except Exception as e:
            print(f'  [오류] 파일 읽기 실패: {e}')
            continue

        df_in = df_in.dropna(subset=['리스계약번호'])
        df_in = df_in[df_in['리스계약번호'].astype(str).str.strip() != '']
        print(f'  계약 수: {len(df_in)}건')

        contracts    = []
        summary_rows = []

        for _, row in df_in.iterrows():
            cid   = str(row['리스계약번호']).strip()
            row_d = row.to_dict()
            print(f'  ▷ {cid}', end='  ')
            try:
                sched, init_liab, init_rou, deposit, deposit_pv, deposit_discount = build_schedule(row_d)
            except Exception as e:
                print(f'[오류] {e}')
                continue

            annual = _annual_summary(sched)

            # 당해연도 집계 (summary_df 컬럼용)
            yr_rows   = sched[sched['연월'].str.startswith(year)]
            dec_row   = sched[sched['연월'] == f'{year}-12']
            has_dec   = not dec_row.empty

            yr_int    = yr_rows['이자비용'].sum()
            yr_dep    = yr_rows['감가상각비'].sum()
            yr_cash   = yr_rows['현금지급액'].sum()
            yr_end_lb = dec_row['기말부채잔액'].values[0]     if has_dec else None
            cur_p     = dec_row['유동성대체대상액'].values[0]    if has_dec else None
            nc_p      = dec_row['비유동성리스부채잔액'].values[0] if has_dec else None

            if has_dec:
                yr_end_rou = dec_row['기말자산잔액'].values[0]
            elif not yr_rows.empty:
                yr_end_rou = yr_rows.iloc[-1]['기말자산잔액']
            else:
                yr_end_rou = 0.0

            annual_r = _safe_float(_col(row_d, '적용할인율(연 이자율)', '적용할인율', '이자율'))
            if annual_r > 1:
                annual_r /= 100

            # 계약 정보 dict (계약별 시트 상단 표시용)
            info = {
                'desc':              str(_col(row_d, '상세정보(호수/차량번호)', '상세정보', '상세') or '').strip(),
                'start':             pd.Timestamp(row['리스개시일']).date(),
                'end':               pd.Timestamp(row['리스종료일']).date(),
                'n_months':          int(_safe_float(_col(row_d, '최종 산정리스기간(월수)'))),
                'payment':           _safe_float(_col(row_d, '정기리스료')),
                'freq':              str(_col(row_d, '지급주기(월/분기/년)', '지급주기(월/분기/반기/년)', '지급주기') or '').strip(),
                'timing':            str(_col(row_d, '지급시점(기초/기말)', '지급시점') or '').strip(),
                'rate':              annual_r,
                'deposit':           deposit,
                'deposit_pv':        deposit_pv,
                'deposit_discount':  deposit_discount,
            }
            contracts.append({'cid': cid, 'info': info, 'sched': sched, 'annual': annual})

            dep_msg = f' / 보증금PV {deposit_pv:>12,.0f}원' if deposit > 0 else ''
            print(f'초기부채 {init_liab:>14,.0f}원 / 사용권자산 {init_rou:>14,.0f}원{dep_msg}')

            summary_row = {
                '리스계약번호':          cid,
                '리스개시일':           pd.Timestamp(row['리스개시일']).date(),
                '리스종료일':           pd.Timestamp(row['리스종료일']).date(),
                '리스기간(월)':         int(_safe_float(_col(row_d, '최종 산정리스기간(월수)'))),
                '정기리스료':           _safe_float(_col(row_d, '정기리스료')),
                '지급주기':             info['freq'],
                '지급시점':             info['timing'],
                '할인율(연)':           annual_r,
                '최초 리스부채':        round(init_liab),
                '사용권자산(최초)':     round(init_rou),
                f'{year}년 이자비용':   round(yr_int),
                f'{year}년 감가상각비': round(yr_dep),
                f'{year}년 리스료지급': round(yr_cash),
                f'{year}년말 리스부채': round(yr_end_lb) if yr_end_lb is not None else 0,
                '유동성대체대상액':      round(cur_p)    if cur_p    is not None else 0,
                '비유동성리스부채잔액':  round(nc_p)     if nc_p     is not None else 0,
                f'{year}년말 사용권자산': round(yr_end_rou) if yr_end_rou is not None else 0,
                '사용권자산 상각누계':   round(init_rou - (yr_end_rou if yr_end_rou is not None else 0)),
            }
            if deposit > 0:
                summary_row['지급보증금']    = round(deposit)
                summary_row['보증금PV']     = round(deposit_pv)
                summary_row['보증금할인차금'] = round(deposit_discount)
            summary_rows.append(summary_row)

        if not contracts:
            print('  [안내] 처리된 계약 없음')
            continue

        summary_df  = pd.DataFrame(summary_rows)
        output_path = os.path.join(OUTPUT_DIR, f'lease_schedule_{company}_{year}.xlsx')
        save_results(summary_df, contracts, output_path)

        # 회사별 results/ 폴더에도 복사
        company_results = os.path.join(PROJECT_DIR, company, 'results')
        if os.path.isdir(company_results):
            copy_path = os.path.join(company_results, f'lease_schedule_{company}_{year}.xlsx')
            shutil.copy2(output_path, copy_path)
            print(f'  → results 복사: {os.path.relpath(copy_path, PROJECT_DIR)}')
        else:
            print(f'  [안내] {company}/results/ 폴더 없음 — results 복사 생략')

        total_months = sum(len(c['sched']) for c in contracts)
        print(f'\n  완료 — 계약 {len(contracts)}건 / 월별행 {total_months:,}행 / 시트 {len(contracts)+1}개')
        print('=' * 60)


if __name__ == '__main__':
    main()
