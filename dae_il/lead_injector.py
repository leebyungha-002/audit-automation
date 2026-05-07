# -*- coding: utf-8 -*-
import sys
import os
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QGroupBox, QPushButton, QLabel, QLineEdit,
    QComboBox, QTextEdit, QFileDialog, QMessageBox, QSizePolicy
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


class LeadInjectorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.source_path = ''
        self.target_path = ''
        self.source_df = None
        self._init_ui()

    def _init_ui(self):
        self.setWindowTitle('리드스케줄 데이터 주입기')
        self.setMinimumWidth(680)
        self.resize(720, 560)

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setSpacing(10)
        root.setContentsMargins(14, 14, 14, 14)

        # ── 1. 파일 선택 ──────────────────────────────────────
        file_group = QGroupBox('① 파일 선택')
        fg = QGridLayout(file_group)
        fg.setColumnStretch(1, 1)

        fg.addWidget(QLabel('소스 (Lead_Summary xlsx):'), 0, 0)
        self.source_edit = QLineEdit()
        self.source_edit.setReadOnly(True)
        self.source_edit.setPlaceholderText('집계된 리드스케줄 파일 선택…')
        fg.addWidget(self.source_edit, 0, 1)
        btn_src = QPushButton('파일 선택')
        btn_src.setFixedWidth(80)
        btn_src.clicked.connect(self._select_source)
        fg.addWidget(btn_src, 0, 2)

        fg.addWidget(QLabel('대상 조서 파일:'), 1, 0)
        self.target_edit = QLineEdit()
        self.target_edit.setReadOnly(True)
        self.target_edit.setPlaceholderText('숫자를 주입할 당기 조서 파일 선택…')
        fg.addWidget(self.target_edit, 1, 1)
        btn_tgt = QPushButton('파일 선택')
        btn_tgt.setFixedWidth(80)
        btn_tgt.clicked.connect(self._select_target)
        fg.addWidget(btn_tgt, 1, 2)

        root.addWidget(file_group)

        # ── 2. 주입 설정 ──────────────────────────────────────
        map_group = QGroupBox('② 주입 설정 (대상 조서 기준)')
        mg = QGridLayout(map_group)
        mg.setColumnStretch(1, 1)
        mg.setColumnStretch(3, 1)

        mg.addWidget(QLabel('대상 시트:'), 0, 0)
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(160)
        mg.addWidget(self.sheet_combo, 0, 1)

        mg.addWidget(QLabel('계정명 컬럼:'), 0, 2, Qt.AlignmentFlag.AlignRight)
        self.col_key = self._col_edit('A')
        mg.addWidget(self.col_key, 0, 3)

        mg.addWidget(QLabel('기초잔액 컬럼:'), 1, 0)
        self.col_start = self._col_edit('C')
        mg.addWidget(self.col_start, 1, 1)

        mg.addWidget(QLabel('차변합계 컬럼:'), 1, 2, Qt.AlignmentFlag.AlignRight)
        self.col_debit = self._col_edit('D')
        mg.addWidget(self.col_debit, 1, 3)

        mg.addWidget(QLabel('대변합계 컬럼:'), 2, 0)
        self.col_credit = self._col_edit('E')
        mg.addWidget(self.col_credit, 2, 1)

        mg.addWidget(QLabel('기말잔액(순액) 컬럼:'), 2, 2, Qt.AlignmentFlag.AlignRight)
        self.col_end = self._col_edit('F')
        mg.addWidget(self.col_end, 2, 3)

        root.addWidget(map_group)

        # ── 3. 실행 버튼 ──────────────────────────────────────
        run_btn = QPushButton('▶  주입 실행')
        run_btn.setMinimumHeight(42)
        run_font = QFont()
        run_font.setPointSize(11)
        run_font.setBold(True)
        run_btn.setFont(run_font)
        run_btn.clicked.connect(self._run_injection)
        root.addWidget(run_btn)

        # ── 4. 로그 ───────────────────────────────────────────
        log_group = QGroupBox('실행 로그')
        lg = QVBoxLayout(log_group)
        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setFont(QFont('Consolas', 9))
        lg.addWidget(self.log_box)
        root.addWidget(log_group)

    # ── 헬퍼 ──────────────────────────────────────────────────
    def _col_edit(self, default: str) -> QLineEdit:
        e = QLineEdit(default)
        e.setMaximumWidth(48)
        e.setAlignment(Qt.AlignmentFlag.AlignCenter)
        return e

    def _log(self, msg: str):
        self.log_box.append(msg)

    # ── 파일 선택 ─────────────────────────────────────────────
    def _select_source(self):
        path, _ = QFileDialog.getOpenFileName(
            self, '소스 파일 선택', '', 'Excel 파일 (*.xlsx *.xlsm)'
        )
        if not path:
            return
        self.source_path = path
        self.source_edit.setText(path)
        try:
            self.source_df = pd.read_excel(path, sheet_name='Lead_Summary')
            self._log(f'[소스] {os.path.basename(path)}  →  {len(self.source_df)}개 행 로드')
            self._log(f'       컬럼: {list(self.source_df.columns)}')
        except Exception as e:
            self._log(f'[오류] 소스 파일 로드 실패: {e}')
            self.source_df = None

    def _select_target(self):
        path, _ = QFileDialog.getOpenFileName(
            self, '대상 조서 파일 선택', '', 'Excel 파일 (*.xlsx *.xlsm)'
        )
        if not path:
            return
        self.target_path = path
        self.target_edit.setText(path)
        try:
            wb = load_workbook(path, read_only=True, data_only=False)
            sheets = wb.sheetnames
            wb.close()
            self.sheet_combo.clear()
            self.sheet_combo.addItems(sheets)
            self._log(f'[대상] {os.path.basename(path)}  →  시트 {len(sheets)}개: {sheets}')
        except Exception as e:
            self._log(f'[오류] 대상 파일 로드 실패: {e}')

    # ── 주입 실행 ─────────────────────────────────────────────
    def _run_injection(self):
        # 유효성 검사
        if self.source_df is None:
            QMessageBox.warning(self, '경고', '소스 파일을 먼저 선택하세요.')
            return
        if not self.target_path:
            QMessageBox.warning(self, '경고', '대상 조서 파일을 먼저 선택하세요.')
            return

        sheet_name = self.sheet_combo.currentText()
        try:
            key_idx   = column_index_from_string(self.col_key.text().strip().upper())
            start_idx = column_index_from_string(self.col_start.text().strip().upper())
            debit_idx = column_index_from_string(self.col_debit.text().strip().upper())
            credit_idx= column_index_from_string(self.col_credit.text().strip().upper())
            end_idx   = column_index_from_string(self.col_end.text().strip().upper())
        except Exception:
            QMessageBox.warning(self, '경고', '컬럼 문자를 올바르게 입력하세요 (예: A, B, AA).')
            return

        self._log('\n── 주입 시작 ──────────────────────────────')
        try:
            # 수식·서식 유지: data_only=False, keep_vba=False
            wb = load_workbook(self.target_path, data_only=False)
            if sheet_name not in wb.sheetnames:
                QMessageBox.critical(self, '오류', f'시트 "{sheet_name}"를 찾을 수 없습니다.')
                wb.close()
                return
            ws = wb[sheet_name]

            # 대상 시트에서 계정명 → 행번호 룩업 빌드
            lookup: dict[str, int] = {}
            for row_cells in ws.iter_rows(min_col=key_idx, max_col=key_idx):
                cell = row_cells[0]
                if cell.value is not None:
                    lookup[str(cell.value).strip()] = cell.row

            # 값 주입 (value only — 수식 덮어쓰지 않음)
            hit, miss = 0, []
            for _, src_row in self.source_df.iterrows():
                acct = str(src_row.get('공시계정명', '')).strip()
                if acct in lookup:
                    r = lookup[acct]
                    ws.cell(row=r, column=start_idx ).value = _to_num(src_row.get('기초잔액'))
                    ws.cell(row=r, column=debit_idx ).value = _to_num(src_row.get('차변합계'))
                    ws.cell(row=r, column=credit_idx).value = _to_num(src_row.get('대변합계'))
                    ws.cell(row=r, column=end_idx   ).value = _to_num(src_row.get('기말잔액(순액)'))
                    hit += 1
                else:
                    miss.append(acct)

            # 원본이름_Updated.xlsx 저장
            base, ext = os.path.splitext(self.target_path)
            out_path = base + '_Updated' + ext
            wb.save(out_path)
            wb.close()

            self._log(f'[결과] 매핑 성공: {hit}개')
            if miss:
                self._log(f'[결과] 미매핑 {len(miss)}개: {", ".join(miss[:8])}{"…" if len(miss) > 8 else ""}')
            self._log(f'[저장] {out_path}')
            self._log('── 완료 ────────────────────────────────────')

            QMessageBox.information(
                self, '주입 완료',
                f'✔ {hit}개 계정 주입 완료\n\n저장 경로:\n{out_path}'
            )

        except Exception as e:
            self._log(f'[오류] {e}')
            QMessageBox.critical(self, '오류', str(e))


def _to_num(val):
    """NaN / None → None, 그 외 → float"""
    try:
        import math
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return None
        return float(val)
    except Exception:
        return None


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    win = LeadInjectorApp()
    win.show()
    sys.exit(app.exec())
