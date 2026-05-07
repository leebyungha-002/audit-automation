# -*- coding: utf-8 -*-
import pandas as pd
import os
from openpyxl.utils import get_column_letter

script_dir = os.path.dirname(os.path.abspath(__file__))

coa_path = os.path.join(script_dir, 'coa_dae_il_25.xlsx')
bal_path = os.path.join(script_dir, 'results', 'dae_il_20260504_일반사항분석.xlsx')
out_path = os.path.join(script_dir, 'test_lead_summary.xlsx')
log_path = os.path.join(script_dir, 'test_lead_summary_log.txt')

NUM_FMT = '#,##0'  # 천 단위 콤마 형식

log_lines = []
def log(msg=''):
    log_lines.append(str(msg))
    print(msg)

# ── 1. 데이터 로드 ──────────────────────────────────────────
df_coa = pd.read_excel(coa_path)
df_bal = pd.read_excel(bal_path)

log(f'COA 행 수: {len(df_coa)},  잔액 행 수: {len(df_bal)}')

# ── 2. 계정코드 문자열 통일 ────────────────────────────────
df_coa['계정코드'] = df_coa['계정코드'].astype(str)
df_bal['계정코드'] = df_bal['계정코드'].astype(str)

# ── 3. Left Join 병합 (공시계정명 포함) ───────────────────
merge_cols = ['계정코드', '대분류 (Level 1)', '리드스케줄 (Level 4)', '공시계정명', '계정성격']
df = pd.merge(df_bal, df_coa[merge_cols], on='계정코드', how='left')

missing = df['대분류 (Level 1)'].isna().sum()
if missing > 0:
    log(f'[경고] COA에 매핑 안 된 계정: {missing}개')
    log(df[df['대분류 (Level 1)'].isna()][['계정코드', '계정과목']].to_string())

# ── 4. 차감 계정 순액 처리 ─────────────────────────────────
df['순기말잔액'] = df['기말잔액'].copy()
df['순기초잔액'] = df['기초잔액'].copy()
df['순차변합계'] = df['차변 합계'].copy()
df['순대변합계'] = df['대변 합계'].copy()

mask_contra = df['계정성격'].str.contains('차감', na=False)
df.loc[mask_contra, '순기말잔액'] = df.loc[mask_contra, '기말잔액']  * -1
df.loc[mask_contra, '순기초잔액'] = df.loc[mask_contra, '기초잔액']  * -1
df.loc[mask_contra, '순차변합계'] = df.loc[mask_contra, '차변 합계'] * -1
df.loc[mask_contra, '순대변합계'] = df.loc[mask_contra, '대변 합계'] * -1

log(f'\n차감 계정 적용: {mask_contra.sum()}개')

# ── 5-A. 요약 데이터: 대분류 × 공시계정명 ────────────────
grp_summary = df.groupby(
    ['대분류 (Level 1)', '공시계정명'],
    dropna=False
).agg(
    기초잔액=('순기초잔액', 'sum'),
    차변합계=('순차변합계', 'sum'),
    대변합계=('순대변합계', 'sum'),
    기말잔액_순액=('순기말잔액', 'sum')
).reset_index()

grp_summary.columns = ['대분류', '공시계정명', '기초잔액', '차변합계', '대변합계', '기말잔액(순액)']

# ── 5-B. 상세 데이터: 공시계정명 × 계정과목 (공시계정명 좌측 배치) ──
grp_detail = df.groupby(
    ['공시계정명', '계정과목'],
    dropna=False
).agg(
    기초잔액=('순기초잔액', 'sum'),
    차변합계=('순차변합계', 'sum'),
    대변합계=('순대변합계', 'sum'),
    기말잔액_순액=('순기말잔액', 'sum')
).reset_index()

grp_detail.columns = ['공시계정명', '계정과목', '기초잔액', '차변합계', '대변합계', '기말잔액(순액)']

# ── 6. 콘솔 출력 ──────────────────────────────────────────
pd.set_option('display.float_format', '{:,.0f}'.format)
pd.set_option('display.max_rows', 200)
pd.set_option('display.width', 200)

log('\n======== [요약] 대분류 × 공시계정명 ========')
log(grp_summary.to_string(index=False))
log('\n======== [상세] 공시계정명 × 계정과목 ========')
log(grp_detail.to_string(index=False))

# ── 7. openpyxl 포맷 적용 헬퍼 ───────────────────────────
def apply_sheet_format(ws, df, num_cols):
    """헤더 자동 필터 + 숫자 컬럼 천 단위 콤마 형식 적용"""
    last_col = get_column_letter(ws.max_column)
    last_row = ws.max_row

    # 헤더 자동 필터
    ws.auto_filter.ref = f'A1:{last_col}1'

    # 숫자 컬럼 인덱스 (1-based) 계산
    header = [cell.value for cell in ws[1]]
    for col_name in num_cols:
        if col_name in header:
            col_idx = header.index(col_name) + 1
            for row in range(2, last_row + 1):
                ws.cell(row=row, column=col_idx).number_format = NUM_FMT

# ── 8. 엑셀 저장: 시트 분리 (Lead_Summary / Lead_Detail) ─
NUM_COLS = ['기초잔액', '차변합계', '대변합계', '기말잔액(순액)']

def write_excel(path):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        grp_summary.to_excel(writer, sheet_name='Lead_Summary', index=False)
        grp_detail.to_excel(writer, sheet_name='Lead_Detail',   index=False)

        apply_sheet_format(writer.sheets['Lead_Summary'], grp_summary, NUM_COLS)
        apply_sheet_format(writer.sheets['Lead_Detail'],  grp_detail,  NUM_COLS)

try:
    write_excel(out_path)
    saved_path = out_path
except PermissionError:
    from datetime import datetime
    fallback = out_path.replace('.xlsx', f'_{datetime.now().strftime("%H%M%S")}.xlsx')
    write_excel(fallback)
    saved_path = fallback
    log(f'[경고] 원본 파일이 열려 있어 대체 파일로 저장: {fallback}')

log(f'\n저장 완료: {saved_path}')
log(f'  - Lead_Summary: {len(grp_summary)}행  (대분류 × 공시계정명)')
log(f'  - Lead_Detail : {len(grp_detail)}행  (공시계정명 × 계정과목)')

# ── 9. 로그 파일 저장 ─────────────────────────────────────
with open(log_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(log_lines))
