"""연월차충당부채 검증앱(잔여연차일수 × 1일통상임금 방식) — 엔진.

input_data/leave_<company>_information_fy<year>.xlsx 를 읽어
결산기준일 현재 재직 중인 임직원별 연차충당부채를 재계산하고,
전기말/당기말 회사계상 충당부채와 대사하는 output/leave_schedule_<company>_<fy>.xlsx 를 생성한다.

핵심 설계: severance_analyzer(퇴직급여충당부채)와 달리 연차는 이월(carryover)이 본질인 값이므로
depreciation_analyzer의 "기초잔액은 입력값 신뢰, 당기분만 계산" 원칙을 그대로 적용한다.
'당기정보'/'전기정보' 두 시트 각각이 자기 시점의 기초 이월연차잔여일수를 독립적으로 입력받고,
그 위에 당기(그 시트가 나타내는 회계연도) 부여일수만 이 앱이 계산해 얹는다.
  당기말 잔여연차일수(계산) = 기초 이월연차잔여일수(입력) + 당기부여일수(계산) − 당기연차사용일수(입력)
  당기말 연차충당부채(계산) = 당기말 잔여연차일수(계산) × 1일 통상임금(입력) × 연차사용촉진 반영 지급률(입력)

연차사용촉진제도(근로기준법 제61조) 반영: 촉진 절차를 적법하게 이행하면 미사용 연차에 대한 금전
지급의무가 면제된다. 인원별 절차 이행 여부를 확인하기 어려운 경우가 많아, '기준정보' 시트의
전사 공통 지급률(%)을 잔여연차 금액에 곱하는 방식으로 단순화했다(잔여'일수'는 그대로 두고 금액만
할인 — 일수 기준 감사 대사는 왜곡하지 않기 위함). 촉진 이행 여부·실제 사용률은 연도마다 다를 수
있어 당기/전기 지급률을 각각 별도로 입력받는다(둘 다 미입력 시 100%, 촉진 미적용 가정).

전기/당기 인원은 severance_analyzer와 동일하게 한 파일 안에 '당기정보'/'전기정보' 두 시트로 나눠
입력받는다. 당기 계산은 '당기정보' 시트만 사용하고, '전기정보'는 사번(없으면 성명) 기준으로
당기 시트와 매칭해 신규입사자/퇴사자 명단을 산출하고, 전기말 잔액(당기 발생액 계산용)을 독립적으로
재계산하는 데 쓰인다.

연차 부여일수 산식('기준정보' 시트의 '연차산정기준' 설정에 따라 근속연수 기산 시점만 달라짐,
근로기준법 제60조 기준):
  공통(근속연수 1년 이상): 부여일수 = min(15 + (근속연수−1)//2, 25)
    (3년 이상 근속 시 최초 1년을 초과하는 계속근로연수 매 2년마다 1일 가산, 25일 한도)
  공통(근속연수 1년 미만·입사연도): 입사 후 매 1개월 경과 시마다 1일씩 발생(최대 11일, 개근 가정)

  [회계기준] 전 직원이 회계연도 시작일(전기 결산기준일 다음날)에 일괄로 근속연수가 갱신되어
    그 시점 근속연수로 위 표를 적용해 한 번에 부여받는다. 입사연도(근속연수 0년차)에는 월단위
    발생 대신 비례연차를 적용한다: 비례연차일수 = floor(15 × 재직개월수 / 12)
    (재직개월수 = 입사월부터 회계연도 종료월까지 달력월 기준 개월수, 상한 12개월).
  [입사기준] 개인별 입사기념일마다 근속연수가 갱신된다. 당기 회계기간 중 도래하는 입사기념일마다
    그 시점 근속연수로 위 표를 적용해 개별 부여한다(정상적인 12개월 회계기간이면 보통 0~1개).
    입사연도(최초 기념일 이전)는 위 '근속연수 1년 미만' 월단위 발생 규정을 그대로 적용한다.

  ※ 근속연수 1년 미만 구간의 월단위 발생분(최대 11일)과 만1년 시점의 15일은 근로기준법상 원래
    별도로 발생하는 값이나, 이 앱은 두 구간을 겹치지 않게(근속연수 0년차는 월단위만, 1년차부터는
    연단위 표만) 계산한다 — 실제 회사 데이터로 대사하며 회사 규정과 다르면 조정한다.

실행 예:
    python leave_schedule.py kyungnam --fiscal-month 12
    python leave_schedule.py --file leave_kyungnam_information_fy2026.xlsx
"""
import argparse
import calendar
import glob
import os
from datetime import date, timedelta

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(HERE, "input_data")
OUTPUT_DIR = os.path.join(HERE, "output")

SIG_THRESHOLD_ABS = 1000      # 유의차이 절대금액 기준(원)
SIG_THRESHOLD_PCT = 0.01      # 유의차이 비율 기준(1%)

COST_TYPES = ["제조원가", "판관비"]
BASIS_KEYS = {
    "제조원가": {
        "전기": "전기말 회사계상 연차충당부채(제조원가분)",
        "당기": "당기말 회사계상 연차충당부채(제조원가분)",
    },
    "판관비": {
        "전기": "전기말 회사계상 연차충당부채(판관비분)",
        "당기": "당기말 회사계상 연차충당부채(판관비분)",
    },
}
DEBIT_BASIS_LABEL = "당기 연차충당부채 차변(당기지급액, 분개장 기준)"

BASIS_MODE_LABEL = "연차산정기준(입사기준/회계기준)"
BASIS_MODE_OPTIONS = ["입사기준", "회계기준"]
BASIS_MODE_DEFAULT = "입사기준"

# 연차사용촉진제도(근로기준법 제61조) 반영 — 촉진 절차를 적법하게 이행하면 미사용 연차에 대한
# 금전 지급의무가 면제된다. 인원별로 절차 이행 여부를 확인하기 어려운 경우가 많아, 이 앱은
# 전사 공통 지급률(%)을 잔여연차 금액에 곱하는 방식으로 단순화해 반영한다(미입력 시 100%, 즉 촉진 미적용 가정).
# 촉진 이행 여부·실제 사용률은 연도마다 다를 수 있어 당기/전기 지급률을 각각 별도로 입력받는다.
PAYOUT_RATE_LABEL_CURRENT = "당기 연차사용촉진 반영 지급률(%, 미입력시 100%)"
PAYOUT_RATE_LABEL_PRIOR = "전기 연차사용촉진 반영 지급률(%, 미입력시 100%)"

CURRENT_SHEET = "당기정보"
PRIOR_SHEET = "전기정보"
LEAVER_SHEET = "당기퇴사자"
PAYROLL_SHEET = "급여대장인원명부"

PAYROLL_COUNT_LABEL = "기말 급여대장상 총인원수(명부 미확보 시 참고용)"

DISPLAY_COLS = ["사업장", "부서", "사번", "성명", "직급", "원가구분", "입사일"]

# 요약표 맨 아래 '참고: 계산 공식' 섹션에 한 줄씩 쓰는 텍스트(compute_employee()와 동일 로직).
FORMULA_NOTE_LINES = [
    ("공통 (근로기준법 제60조)", True),
    ("근속연수 1년 이상: 부여일수 = min(15 + (근속연수−1)//2, 25)   (3년 이상부터 매 2년마다 1일 가산, 25일 한도)", False),
    ("근속연수 1년 미만(입사연도): 입사 후 매 1개월 경과 시마다 1일씩 발생(최대 11일, 개근 가정)", False),
    ("당기말 잔여연차일수(계산) = 기초 이월연차잔여일수(입력) + 당기부여일수(계산) − 당기연차사용일수(입력)", False),
    ("당기말 연차충당부채(계산) = 당기말 잔여연차일수(계산) × 1일 통상임금(입력) × 당기 지급률(입력, 미입력시 100%)", False),
    ("전기말 연차충당부채(계산)도 동일 산식이되 '전기 지급률'을 곱한다(당기/전기 지급률은 서로 다를 수 있어 별도 입력)", False),
    ("※ 연차사용촉진(근로기준법 제61조) 절차를 적법 이행하면 미사용 연차의 금전 지급의무가 면제되므로, "
     "'기준정보' 시트의 지급률(%)로 잔여일수는 그대로 두고 충당부채 금액만 낮춘다(전사 공통 비율로 단순화)", False),
    ("", False),
    ("[회계기준] 근속연수 기산 시점 = 당기 회계연도 시작일(전기 결산기준일 + 1일), 전 직원 동일 시점에 일괄 부여", True),
    ("입사연도(근속연수 0년차) 비례연차 = floor(15 × 재직개월수 ÷ 12)   "
     "(재직개월수 = 입사월~회계연도 종료월, 달력월 기준 상한 12개월)", False),
    ("", False),
    ("[입사기준] 근속연수 기산 시점 = 개인별 입사기념일, 당기 회계기간 중 도래하는 기념일마다 개별 부여", True),
    ("입사연도(최초 기념일 전)는 위 '근속연수 1년 미만' 월단위 발생 규정 적용", False),
]


# ── 공용 헬퍼 ────────────────────────────────────────────────────────────────

def _safe_float(v, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        if pd.isna(v):
            return default
    except (TypeError, ValueError):
        pass
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _safe_date(v):
    if v is None or v == "":
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    ts = pd.Timestamp(v)
    return ts.date() if not pd.isna(ts) else None


def _cost_type(emp: dict) -> str:
    return str(emp.get("원가구분(제조원가/판관비)") or "").strip() or "(미분류)"


def _employee_key(emp: dict):
    """전기/당기 인원 매칭 키. 사번이 있으면 사번, 없으면 성명으로 매칭한다."""
    사번 = emp.get("사번")
    if 사번 not in (None, ""):
        return f"ID::{str(사번).strip()}"
    성명 = emp.get("성명")
    if 성명 not in (None, ""):
        return f"NAME::{str(성명).strip()}"
    return None


def _fy_bounds(target_fy: str, fiscal_month: int) -> tuple:
    """대상 회계연도(target_fy)의 시작월/종료월(YYYY-MM 문자열) 반환.
    depreciation_schedule.py/severance_schedule.py 와 동일 규칙."""
    fy = int(target_fy)
    if fiscal_month == 12:
        return f"{fy}-01", f"{fy}-12"
    return f"{fy - 1}-{fiscal_month + 1:02d}", f"{fy}-{fiscal_month:02d}"


def _apply_interim(fy_end: str, target_fy: str, interim_month: int = None) -> str:
    if not interim_month:
        return fy_end
    interim_ym = f"{target_fy}-{interim_month:02d}"
    return min(fy_end, interim_ym)


def _ym_to_end_date(ym: str) -> date:
    y, m = int(ym[:4]), int(ym[5:7])
    last_day = calendar.monthrange(y, m)[1]
    return date(y, m, last_day)


def _find_input_file(company: str = None, file: str = None) -> str:
    if file:
        path = file if os.path.isabs(file) else os.path.join(INPUT_DIR, file)
        if not os.path.exists(path):
            raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {path}")
        return path

    if company:
        pattern = os.path.join(INPUT_DIR, f"leave_{company}_information_fy*.xlsx")
    else:
        pattern = os.path.join(INPUT_DIR, "leave_*_information_fy*.xlsx")

    matches = [p for p in glob.glob(pattern) if "template" not in os.path.basename(p)]
    if not matches:
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {pattern}")
    if len(matches) > 1 and not company:
        raise ValueError(f"회사를 특정해주세요. 후보 파일 여러 개: {matches}")
    return matches[0]


def _is_significant(diff, base) -> bool:
    if diff is None or pd.isna(diff):
        return False
    if abs(diff) >= SIG_THRESHOLD_ABS and (base in (None, 0) or abs(diff) >= abs(base) * SIG_THRESHOLD_PCT):
        return True
    return False


def _add_months(d: date, n: int) -> date:
    """d로부터 n개월 뒤 날짜. 대상 월에 그 일자가 없으면(예: 1/31 + 1개월) 그 달의 말일로 보정."""
    total = d.month - 1 + n
    y = d.year + total // 12
    m = total % 12 + 1
    last = calendar.monthrange(y, m)[1]
    return date(y, m, min(d.day, last))


def _add_years(d: date, n: int) -> date:
    """d로부터 n년 뒤 날짜(입사기념일). 2/29 등 대상 연도에 없는 날짜는 2/28로 보정."""
    try:
        return d.replace(year=d.year + n)
    except ValueError:
        return d.replace(year=d.year + n, day=28)


def _years_between(start: date, end: date) -> int:
    """start부터 end까지의 만년수(생일 계산과 동일한 정수 근속연수). end < start면 음수가 될 수 있다."""
    years = end.year - start.year
    if (end.month, end.day) < (start.month, start.day):
        years -= 1
    return years


def _months_between_inclusive(start: date, end: date) -> int:
    """start가 속한 달부터 end가 속한 달까지의 달력월 개월수(양끝 포함), 최대 12·최소 0."""
    if end < start:
        return 0
    months = (end.year - start.year) * 12 + (end.month - start.month) + 1
    return max(0, min(months, 12))


def _monthly_accrual_days(입사일: date, anchor_start: date, anchor_end: date) -> int:
    """근속연수 1년 미만 구간의 월단위 발생(근로기준법상 매 1개월 개근 시 1일, 최대 11일).
    입사 후 n개월째(n=1..11) 되는 날짜가 이번 회계기간 (anchor_start, anchor_end] 안에 있으면 카운트."""
    count = 0
    for n in range(1, 12):
        d = _add_months(입사일, n)
        if d > anchor_end:
            break
        if d > anchor_start:
            count += 1
    return count


def _entitlement_by_tenure(근속연수: int) -> int:
    """근로기준법 제60조: 근속연수(1년 이상) → 법정 부여일수(3년 이상부터 매 2년마다 1일 가산, 25일 한도)."""
    가산 = (근속연수 - 1) // 2
    return min(15 + 가산, 25)


def _anniversaries_in_range(입사일: date, start_excl: date, end_incl: date) -> list:
    """입사기념일 중 (start_excl, end_incl] 구간에 속하는 것들을 [(근속연수, 날짜), ...]로 반환."""
    if 입사일 is None or 입사일 > end_incl:
        return []
    results = []
    n = max(1, start_excl.year - 입사일.year)
    while True:
        d = _add_years(입사일, n)
        if d > end_incl:
            break
        if d > start_excl:
            results.append((n, d))
        n += 1
    return results


# ── 입력 로딩 ────────────────────────────────────────────────────────────────

def load_employees(path: str, sheet_name: str) -> list:
    """지정한 인원 시트(1~2행 헤더, 3행부터 데이터)를 읽어 dict 목록으로 반환.
    시트가 없으면 빈 목록을 반환한다(예: '전기정보' 미작성 시 인원변동 명단만 생략됨)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [c.value for c in ws[2]]
    employees = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        rec = dict(zip(headers, row))
        if not rec.get("사번") and not rec.get("성명"):
            continue
        employees.append(rec)
    return employees


def load_basis(path: str) -> dict:
    """'기준정보' 시트를 라벨(A열) 기준으로 읽어 {라벨: 원본값} dict 반환(문자/숫자 모두 그대로 보존).
    금액 항목은 사용하는 쪽에서 _basis_float()로 변환한다(연차산정기준 행은 문자열이라 그대로 둬야 함)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "기준정보" not in wb.sheetnames:
        return {}
    ws = wb["기준정보"]
    basis = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        label, value = row[0], row[1]
        if not isinstance(label, str):
            continue
        label = label.strip()
        if not label:
            continue
        basis[label] = value
    return basis


def _basis_float(basis: dict, label: str):
    v = basis.get(label)
    return None if v in (None, "") else _safe_float(v)


def leave_basis_mode(basis: dict) -> str:
    raw = str(basis.get(BASIS_MODE_LABEL) or "").strip()
    return raw if raw in BASIS_MODE_OPTIONS else BASIS_MODE_DEFAULT


def leave_payout_rate(basis: dict, label: str) -> float:
    """'당기/전기 연차사용촉진 반영 지급률(%)' — 0~100 사이 값을 0~1 배수로 변환. 미입력·범위 밖이면 100%(1.0)."""
    v = _basis_float(basis, label)
    if v is None:
        return 1.0
    if v < 0 or v > 100:
        return 1.0
    return v / 100.0


# ── 전기/당기 인원 매칭 (신규입사자/퇴사자 판정) ──────────────────────────────

def match_periods(당기_employees: list, 전기_employees: list) -> dict:
    """당기/전기 인원을 사번(없으면 성명) 기준으로 매칭해 신규입사자/퇴사자를 산출한다."""
    당기_by_key, 전기_by_key = {}, {}
    for e in 당기_employees:
        k = _employee_key(e)
        if k is not None:
            당기_by_key.setdefault(k, e)
    for e in 전기_employees:
        k = _employee_key(e)
        if k is not None:
            전기_by_key.setdefault(k, e)

    신규입사_keys = set(당기_by_key) - set(전기_by_key)
    퇴사_keys = set(전기_by_key) - set(당기_by_key)

    return {
        "신규입사_keys": 신규입사_keys,
        "신규입사자": [당기_by_key[k] for k in 신규입사_keys],
        "퇴사자": [전기_by_key[k] for k in 퇴사_keys],
        "전기_by_key": 전기_by_key,
        "당기_by_key": 당기_by_key,
        "전기인원수": len(전기_by_key),
        "당기인원수": len(당기_by_key),
    }


def compute_prior_balances(전기_by_key: dict, 전기_anchor_start: date, 전기결산일: date,
                            mode: str, payout_rate: float = 1.0) -> dict:
    """'전기정보' 인원 각각의 전기 결산기준일 시점 연차충당부채를 사번(없으면 성명) 키로 반환.
    당기 인원별 표에서 '당기말 - 전기말' 차이를 인별로 계산하는 데 쓰인다."""
    return {k: compute_employee(e, 전기_anchor_start, 전기결산일, mode, payout_rate)["당기말충당부채"]
            for k, e in 전기_by_key.items()}


def _to_display_df(records: list) -> pd.DataFrame:
    rows = []
    for e in records:
        rows.append({
            "사업장": e.get("사업장") or "",
            "부서": e.get("부서") or "",
            "사번": e.get("사번"),
            "성명": e.get("성명"),
            "직급": e.get("직급"),
            "원가구분": _cost_type(e),
            "입사일": e.get("입사일"),
        })
    if not rows:
        return pd.DataFrame(columns=DISPLAY_COLS)
    return pd.DataFrame(rows)[DISPLAY_COLS]


LEAVER_MATCH_COLS = ["사업장", "부서", "사번", "직급", "원가구분",
                      "전기정보 있으나 당기정보 없음", "실제 퇴사자", "비고"]


def _build_leaver_match_df(전기_by_key: dict, 당기_by_key: dict, 퇴사자_recs: list,
                            leaver_payments: list) -> pd.DataFrame:
    """'전기정보-당기정보 자동 비교로 산출된 퇴사자 명단'과 '당기퇴사자' 시트(사용자 입력) 두 명단을
    사번(없으면 성명) 기준으로 나란히 비교하는 통합 표. 양쪽에 모두 있으면 '이상없음',
    한쪽에만 있으면 그 원인을 비고에 표시한다."""
    퇴사_keys = {_employee_key(e) for e in 퇴사자_recs if _employee_key(e) is not None}

    당기퇴사자_by_key: dict = {}
    for rec in leaver_payments:
        if rec.get("실제지급액(원)") in (None, ""):
            continue
        k = _employee_key(rec)
        if k is not None:
            당기퇴사자_by_key.setdefault(k, []).append(rec)

    rows = []
    for key in 퇴사_keys | set(당기퇴사자_by_key.keys()):
        전기레코드 = 전기_by_key.get(key)
        당기퇴사자_recs = 당기퇴사자_by_key.get(key, [])
        in_auto = key in 퇴사_keys
        in_actual = len(당기퇴사자_recs) > 0
        source = 전기레코드 or 당기_by_key.get(key)

        tags = []
        if in_auto and in_actual:
            if key in 당기_by_key:
                tags.append("⚠ '당기정보'에도 그대로 존재함 — 실제 퇴사 여부 확인 필요")
            if len(당기퇴사자_recs) > 1:
                tags.append("⚠ 이중기입의심(동일 인원으로 보이는 항목이 '당기퇴사자' 시트에 여러 번 입력됨 — 동명이인 여부 확인 필요)")
            if not tags:
                tags.append("이상없음")
        elif in_auto and not in_actual:
            tags.append("⚠ '당기퇴사자' 시트에 지급액 입력 없음")
        else:  # in_actual and not in_auto
            if 전기레코드 is not None:
                tags.append("⚠ '당기정보'에도 그대로 존재함 — 실제 퇴사 여부 확인 필요")
            else:
                tags.append("⚠ '전기정보'에서 매칭되는 인원을 찾지 못함(확인 필요)")
            if len(당기퇴사자_recs) > 1:
                tags.append("⚠ 이중기입의심(동일 인원으로 보이는 항목이 '당기퇴사자' 시트에 여러 번 입력됨 — 동명이인 여부 확인 필요)")

        성명_auto = ((전기레코드 or {}).get("성명") or (전기레코드 or {}).get("사번") or "") if in_auto else ""
        성명_actual = ((당기퇴사자_recs[0].get("성명") or 당기퇴사자_recs[0].get("사번") or "") if 당기퇴사자_recs else "")
        사번 = (전기레코드 or {}).get("사번") or (당기퇴사자_recs[0].get("사번") if 당기퇴사자_recs else None)

        rows.append({
            "사업장": (source or {}).get("사업장") or "",
            "부서": (source or {}).get("부서") or "",
            "사번": 사번,
            "직급": (source or {}).get("직급") or "",
            "원가구분": _cost_type(source) if source else "(미상)",
            "전기정보 있으나 당기정보 없음": 성명_auto,
            "실제 퇴사자": 성명_actual,
            "비고": " / ".join(tags),
        })

    if not rows:
        return pd.DataFrame(columns=LEAVER_MATCH_COLS)
    df = pd.DataFrame(rows)[LEAVER_MATCH_COLS]
    return df.sort_values(
        ["사업장", "부서", "전기정보 있으나 당기정보 없음", "실제 퇴사자"], na_position="last"
    ).reset_index(drop=True)


PAYROLL_MATCH_COLS = ["사업장", "부서", "사번", "직급", "원가구분",
                       "급여대장인원명부에만 존재", "연차정보(당기정보)에만 존재", "비고"]


def _build_payroll_match_df(당기_by_key: dict, payroll_employees: list) -> pd.DataFrame:
    """연차수당 대상인원(당기정보)과 기말 급여대장상 인원명부(선택 입력, '급여대장인원명부' 시트)를
    사번(없으면 성명) 기준으로 대사한다. 급여대장에는 있는데 연차정보에 없으면 연차 대상 인원이
    누락됐을 가능성, 반대로 연차정보에만 있으면 이미 퇴사했는데 남아있는 등의 확인이 필요하다는 뜻."""
    payroll_by_key: dict = {}
    for rec in payroll_employees:
        k = _employee_key(rec)
        if k is not None:
            payroll_by_key.setdefault(k, rec)

    rows = []
    for key in set(payroll_by_key) | set(당기_by_key):
        급여대장레코드 = payroll_by_key.get(key)
        연차정보레코드 = 당기_by_key.get(key)
        in_payroll = 급여대장레코드 is not None
        in_leave = 연차정보레코드 is not None
        source = 급여대장레코드 or 연차정보레코드

        if in_payroll and in_leave:
            tags = ["이상없음"]
        elif in_payroll and not in_leave:
            tags = ["⚠ 급여대장에는 있으나 연차정보(당기정보)에 없음 — 연차 대상 인원 누락 가능"]
        else:
            tags = ["⚠ 연차정보(당기정보)에는 있으나 급여대장에 없음 — 당기 중 퇴사 등 확인 필요"]

        성명_급여대장 = (급여대장레코드.get("성명") or 급여대장레코드.get("사번") or "") if in_payroll else ""
        성명_연차정보 = (연차정보레코드.get("성명") or 연차정보레코드.get("사번") or "") if in_leave else ""
        사번 = (source or {}).get("사번")

        rows.append({
            "사업장": (source or {}).get("사업장") or "",
            "부서": (source or {}).get("부서") or "",
            "사번": 사번,
            "직급": (source or {}).get("직급") or "",
            "원가구분": _cost_type(source) if source else "(미상)",
            "급여대장인원명부에만 존재": 성명_급여대장,
            "연차정보(당기정보)에만 존재": 성명_연차정보,
            "비고": " / ".join(tags),
        })

    if not rows:
        return pd.DataFrame(columns=PAYROLL_MATCH_COLS)
    df = pd.DataFrame(rows)[PAYROLL_MATCH_COLS]
    return df.sort_values(["비고", "사업장", "부서"], na_position="last").reset_index(drop=True)


def _build_group_summary(d: pd.DataFrame, 전기_calc_df: pd.DataFrame, group_col: str) -> list:
    """사업장별/부서별 요약표. 당기말은 당기_df(d, 당기 재직자)에서, 전기말은 전기_calc_df
    (전기 결산기준일 현재 재직 중이던 전체 인원 — 당기 중 퇴사한 사람 포함)에서 각각 집계한다."""
    if (group_col not in d.columns or d[group_col].astype(str).str.strip().eq("").all()) and \
       (group_col not in 전기_calc_df.columns or 전기_calc_df.empty or
            전기_calc_df[group_col].astype(str).str.strip().eq("").all()):
        return []

    당기_by_group = {}
    if not d.empty:
        for g, gdf in d.groupby(group_col, sort=False):
            if str(g).strip() == "":
                continue
            당기_by_group[g] = {
                "인원수": int(len(gdf)),
                "당기말": float(gdf["당기말 연차충당부채(계산)"].sum()),
            }
    전기_by_group = {}
    if not 전기_calc_df.empty:
        for g, gdf in 전기_calc_df.groupby(group_col, sort=False):
            if str(g).strip() == "":
                continue
            전기_by_group[g] = float(gdf["전기말 연차충당부채(계산)"].sum())

    rows = []
    for g in sorted(set(당기_by_group) | set(전기_by_group), key=lambda x: str(x)):
        rows.append({
            group_col: g,
            "당기인원수": 당기_by_group.get(g, {}).get("인원수", 0),
            "전기말(재계산)": 전기_by_group.get(g, 0.0),
            "당기말(재계산)": 당기_by_group.get(g, {}).get("당기말", 0.0),
            "당기 연차수당비용(재계산)": 당기_by_group.get(g, {}).get("당기말", 0.0) - 전기_by_group.get(g, 0.0),
        })
    return rows


# ── 인원별 연차충당부채 계산 ('당기정보' 시트 기준) ───────────────────────────

def compute_employee(emp: dict, anchor_start: date, anchor_end: date, mode: str,
                      payout_rate: float = 1.0) -> dict:
    """anchor_start(회계연도 시작일 전날 기준, 미포함) ~ anchor_end(결산기준일, 포함) 기간에 대한
    당기 연차 부여일수와 당기말 잔여연차일수·연차충당부채를 계산한다. '당기정보'/'전기정보' 두 시트
    어느 쪽 인원이든 각자의 기간(anchor)을 넣어 동일 로직으로 계산할 수 있도록 인자로 받는다.
    payout_rate(연차사용촉진 반영 지급률, 0~1): 잔여연차일수는 그대로 두고 금액(충당부채)에만 곱한다
    — 촉진 절차로 실제 현금 정산될 것으로 예상되는 비율만큼만 부채로 인식한다는 뜻."""
    입사일 = _safe_date(emp.get("입사일"))
    기초이월 = _safe_float(emp.get("기초 이월연차잔여일수(일)"))
    당기사용 = _safe_float(emp.get("당기 연차사용일수(일)"))
    통상임금 = _safe_float(emp.get("1일 통상임금(원)"))

    result = {
        "근속연수": None, "당기부여일수": 0.0, "당기말잔여일수": None, "당기말충당부채": 0.0,
        "warning": None,
    }

    def _add_warning(w):
        result["warning"] = f"{result['warning']} / {w}".strip(" /") if result["warning"] else w

    if 입사일 is None:
        _add_warning("입사일 미입력 — 당기부여일수는 0으로 계산됨(이월잔여·사용실적만 반영)")
        당기말잔여 = 기초이월 - 당기사용
        result["당기말잔여일수"] = 당기말잔여
        result["당기말충당부채"] = 당기말잔여 * 통상임금 * payout_rate
        return result

    if 입사일 > anchor_end:
        _add_warning("입사일이 결산기준일 이후 — 확인 필요(해당 기준일 시점에 아직 미입사)")
        return result

    if mode == "회계기준":
        if 입사일 <= anchor_start:
            근속연수 = _years_between(입사일, anchor_start)
        else:
            근속연수 = -1  # 당기 중 입사

        if 근속연수 >= 1:
            부여일수 = _entitlement_by_tenure(근속연수)
            근속연수_표시 = 근속연수
        elif 근속연수 == 0:
            재직개월수 = _months_between_inclusive(입사일, anchor_end)
            부여일수 = (15 * 재직개월수) // 12
            근속연수_표시 = 0
        else:  # 당기 중 입사(회계기준이라도 최초 근속연도는 입사일 기준 월단위 발생 적용)
            부여일수 = _monthly_accrual_days(입사일, anchor_start, anchor_end)
            근속연수_표시 = 0
    else:  # 입사기준
        if 입사일 > anchor_start:
            부여일수 = _monthly_accrual_days(입사일, anchor_start, anchor_end)
            근속연수_표시 = 0
        else:
            anniversaries = _anniversaries_in_range(입사일, anchor_start, anchor_end)
            if anniversaries:
                부여일수 = sum(_entitlement_by_tenure(n) for n, _ in anniversaries)
                근속연수_표시 = anniversaries[-1][0]
            else:
                근속연수 = _years_between(입사일, anchor_start)
                if 근속연수 == 0:
                    부여일수 = _monthly_accrual_days(입사일, anchor_start, anchor_end)
                    근속연수_표시 = 0
                else:
                    부여일수 = 0
                    근속연수_표시 = 근속연수

    당기말잔여 = 기초이월 + 부여일수 - 당기사용
    if 당기말잔여 < 0:
        _add_warning(f"당기말 잔여연차일수가 음수({당기말잔여:g}일) — 사용일수 입력 확인 필요")
    if 통상임금 <= 0:
        _add_warning("1일 통상임금 미입력 — 충당부채 0으로 계산됨")

    result["근속연수"] = 근속연수_표시
    result["당기부여일수"] = float(부여일수)
    result["당기말잔여일수"] = 당기말잔여
    result["당기말충당부채"] = 당기말잔여 * 통상임금 * payout_rate
    return result


# ── 명세서 구성 ('당기정보' 시트 기준) ────────────────────────────────────

def build_schedule_table(employees: list, anchor_start: date, anchor_end: date,
                          신규입사_keys: set, 전기말_balances: dict, mode: str,
                          payout_rate: float = 1.0) -> pd.DataFrame:
    """전기말_balances: {사번(또는 성명) 키: 전기 결산기준일 시점 연차충당부채}. compute_prior_balances()로 생성.
    신규입사자 등 전기 대응값이 없는 인원은 전기말 연차충당부채를 0으로 본다(전기 시점 미재직)."""
    rows = []
    for e in employees:
        r = compute_employee(e, anchor_start, anchor_end, mode, payout_rate)
        원가구분 = _cost_type(e)
        key = _employee_key(e)
        신규입사 = key in 신규입사_keys
        전기말충당부채 = 전기말_balances.get(key, 0.0)
        당기연차수당비용 = r["당기말충당부채"] - 전기말충당부채
        회사계상_raw = e.get("회사계상 기말 연차충당부채(원)")
        당기회사계상 = _safe_float(회사계상_raw) if 회사계상_raw not in (None, "") else None
        당기말차이 = None if 당기회사계상 is None else r["당기말충당부채"] - 당기회사계상

        비고 = e.get("비고") or ""
        tags = []
        if r["warning"]:
            tags.append(r["warning"])
        if 신규입사:
            tags.append("당기 신규입사(전기정보에는 없음)")
        if tags:
            비고 = f"{비고} / {' / '.join(tags)}".strip(" /")

        rows.append({
            "사업장": e.get("사업장") or "",
            "부서": e.get("부서") or "",
            "사번": e.get("사번"),
            "성명": e.get("성명"),
            "직급": e.get("직급"),
            "원가구분": 원가구분,
            "입사일": e.get("입사일"),
            "기초 이월연차잔여일수(일)": _safe_float(e.get("기초 이월연차잔여일수(일)")),
            "당기 연차사용일수(일)": _safe_float(e.get("당기 연차사용일수(일)")),
            "1일 통상임금(원)": _safe_float(e.get("1일 통상임금(원)")),
            "근속연수(당기말기준)": r["근속연수"],
            "당기부여일수(계산)": r["당기부여일수"],
            "당기말 잔여연차일수(계산)": r["당기말잔여일수"],
            "전기말 연차충당부채(계산)": 전기말충당부채,
            "당기말 연차충당부채(계산)": r["당기말충당부채"],
            "당기말 회사계상 연차충당부채(원)": 당기회사계상,
            "당기말 차이(계산-회사계상)": 당기말차이,
            "당기 연차수당비용(계산)": 당기연차수당비용,
            "당기신규입사": 신규입사,
            "비고": 비고,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    return df.sort_values(["사업장", "부서", "원가구분", "사번"], na_position="last").reset_index(drop=True)


def build_prior_schedule_table(전기_employees: list, anchor_start: date, anchor_end: date, mode: str,
                                payout_rate: float = 1.0) -> pd.DataFrame:
    """'전기정보' 시트 인원별 전기 결산기준일 시점 연차충당부채와, 입력된 '회사계상 기말 연차충당부채(원)'을
    인별로 대사한 명세를 만든다. build_schedule_table()의 전기(당기 대신) 버전."""
    rows = []
    for e in 전기_employees:
        r = compute_employee(e, anchor_start, anchor_end, mode, payout_rate)
        원가구분 = _cost_type(e)
        회사계상_raw = e.get("회사계상 기말 연차충당부채(원)")
        회사계상 = _safe_float(회사계상_raw) if 회사계상_raw not in (None, "") else None
        차이 = None if 회사계상 is None else r["당기말충당부채"] - 회사계상

        비고 = e.get("비고") or ""
        if r["warning"]:
            비고 = f"{비고} / {r['warning']}".strip(" /")

        rows.append({
            "사업장": e.get("사업장") or "",
            "부서": e.get("부서") or "",
            "사번": e.get("사번"),
            "성명": e.get("성명"),
            "직급": e.get("직급"),
            "원가구분": 원가구분,
            "입사일": e.get("입사일"),
            "기초 이월연차잔여일수(일)": _safe_float(e.get("기초 이월연차잔여일수(일)")),
            "당기 연차사용일수(일)": _safe_float(e.get("당기 연차사용일수(일)")),
            "1일 통상임금(원)": _safe_float(e.get("1일 통상임금(원)")),
            "근속연수(전기말기준)": r["근속연수"],
            "전기부여일수(계산)": r["당기부여일수"],
            "전기말 잔여연차일수(계산)": r["당기말잔여일수"],
            "전기말 연차충당부채(계산)": r["당기말충당부채"],
            "전기말 회사계상 연차충당부채(원)": 회사계상,
            "전기말 차이(계산-회사계상)": 차이,
            "비고": 비고,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    return df.sort_values(["사업장", "부서", "원가구분", "사번"], na_position="last").reset_index(drop=True)


# ── 요약(원가구분별 대사 + 인원변동) ──────────────────────────────────────────

def build_summary(당기_df: pd.DataFrame, 전기_employees: list,
                   신규입사자_recs: list, 퇴사자_recs: list, basis: dict,
                   전기_by_key: dict, 전기말_balances: dict, leaver_payments: list,
                   당기_by_key: dict, payroll_employees: list = None) -> dict:
    if 당기_df.empty and not 전기_employees and not 신규입사자_recs and not 퇴사자_recs:
        return {}

    d = 당기_df.copy()

    # 전기말(재계산)은 반드시 '전기 결산기준일 현재 재직 중이던 전체 인원'(전기_by_key) 기준으로 집계해야 한다.
    전기_calc_rows = [
        {
            "사업장": e.get("사업장") or "",
            "부서": e.get("부서") or "",
            "원가구분": _cost_type(e),
            "전기말 연차충당부채(계산)": 전기말_balances.get(k, 0.0),
        }
        for k, e in 전기_by_key.items()
    ]
    전기_calc_df = pd.DataFrame(전기_calc_rows) if 전기_calc_rows else pd.DataFrame(
        columns=["사업장", "부서", "원가구분", "전기말 연차충당부채(계산)"]
    )

    by_cost = []
    총_전기재계산 = 0.0
    총_당기재계산 = 0.0
    총_전기입력 = None
    총_당기회사계상 = None
    for cost in COST_TYPES:
        cdf = d[d["원가구분"] == cost] if not d.empty else d
        전기_cdf = 전기_calc_df[전기_calc_df["원가구분"] == cost] if not 전기_calc_df.empty else 전기_calc_df
        전기재계산 = float(전기_cdf["전기말 연차충당부채(계산)"].sum()) if not 전기_cdf.empty else 0.0
        당기재계산 = float(cdf["당기말 연차충당부채(계산)"].sum()) if not cdf.empty else 0.0
        당기연차수당비용 = float(cdf["당기 연차수당비용(계산)"].sum()) if not cdf.empty else 0.0
        전기입력 = _basis_float(basis, BASIS_KEYS[cost]["전기"])
        당기회사계상 = _basis_float(basis, BASIS_KEYS[cost]["당기"])
        by_cost.append({
            "구분": cost,
            "전기말(재계산)": 전기재계산,
            "전기말(회사계상)": 전기입력,
            "당기말(재계산)": 당기재계산,
            "당기말(회사계상)": 당기회사계상,
            "당기 연차수당비용(재계산)": 당기연차수당비용,
            "대사차이(재계산-회사계상)": None if 당기회사계상 is None else 당기재계산 - 당기회사계상,
        })
        총_전기재계산 += 전기재계산
        총_당기재계산 += 당기재계산
        if 전기입력 is not None:
            총_전기입력 = (총_전기입력 or 0.0) + 전기입력
        if 당기회사계상 is not None:
            총_당기회사계상 = (총_당기회사계상 or 0.0) + 당기회사계상

    미분류_당기 = d[~d["원가구분"].isin(COST_TYPES)] if not d.empty else d
    미분류_전기 = 전기_calc_df[~전기_calc_df["원가구분"].isin(COST_TYPES)] if not 전기_calc_df.empty else 전기_calc_df
    if not 미분류_당기.empty or not 미분류_전기.empty:
        미분류_전기재계산 = float(미분류_전기["전기말 연차충당부채(계산)"].sum()) if not 미분류_전기.empty else 0.0
        미분류_당기재계산 = float(미분류_당기["당기말 연차충당부채(계산)"].sum()) if not 미분류_당기.empty else 0.0
        미분류_당기연차수당비용 = float(미분류_당기["당기 연차수당비용(계산)"].sum()) if not 미분류_당기.empty else 0.0
        by_cost.append({
            "구분": "(미분류)",
            "전기말(재계산)": 미분류_전기재계산,
            "전기말(회사계상)": None,
            "당기말(재계산)": 미분류_당기재계산,
            "당기말(회사계상)": None,
            "당기 연차수당비용(재계산)": 미분류_당기연차수당비용,
            "대사차이(재계산-회사계상)": None,
        })
        총_전기재계산 += 미분류_전기재계산
        총_당기재계산 += 미분류_당기재계산

    총_당기연차수당비용 = float(d["당기 연차수당비용(계산)"].sum()) if not d.empty else 0.0
    total_row = {
        "구분": "합계",
        "전기말(재계산)": 총_전기재계산,
        "전기말(회사계상)": 총_전기입력,
        "당기말(재계산)": 총_당기재계산,
        "당기말(회사계상)": 총_당기회사계상,
        "당기 연차수당비용(재계산)": 총_당기연차수당비용,
        "대사차이(재계산-회사계상)": None if 총_당기회사계상 is None else 총_당기재계산 - 총_당기회사계상,
    }

    전기_df = _to_display_df(전기_employees)
    신규입사자_df = _to_display_df(신규입사자_recs)

    headcount = {
        "전기말인원수": int(len(전기_df)),
        "신규입사인원수": int(len(신규입사자_df)),
        "퇴사인원수": len(퇴사자_recs),
        "당기말인원수": int(len(d)),
    }
    headcount_by_cost = []
    for cost in COST_TYPES:
        headcount_by_cost.append({
            "원가구분": cost,
            "전기말인원수": int(len(전기_df[전기_df["원가구분"] == cost])) if not 전기_df.empty else 0,
            "신규입사인원수": int(len(신규입사자_df[신규입사자_df["원가구분"] == cost])) if not 신규입사자_df.empty else 0,
            "퇴사인원수": sum(1 for e in 퇴사자_recs if _cost_type(e) == cost),
            "당기말인원수": int(len(d[d["원가구분"] == cost])) if not d.empty else 0,
        })

    site_summary = _build_group_summary(d, 전기_calc_df, "사업장")
    dept_summary = _build_group_summary(d, 전기_calc_df, "부서")

    # 연차수당 대상인원 대사 — 급여대장인원명부(선택)가 있으면 인별 대사, 없으면 '기준정보'에 입력된
    # 총인원수(선택)와 당기말인원수만 비교(명부를 못 받았을 때 수기 검증용 참고치).
    payroll_employees = payroll_employees or []
    payroll_match_df = None
    payroll_count_check = None
    if payroll_employees:
        payroll_match_df = _build_payroll_match_df(당기_by_key, payroll_employees)
    else:
        급여대장총인원수 = _basis_float(basis, PAYROLL_COUNT_LABEL)
        if 급여대장총인원수 is not None:
            payroll_count_check = {
                "연차수당 대상인원수(당기정보)": headcount["당기말인원수"],
                "기말 급여대장상 총인원수(입력)": 급여대장총인원수,
                "차이": headcount["당기말인원수"] - 급여대장총인원수,
            }

    # T계정 검증(tie-out): 전기말(회사계상) + 당기 연차수당비용(재계산) - 당기지급액(분개장, 입력) =? 당기말(회사계상)
    당기지급액 = _basis_float(basis, DEBIT_BASIS_LABEL)
    tie_out = None
    if 총_전기입력 is not None and 당기지급액 is not None and 총_당기회사계상 is not None:
        계산상기말 = 총_전기입력 + total_row["당기 연차수당비용(재계산)"] - 당기지급액
        tie_out = {
            "전기말(회사계상)": 총_전기입력,
            "당기 연차수당비용(재계산)": total_row["당기 연차수당비용(재계산)"],
            "당기지급액(분개장)": 당기지급액,
            "계산상 당기말": 계산상기말,
            "당기말(회사계상)": 총_당기회사계상,
            "차이(계산상당기말-회사계상)": 계산상기말 - 총_당기회사계상,
        }

    # 퇴사자 명단 대사 — 자동 산출(전기정보-당기정보 차이) vs '당기퇴사자' 시트(사용자 입력) 통합 비교표
    leaver_match_df = _build_leaver_match_df(전기_by_key, 당기_by_key, 퇴사자_recs, leaver_payments)

    # 퇴사자 금액차이 분석 — '당기퇴사자' 시트에 입력된 경우만, 전기말 연차충당부채와 실제지급액을 금액으로만 비교
    leaver_recon_rows = []
    for rec in leaver_payments:
        실제지급액_raw = rec.get("실제지급액(원)")
        if 실제지급액_raw in (None, ""):
            continue
        key = _employee_key(rec)
        실제지급액 = _safe_float(실제지급액_raw)
        전기말충당부채 = 전기말_balances.get(key)
        전기레코드 = 전기_by_key.get(key)
        leaver_recon_rows.append({
            "사업장": (전기레코드 or {}).get("사업장") or "",
            "부서": (전기레코드 or {}).get("부서") or "",
            "사번": rec.get("사번"),
            "성명": rec.get("성명"),
            "직급": (전기레코드 or {}).get("직급") or "",
            "원가구분": _cost_type(전기레코드) if 전기레코드 else "(미상)",
            "전기말 연차충당부채(계산)": 전기말충당부채,
            "실제지급액(입력)": 실제지급액,
            "차이(연차충당부채-실제지급액)": None if 전기말충당부채 is None else 전기말충당부채 - 실제지급액,
        })
    leaver_recon_df = pd.DataFrame(leaver_recon_rows) if leaver_recon_rows else pd.DataFrame(
        columns=["사업장", "부서", "사번", "성명", "직급", "원가구분",
                 "전기말 연차충당부채(계산)", "실제지급액(입력)", "차이(연차충당부채-실제지급액)"]
    )

    return {
        "by_cost": by_cost,
        "total_row": total_row,
        "site_summary": site_summary,
        "dept_summary": dept_summary,
        "tie_out": tie_out,
        "leaver_match": leaver_match_df,
        "leaver_recon": leaver_recon_df,
        "신규입사자": 신규입사자_df[["사업장", "부서", "사번", "성명", "직급", "원가구분", "입사일"]],
        "headcount": headcount,
        "headcount_by_cost": headcount_by_cost,
        "payroll_match": payroll_match_df,
        "payroll_count_check": payroll_count_check,
    }


def write_summary_sheet(ws, summary: dict, company: str, target_fy: str,
                         당기결산일: date, 전기결산일: date, mode: str, interim_month: int = None,
                         payout_rate_current: float = 1.0, payout_rate_prior: float = 1.0):
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="203864")
    section_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill("solid", fgColor="9DC3E6")
    sig_fill = PatternFill("solid", fgColor="FFFF00")
    bold = Font(bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    period_note = f", ~{interim_month}월 중간결산(반기 등)" if interim_month else ""
    rate_parts = []
    if payout_rate_current != 1.0:
        rate_parts.append(f"당기 지급률 {payout_rate_current * 100:g}%")
    if payout_rate_prior != 1.0:
        rate_parts.append(f"전기 지급률 {payout_rate_prior * 100:g}%")
    payout_note = ", " + ", ".join(rate_parts) if rate_parts else ""
    ws.cell(row=1, column=1,
            value=(f"연월차충당부채 요약표 (회사: {company}, 회계연도: {target_fy}{period_note}, "
                   f"연차산정기준: {mode}{payout_note}, 전기말: {전기결산일}, 당기말: {당기결산일})")).font = Font(bold=True, size=13)
    for col in "ABCDEFG":
        ws.column_dimensions[col].width = 18

    r = 3
    if not summary:
        ws.cell(row=r, column=1, value="(인원 데이터 없음)")
        return

    # 1) 원가구분별 전기·당기 비교 + 대사
    ws.cell(row=r, column=1, value="■ 연차충당부채 원가구분별 전기·당기 비교 (대사)").fill = section_fill
    ws.cell(row=r, column=1).font = section_font
    for c in range(2, 8):
        ws.cell(row=r, column=c).fill = section_fill
    r += 2

    cost_headers = ["구분", "전기말(재계산)", "전기말(회사계상)", "당기말(재계산)", "당기말(회사계상)",
                     "당기 연차수당비용(재계산)", "대사차이(재계산-회사계상)"]
    for i, h in enumerate(cost_headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    r += 1

    for row_data in summary["by_cost"] + [summary["total_row"]]:
        is_total = row_data["구분"] == "합계"
        cell = ws.cell(row=r, column=1, value=row_data["구분"])
        cell.border = border
        vals = [row_data["전기말(재계산)"], row_data["전기말(회사계상)"], row_data["당기말(재계산)"],
                row_data["당기말(회사계상)"], row_data["당기 연차수당비용(재계산)"],
                row_data["대사차이(재계산-회사계상)"]]
        if is_total:
            cell.font = bold
            cell.fill = total_fill
        for i, v in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            if v is not None:
                cell.number_format = "#,##0"
            if is_total:
                cell.font = bold
                cell.fill = total_fill
            elif i == 7 and _is_significant(v, row_data["당기말(회사계상)"]):
                cell.fill = sig_fill
        r += 1
    r += 2

    # 1-1) T계정 검증 (분개장 당기지급액 입력 시에만 표시)
    tie_out = summary.get("tie_out")
    if tie_out is not None:
        ws.cell(row=r, column=1,
                value="■ 연차충당부채 T계정 검증 (기초+당기전입-당기지급액 =? 기말, 전액 회사계상·분개장 기준)").fill = section_fill
        ws.cell(row=r, column=1).font = section_font
        for c in range(2, 7):
            ws.cell(row=r, column=c).fill = section_fill
        r += 2

        tie_headers = ["전기말(회사계상)", "당기 연차수당비용(재계산)", "당기지급액(분개장)",
                        "계산상 당기말", "당기말(회사계상)", "차이(계산상당기말-회사계상)"]
        for i, h in enumerate(tie_headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1
        vals = [tie_out["전기말(회사계상)"], tie_out["당기 연차수당비용(재계산)"], tie_out["당기지급액(분개장)"],
                tie_out["계산상 당기말"], tie_out["당기말(회사계상)"], tie_out["차이(계산상당기말-회사계상)"]]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            cell.number_format = "#,##0"
            if i == 6 and _is_significant(v, tie_out["당기말(회사계상)"]):
                cell.fill = sig_fill
        r += 2

    # 1-2)/1-3) 사업장별/부서별 요약
    for group_col, section_title in (("사업장", "■ 사업장별 요약"), ("부서", "■ 부서별 요약")):
        rows_data = summary.get("site_summary" if group_col == "사업장" else "dept_summary") or []
        if not rows_data:
            continue

        ws.cell(row=r, column=1, value=section_title).fill = section_fill
        ws.cell(row=r, column=1).font = section_font
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = section_fill
        r += 2

        grp_headers = [group_col, "당기인원수", "전기말(재계산)", "당기말(재계산)", "당기 연차수당비용(재계산)"]
        for i, h in enumerate(grp_headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1

        총인원수 = 총전기말 = 총당기말 = 총연차수당비용 = 0
        for row_data in rows_data:
            cell = ws.cell(row=r, column=1, value=row_data[group_col])
            cell.border = border
            vals = [row_data["당기인원수"], row_data["전기말(재계산)"], row_data["당기말(재계산)"],
                    row_data["당기 연차수당비용(재계산)"]]
            for i, v in enumerate(vals, start=2):
                cell = ws.cell(row=r, column=i, value=v)
                cell.border = border
                if i > 2:
                    cell.number_format = "#,##0"
            총인원수 += row_data["당기인원수"]
            총전기말 += row_data["전기말(재계산)"]
            총당기말 += row_data["당기말(재계산)"]
            총연차수당비용 += row_data["당기 연차수당비용(재계산)"]
            r += 1

        cell = ws.cell(row=r, column=1, value="합계")
        cell.font = bold
        cell.fill = total_fill
        cell.border = border
        for i, v in enumerate([총인원수, 총전기말, 총당기말, 총연차수당비용], start=2):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            cell.font = bold
            cell.fill = total_fill
            if i > 2:
                cell.number_format = "#,##0"
        r += 3

    # 2) 인원 변동 요약
    ws.cell(row=r, column=1, value="■ 인원 변동 요약").fill = section_fill
    ws.cell(row=r, column=1).font = section_font
    for c in range(2, 6):
        ws.cell(row=r, column=c).fill = section_fill
    r += 2

    hc_headers = ["구분", "전기말인원수", "신규입사인원수", "퇴사인원수", "당기말인원수"]
    for i, h in enumerate(hc_headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    r += 1

    for hc in summary["headcount_by_cost"]:
        cell = ws.cell(row=r, column=1, value=hc["원가구분"])
        cell.border = border
        for i, k in enumerate(["전기말인원수", "신규입사인원수", "퇴사인원수", "당기말인원수"], start=2):
            cell = ws.cell(row=r, column=i, value=hc[k])
            cell.border = border
        r += 1
    hc = summary["headcount"]
    cell = ws.cell(row=r, column=1, value="합계")
    cell.font = bold
    cell.fill = total_fill
    cell.border = border
    for i, k in enumerate(["전기말인원수", "신규입사인원수", "퇴사인원수", "당기말인원수"], start=2):
        cell = ws.cell(row=r, column=i, value=hc[k])
        cell.border = border
        cell.font = bold
        cell.fill = total_fill
    r += 3

    # 2-1) 연차수당 대상인원 대사 (급여대장인원명부 vs 연차정보) — 둘 다 없으면 섹션 자체를 생략
    payroll_match = summary.get("payroll_match")
    payroll_count_check = summary.get("payroll_count_check")
    if payroll_match is not None:
        ws.cell(row=r, column=1, value="■ 연차수당 대상인원 대사 (급여대장인원명부 vs 연차정보)").fill = section_fill
        ws.cell(row=r, column=1).font = section_font
        for c in range(2, 9):
            ws.cell(row=r, column=c).fill = section_fill
        r += 1
        ws.cell(row=r, column=1,
                value="※ 양쪽 명단에 모두 있으면 '이상없음', 한쪽에만 있으면 비고에 원인을 표시합니다. "
                      "사번(없으면 성명) 기준 매칭입니다.").font = Font(italic=True, color="808080", size=9)
        r += 1

        pm_headers = ["사업장", "부서", "사번", "직급", "원가구분",
                      "급여대장인원명부에만 존재", "연차정보(당기정보)에만 존재", "비고"]
        for i, h in enumerate(pm_headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            if h == "비고":
                ws.column_dimensions[get_column_letter(i)].width = 46
        r += 1
        if payroll_match.empty:
            ws.cell(row=r, column=1, value="(해당 없음)")
            for c in range(1, 9):
                ws.cell(row=r, column=c).border = border
            r += 1
        else:
            for _, row in payroll_match.iterrows():
                for i, h in enumerate(pm_headers, start=1):
                    val = row[h]
                    cell = ws.cell(row=r, column=i, value=(None if pd.isna(val) or val == "" else val))
                    cell.border = border
                    if h == "비고" and val:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        if val != "이상없음":
                            cell.fill = sig_fill
                r += 1

            급여대장만_인원수 = int(payroll_match["비고"].str.contains("급여대장에는 있으나", na=False).sum())
            연차정보만_인원수 = int(payroll_match["비고"].str.contains("연차정보\\(당기정보\\)에는 있으나", na=False).sum())
            cell = ws.cell(row=r, column=1, value="불일치 인원수")
            cell.font = bold
            cell.fill = total_fill
            cell.border = border
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = total_fill
                ws.cell(row=r, column=c).border = border
            cell = ws.cell(row=r, column=6, value=급여대장만_인원수)
            cell.font = bold
            cell.alignment = center
            cell = ws.cell(row=r, column=7, value=연차정보만_인원수)
            cell.font = bold
            cell.alignment = center
            r += 1
        r += 2
    elif payroll_count_check is not None:
        ws.cell(row=r, column=1,
                value="■ 연차수당 대상인원 대사 (급여대장 총인원수 vs 연차정보 — 인원명부 미확보)").fill = section_fill
        ws.cell(row=r, column=1).font = section_font
        for c in range(2, 4):
            ws.cell(row=r, column=c).fill = section_fill
        r += 1
        ws.cell(row=r, column=1,
                value="※ 급여대장 인원명부를 확보하지 못해 총인원수만 비교합니다 — 차이가 있으면 수기로 원인을 확인하세요."
                ).font = Font(italic=True, color="808080", size=9)
        r += 1
        pc_headers = ["연차수당 대상인원수(당기정보)", "기말 급여대장상 총인원수(입력)", "차이"]
        for i, h in enumerate(pc_headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1
        vals = [payroll_count_check["연차수당 대상인원수(당기정보)"],
                payroll_count_check["기말 급여대장상 총인원수(입력)"], payroll_count_check["차이"]]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            if i == 3 and v != 0:
                cell.fill = sig_fill
                cell.font = bold
        r += 2

    # 3) 신규입사자 명단
    ws.cell(row=r, column=1,
            value="■ 신규입사자 명단 ('당기정보'에는 있으나 '전기정보'에는 없음)").fill = section_fill
    ws.cell(row=r, column=1).font = section_font
    for c in range(2, 8):
        ws.cell(row=r, column=c).fill = section_fill
    r += 2

    new_headers = ["사업장", "부서", "사번", "성명", "직급", "원가구분", "입사일"]
    for i, h in enumerate(new_headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    r += 1
    신규입사자 = summary["신규입사자"]
    if 신규입사자.empty:
        ws.cell(row=r, column=1, value="(해당 없음)")
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = border
        r += 1
    else:
        for _, row in 신규입사자.iterrows():
            for i, h in enumerate(new_headers, start=1):
                val = row[h]
                cell = ws.cell(row=r, column=i, value=(None if pd.isna(val) else val))
                cell.border = border
                if h == "입사일" and val is not None and not pd.isna(val):
                    cell.number_format = "yyyy-mm-dd"
            r += 1
    r += 2

    # 4) 퇴사자 명단 대사
    ws.cell(row=r, column=1,
            value="■ 퇴사자 명단 대사 (자동 산출 명단 vs '당기퇴사자' 시트 입력 명단 비교)").fill = section_fill
    ws.cell(row=r, column=1).font = section_font
    for c in range(2, 9):
        ws.cell(row=r, column=c).fill = section_fill
    r += 1
    ws.cell(row=r, column=1,
            value="※ 양쪽 명단에 모두 있으면 '이상없음', 한쪽에만 있으면 비고에 원인을 표시합니다. "
                  "사번(없으면 성명) 기준 매칭입니다.").font = Font(italic=True, color="808080", size=9)
    r += 1

    match_headers = ["사업장", "부서", "사번", "직급", "원가구분",
                      "전기정보 있으나 당기정보 없음", "실제 퇴사자", "비고"]
    for i, h in enumerate(match_headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        if h == "비고":
            ws.column_dimensions[get_column_letter(i)].width = 46
    r += 1
    leaver_match = summary.get("leaver_match")
    if leaver_match is None or leaver_match.empty:
        ws.cell(row=r, column=1, value="(해당 없음)")
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = border
        r += 1
    else:
        for _, row in leaver_match.iterrows():
            for i, h in enumerate(match_headers, start=1):
                val = row[h]
                cell = ws.cell(row=r, column=i, value=(None if pd.isna(val) or val == "" else val))
                cell.border = border
                if h == "비고" and val:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if val != "이상없음":
                        cell.fill = sig_fill
            r += 1

        전기없음_인원수 = int((leaver_match["전기정보 있으나 당기정보 없음"] != "").sum())
        실제퇴사_인원수 = int((leaver_match["실제 퇴사자"] != "").sum())
        cell = ws.cell(row=r, column=1, value="합계(인원수)")
        cell.font = bold
        cell.fill = total_fill
        cell.border = border
        for c in range(2, 9):
            ws.cell(row=r, column=c).fill = total_fill
            ws.cell(row=r, column=c).border = border
        cell = ws.cell(row=r, column=6, value=전기없음_인원수)
        cell.font = bold
        cell.alignment = center
        cell = ws.cell(row=r, column=7, value=실제퇴사_인원수)
        cell.font = bold
        cell.alignment = center
        r += 1
    r += 2

    # 5) 퇴사자 금액차이 분석
    leaver_recon = summary.get("leaver_recon")
    if leaver_recon is not None and not leaver_recon.empty:
        ws.cell(row=r, column=1,
                value="■ 퇴사자 금액차이 분석 ('당기퇴사자' 시트 입력분 — 전기말 연차충당부채 vs 실제지급액)").fill = section_fill
        ws.cell(row=r, column=1).font = section_font
        for c in range(2, 10):
            ws.cell(row=r, column=c).fill = section_fill
        r += 2

        recon_headers = ["사업장", "부서", "사번", "성명", "직급", "원가구분",
                          "전기말 연차충당부채(계산)", "실제지급액(입력)", "차이(연차충당부채-실제지급액)"]
        for i, h in enumerate(recon_headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1
        for _, row in leaver_recon.iterrows():
            for i, h in enumerate(recon_headers, start=1):
                val = row[h]
                cell = ws.cell(row=r, column=i, value=(None if pd.isna(val) or val == "" else val))
                cell.border = border
                if h in ("전기말 연차충당부채(계산)", "실제지급액(입력)", "차이(연차충당부채-실제지급액)") \
                        and val is not None and not pd.isna(val):
                    cell.number_format = "#,##0"
                    if h == "차이(연차충당부채-실제지급액)" and _is_significant(val, row["전기말 연차충당부채(계산)"]):
                        cell.fill = sig_fill
            r += 1

        총전기말충당부채 = float(leaver_recon["전기말 연차충당부채(계산)"].sum(skipna=True))
        총실제지급액 = float(leaver_recon["실제지급액(입력)"].sum(skipna=True))
        cell = ws.cell(row=r, column=1, value="합계")
        cell.font = bold
        cell.fill = total_fill
        cell.border = border
        for c in range(2, 10):
            ws.cell(row=r, column=c).fill = total_fill
            ws.cell(row=r, column=c).border = border
        for col, v in ((7, 총전기말충당부채), (8, 총실제지급액), (9, 총전기말충당부채 - 총실제지급액)):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = bold
            cell.number_format = "#,##0"
        r += 1
    r += 2

    # 6) 참고: 계산 공식
    ws.cell(row=r, column=1, value="■ 참고: 연차충당부채 계산 공식 (수기 검증용)").fill = section_fill
    ws.cell(row=r, column=1).font = section_font
    for c in range(2, 8):
        ws.cell(row=r, column=c).fill = section_fill
    r += 2
    for text, is_bold in FORMULA_NOTE_LINES:
        if text:
            cell = ws.cell(row=r, column=1, value=text if is_bold else f"    {text}")
            cell.font = Font(bold=True, size=11) if is_bold else Font(size=10)
        r += 1


# ── 엑셀 저장 ────────────────────────────────────────────────────────────────

MONEY_COLS = ["1일 통상임금(원)", "전기말 연차충당부채(계산)", "당기말 연차충당부채(계산)",
              "당기말 회사계상 연차충당부채(원)", "당기말 차이(계산-회사계상)", "당기 연차수당비용(계산)"]
DAY_COLS = ["기초 이월연차잔여일수(일)", "당기 연차사용일수(일)", "당기부여일수(계산)", "당기말 잔여연차일수(계산)"]
DATE_COLS = ["입사일"]
DIFF_BASE_COLS = {"당기말 차이(계산-회사계상)": "당기말 회사계상 연차충당부채(원)"}
COUNT_COLS = ["당기신규입사"]

PRIOR_MONEY_COLS = ["1일 통상임금(원)", "전기말 연차충당부채(계산)",
                     "전기말 회사계상 연차충당부채(원)", "전기말 차이(계산-회사계상)"]
PRIOR_DAY_COLS = ["기초 이월연차잔여일수(일)", "당기 연차사용일수(일)", "전기부여일수(계산)", "전기말 잔여연차일수(계산)"]
PRIOR_DIFF_BASE_COLS = {"전기말 차이(계산-회사계상)": "전기말 회사계상 연차충당부채(원)"}


def save_results(df: pd.DataFrame, output_path: str, company: str, target_fy: str,
                  당기결산일: date, 전기결산일: date, 전기_employees: list,
                  신규입사자_recs: list, 퇴사자_recs: list, basis: dict,
                  전기_by_key: dict, 전기말_balances: dict, leaver_payments: list,
                  당기_by_key: dict, mode: str, 전기_anchor_start: date, interim_month: int = None,
                  payroll_employees: list = None, payout_rate_current: float = 1.0,
                  payout_rate_prior: float = 1.0):
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약표"
    summary = build_summary(df, 전기_employees, 신규입사자_recs, 퇴사자_recs, basis,
                             전기_by_key, 전기말_balances, leaver_payments, 당기_by_key, payroll_employees)
    write_summary_sheet(ws_summary, summary, company, target_fy, 당기결산일, 전기결산일, mode, interim_month,
                         payout_rate_current, payout_rate_prior)

    ws = wb.create_sheet("인원별추계명세")

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    subtotal_fill = PatternFill("solid", fgColor="D9E1F2")
    total_fill = PatternFill("solid", fgColor="9DC3E6")
    sig_fill = PatternFill("solid", fgColor="FFFF00")
    bold = Font(bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    period_note = f", ~{interim_month}월 중간결산(반기 등)" if interim_month else ""
    ws.cell(row=1, column=1,
            value=(f"인원별 연차충당부채 명세서 (회사: {company}, 회계연도: {target_fy}{period_note}, "
                   f"연차산정기준: {mode}, 당기말 결산기준일: {당기결산일})")).font = Font(bold=True, size=13)

    headers = list(df.columns) if not df.empty else []
    header_row = 3
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = 16 if h != "비고" else 40
    ws.freeze_panes = f"A{header_row + 1}"

    r = header_row + 1
    if df.empty:
        ws.cell(row=r, column=1, value="(인원 데이터 없음)")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        return

    totals = {c: 0.0 for c in MONEY_COLS + DAY_COLS}
    grand_totals = {c: 0.0 for c in MONEY_COLS + DAY_COLS}
    counts = {c: 0 for c in COUNT_COLS}
    grand_counts = {c: 0 for c in COUNT_COLS}

    for cost, gdf in df.groupby("원가구분", sort=False):
        for _, row in gdf.iterrows():
            for i, h in enumerate(headers, start=1):
                val = row[h]
                cell = ws.cell(row=r, column=i, value=(None if pd.isna(val) else val))
                cell.border = border
                if h in DATE_COLS and val is not None and not pd.isna(val):
                    cell.number_format = "yyyy-mm-dd"
                if h in MONEY_COLS:
                    cell.number_format = "#,##0"
                if h in DAY_COLS:
                    cell.number_format = "0.0"
                if h == "당기신규입사":
                    cell.value = "O" if val is True else None
                    cell.alignment = center
                if h in DIFF_BASE_COLS and _is_significant(val, row.get(DIFF_BASE_COLS[h])):
                    cell.fill = sig_fill
            for c in MONEY_COLS + DAY_COLS:
                v = row.get(c)
                if v is not None and not pd.isna(v):
                    totals[c] += v
                    grand_totals[c] += v
            for c in COUNT_COLS:
                if row.get(c) is True:
                    counts[c] += 1
                    grand_counts[c] += 1
            r += 1

        ws.cell(row=r, column=1, value=f"[{cost} 소계]").font = bold
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=i)
            cell.fill = subtotal_fill
            cell.border = border
            if h in MONEY_COLS:
                cell.value = totals[h]
                cell.number_format = "#,##0"
                cell.font = bold
            elif h in DAY_COLS:
                cell.value = totals[h]
                cell.number_format = "0.0"
                cell.font = bold
            elif h in COUNT_COLS and counts[h] > 0:
                cell.value = f"{counts[h]}명"
                cell.font = bold
                cell.alignment = center
        r += 1
        totals = {c: 0.0 for c in MONEY_COLS + DAY_COLS}
        counts = {c: 0 for c in COUNT_COLS}

    ws.cell(row=r, column=1, value="총계").font = Font(bold=True, size=11)
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=i)
        cell.fill = total_fill
        cell.border = border
        if h in MONEY_COLS:
            cell.value = grand_totals[h]
            cell.number_format = "#,##0"
            cell.font = bold
        elif h in DAY_COLS:
            cell.value = grand_totals[h]
            cell.number_format = "0.0"
            cell.font = bold
        elif h in COUNT_COLS and grand_counts[h] > 0:
            cell.value = f"{grand_counts[h]}명"
            cell.font = Font(bold=True, size=11)
            cell.alignment = center

    # 전기인원별추계명세 — '전기정보' 인원별 전기말 재계산액과 '회사계상 기말 연차충당부채(원)' 대사
    prior_df = build_prior_schedule_table(전기_employees, 전기_anchor_start, 전기결산일, mode, payout_rate_prior)
    ws_prior = wb.create_sheet("전기인원별추계명세")
    ws_prior.cell(row=1, column=1,
                  value=(f"전기인원별 연차충당부채 명세서 (회사: {company}, 회계연도: {target_fy}, "
                         f"연차산정기준: {mode}, 전기말 결산기준일: {전기결산일})")).font = Font(bold=True, size=13)

    prior_headers = list(prior_df.columns) if not prior_df.empty else []
    for i, h in enumerate(prior_headers, start=1):
        c = ws_prior.cell(row=header_row, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
        ws_prior.column_dimensions[get_column_letter(i)].width = 16 if h != "비고" else 40
    ws_prior.freeze_panes = f"A{header_row + 1}"

    pr = header_row + 1
    if prior_df.empty:
        ws_prior.cell(row=pr, column=1, value="(인원 데이터 없음)")
    else:
        p_totals = {c: 0.0 for c in PRIOR_MONEY_COLS + PRIOR_DAY_COLS}
        p_grand_totals = {c: 0.0 for c in PRIOR_MONEY_COLS + PRIOR_DAY_COLS}
        for cost, gdf in prior_df.groupby("원가구분", sort=False):
            for _, row in gdf.iterrows():
                for i, h in enumerate(prior_headers, start=1):
                    val = row[h]
                    cell = ws_prior.cell(row=pr, column=i, value=(None if pd.isna(val) else val))
                    cell.border = border
                    if h in DATE_COLS and val is not None and not pd.isna(val):
                        cell.number_format = "yyyy-mm-dd"
                    if h in PRIOR_MONEY_COLS:
                        cell.number_format = "#,##0"
                    if h in PRIOR_DAY_COLS:
                        cell.number_format = "0.0"
                    if h in PRIOR_DIFF_BASE_COLS and _is_significant(val, row.get(PRIOR_DIFF_BASE_COLS[h])):
                        cell.fill = sig_fill
                for c in PRIOR_MONEY_COLS + PRIOR_DAY_COLS:
                    v = row.get(c)
                    if v is not None and not pd.isna(v):
                        p_totals[c] += v
                        p_grand_totals[c] += v
                pr += 1

            ws_prior.cell(row=pr, column=1, value=f"[{cost} 소계]").font = bold
            for i, h in enumerate(prior_headers, start=1):
                cell = ws_prior.cell(row=pr, column=i)
                cell.fill = subtotal_fill
                cell.border = border
                if h in PRIOR_MONEY_COLS:
                    cell.value = p_totals[h]
                    cell.number_format = "#,##0"
                    cell.font = bold
                elif h in PRIOR_DAY_COLS:
                    cell.value = p_totals[h]
                    cell.number_format = "0.0"
                    cell.font = bold
            pr += 1
            p_totals = {c: 0.0 for c in PRIOR_MONEY_COLS + PRIOR_DAY_COLS}

        ws_prior.cell(row=pr, column=1, value="총계").font = Font(bold=True, size=11)
        for i, h in enumerate(prior_headers, start=1):
            cell = ws_prior.cell(row=pr, column=i)
            cell.fill = total_fill
            cell.border = border
            if h in PRIOR_MONEY_COLS:
                cell.value = p_grand_totals[h]
                cell.number_format = "#,##0"
                cell.font = bold
            elif h in PRIOR_DAY_COLS:
                cell.value = p_grand_totals[h]
                cell.number_format = "0.0"
                cell.font = bold

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="연월차충당부채 검증앱(잔여연차일수×1일통상임금 방식)")
    parser.add_argument("company", nargs="?", default=None, help="처리할 회사명 (생략 시 파일 자동 탐색)")
    parser.add_argument("--file", default=None, help="처리할 특정 입력 파일명 (input_data/ 기준)")
    parser.add_argument("--fiscal-month", type=int, default=12, help="결산월 (기본 12월). 예: 6월 결산법인이면 6")
    parser.add_argument("--fiscal-year", default=None, help="검증 대상 회계연도 (예: 2026). 생략 시 입력파일명의 fy 뒤 숫자 사용")
    parser.add_argument("--interim-month", type=int, default=None,
                         help="반기 등 중간결산 검토월 (예: 6). 당기말 결산기준일만 해당 월말로 앞당기고, "
                              "전기말은 직전 회계연도 결산기준일 그대로 사용")
    args = parser.parse_args()

    input_path = _find_input_file(args.company, args.file)
    company = args.company or os.path.basename(input_path).split("_")[1]

    target_fy = args.fiscal_year
    if not target_fy:
        base = os.path.basename(input_path)
        idx = base.lower().find("fy")
        if idx == -1:
            raise ValueError("--fiscal-year 를 지정하거나 파일명에 'fy<연도>'를 포함하세요.")
        digits = ""
        for ch in base[idx + 2:]:
            if ch.isdigit():
                digits += ch
            else:
                break
        target_fy = f"20{digits}" if len(digits) == 2 else digits

    _, fy_end_ym = _fy_bounds(target_fy, args.fiscal_month)
    fy_end_ym = _apply_interim(fy_end_ym, target_fy, args.interim_month)
    당기결산일 = _ym_to_end_date(fy_end_ym)

    prior_fy = str(int(target_fy) - 1)
    _, prior_fy_end_ym = _fy_bounds(prior_fy, args.fiscal_month)
    전기결산일 = _ym_to_end_date(prior_fy_end_ym)

    prior2_fy = str(int(target_fy) - 2)
    _, prior2_fy_end_ym = _fy_bounds(prior2_fy, args.fiscal_month)
    전전기결산일 = _ym_to_end_date(prior2_fy_end_ym)

    당기_anchor_start = 전기결산일 + timedelta(days=1)
    전기_anchor_start = 전전기결산일 + timedelta(days=1)

    print(f"[입력] {input_path}")
    interim_note = f", 중간결산월={args.interim_month}(반기 등)" if args.interim_month else ""
    print(f"[대상] 회사={company}, 회계연도={target_fy}, 결산월={args.fiscal_month}{interim_note}")
    print(f"[기준일] 당기말={당기결산일}, 전기말={전기결산일}, 전전기말={전전기결산일}")

    당기_employees = load_employees(input_path, CURRENT_SHEET)
    전기_employees = load_employees(input_path, PRIOR_SHEET)
    leaver_payments = load_employees(input_path, LEAVER_SHEET)
    payroll_employees = load_employees(input_path, PAYROLL_SHEET)
    basis = load_basis(input_path)
    mode = leave_basis_mode(basis)
    payout_rate_current = leave_payout_rate(basis, PAYOUT_RATE_LABEL_CURRENT)
    payout_rate_prior = leave_payout_rate(basis, PAYOUT_RATE_LABEL_PRIOR)
    print(f"[연차산정기준] {mode}")
    if payout_rate_current != 1.0 or payout_rate_prior != 1.0:
        print(f"[연차사용촉진 반영 지급률] 당기={payout_rate_current * 100:g}%, 전기={payout_rate_prior * 100:g}%")
    print(f"[인원 수] 당기={len(당기_employees)}건, 전기={len(전기_employees)}건, 당기퇴사자(지급액 입력)={len(leaver_payments)}건, "
          f"급여대장인원명부={len(payroll_employees)}건")

    matched = match_periods(당기_employees, 전기_employees)
    전기말_balances = compute_prior_balances(matched["전기_by_key"], 전기_anchor_start, 전기결산일, mode, payout_rate_prior)
    df = build_schedule_table(당기_employees, 당기_anchor_start, 당기결산일,
                               matched["신규입사_keys"], 전기말_balances, mode, payout_rate_current)

    suffix = f"_interim{args.interim_month:02d}" if args.interim_month else ""
    output_path = os.path.join(OUTPUT_DIR, f"leave_schedule_{company}_{target_fy}{suffix}.xlsx")
    save_results(df, output_path, company, target_fy, 당기결산일, 전기결산일,
                 전기_employees, matched["신규입사자"], matched["퇴사자"], basis,
                 matched["전기_by_key"], 전기말_balances, leaver_payments,
                 matched["당기_by_key"], mode, 전기_anchor_start, args.interim_month, payroll_employees,
                 payout_rate_current, payout_rate_prior)
    print(f"[완료] {output_path}")


if __name__ == "__main__":
    main()
