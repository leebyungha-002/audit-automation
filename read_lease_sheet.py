#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""감사조서 사용권자산_리스부채 시트 구조 확인용 임시 스크립트"""
import sys, os
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
path = os.path.join(BASE, 'graphy', '감사조서', '당기_graphy_조서_25년.xlsx')
wb = openpyxl.load_workbook(path, data_only=True)

print('=== 시트 목록 ===')
for s in wb.sheetnames:
    print(' ', s)

# 사용권자산/리스부채 시트 탐색
target = None
for s in wb.sheetnames:
    if '사용권' in s or '리스부채' in s:
        target = s
        break

if not target:
    print('[없음] 사용권자산 시트를 찾을 수 없음')
    wb.close()
    sys.exit()

print(f'\n=== [{target}] 시트 내용 (1~35행) ===')
ws = wb[target]
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=35, values_only=True), 1):
    if any(v is not None for v in row):
        print(f'  row{i:>2}: {row}')
wb.close()
