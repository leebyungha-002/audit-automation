#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""매핑 리스트 기반 데이터 주입 엔진
Usage: python data_injector.py <company_name>

기준 폴더: ../<company>/감사조서/
- 매핑 파일:  ../<company>/감사조서/<company>_mapping_list*.xlsx
- 대상 조서:  ../<company>/감사조서/ (매핑의 '대상 조서 파일명' 키워드로 탐색)
- 소스 데이터: ../<company>/results/ → raw_data/ → <company>/ 순서로 탐색
"""

import sys
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

try:
    from PIL import Image as _PilImage  # noqa: F401
    _PILLOW_OK = True
except ImportError:
    _PILLOW_OK = False
    print('[경고] Pillow 미설치 — ws._images 처리 불가. pip install Pillow')

try:
    import xlwings as xw
    _XLWINGS_OK = True
except ImportError:
    _XLWINGS_OK = False

# Windows 콘솔 한글·특수문자 출력 보장
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)


# ─── 파일명 정규화 ────────────────────────────────────────────────────────────

def _normalize(fname_no_ext):
    """YYYYMMDD(8자리) 날짜 토큰만 제거 후 연속 _ 정리."""
    s = re.sub(r'_\d{8}(?=_|$)', '', fname_no_ext)
    return re.sub(r'_+', '_', s).strip('_')


def _keyword_matches(keyword, normalized_fname):
    """keyword가 정규화된 파일명 내에 _ 경계로 정확히 포함되는지 확인.

    keyword='dae_il_외상매출금'
      normalized='dae_il_외상매출금'             → True  (정확일치)
      normalized='dae_il_외상매출금_상세'         → True  (suffix 허용)
      normalized='dae_il_벤포드_외상매출금'       → False (앞에 다른 토큰)
    """
    text    = '_' + normalized_fname + '_'
    pattern = '_' + re.escape(keyword) + '_'
    return bool(re.search(pattern, text, re.IGNORECASE))


# ─── 파일 탐색 ────────────────────────────────────────────────────────────────

def find_file_by_keyword(directories, keyword, exclude_suffixes=None):
    """keyword를 파일명(날짜 정규화 후)에서 _ 경계로 탐색.

    여러 개 발견 시 가장 최근 수정 파일 반환.
    """
    if isinstance(directories, str):
        directories = [directories]
    if exclude_suffixes is None:
        exclude_suffixes = ['~$', '_updated']

    matches = []
    for directory in directories:
        if not os.path.isdir(directory):
            continue
        for fname in sorted(os.listdir(directory)):
            if any(ex in fname for ex in exclude_suffixes):
                continue
            if not fname.lower().endswith('.xlsx'):
                continue
            normalized = _normalize(os.path.splitext(fname)[0])
            if _keyword_matches(keyword, normalized):
                matches.append(os.path.join(directory, fname))

    if not matches:
        return None
    if len(matches) == 1:
        return matches[0]

    matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    names = [os.path.basename(p) for p in matches]
    print(f"    [주의] '{keyword}' 키워드로 {len(matches)}개 파일 발견:")
    for n in names:
        print(f"           {n}")
    print(f"           → 최근 파일 선택: {names[0]}")
    return matches[0]


# ─── 시트 탐색 ────────────────────────────────────────────────────────────────

def resolve_sheet(sheetnames, keyword):
    """정확히 일치 → 대소문자 무시 일치 → keyword 포함 첫 번째 시트 순으로 탐색."""
    if keyword in sheetnames:
        return keyword
    kw_lower = keyword.lower()
    for s in sheetnames:
        if s.lower() == kw_lower:
            return s
    matched = [s for s in sheetnames if kw_lower in s.lower()]
    return matched[0] if matched else None


# ─── 셀 좌표 / 범위 파싱 ─────────────────────────────────────────────────────

def _parse_cell(cell_ref):
    """'A7' → (row=7, col=1)  /  대소문자 무관. 범위(A1:I400) 입력 시 시작 셀만 사용."""
    cell_ref = cell_ref.strip().split(':')[0]  # 범위 형식이면 시작 셀만 취함
    m = re.match(r'^([A-Za-z]+)(\d+)$', cell_ref)
    if not m:
        raise ValueError(f"잘못된 셀 좌표: {cell_ref}")
    return int(m.group(2)), column_index_from_string(m.group(1).upper())


def _parse_range(range_str):
    """'B2:C13' → (min_row=2, min_col=2, max_row=13, max_col=3)."""
    parts = range_str.strip().upper().split(':')
    if len(parts) != 2:
        raise ValueError(f"잘못된 범위: {range_str}  (형식 예: B2:C13)")
    min_row, min_col = _parse_cell(parts[0])
    max_row, max_col = _parse_cell(parts[1])
    return min_row, min_col, max_row, max_col


# ─── 데이터 주입 ──────────────────────────────────────────────────────────────

def inject_data(ws_src, ws_tgt, start_cell, src_range=None):
    """소스 시트 데이터를 대상 시트의 start_cell 부터 값만 주입 (서식·수식 보존).

    src_range 지정 ('B2:C13') : 해당 영역만 추출하여 주입
    src_range 미지정 (None)   : 소스 시트 used range 전체 주입
    행·열 구조(Matrix)는 그대로 유지.  반환: 주입된 셀 수
    """
    start_row, start_col = _parse_cell(start_cell)

    if src_range:
        min_row, min_col, max_row, max_col = _parse_range(src_range)
        src_rows = ws_src.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col,
            values_only=True,
        )
    else:
        src_rows = ws_src.iter_rows(values_only=True)

    count = 0
    for r_idx, row in enumerate(src_rows):
        for c_idx, value in enumerate(row):
            if value is not None:
                ws_tgt.cell(row=start_row + r_idx,
                            column=start_col + c_idx).value = value
                count += 1
    return count


# ─── 잔존 데이터 클리어 ──────────────────────────────────────────────────────
# PIVOT_AGING / ANALYSIS_INJECT 처럼 매번 가변 행수를 주입하는 경우, 재실행 시
# 이전 결과보다 행/열이 줄어들면 끝부분 잔존 데이터가 남을 수 있다.
# 주입 전 고정 범위를 비워(clear) 이를 방지한다.

_CLEAR_MAX_ROWS = 200
_CLEAR_MAX_COLS = 30
_NO_FILL = PatternFill(fill_type=None)


def _clear_range(ws, start_row, start_col, n_rows=_CLEAR_MAX_ROWS, n_cols=_CLEAR_MAX_COLS, reset_fill=False):
    """start_row~start_row+n_rows-1 행, start_col~start_col+n_cols-1 열의 셀 값을 비운다.

    reset_fill=True 이면 강조 서식(예: _YELLOW_FILL)도 함께 초기화한다.
    """
    for r in range(start_row, start_row + n_rows):
        for c in range(start_col, start_col + n_cols):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            if reset_fill:
                cell.fill = _NO_FILL


# ─── 매핑 파일 로드 ──────────────────────────────────────────────────────────

def load_mapping(mapping_path):
    """<회사>_mapping_list*.xlsx 를 읽어 매핑 행 리스트 반환.

    컬럼 순서 (A~I):
      A 계정과목(label) / B 소스파일명(src_kw) / C 소스시트(src_sheet)
      D 소스 데이터 범위(src_range, 선택 — 예: B2:C13)
      E 대상파일명(tgt_kw) / F 대상시트(tgt_sheet) / G 시작셀(start_cell)
      H 기준금액(threshold, 선택 — ANALYSIS_INJECT 유의적 변동 판단 기준)
      I 비고(remarks, 선택 — 예: PIVOT_AGING / MOVE_IMAGE / ANALYSIS_INJECT / LEASE_INJECT / AI_INJECT)
    """
    wb = load_workbook(mapping_path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        padded = list(row) + [None] * 9
        label, src_kw, src_sheet, src_range, tgt_kw, tgt_sheet, start_cell, threshold, remarks = padded[:9]
        if not src_kw or not tgt_kw or not start_cell:
            continue
        try:
            threshold_val = float(re.sub(r'[,\s]', '', str(threshold))) if threshold else 0.0
        except (ValueError, TypeError):
            threshold_val = 0.0
        rows.append({
            'label':      str(label      or '').strip(),
            'src_kw':     str(src_kw    ).strip(),
            'src_sheet':  str(src_sheet ).strip(),
            'tgt_kw':     str(tgt_kw    ).strip(),
            'tgt_sheet':  str(tgt_sheet ).strip(),
            'start_cell': str(start_cell).strip().upper(),
            'src_range':  str(src_range ).strip().upper() if src_range else '',
            'remarks':    str(remarks   ).strip().upper() if remarks else '',
            'threshold':  threshold_val,
        })
    return rows


# ─── Aging 피벗 ──────────────────────────────────────────────────────────────

def build_pivot_aging(src_path, sheet_name):
    """pandas(calamine 우선)로 소스 파일을 읽어 거래처명 × 월별 차변금액 피벗을 생성.

    Returns (headers, data_rows):
      headers   = ['거래처명', '2025-01', ..., '합계']
      data_rows = [['거래처A', 100000, None, ..., 100000], ..., ['합계', ...]]
    """
    def _read(engine, **kw):
        return pd.read_excel(src_path, sheet_name=sheet_name, engine=engine, **kw)

    # ── 1. 엔진 선택 + 헤더 컬럼 확인 (nrows=0 으로 빠르게) ──────────────
    try:
        df_head = _read('calamine', nrows=0)
        engine  = 'calamine'
    except Exception:
        df_head = _read('openpyxl', nrows=0)
        engine  = 'openpyxl'

    def find_col(*keywords):
        for c in df_head.columns:
            if any(kw in str(c) for kw in keywords):
                return c
        return None

    col_cust = find_col('거래처')
    col_date = find_col('전표날짜', '날짜', '일자')
    col_amt  = find_col('차변금액', '차변', '금액')

    missing = [n for n, c in [('거래처명', col_cust), ('전표날짜', col_date), ('차변금액', col_amt)] if c is None]
    if missing:
        raise ValueError(f"필수 컬럼을 찾을 수 없습니다: {', '.join(missing)}")

    # ── 2. 필요 컬럼만 로드 (usecols 로 I/O 최소화) ──────────────────────
    df = _read(engine, usecols=[col_cust, col_date, col_amt])
    df = df.rename(columns={col_cust: '거래처명', col_date: '_date', col_amt: '차변금액'})

    # ── 3. 전처리 ─────────────────────────────────────────────────────────
    df['차변금액'] = pd.to_numeric(
        df['차변금액'].astype(str).str.replace(r'[,원\s]', '', regex=True),
        errors='coerce',
    ).fillna(0)
    df['_month'] = pd.to_datetime(df['_date'], errors='coerce').dt.strftime('%Y-%m')
    df = df.dropna(subset=['거래처명', '_month'])
    df = df[df['거래처명'].astype(str).str.strip().ne('')]

    if df.empty:
        raise ValueError("피벗 데이터 없음 — 유효한 거래처명/날짜 행이 없습니다.")

    # ── 4. 피벗 집계 ─────────────────────────────────────────────────────
    pivot = df.pivot_table(
        index='거래처명',
        columns='_month',
        values='차변금액',
        aggfunc='sum',
        fill_value=0,
    ).sort_index()
    pivot.columns.name = None

    # ── 5. 합계 행/열 추가 ───────────────────────────────────────────────
    pivot['합계'] = pivot.sum(axis=1)
    total = pivot.sum(axis=0).rename('합계')
    pivot = pd.concat([pivot, total.to_frame().T])

    # ── 6. (headers, data_rows) 포맷 변환 ───────────────────────────────
    month_cols = [c for c in pivot.columns if c != '합계']
    headers    = ['거래처명'] + month_cols + ['합계']

    data_rows = []
    for cust, row in pivot.iterrows():
        vals = [cust] + [float(row[m]) if row[m] != 0 else None for m in month_cols]
        tot  = row['합계']
        vals.append(float(tot) if tot != 0 else None)
        data_rows.append(vals)

    return headers, data_rows


def inject_pivot_aging(src_path, src_sheet, wb_tgt, tgt_sheet_name, start_cell):
    """피벗 Aging 테이블을 대상 워크북의 tgt_sheet_name 시트에 주입한다.

    추가로 Aging_분석 시트 A5부터 거래처 리스트를 세로로 업데이트한다.
    시트가 없으면 새로 생성한다. 반환값: 주입된 데이터 행 수.
    """
    headers, data_rows = build_pivot_aging(src_path, src_sheet)

    # ── 1) Aging_Source: 피벗 테이블 전체 주입 ───────────────────────────────
    if tgt_sheet_name in wb_tgt.sheetnames:
        ws_aging = wb_tgt[tgt_sheet_name]
    else:
        ws_aging = wb_tgt.create_sheet(title=tgt_sheet_name)
        print(f'    [Aging] 시트 신규 생성: {tgt_sheet_name}')

    start_row, start_col = _parse_cell(start_cell)

    # 재실행 시 이전 결과보다 행/열이 줄어들 경우 잔존 데이터 제거
    _clear_range(ws_aging, start_row, start_col)

    for c_idx, h in enumerate(headers):
        ws_aging.cell(row=start_row, column=start_col + c_idx).value = h

    for r_idx, row in enumerate(data_rows, start=1):
        for c_idx, val in enumerate(row):
            ws_aging.cell(row=start_row + r_idx, column=start_col + c_idx).value = val

    # ── 2) Aging_분석: A5부터 거래처 리스트 세로 주입 ────────────────────────
    # data_rows 마지막 행은 '합계' 행이므로 제외
    customer_list = [row[0] for row in data_rows[:-1]]

    analysis_sheet = 'Aging_분석'
    if analysis_sheet in wb_tgt.sheetnames:
        ws_analysis = wb_tgt[analysis_sheet]
    else:
        ws_analysis = wb_tgt.create_sheet(title=analysis_sheet)
        print(f'    [Aging] 시트 신규 생성: {analysis_sheet}')

    # 재실행 시 잔존 데이터 제거 (B4↑ 월 헤더 행 / A5↓ 거래처 리스트 열)
    _clear_range(ws_analysis, 4, 2, n_rows=1)
    _clear_range(ws_analysis, 5, 1, n_cols=1)

    month_list = headers[1:-1]  # '거래처명'·'합계' 제외한 월 헤더
    for c_idx, month in enumerate(month_list):
        ws_analysis.cell(row=4, column=2 + c_idx).value = month

    for r_idx, name in enumerate(customer_list):
        ws_analysis.cell(row=5 + r_idx, column=1).value = name
    print(f'    [Aging] {analysis_sheet} B4→ 월 {len(month_list)}개 / A5↓ 거래처 {len(customer_list)}개 주입')

    return len(data_rows)


# ─── 일반사항분석 주입 ───────────────────────────────────────────────────────

_YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


def inject_analysis_result(src_path, src_sheet, wb_tgt, tgt_sheet_name, start_cell, threshold):
    """pandas로 일반사항분석 파일을 읽어 변동비율·유의적변동 컬럼을 추가 후 조서에 주입.

    - 변동금액    = 당기잔액 − 전기잔액
    - 변동비율(%) = 변동금액 / 전기잔액 × 100  (전기잔액 0이면 None)
    - 유의적변동  = abs(변동금액) >= threshold → '유의' (threshold=0이면 판단 생략)
    유의적 행은 _YELLOW_FILL 으로 강조. 시트 없으면 신규 생성.
    Returns: 주입된 데이터 행 수.
    """
    def _read(engine):
        return pd.read_excel(src_path, sheet_name=src_sheet, engine=engine)

    try:
        df = _read('calamine')
    except Exception:
        df = _read('openpyxl')

    df = df.dropna(how='all').reset_index(drop=True)

    # ── 잔액 컬럼 탐색 ────────────────────────────────────────────────
    def find_col(*keywords):
        for c in df.columns:
            if any(kw in str(c) for kw in keywords):
                return c
        return None

    col_curr = find_col('당기', '금기', '현재')
    col_prev = find_col('전기', '전년', '비교', '기초')

    missing = [n for n, c in [('당기잔액', col_curr), ('전기잔액', col_prev)] if c is None]
    if missing:
        raise ValueError(f"잔액 컬럼을 찾을 수 없습니다: {', '.join(missing)}")

    # ── 숫자 정제 ─────────────────────────────────────────────────────
    for c in [col_curr, col_prev]:
        df[c] = pd.to_numeric(
            df[c].astype(str).str.replace(r'[,원\s]', '', regex=True),
            errors='coerce',
        ).fillna(0)

    # ── 변동 계산 ─────────────────────────────────────────────────────
    df['변동금액'] = df[col_curr] - df[col_prev]
    df['변동비율(%)'] = df.apply(
        lambda r: round(r['변동금액'] / r[col_prev] * 100, 1) if r[col_prev] != 0 else None,
        axis=1,
    )
    df['유의적변동'] = df['변동금액'].abs().apply(
        lambda v: '유의' if threshold > 0 and v >= threshold else ''
    )

    # ── 대상 시트 확보 ────────────────────────────────────────────────
    if tgt_sheet_name in wb_tgt.sheetnames:
        ws = wb_tgt[tgt_sheet_name]
    else:
        ws = wb_tgt.create_sheet(title=tgt_sheet_name)
        print(f'    [Analysis] 시트 신규 생성: {tgt_sheet_name}')

    start_row, start_col = _parse_cell(start_cell)

    # 재실행 시 이전 결과보다 행이 줄어들 경우 잔존 데이터·강조 서식 제거
    _clear_range(ws, start_row, start_col, n_rows=1)
    _clear_range(ws, start_row + 1, start_col, reset_fill=True)

    # ── 헤더 주입 ─────────────────────────────────────────────────────
    for c_idx, col_name in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + c_idx).value = col_name

    # ── 데이터 주입 + 유의적 행 강조 ─────────────────────────────────
    sig_count = 0
    for r_idx, (_, row_data) in enumerate(df.iterrows(), start=1):
        is_sig = row_data['유의적변동'] == '유의'
        if is_sig:
            sig_count += 1
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx)
            cell.value = None if pd.isna(val) else val
            if is_sig:
                cell.fill = _YELLOW_FILL

    print(f'    [Analysis] 유의적 변동 {sig_count}행 강조'
          + (f' (기준금액 {threshold:,.0f}원 이상)' if threshold > 0 else ' (기준금액 미설정)'))
    return len(df)


# ─── AI 계정별 검토결과 주입 ─────────────────────────────────────────────────

def inject_ai_result(src_path, src_sheet, wb_tgt, tgt_sheet_name, start_cell):
    """journal_analyzer 메뉴26(AI계정별분석_실행)의 AI검토결과 표를 감사조서에 주입.

    '위험평가'=='높음' 또는 '결론'=='추가확인필요'인 행은 _YELLOW_FILL 강조.
    시트 없으면 신규 생성. Returns: 주입된 데이터 행 수.
    """
    def _read(engine):
        return pd.read_excel(src_path, sheet_name=src_sheet, engine=engine)

    try:
        df = _read('calamine')
    except Exception:
        df = _read('openpyxl')

    df = df.dropna(how='all').reset_index(drop=True)

    if tgt_sheet_name in wb_tgt.sheetnames:
        ws = wb_tgt[tgt_sheet_name]
    else:
        ws = wb_tgt.create_sheet(title=tgt_sheet_name)
        print(f'    [AI] 시트 신규 생성: {tgt_sheet_name}')

    start_row, start_col = _parse_cell(start_cell)

    # 재실행 시 이전 결과보다 행이 줄어들 경우 잔존 데이터·강조 서식 제거
    _clear_range(ws, start_row, start_col, n_rows=1)
    _clear_range(ws, start_row + 1, start_col, reset_fill=True)

    # ── 헤더 주입 ─────────────────────────────────────────────────────
    for c_idx, col_name in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + c_idx).value = col_name

    # ── 데이터 주입 + 위험/추가확인 행 강조 ──────────────────────────
    sig_count = 0
    for r_idx, (_, row_data) in enumerate(df.iterrows(), start=1):
        is_flag = (str(row_data.get('위험평가', '')).strip() == '높음'
                   or str(row_data.get('결론', '')).strip() == '추가확인필요')
        if is_flag:
            sig_count += 1
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx)
            cell.value = None if pd.isna(val) else val
            if is_flag:
                cell.fill = _YELLOW_FILL

    print(f'    [AI] 위험평가 높음/추가확인필요 {sig_count}행 강조')
    return len(df)


# ─── 리스 스케줄 주입 ────────────────────────────────────────────────────────

def inject_lease_schedule(src_path, src_sheet, wb_tgt, tgt_sheet_name, start_cell):
    """리스 스케줄 요약(계약별 요약 시트)을 감사조서 사용권자산_리스부채 시트에 주입.

    매핑 규칙:
      I  = 납부액(당기)     ← {year}년 리스료지급
      J  = 리스부채(유동)   ← 유동성대체대상액
      K  = 리스부채(비유동) ← 비유동성리스부채잔액
      L  = 이자비용(당기)   ← {year}년 이자비용
      P  = 사용권자산취득가  ← 사용권자산(최초)
      Q  = 상각누계         ← 사용권자산 상각누계
      R  = 감가상각비(당기)  ← {year}년 감가상각비
    행 매칭: 감사조서 B열(개시일) == 리스개시일
    """
    # year: 파일명 lease_schedule_{company}_{year}.xlsx 에서 추출
    m = re.search(r'_(\d{4})\.xlsx$', os.path.basename(src_path), re.IGNORECASE)
    year = m.group(1) if m else str(pd.Timestamp.now().year)

    df = pd.read_excel(src_path, sheet_name=src_sheet)
    if df.empty:
        raise ValueError('리스 스케줄 요약 시트가 비어 있습니다.')

    resolved_tgt = resolve_sheet(wb_tgt.sheetnames, tgt_sheet_name)
    if not resolved_tgt:
        raise ValueError(f'대상 시트 없음: {tgt_sheet_name}')
    ws = wb_tgt[resolved_tgt]

    # start_cell(예: B6)에서 검색 시작 행 결정
    search_start_row, _ = _parse_cell(start_cell)

    # 컬럼 인덱스 (1-based)
    COL_B = 2   # 개시일
    COL_I = 9   # 납부액(당기)
    COL_J = 10  # 리스부채(유동)
    COL_K = 11  # 리스부채(비유동)
    COL_L = 12  # 이자비용(당기)
    COL_P = 16  # 사용권자산취득가
    COL_Q = 17  # 상각누계
    COL_R = 18  # 감가상각비(당기)

    def _v(row, key, default=0):
        val = row.get(key, default)
        return 0 if (val is None or (isinstance(val, float) and val != val)) else val

    matched = 0
    for _, contract in df.iterrows():
        start_date = contract.get('리스개시일')
        if pd.isna(start_date):
            continue
        start_ts = pd.Timestamp(start_date).date()

        for row_idx in range(search_start_row, search_start_row + 50):
            cell_b = ws.cell(row=row_idx, column=COL_B).value
            if cell_b is None:
                continue
            try:
                if pd.Timestamp(cell_b).date() == start_ts:
                    cid = contract.get('리스계약번호', '')
                    ws.cell(row=row_idx, column=COL_I).value = _v(contract, f'{year}년 리스료지급')
                    ws.cell(row=row_idx, column=COL_J).value = _v(contract, '유동성대체대상액')
                    ws.cell(row=row_idx, column=COL_K).value = _v(contract, '비유동성리스부채잔액')
                    ws.cell(row=row_idx, column=COL_L).value = _v(contract, f'{year}년 이자비용')
                    ws.cell(row=row_idx, column=COL_P).value = _v(contract, '사용권자산(최초)')
                    ws.cell(row=row_idx, column=COL_Q).value = _v(contract, '사용권자산 상각누계')
                    ws.cell(row=row_idx, column=COL_R).value = _v(contract, f'{year}년 감가상각비')
                    print(f'    [LEASE] 계약 {cid} ({start_ts}) → row{row_idx} 주입 완료')
                    matched += 1
                    break
            except Exception:
                continue

    unmatched = len(df) - matched
    if unmatched > 0:
        print(f'    [LEASE] 주의: {unmatched}건 매칭 실패 (개시일 불일치)')
    print(f'    [LEASE] {matched}/{len(df)}건 주입 완료 (연도: {year})')
    return matched


# ─── 이미지 복사 ─────────────────────────────────────────────────────────────

def _extract_first_image_zip(src_path, sheet_name):
    """xlsx ZIP 내부 drawing XML을 직접 파싱해 첫 번째 이미지 바이트와 표시 크기를 추출.

    ws._images 가 비어있는 경우(EMF 등)의 폴백용.
    Returns (img_bytes, ext_lower, width_px, height_px) 또는 (None, None, None, None).
    """
    NS_R   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    NS_REL = 'http://schemas.openxmlformats.org/package/2006/relationships'
    NS_SS  = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    NS_A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    def _tag(ns, local): return f'{{{ns}}}{local}'

    def _resolve(base, target):
        parts = (base.rsplit('/', 1)[0] + '/' + target).split('/')
        out = []
        for p in parts:
            if p == '..':
                if out: out.pop()
            elif p and p != '.':
                out.append(p)
        return '/'.join(out)

    def _rels(path):
        d, f = path.rsplit('/', 1)
        return f'{d}/_rels/{f}.rels'

    def _iter_rels(xml_el):
        """rels 파일의 Relationship 요소 반복 — 패키지 ns 우선, 없으면 ns 없는 태그."""
        items = list(xml_el.iter(_tag(NS_REL, 'Relationship')))
        return items if items else list(xml_el.iter('Relationship'))

    try:
        with zipfile.ZipFile(src_path, 'r') as zf:
            znames = set(zf.namelist())
            def rxl(p): return ET.fromstring(zf.read(p))

            # 1. workbook → 시트 파일 경로 (rels: NS_REL / sheet r:id: NS_R)
            rid_map = {r.get('Id'): r.get('Target')
                       for r in _iter_rels(rxl('xl/_rels/workbook.xml.rels'))}
            sheet_file = None
            for s in rxl('xl/workbook.xml').iter(_tag(NS_SS, 'sheet')):
                if s.get('name') == sheet_name:
                    sheet_file = _resolve('xl/workbook.xml', rid_map.get(s.get(_tag(NS_R, 'id')), ''))
                    break
            if not sheet_file or sheet_file not in znames:
                return None, None

            # 2. 시트 → drawing rId
            drawing_rid = None
            for el in rxl(sheet_file).iter():
                if el.tag.endswith('}drawing'):
                    drawing_rid = el.get(_tag(NS_R, 'id'))
                    break
            if not drawing_rid:
                return None, None

            # 3. 시트 rels → drawing 파일
            srels_path = _rels(sheet_file)
            if srels_path not in znames:
                return None, None
            drawing_file = None
            for r in _iter_rels(rxl(srels_path)):
                if r.get('Id') == drawing_rid:
                    drawing_file = _resolve(sheet_file, r.get('Target'))
                    break
            if not drawing_file or drawing_file not in znames:
                return None, None

            # 4. drawing → 첫 번째 blip rId + 표시 크기 (cx/cy in EMU)
            #    blip 없으면 Chart/Shape 객체일 가능성 → 'no_blip' 마커 반환
            NS_XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
            draw_tree = rxl(drawing_file)
            img_rid = None
            for blip in draw_tree.iter(_tag(NS_A, 'blip')):
                img_rid = blip.get(_tag(NS_R, 'embed'))
                if img_rid:
                    break
            if not img_rid:
                return None, 'no_blip', None, None

            # drawing XML에서 표시 크기 추출 (xdr:ext → oneCellAnchor, a:ext → spPr)
            EMU_PER_PX = 9525  # 914400 EMU/inch ÷ 96 DPI
            width_px = height_px = None
            for ext_el in draw_tree.iter(_tag(NS_XDR, 'ext')):
                cx, cy = ext_el.get('cx'), ext_el.get('cy')
                if cx and cy:
                    try:
                        width_px = round(int(cx) / EMU_PER_PX)
                        height_px = round(int(cy) / EMU_PER_PX)
                    except ValueError:
                        pass
                    break
            if width_px is None:
                for ext_el in draw_tree.iter(_tag(NS_A, 'ext')):
                    cx, cy = ext_el.get('cx'), ext_el.get('cy')
                    if cx and cy:
                        try:
                            width_px = round(int(cx) / EMU_PER_PX)
                            height_px = round(int(cy) / EMU_PER_PX)
                        except ValueError:
                            pass
                        break

            # 5. drawing rels → 이미지 파일
            drels_path = _rels(drawing_file)
            if drels_path not in znames:
                return None, None, None, None
            img_file = None
            for r in _iter_rels(rxl(drels_path)):
                if r.get('Id') == img_rid:
                    img_file = _resolve(drawing_file, r.get('Target'))
                    break
            if not img_file or img_file not in znames:
                return None, None, None, None

            ext = img_file.rsplit('.', 1)[-1].lower()
            return zf.read(img_file), ext, width_px, height_px

    except Exception as e:
        print(f'    [MOVE_IMAGE] ZIP 추출 오류: {e}')
        return None, None, None, None


def _remove_images_at(ws, start_cell):
    """대상 시트에서 start_cell과 동일한 anchor 위치의 기존 이미지를 제거한다.

    재실행 시 그래프가 중첩되지 않도록 새 이미지 추가 전에 호출한다.
    Returns: 제거된 이미지 수.
    """
    start_row, start_col = _parse_cell(start_cell)
    target_pos = (start_row - 1, start_col - 1)  # openpyxl anchor는 0-based

    def _anchor_pos(img):
        anchor = img.anchor
        if isinstance(anchor, str):
            try:
                r, c = _parse_cell(anchor)
                return (r - 1, c - 1)
            except ValueError:
                return None
        _from = getattr(anchor, '_from', None)
        return (_from.row, _from.col) if _from is not None else None

    images = getattr(ws, '_images', [])
    kept = [img for img in images if _anchor_pos(img) != target_pos]
    removed = len(images) - len(kept)
    ws._images = kept
    return removed


def inject_image(ws_src, src_path, src_sheet, ws_tgt, start_cell):
    """소스 시트의 첫 번째 이미지를 대상 시트의 start_cell 위치에 복사한다.

    시도 순서: ws._images → ZIP 직접 추출(PNG/JPEG) → win32com 후처리 예약.
    주입 전 동일 위치의 기존 이미지는 제거한다(재실행 시 중첩 방지).
    Returns (주입개수: int, win32com_필요: bool).
    """
    removed = _remove_images_at(ws_tgt, start_cell)
    if removed:
        print(f'    [MOVE_IMAGE] 기존 이미지 {removed}개 제거 (재주입 전)')

    # ── 1. ws._images 경로 (Pillow 필요) ────────────────────────────────
    if _PILLOW_OK and getattr(ws_src, '_images', None):
        orig = ws_src._images[0]
        new_img = XLImage(BytesIO(orig._data()))
        new_img.anchor = start_cell
        # 원본 표시 크기 보존 (없으면 PIL 기본값 유지)
        if getattr(orig, 'width', None):
            new_img.width = orig.width
        if getattr(orig, 'height', None):
            new_img.height = orig.height
        ws_tgt.add_image(new_img)
        print(f'    [MOVE_IMAGE] ws._images 경로로 복사 완료 ({new_img.width}×{new_img.height}px)')
        return 1, False

    print('    [MOVE_IMAGE] ws._images 비어있음 — ZIP 직접 추출 시도')

    # ── 2. ZIP/XML 직접 추출 ────────────────────────────────────────────
    img_bytes, ext, width_px, height_px = _extract_first_image_zip(src_path, src_sheet)

    if img_bytes is None:
        if ext == 'no_blip':
            # Drawing XML은 있으나 래스터 이미지 없음 → Chart/Shape → win32com 시도
            print('    [MOVE_IMAGE] Chart/Shape 객체 감지 — win32com 후처리로 전환')
            return 0, True
        print('    [MOVE_IMAGE] ZIP 추출 실패 — drawing 없음')
        return 0, False

    print(f'    [MOVE_IMAGE] ZIP 추출 성공 ({ext.upper()}, {len(img_bytes):,} bytes)')

    if ext in ('emf', 'wmf'):
        print(f'    [MOVE_IMAGE] {ext.upper()} 포맷은 openpyxl 미지원 — win32com 후처리로 전환')
        return 0, True

    new_img = XLImage(BytesIO(img_bytes))
    new_img.anchor = start_cell
    # drawing XML에서 추출한 원본 표시 크기 적용
    if width_px:
        new_img.width = width_px
    if height_px:
        new_img.height = height_px
    ws_tgt.add_image(new_img)
    size_info = f'{width_px}×{height_px}px' if width_px else '크기미확인'
    print(f'    [MOVE_IMAGE] ZIP 경로로 복사 완료 ({size_info})')
    return 1, False


def inject_image_win32com(src_path, src_sheet, tgt_path, tgt_sheet, start_cell):
    """win32com(Excel COM)으로 소스의 첫 번째 Shape를 대상 파일에 복사·붙여넣기.

    tgt_path 는 이미 저장된 _updated 파일이어야 한다.
    """
    try:
        import win32com.client
    except ImportError:
        raise RuntimeError('pywin32 미설치 — pip install pywin32')

    xl = win32com.client.Dispatch('Excel.Application')
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb_src = xl.Workbooks.Open(src_path)
        ws_s = next((wb_src.Sheets(i) for i in range(1, wb_src.Sheets.Count + 1)
                     if wb_src.Sheets(i).Name == src_sheet), None)
        if ws_s is None or ws_s.Shapes.Count == 0:
            wb_src.Close(False)
            return 0

        ws_s.Shapes(1).Copy()

        wb_tgt = xl.Workbooks.Open(tgt_path)
        ws_t = next((wb_tgt.Sheets(i) for i in range(1, wb_tgt.Sheets.Count + 1)
                     if wb_tgt.Sheets(i).Name == tgt_sheet), None)
        if ws_t is None:
            wb_src.Close(False)
            wb_tgt.Close(False)
            return 0

        # 기존 그래프 제거 (재실행 시 중첩 방지)
        target_addr = start_cell.upper().replace('$', '')
        for i in range(ws_t.Shapes.Count, 0, -1):
            try:
                addr = ws_t.Shapes(i).TopLeftCell.Address(False, False)
            except Exception:
                continue
            if addr.upper() == target_addr:
                ws_t.Shapes(i).Delete()

        ws_t.Range(start_cell).Select()
        ws_t.Paste()
        xl.CutCopyMode = False
        wb_tgt.Save()
        wb_tgt.Close(False)
        wb_src.Close(False)
        return 1
    finally:
        try: xl.Quit()
        except: pass


# ─── 경로 헬퍼 ───────────────────────────────────────────────────────────────

def updated_path(original_path):
    """파일명 뒤에 _updated 를 붙인 경로 반환 (이미 있으면 그대로)."""
    base, ext = os.path.splitext(original_path)
    return original_path if base.endswith('_updated') else f'{base}_updated{ext}'


# ─── 메인 ────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(description='매핑 리스트 기반 데이터 주입 엔진')
    parser.add_argument('company', help='회사명')
    parser.add_argument('--base', default=None,
                        help='루트 하위 기준 폴더 (예: --base journal_analyzer)')
    args = parser.parse_args()

    company    = args.company
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir   = os.path.normpath(os.path.join(script_dir, '..'))

    if args.base:
        company_dir = os.path.join(root_dir, args.base, company)
    else:
        company_dir = os.path.join(root_dir, company)

    audit_dir   = os.path.join(company_dir, '감사조서')
    results_dir = os.path.join(company_dir, 'results')
    raw_dir     = os.path.join(company_dir, 'raw_data')

    print(f'[{company}] 데이터 주입 엔진 시작')
    print(f'  감사조서 폴더 : {audit_dir}')
    print(f'  소스(results) : {results_dir}')
    print(f'  소스(raw_data): {raw_dir}')

    # ── 1. 매핑 파일 탐색 ────────────────────────────────────────────────────
    mapping_path = find_file_by_keyword(audit_dir, f'{company}_mapping_list')
    if not mapping_path:
        print(f'\n[오류] 매핑 파일 없음. 키워드: {company}_mapping_list  폴더: {audit_dir}')
        sys.exit(1)
    print(f'  매핑 파일     : {os.path.basename(mapping_path)}')

    # ── 2. 매핑 읽기 ─────────────────────────────────────────────────────────
    mapping_rows = load_mapping(mapping_path)
    print(f'  매핑 항목 수  : {len(mapping_rows)}건\n')

    # ── 3. 대상 워크북 캐시 (동일 파일 중복 로드 방지) ──────────────────────
    tgt_book_cache   = {}   # real_path → Workbook
    tgt_path_cache   = {}   # keyword   → real_path
    win32com_pending = []   # (src_path, src_sheet, tgt_updated_path, tgt_sheet, start_cell, label)

    errors  = []
    success = 0

    # ── 4. 매핑 처리 ─────────────────────────────────────────────────────────
    for row in mapping_rows:
        label      = row['label']
        src_kw     = row['src_kw']
        src_sheet  = row['src_sheet']
        tgt_kw     = row['tgt_kw']
        tgt_sheet  = row['tgt_sheet']
        start_cell = row['start_cell']
        src_range  = row['src_range']
        remarks    = row['remarks']
        threshold  = row['threshold']

        mode_tag = f' [{remarks}]' if remarks else ''
        print(f'  [{label}]{mode_tag} {src_kw}!{src_sheet} → {tgt_kw}!{tgt_sheet} @ {start_cell}')

        # ── 소스 파일 탐색 ─────────────────────────────────────────────────
        src_path = find_file_by_keyword([results_dir, raw_dir, company_dir], src_kw)
        if not src_path:
            msg = f'소스 파일 없음: {src_kw}'
            print(f'    [오류] {msg}')
            errors.append(f'[{label}] {msg}')
            continue
        print(f'    매칭 성공 (소스) : {src_kw}')
        print(f'                    → {os.path.relpath(src_path, company_dir)}')

        # ── pandas 직접 처리 조기 분기 (PIVOT_AGING / ANALYSIS_INJECT / LEASE_INJECT / AI_INJECT) ─
        if remarks in ('PIVOT_AGING', 'ANALYSIS_INJECT', 'LEASE_INJECT', 'AI_INJECT'):
            if tgt_kw not in tgt_path_cache:
                tgt_path = find_file_by_keyword(audit_dir, tgt_kw)
                if not tgt_path:
                    msg = f'대상 조서 파일 없음: {tgt_kw}'
                    print(f'    [오류] {msg}')
                    errors.append(f'[{label}] {msg}')
                    continue
                tgt_path_cache[tgt_kw] = tgt_path
                print(f'    매칭 성공 (대상) : {tgt_kw}')
                print(f'                    → {os.path.relpath(tgt_path, company_dir)}')
            else:
                tgt_path = tgt_path_cache[tgt_kw]
            if tgt_path not in tgt_book_cache:
                try:
                    tgt_book_cache[tgt_path] = load_workbook(tgt_path)
                except Exception as e:
                    msg = f'대상 파일 오픈 실패: {e}'
                    print(f'    [오류] {msg}')
                    errors.append(f'[{label}] {msg}')
                    continue
            wb_tgt = tgt_book_cache[tgt_path]
            try:
                if remarks == 'PIVOT_AGING':
                    print(f'    [Aging] 피벗 생성 → {tgt_sheet} @ {start_cell}')
                    injected = inject_pivot_aging(src_path, src_sheet, wb_tgt, tgt_sheet, start_cell)
                    print(f'    [완료] 피벗 {injected}행 주입')
                elif remarks == 'LEASE_INJECT':
                    print(f'    [Lease] 리스 스케줄 주입 → {tgt_sheet} @ {start_cell}')
                    injected = inject_lease_schedule(src_path, src_sheet, wb_tgt, tgt_sheet, start_cell)
                    print(f'    [완료] 리스 {injected}건 주입')
                elif remarks == 'AI_INJECT':
                    print(f'    [AI] AI검토결과 주입 → {tgt_sheet} @ {start_cell}')
                    injected = inject_ai_result(src_path, src_sheet, wb_tgt, tgt_sheet, start_cell)
                    print(f'    [완료] AI검토결과 {injected}행 주입')
                else:  # ANALYSIS_INJECT
                    print(f'    [Analysis] 변동분석 주입 → {tgt_sheet} @ {start_cell}'
                          + (f'  기준금액: {threshold:,.0f}' if threshold else ''))
                    injected = inject_analysis_result(
                        src_path, src_sheet, wb_tgt, tgt_sheet, start_cell, threshold)
                    print(f'    [완료] 분석결과 {injected}행 주입')
                success += 1
            except Exception as e:
                msg = f'데이터 주입 오류: {e}'
                print(f'    [오류] {msg}')
                errors.append(f'[{label}] {msg}')
            continue

        # ── 소스 시트 로드 ─────────────────────────────────────────────────
        try:
            if remarks == 'MOVE_IMAGE':
                # read_only 모드에서는 ws._images 가 채워지지 않으므로 full 모드로 열기
                wb_src = load_workbook(src_path, data_only=True)
            else:
                wb_src = load_workbook(src_path, data_only=True, read_only=True)
        except Exception as e:
            msg = f'소스 파일 오픈 실패: {e}'
            print(f'    [오류] {msg}')
            errors.append(f'[{label}] {msg}')
            continue

        resolved_src = resolve_sheet(wb_src.sheetnames, src_sheet)
        if not resolved_src:
            msg = f'소스 시트 없음: {src_sheet}  (파일: {os.path.basename(src_path)})'
            print(f'    [오류] {msg}')
            errors.append(f'[{label}] {msg}')
            wb_src.close()
            continue
        if resolved_src != src_sheet:
            print(f'    시트 매칭 (소스) : {src_sheet} → {resolved_src}')
        ws_src = wb_src[resolved_src]

        # ── 대상 파일 탐색 (캐시) ─────────────────────────────────────────
        if tgt_kw not in tgt_path_cache:
            tgt_path = find_file_by_keyword(audit_dir, tgt_kw)
            if not tgt_path:
                msg = f'대상 조서 파일 없음: {tgt_kw}'
                print(f'    [오류] {msg}')
                errors.append(f'[{label}] {msg}')
                wb_src.close()
                continue
            tgt_path_cache[tgt_kw] = tgt_path
            print(f'    매칭 성공 (대상) : {tgt_kw}')
            print(f'                    → {os.path.relpath(tgt_path, company_dir)}')
        else:
            tgt_path = tgt_path_cache[tgt_kw]

        # ── 대상 워크북 로드 (캐시) ───────────────────────────────────────
        if tgt_path not in tgt_book_cache:
            try:
                tgt_book_cache[tgt_path] = load_workbook(tgt_path, keep_links=False)
            except Exception as e:
                msg = f'대상 파일 오픈 실패: {e}'
                print(f'    [오류] {msg}')
                errors.append(f'[{label}] {msg}')
                wb_src.close()
                continue

        wb_tgt = tgt_book_cache[tgt_path]

        # ── 대상 시트 확인 (없으면 신규 생성) ───────────────────────────────
        resolved_tgt = resolve_sheet(wb_tgt.sheetnames, tgt_sheet)
        if not resolved_tgt:
            print(f'    [안내] 대상 시트 없음 → 신규 생성: {tgt_sheet}')
            wb_tgt.create_sheet(tgt_sheet)
            resolved_tgt = tgt_sheet
        if resolved_tgt != tgt_sheet:
            print(f'    시트 매칭 (대상) : {tgt_sheet} → {resolved_tgt}')
        ws_tgt = wb_tgt[resolved_tgt]

        # ── 데이터 주입 ───────────────────────────────────────────────────
        try:
            if remarks == 'MOVE_IMAGE':
                print(f'    [Image] 이미지 복사 → {tgt_sheet} @ {start_cell}')
                injected, need_win32 = inject_image(ws_src, src_path, resolved_src, ws_tgt, start_cell)
                if need_win32:
                    win32com_pending.append((src_path, resolved_src,
                                             updated_path(tgt_path), resolved_tgt,
                                             start_cell, label))
                success += 1
                suffix = ' (win32com 후처리 예정)' if need_win32 else ''
                print(f'    [완료] 이미지 {injected}개 복사{suffix}')
            else:
                if src_range:
                    print(f'    소스 범위 지정 : {src_range}')
                injected = inject_data(ws_src, ws_tgt, start_cell, src_range or None)
                success += 1
                print(f'    [완료] {injected}개 셀 주입')
        except Exception as e:
            msg = f'데이터 주입 오류: {e}'
            print(f'    [오류] {msg}')
            errors.append(f'[{label}] {msg}')

        wb_src.close()

    # ── 5. 결과 저장 ─────────────────────────────────────────────────────────
    print('\n─── 저장 ───')
    saved_paths = []
    for tgt_path, wb in tgt_book_cache.items():
        out_path = updated_path(tgt_path)
        try:
            wb.save(out_path)
            print(f'  저장 완료: {os.path.relpath(out_path, company_dir)}')
            saved_paths.append(out_path)
        except Exception as e:
            print(f'  [오류] 저장 실패 ({os.path.basename(tgt_path)}): {e}')

    # ── 5-1. xlwings 재저장 (Named Range·Drawing 손상 복구) ──────────────────
    if _XLWINGS_OK and saved_paths:
        print('\n─── xlwings 재저장 (XML 정합성 복구) ───')
        xl_app = xw.App(visible=False, add_book=False)
        try:
            for out_path in saved_paths:
                try:
                    wb_xw = xl_app.books.open(out_path)
                    wb_xw.save()
                    wb_xw.close()
                    print(f'  재저장 완료: {os.path.basename(out_path)}')
                except Exception as e:
                    print(f'  [경고] xlwings 재저장 실패 ({os.path.basename(out_path)}): {e}')
        finally:
            xl_app.quit()

    # ── 6. win32com 후처리 (EMF/WMF 이미지) ─────────────────────────────────
    if win32com_pending:
        print('\n─── win32com 이미지 후처리 ───')
        for src_p, src_s, tgt_p, tgt_s, cell, lbl in win32com_pending:
            if not os.path.exists(tgt_p):
                print(f'  [{lbl}] 대상 파일 없음 (저장 실패?): {os.path.basename(tgt_p)}')
                continue
            try:
                cnt = inject_image_win32com(src_p, src_s, tgt_p, tgt_s, cell)
                print(f'  [{lbl}] win32com 복사 완료 ({cnt}개)')
            except Exception as e:
                print(f'  [{lbl}] win32com 오류: {e}')
                errors.append(f'[{lbl}] win32com 이미지 오류: {e}')

    # ── 7. 요약 ──────────────────────────────────────────────────────────────
    print('\n─── 작업 요약 ───')
    print(f'  성공 : {success}/{len(mapping_rows)}건')
    if errors:
        print(f'  오류 ({len(errors)}건):')
        for err in errors:
            print(f'    - {err}')
    else:
        print('  오류 없음')


if __name__ == '__main__':
    main()
