#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""현금및현금성자산 감사조서 리드 생성기
Usage: python cash_lead_generator.py <company_name>
"""

import sys, os, re, glob
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import xlrd

# ─── 계정 분류 상수 ────────────────────────────────────────────────────
CASH_MAJORS = ['현금', '보통예금', '당좌예금', '외화예금', '기업자유예금']
FIN_MAJORS  = ['정기예금', '단기금융상품', '단기투자증권', '장기금융상품', '적금', '부금']

PRIOR_KEYWORDS = ['현금', '예금', '금융', '당좌', '외화', '부금', '적금']
PRIOR_EXCLUDES = ['기부금', '보증', '보험', '이자수익', '연금']

NUM_FMT  = '#,##0'
ACCT_FMT = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'


# ─── 유틸리티 ─────────────────────────────────────────────────────────
def normalize_name(name: str) -> str:
    """계정명 정규화: 번호접두사/코드제거, '/' → ' '"""
    name = re.sub(r'^\d+_', '', str(name).strip())
    name = re.sub(r'\s*\(\d+\)\s*$', '', name)
    return name.replace('/', ' ').strip()


def classify_account(name: str):
    """(category, major) 반환. category: 'cash' | 'fin' | None"""
    if '기부금' in name or '보증금' in name:
        return None, None
    for major in CASH_MAJORS:
        if name.startswith(major):
            return 'cash', major
    for major in FIN_MAJORS:
        if name.startswith(major):
            return 'fin', major
        # '은행/상품명' 형식 처리
        if '/' in name:
            after = name.split('/', 1)[1]
            if major in after:
                return 'fin', major
        if major in name:
            return 'fin', major
    return None, None


def find_file(directory: str, *keywords) -> str:
    if not os.path.isdir(directory):
        return ''
    for fname in sorted(os.listdir(directory)):
        if all(k in fname for k in keywords):
            return os.path.join(directory, fname)
    return ''


# ─── 데이터 로딩 ─────────────────────────────────────────────────────
def load_prior_balances(prior_file: str) -> dict:
    """전기 원장 파일에서 {normalized_name: 기말잔액}"""
    result = {}
    if not prior_file or not os.path.exists(prior_file):
        return result
    try:
        if prior_file.lower().endswith('.xls'):
            wb = xlrd.open_workbook(prior_file)
            sheet_names = wb.sheet_names()
        else:
            wb_ox = load_workbook(prior_file, data_only=True, read_only=True)
            sheet_names = wb_ox.sheetnames

        for shname in sheet_names:
            norm = normalize_name(shname)
            if not any(k in norm for k in PRIOR_KEYWORDS):
                continue
            if any(e in norm for e in PRIOR_EXCLUDES):
                continue

            last_balance = None
            if prior_file.lower().endswith('.xls'):
                ws = wb.sheet_by_name(shname)
                for r in range(ws.nrows - 1, 0, -1):
                    v = ws.cell_value(r, 6)
                    if isinstance(v, (int, float)):
                        last_balance = float(v)
                        break
            else:
                ws = wb_ox[shname]
                for row in reversed(list(ws.iter_rows(min_row=2, values_only=True))):
                    v = row[6] if len(row) > 6 else None
                    if isinstance(v, (int, float)):
                        last_balance = float(v)
                        break

            if last_balance is not None and last_balance != 0:
                result[norm] = last_balance
    except Exception as e:
        print(f'  [경고] 전기 파일 오류: {e}')
    return result


def load_current_accounts(current_file: str) -> dict:
    """일반사항분析 → {name: {cat, major, 기초, 차변, 대변, 기말}}"""
    result = {}
    if not current_file or not os.path.exists(current_file):
        return result
    try:
        wb = load_workbook(current_file, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            name = str(row[0]).strip()
            cat, major = classify_account(name)
            if cat:
                result[name] = {
                    'cat': cat, 'major': major,
                    '기초': float(row[2] or 0),
                    '차변': float(row[3] or 0),
                    '대변': float(row[4] or 0),
                    '기말': float(row[6] or 0),
                }
    except Exception as e:
        print(f'  [경고] 당기 파일 오류: {e}')
    return result


def build_sections(current_accounts: dict, prior_balances: dict):
    """
    (cash_list, fin_list) 반환.
    각 항목: (name, 당기기말, 전기기말, 차변합계, 대변합계)
    """
    cash_list, fin_list = [], []
    processed = set()

    for name, data in current_accounts.items():
        norm = normalize_name(name)
        # 전기잔액: 정규화명으로 prior 조회 → 없으면 기초잔액
        prior = prior_balances.get(norm, data['기초'])
        processed.add(norm)

        if data['기말'] == 0 and prior == 0 and data['기초'] == 0:
            continue

        entry = (name, data['기말'], prior, data['차변'], data['대변'])
        if data['cat'] == 'cash':
            cash_list.append(entry)
        else:
            fin_list.append(entry)

    # 전기에만 존재하는 계정 추가
    for norm, bal in prior_balances.items():
        if norm in processed or abs(bal) < 1000:
            continue
        cat, major = classify_account(norm)
        if not cat:
            continue
        entry = (norm, 0, bal, 0, 0)
        if cat == 'cash':
            cash_list.append(entry)
        else:
            fin_list.append(entry)

    cash_list.sort(key=lambda x: -abs(x[1]))
    fin_list.sort(key=lambda x: -abs(x[1]))
    return cash_list, fin_list


# ─── 워크시트 편집 헬퍼 ──────────────────────────────────────────────
def copy_row_style(ws, src_row: int, dst_row: int):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)


def adjust_section(ws, data_start: int, n_template: int, n_needed: int):
    """섹션 데이터 행 수를 n_needed로 조정"""
    delta = n_needed - n_template
    if delta > 0:
        insert_at = data_start + n_template
        ws.insert_rows(insert_at, delta)
        for i in range(delta):
            copy_row_style(ws, data_start, insert_at + i)
    elif delta < 0:
        ws.delete_rows(data_start + n_needed, -delta)


def write_data_rows(ws, data_start: int, accounts: list):
    if not accounts:
        # 빈 섹션: 템플릿 잔류 값 제거
        ws.cell(row=data_start, column=1).value = None
        ws.cell(row=data_start, column=2).value = None
        ws.cell(row=data_start, column=3).value = None
        return
    for i, row_data in enumerate(accounts):
        name, curr, prev = row_data[0], row_data[1], row_data[2]
        r = data_start + i
        ws.cell(row=r, column=1).value = name
        bc = ws.cell(row=r, column=2)
        cc = ws.cell(row=r, column=3)
        bc.value = curr
        cc.value = prev
        bc.number_format = NUM_FMT
        cc.number_format = NUM_FMT


def write_sum_row(ws, sum_row: int, data_start: int, data_end: int):
    ws.cell(row=sum_row, column=1).value = '합계'
    b = ws.cell(row=sum_row, column=2)
    c = ws.cell(row=sum_row, column=3)
    b.value = f'=SUM(B{data_start}:B{data_end})'
    c.value = f'=SUM(C{data_start}:C{data_end})'
    b.number_format = NUM_FMT
    c.number_format = NUM_FMT


def add_deposit_schedule(ws, start_row: int, cash_list: list):
    """예금명세 섹션 추가 — 당기 잔액이 있는 예금 계정만 표시"""
    deposit_kw = ['보통예금', '당좌예금', '외화예금']
    # 당기 잔액이 있는 예금 계정 필터링
    deposit_rows = [
        (name, curr, prev)
        for name, curr, prev, _, _ in cash_list
        if any(name.startswith(k) for k in deposit_kw) and curr != 0
    ]
    if not deposit_rows:
        return start_row

    ws.cell(row=start_row,     column=1).value = '예금명세>'
    ws.cell(row=start_row + 1, column=1).value = '구분'
    ws.cell(row=start_row + 1, column=2).value = '금융기관'
    ws.cell(row=start_row + 1, column=3).value = '계좌번호 (수기 입력)'
    ws.cell(row=start_row + 1, column=4).value = '당기 기말잔액'
    ws.cell(row=start_row + 1, column=5).value = '전기 기말잔액'
    ws.cell(row=start_row + 1, column=6).value = '비고'

    r = start_row + 2
    for name, curr, prev in deposit_rows:
        bank = _extract_bank_name(name)
        ws.cell(row=r, column=1).value = name
        ws.cell(row=r, column=2).value = bank if bank else None
        ws.cell(row=r, column=3).value = None       # 계좌번호: 수기 입력
        c4 = ws.cell(row=r, column=4)
        c4.value = curr
        c4.number_format = NUM_FMT
        c5 = ws.cell(row=r, column=5)
        c5.value = prev if prev != 0 else None
        c5.number_format = NUM_FMT
        r += 1

    ws.cell(row=r, column=1).value = '합계'
    s4 = ws.cell(row=r, column=4)
    s5 = ws.cell(row=r, column=5)
    s4.value = f'=SUM(D{start_row+2}:D{r-1})'
    s5.value = f'=SUM(E{start_row+2}:E{r-1})'
    s4.number_format = NUM_FMT
    s5.number_format = NUM_FMT
    return r


def _extract_bank_name(account_name: str) -> str:
    """계정명에서 금융기관명 추출"""
    # '보통예금/신한수금' → '신한'
    known = ['기은', '신한', '국민', '하나', '우리', '농협', '기업', 'IBK']
    if '/' in account_name:
        part = account_name.split('/', 1)[1]
        for b in known:
            if b in part:
                return b
        return part.split('/')[0].rstrip('수금').rstrip('대출').rstrip('보조금').strip()
    for b in known:
        if b in account_name:
            return b
    return ''


# ─── 메인 ────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print('Usage: python cash_lead_generator.py <company_name>')
        sys.exit(1)

    company = sys.argv[1]
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    company_dir = os.path.normpath(os.path.join(script_dir, '..', company))
    raw_dir     = os.path.join(company_dir, 'raw_data')
    results_dir = os.path.join(company_dir, 'results')
    template_path = os.path.join(script_dir, 'templet', 'temp_cash.xlsx')
    output_path   = os.path.join(company_dir, f'{company}_2025_현금및현금성자산_리드.xlsx')

    print(f'[{company}] 현금및현금성자산 리드 생성 시작')

    # ─── 파일 탐색 ────────────────────────────────────────────────
    prior_file   = find_file(raw_dir, '전기')
    current_file = (find_file(results_dir, '일반사항분析') or
                    find_file(results_dir, '일반사항분석'))

    print(f'  전기 파일: {os.path.basename(prior_file) if prior_file else "없음"}')
    print(f'  당기 파일: {os.path.basename(current_file) if current_file else "없음"}')

    # ─── 데이터 로딩 ──────────────────────────────────────────────
    prior_balances   = load_prior_balances(prior_file)
    current_accounts = load_current_accounts(current_file)

    print(f'  전기 계정 수: {len(prior_balances)}개')
    print(f'  당기 현금/금융 계정 수: {len(current_accounts)}개')

    # ─── 계정 리스트 구성 ─────────────────────────────────────────
    cash_list, fin_list = build_sections(current_accounts, prior_balances)
    n_cash = max(len(cash_list), 1)
    n_fin  = max(len(fin_list), 1)

    print(f'  현금及현금성자산 계정: {n_cash}개  단기금융자산 계정: {n_fin}개')

    # ─── 워크북 로드 ──────────────────────────────────────────────
    wb = load_workbook(template_path)
    ws = wb.active
    ws.title = '현금및현금성자산'

    # ─── 클라이언트 정보 ──────────────────────────────────────────
    ws['A3'] = company
    ws['B3'] = '2025.12.31'
    ws['C3'] = '현금및현금성자산 Lead Schedule'

    # ─── 템플릿 고정 행 위치 ──────────────────────────────────────
    CASH_DATA_START   = 15
    N_TMPL_CASH       = 8     # rows 15-22
    # 단기금융 섹션 (변경 전 기준)
    FIN_DATA_TMPL_START = 27
    N_TMPL_FIN        = 3     # rows 27-29
    # FIN_TMPL_SUM      = 30  (참고용)

    # ─── Step 1: 현금 섹션 행 수 조정 ────────────────────────────
    adjust_section(ws, CASH_DATA_START, N_TMPL_CASH, n_cash)
    delta_cash = n_cash - N_TMPL_CASH

    cash_data_end = CASH_DATA_START + n_cash - 1
    cash_sum_row  = cash_data_end + 1   # 합계 행 위치

    # ─── Step 2: 현금 데이터 작성 ────────────────────────────────
    write_data_rows(ws, CASH_DATA_START, cash_list)
    write_sum_row(ws, cash_sum_row, CASH_DATA_START, cash_data_end)

    # 요약 행 8 업데이트
    ws.cell(row=8, column=2).value = f'=C{cash_sum_row}'
    ws.cell(row=8, column=3).value = sum(a[3] for a in cash_list)
    ws.cell(row=8, column=4).value = sum(a[4] for a in cash_list)
    for col in (2, 3, 4):
        ws.cell(row=8, column=col).number_format = ACCT_FMT

    # ─── Step 3: 단기금융 섹션 행 수 조정 ───────────────────────
    # delta_cash 만큼 이미 행이 이동됨
    fin_data_start = FIN_DATA_TMPL_START + delta_cash

    adjust_section(ws, fin_data_start, N_TMPL_FIN, n_fin)
    delta_fin = n_fin - N_TMPL_FIN

    fin_data_end = fin_data_start + n_fin - 1
    fin_sum_row  = fin_data_end + 1

    # ─── Step 4: 단기금융 데이터 작성 ────────────────────────────
    write_data_rows(ws, fin_data_start, fin_list)
    write_sum_row(ws, fin_sum_row, fin_data_start, fin_data_end)

    # 요약 행 9 업데이트
    ws.cell(row=9, column=2).value = f'=C{fin_sum_row}'
    ws.cell(row=9, column=3).value = sum(a[3] for a in fin_list)
    ws.cell(row=9, column=4).value = sum(a[4] for a in fin_list)
    for col in (2, 3, 4):
        ws.cell(row=9, column=col).number_format = ACCT_FMT

    # 단기금융 합계 행 (원래 row 30 위치): SUM 공식 재작성
    # (adjust_section으로 이미 이동된 상태 → fin_sum_row에 작성 완료)

    # ─── Step 5: 예금 명세표 추가 ────────────────────────────────
    sched_start = fin_sum_row + 2
    add_deposit_schedule(ws, sched_start, cash_list)

    # ─── 틀 고정 ─────────────────────────────────────────────────
    ws.freeze_panes = 'A7'

    # ─── 저장 ────────────────────────────────────────────────────
    wb.save(output_path)

    # 검증 출력
    curr_total = sum(a[1] for a in cash_list)
    prev_total = sum(a[2] for a in cash_list)
    print(f'\n  현금및현금성자산  당기: {curr_total:>18,.0f}  전기: {prev_total:>18,.0f}')
    curr_fin = sum(a[1] for a in fin_list)
    prev_fin = sum(a[2] for a in fin_list)
    print(f'  단기금융자산      당기: {curr_fin:>18,.0f}  전기: {prev_fin:>18,.0f}')
    print(f'\n  저장 완료: {output_path}')


if __name__ == '__main__':
    main()
