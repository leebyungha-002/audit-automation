#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""매핑 공통 라이브러리

다른 엔진(data_injector.py, cash_lead_generator.py 등)에서 import하여 사용.

주요 기능
- find_file_by_keyword : 날짜 토큰을 무시한 스마트 파일 탐색
- resolve_sheet        : 시트명 키워드 매칭
- load_mapping         : <회사>_mapping_list*.xlsx 파싱
- inject_data          : 소스 시트 → 대상 시트 값 주입 (서식 보존)
- updated_path         : _updated 저장 경로 생성
"""

import os
import re
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# ─── 파일명 정규화 ────────────────────────────────────────────────────────────

def _normalize(fname_no_ext):
    """YYYYMMDD(8자리) 날짜 토큰만 제거 후 연속 _ 정리."""
    s = re.sub(r'_\d{8}(?=_|$)', '', fname_no_ext)
    return re.sub(r'_+', '_', s).strip('_')


def _keyword_matches(keyword, normalized_fname):
    """keyword가 정규화된 파일명 내에 _ 경계로 정확히 포함되는지 확인.

    keyword='dae_il_외상매출금'
      normalized='dae_il_외상매출금'               → True  (정확일치)
      normalized='dae_il_외상매출금_상세'           → True  (suffix 허용)
      normalized='dae_il_벤포드_외상매출금'         → False (앞에 다른 토큰)
      normalized='dae_il_월별트렌드분析_외상매출금' → False
    """
    text    = '_' + normalized_fname + '_'
    pattern = '_' + re.escape(keyword) + '_'
    return bool(re.search(pattern, text))


# ─── 파일 탐색 ────────────────────────────────────────────────────────────────

def find_file_by_keyword(directories, keyword, exclude_suffixes=None):
    """keyword를 파일명(날짜 정규화 후)에서 _ 경계로 탐색.

    - directories: 문자열 또는 리스트 — 앞쪽 폴더부터 우선 탐색
    - 여러 개 발견 시 가장 최근 수정 파일 반환
    - exclude_suffixes: 파일명에 포함되면 제외 (기본: ~$, _updated)
    반환: 절대 경로 문자열, 없으면 None
    """
    if isinstance(directories, str):
        directories = [directories]
    if exclude_suffixes is None:
        exclude_suffixes = ['~$', '_updated']

    matches = []
    for directory in directories:
        if not os.path.isdir(directory):
            continue
        for fname in sorted(os.listdir(directory)):
            if any(ex in fname for ex in exclude_suffixes):
                continue
            if not fname.lower().endswith('.xlsx'):
                continue
            normalized = _normalize(os.path.splitext(fname)[0])
            if _keyword_matches(keyword, normalized):
                matches.append(os.path.join(directory, fname))

    if not matches:
        return None
    if len(matches) == 1:
        return matches[0]

    matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    names = [os.path.basename(p) for p in matches]
    print(f"    [주의] '{keyword}' 키워드로 {len(matches)}개 파일 발견:")
    for n in names:
        print(f"           {n}")
    print(f"           → 최근 파일 선택: {names[0]}")
    return matches[0]


# ─── 시트 탐색 ────────────────────────────────────────────────────────────────

def resolve_sheet(sheetnames, keyword):
    """정확히 일치 → keyword 포함 첫 번째 시트 순으로 탐색.

    반환: 시트명 문자열, 없으면 None
    """
    if keyword in sheetnames:
        return keyword
    matched = [s for s in sheetnames if keyword in s]
    return matched[0] if matched else None


# ─── 셀 좌표 파싱 ─────────────────────────────────────────────────────────────

def parse_cell(cell_ref):
    """'A7' → (row=7, col=1)  /  대소문자 무관."""
    m = re.match(r'^([A-Za-z]+)(\d+)$', cell_ref.strip())
    if not m:
        raise ValueError(f"잘못된 셀 좌표: {cell_ref}")
    return int(m.group(2)), column_index_from_string(m.group(1).upper())


# ─── 데이터 주입 ──────────────────────────────────────────────────────────────

def inject_data(ws_src, ws_tgt, start_cell):
    """ws_src 전체를 ws_tgt 의 start_cell 부터 값만 주입 (서식·수식 보존).

    반환: 주입된 셀 수
    """
    start_row, start_col = parse_cell(start_cell)
    count = 0
    for r_idx, row in enumerate(ws_src.iter_rows(values_only=True)):
        for c_idx, value in enumerate(row):
            if value is not None:
                ws_tgt.cell(row=start_row + r_idx,
                            column=start_col + c_idx).value = value
                count += 1
    return count


# ─── 매핑 파일 로드 ──────────────────────────────────────────────────────────

def load_mapping(mapping_path):
    """<회사>_mapping_list*.xlsx 를 읽어 매핑 행 리스트 반환.

    각 행은 dict:
      label, src_kw, src_sheet, tgt_kw, tgt_sheet, start_cell
    필수 필드(src_kw / tgt_kw / start_cell) 가 빈 행은 자동 스킵.
    """
    wb  = load_workbook(mapping_path, data_only=True)
    ws  = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        padded = list(row) + [None] * 7
        label, src_kw, src_sheet, tgt_kw, tgt_sheet, start_cell = padded[:6]
        if not src_kw or not tgt_kw or not start_cell:
            continue
        rows.append({
            'label':      str(label      or '').strip(),
            'src_kw':     str(src_kw    ).strip(),
            'src_sheet':  str(src_sheet ).strip(),
            'tgt_kw':     str(tgt_kw    ).strip(),
            'tgt_sheet':  str(tgt_sheet ).strip(),
            'start_cell': str(start_cell).strip().upper(),
        })
    return rows


# ─── 경로 헬퍼 ───────────────────────────────────────────────────────────────

def updated_path(original_path):
    """파일명 뒤에 _updated 를 붙인 경로 반환 (이미 있으면 그대로)."""
    base, ext = os.path.splitext(original_path)
    return original_path if base.endswith('_updated') else f'{base}_updated{ext}'
