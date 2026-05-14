#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sheet_splitter/split_sheets.py

Usage:
    # [기본] 시트명 접두어 기준 그룹핑 — win32com 네이티브 복사
    python split_sheets.py
    python split_sheets.py --mode sheet --input 통합조서.xlsx

    # [컬럼값 기준 분리] — pandas/openpyxl 기반 (조서번호_시트명 컬럼 존재 시)
    python split_sheets.py --mode col --col 조서번호_시트명
"""

import sys
import os
import glob
import argparse
import re

# ── win32com (--mode sheet) ───────────────────────────────────────────────────
try:
    import win32com.client
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False

# ── pandas / openpyxl (--mode col) ───────────────────────────────────────────
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR  = os.path.join(BASE_DIR, 'input_data')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

# ── 공통 상수 ──────────────────────────────────────────────────────────────────
DEFAULT_TARGET_COL    = '조서번호_시트명'
DEFAULT_NAN_LABEL     = '기타_분류안됨'
OUTPUT_FILENAME_COL   = '조서분리완료_결과물.xlsx'
OUTPUT_FILENAME       = OUTPUT_FILENAME_COL   # 하위호환

NOTES_SHEET_KEYWORD   = '보고서주석'
OUTPUT_FILENAME_NOTES = '보고서주석.xlsx'

# Excel COM 상수
XL_PASTE_VALUES = -4163   # xlPasteValues
XL_EXCEL_LINKS  = 1       # xlExcelLinks
XL_XLSX_FORMAT  = 51      # xlOpenXMLWorkbook (.xlsx)

MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 8
HDR_EXTRA     = 4


# ── 파일 탐색 ──────────────────────────────────────────────────────────────────
def find_input_file(directory: str, filename: str | None = None) -> str:
    """input_data/ 에서 .xlsx 탐색. 파일명 지정 시 해당 파일, 없으면 최근 파일 자동 선택."""
    if filename:
        path = os.path.join(directory, filename)
        if os.path.isfile(path):
            return path
        raise FileNotFoundError(f'지정 파일 없음: {path}')

    candidates = sorted(
        [f for f in glob.glob(os.path.join(directory, '*.xlsx'))
         if not os.path.basename(f).startswith('~$')],
        key=os.path.getmtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(
            f'input_data/ 에 .xlsx 파일이 없습니다.\n경로: {directory}'
        )
    if len(candidates) > 1:
        print(f'[주의] 파일 {len(candidates)}개 발견 → 최근 파일 사용: '
              f'{os.path.basename(candidates[0])}')
    return candidates[0]


# ── 이름 정제 ──────────────────────────────────────────────────────────────────
def sanitize_sheet_name(name: str) -> str:
    """엑셀 시트명 제약: 금지 문자 제거, 31자 초과 절삭."""
    name = re.sub(r'[\\/:*?\[\]]', '_', str(name)).strip()
    return name[:31] if name else '시트'


def sanitize_filename(name: str) -> str:
    """Windows 파일명 제약: 시트명 금지 문자 + < > \" | 제거, 연속 _ 정리."""
    name = re.sub(r'[\\/:*?"<>|]', '_', str(name)).strip()
    name = re.sub(r'_+', '_', name).strip('_')
    return name[:60] if name else '파일'


# ── win32com 헬퍼 (--mode sheet) ──────────────────────────────────────────────
def _start_excel():
    """Excel COM 객체 초기화. 백그라운드 실행, 화면·경고 비활성화."""
    app = win32com.client.Dispatch('Excel.Application')
    app.Visible        = False
    app.DisplayAlerts  = False
    app.ScreenUpdating = False
    return app


def _copy_sheets_to_new_wb(src_wb, sheet_names: list[str], app):
    """
    sheet_names 를 새 통합문서에 이름 오름차순으로 네이티브 복사.

    - Sheet.Copy(After=...) : VBA의 Sheet.Copy After:=... 와 동일
    - 틀 고정, 셀 병합, 색상, 열 너비 등 서식 100% 유지
    - 새 통합문서의 기본 시트(Sheet1 등)는 복사 후 삭제
    """
    new_wb   = app.Workbooks.Add()
    defaults = [new_wb.Worksheets(i + 1).Name
                for i in range(new_wb.Worksheets.Count)]

    for name in sorted(sheet_names):
        src_wb.Worksheets(name).Copy(
            After=new_wb.Worksheets(new_wb.Worksheets.Count)
        )

    for name in defaults:
        try:
            new_wb.Worksheets(name).Delete()
        except Exception:
            pass

    return new_wb


def _formulas_to_values(wb):
    """
    모든 시트의 수식을 현재 계산값으로 직접 대체.
    UsedRange.Value = UsedRange.Value 로 일괄 처리 →
    외부 링크 포함 모든 수식 제거, 데이터 연결 오류 없음 보장.
    """
    for ws in wb.Worksheets:
        try:
            used = ws.UsedRange
            val  = used.Value
            if val is not None:
                used.Value = val
        except Exception as e:
            print(f'    [경고] 수식변환 실패 ({ws.Name}): {e}')


def _save_group_file(src_wb, sheet_list: list[str], out_path: str, app):
    """
    그룹 시트들을 새 파일로 저장하는 전체 파이프라인:
    네이티브 복사 → 수식/링크 제거(값으로 변환) → 저장 → 닫기
    """
    new_wb = None
    try:
        new_wb = _copy_sheets_to_new_wb(src_wb, sheet_list, app)
        _formulas_to_values(new_wb)
        new_wb.SaveAs(out_path, FileFormat=XL_XLSX_FORMAT)
    finally:
        if new_wb:
            try:
                new_wb.Close(SaveChanges=False)
            except Exception:
                pass


# ── 시트명 접두어 그룹핑 (--mode sheet) ───────────────────────────────────────
def group_by_sheet_prefix(
    input_file: str | None = None,
    sep:        str        = '_',
):
    """
    입력 파일의 시트를 '_' 앞 접두어 기준으로 그룹핑하고
    그룹마다 별도 xlsx 파일로 저장.

    [처리 흐름]
    1. 백그라운드 Excel 실행 → 원본 파일 열기 (ReadOnly)
    2. 시트명 '보고서주석' 발견 시 → 해당 시트 이후를 보고서주석.xlsx 로 분리
    3. 나머지 시트를 접두어별로 그룹핑
    4. 그룹마다 새 통합문서 생성 → 네이티브 Sheet.Copy → 수식/링크 제거 → 저장
    5. 모든 작업 후 Excel 종료 (try-finally 자원 해제)

    [출력 예]
    output/4300.xlsx   (4300_매출채권, 4300_연령분석, … 탭)
    output/재무제표.xlsx (재무제표_합잔_당기, 재무제표_합잔_전기, … 탭)
    output/보고서주석.xlsx
    """
    if not HAS_WIN32COM:
        raise RuntimeError(
            'pywin32 가 설치되어 있지 않습니다.\n'
            '설치: pip install pywin32'
        )

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.abspath(find_input_file(INPUT_DIR, input_file))
    print(f'\n  파일: {os.path.basename(filepath)}')

    app    = None
    src_wb = None
    try:
        app    = _start_excel()
        src_wb = app.Workbooks.Open(filepath, ReadOnly=True, UpdateLinks=False)

        all_sheets = [src_wb.Worksheets(i + 1).Name
                      for i in range(src_wb.Worksheets.Count)]

        # 보고서주석 분리
        notes_idx = next(
            (i for i, sn in enumerate(all_sheets) if NOTES_SHEET_KEYWORD in sn),
            None,
        )
        if notes_idx is not None:
            main_sheets  = all_sheets[:notes_idx]
            notes_sheets = all_sheets[notes_idx:]
            print(f'  총 시트 수     : {len(all_sheets)}개')
            print(f'  감사조서 시트  : {len(main_sheets)}개')
            print(f'  보고서주석 시트: {len(notes_sheets)}개')
        else:
            main_sheets  = list(all_sheets)
            notes_sheets = []
            print(f'  총 시트 수: {len(all_sheets)}개  (보고서주석 시트 없음)')

        # 접두어별 그룹핑
        groups: dict[str, list[str]] = {}
        for sn in main_sheets:
            prefix = sn.split(sep, 1)[0] if sep in sn else sn
            groups.setdefault(prefix, []).append(sn)

        print(f'\n  그룹 수: {len(groups)}개  →  output/ 폴더에 파일별 저장\n')

        # 감사조서 그룹 → 파일별 저장
        for prefix in sorted(groups.keys()):
            sheets_in_group = groups[prefix]
            out_name = sanitize_filename(prefix) + '.xlsx'
            out_path = os.path.join(OUTPUT_DIR, out_name)

            try:
                _save_group_file(src_wb, sheets_in_group, out_path, app)
                preview = ', '.join(sheets_in_group[:3])
                if len(sheets_in_group) > 3:
                    preview += f' 외 {len(sheets_in_group)-3}개'
                print(f'  [{out_name:<35}]  {len(sheets_in_group)}탭  ← {preview}')
            except Exception as e:
                print(f'  [오류] {out_name}: {e}')

        # 보고서주석 → 별도 파일
        if notes_sheets:
            out_path = os.path.join(OUTPUT_DIR, OUTPUT_FILENAME_NOTES)
            try:
                _save_group_file(src_wb, notes_sheets, out_path, app)
                print(f'\n  [{OUTPUT_FILENAME_NOTES}]  {len(notes_sheets)}탭 저장 완료')
            except Exception as e:
                print(f'  [오류] {OUTPUT_FILENAME_NOTES}: {e}')

        print(f'\n  전체 완료 — {len(groups)}개 파일 생성')

    finally:
        if src_wb:
            try:
                src_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if app:
            try:
                app.Quit()
            except Exception:
                pass


# ── pandas/openpyxl 헬퍼 (--mode col) ────────────────────────────────────────
def _dedup_columns(headers: list[str]) -> list[str]:
    """중복 컬럼명에 .1 .2 ... 접미사를 붙여 유일하게 만든다."""
    seen: dict[str, int] = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f'{h}.{seen[h]}')
        else:
            seen[h] = 0
            result.append(h)
    return result


def load_as_values(filepath: str, sheet_name: str | None = None) -> pd.DataFrame:
    """openpyxl data_only=True 로 수식 배제, 값만 DataFrame 반환."""
    print(f'  파일 로드 중: {os.path.basename(filepath)}')
    wb = load_workbook(filepath, data_only=True, read_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f'시트 "{sheet_name}" 없음. 사용 가능: {wb.sheetnames}'
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active
        print(f'  활성 시트: {ws.title}')

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError('시트에 데이터가 없습니다.')

    raw_headers = [
        str(c).strip() if c is not None else f'열_{i}'
        for i, c in enumerate(rows[0], 1)
    ]
    headers = _dedup_columns(raw_headers)
    data    = [list(r) for r in rows[1:] if any(c is not None for c in r)]
    df      = pd.DataFrame(data, columns=headers)
    print(f'  → {len(df):,}행 × {len(df.columns)}열 로드 완료')
    return df


def apply_header_style(ws):
    """1행 헤더: 진한 파랑 배경 + 흰 볼드 + 틀 고정."""
    HDR_FILL = PatternFill('solid', fgColor='1F497D')
    HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
    for cell in ws[1]:
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'


def auto_col_width(ws,
                   min_width: int = MIN_COL_WIDTH,
                   max_width: int = MAX_COL_WIDTH,
                   extra:     int = HDR_EXTRA):
    """컬럼 데이터 최대 길이 기반 열 너비 자동 조절. 한글은 2폭으로 계산."""
    for col_cells in ws.iter_cols():
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            text  = str(cell.value)
            width = sum(2 if ord(c) > 0xFF else 1 for c in text)
            if width > max_len:
                max_len = width
        ws.column_dimensions[col_letter].width = min(
            max(max_len + extra, min_width), max_width
        )


# ── 컬럼값 기준 분리 (--mode col) ────────────────────────────────────────────
def split_sheets(
    target_col: str        = DEFAULT_TARGET_COL,
    nan_label:  str        = DEFAULT_NAN_LABEL,
    input_file: str | None = None,
    sheet_name: str | None = None,
):
    """단일 시트의 특정 컬럼값 기준으로 행 분리 → 컬럼값별 시트 저장."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = find_input_file(INPUT_DIR, input_file)
    df = load_as_values(filepath, sheet_name)

    if target_col not in df.columns:
        close = [c for c in df.columns
                 if target_col in str(c) or str(c) in target_col]
        if close:
            print(f'[안내] 컬럼 "{target_col}" 없음 → 유사 컬럼 "{close[0]}" 사용')
            target_col = close[0]
        else:
            raise ValueError(
                f'분리 기준 컬럼 "{target_col}"을 찾을 수 없습니다.\n'
                f'사용 가능한 컬럼 목록:\n  {list(df.columns)}'
            )

    df[target_col] = df[target_col].fillna(nan_label).astype(str).str.strip()
    df.loc[df[target_col] == '', target_col] = nan_label

    unique_vals = sorted(df[target_col].unique())
    print(f'\n  분리 기준 컬럼 : {target_col}')
    print(f'  고유값 수       : {len(unique_vals)}개')
    print(f'  출력 파일       : {OUTPUT_FILENAME_COL}\n')

    output_path = os.path.join(OUTPUT_DIR, OUTPUT_FILENAME_COL)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for val in unique_vals:
            subset = df[df[target_col] == val].copy()
            sn     = sanitize_sheet_name(val)
            subset.to_excel(writer, sheet_name=sn, index=False)
            ws = writer.sheets[sn]
            apply_header_style(ws)
            auto_col_width(ws)
            print(f'  [{sn:<31}]  {len(subset):>7,}행')

    print(f'\n  완료 — {len(unique_vals)}개 시트 → {output_path}')


# ── 진입점 ────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description=(
            'sheet 모드(기본): 시트명 "_" 앞 접두어로 그룹핑 → 그룹별 파일 저장 [win32com 네이티브]\n'
            'col   모드       : 단일 시트의 컬럼값 기준 행 분리 [pandas]'
        ),
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument('--mode',  choices=['sheet', 'col'], default='sheet',
                        help='sheet(기본값) | col')
    parser.add_argument('--col',   default=DEFAULT_TARGET_COL,
                        help=f'[col] 분리 기준 컬럼명 (기본값: "{DEFAULT_TARGET_COL}")')
    parser.add_argument('--nan',   default=DEFAULT_NAN_LABEL,
                        help=f'[col] 결측치 레이블 (기본값: "{DEFAULT_NAN_LABEL}")')
    parser.add_argument('--input', default=None,
                        help='input_data/ 내 파일명. 생략 시 최근 .xlsx 자동 선택')
    parser.add_argument('--sheet', default=None,
                        help='[col] 읽을 시트명. 생략 시 활성 시트')
    parser.add_argument('--sep',   default='_',
                        help='[sheet] 그룹 구분자 (기본값: "_")')
    args = parser.parse_args()

    if args.mode == 'sheet':
        group_by_sheet_prefix(
            input_file=args.input,
            sep=args.sep,
        )
    else:
        split_sheets(
            target_col=args.col,
            nan_label=args.nan,
            input_file=args.input,
            sheet_name=args.sheet,
        )


if __name__ == '__main__':
    main()
