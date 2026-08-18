"""퇴직급여충당부채 검증앱(일반기업회계기준·퇴직금 추계액 방식) — 엔진.

input_data/severance_<company>_information_fy<year>.xlsx 를 읽어
결산기준일 현재 재직 중인 임직원별 퇴직금 추계액을 재계산하고,
전기말/당기말 회사계상 충당부채와 대사하는 output/severance_schedule_<company>_<fy>.xlsx 를 생성한다.

핵심 설계: depreciation_analyzer와 달리 기초(전기말) 잔액을 신뢰하지 않는다. 당기말 퇴충잔액은
결산기준일 현재 재직 중인 임직원의 재직일수·급여만으로 전액 독립 재계산하며, 회사계상 전기말/당기말
잔액은 오직 대사(비교)용 참고값으로만 쓰인다. 결산기준일은 입력파일에 직접 적지 않고
파일명(fy<연도>)과 --fiscal-month 로 depreciation_analyzer와 동일한 규칙으로 계산한다.

전기/당기 인원은 회사별로 별도 파일을 만드는 대신, 한 파일 안에 '전기정보'/'당기정보'
두 시트로 나눠 입력받는다(회사가 매년 스냅샷 형태로 제공하는 인원현황을 그대로 붙여넣는 것을 전제).
당기 계산(추계액)은 '당기정보' 시트만 사용하고, '전기정보'는 사번(없으면 성명) 기준으로
당기 시트와 매칭해 신규입사자/퇴사자 명단을 산출하는 데에만 쓰인다 — 실제 퇴사일 필드가 없어도
"전기에는 있었는데 당기에는 없는 사번"으로 진짜 퇴사자를 식별할 수 있다.

퇴직금 추계액 산식('당기정보' 시트 기준, 일반직원/임원 산식이 다름):
  공통: 급여기준액 = 월평균급여 + 상여금(연간) × 3 / 12   (임원 산식·화면 표시용. 상여금 3개월분 상당액을 가산)
        기산일 = 중간정산일(퇴직금 중간정산 이력이 있는 경우) 있으면 그 날짜, 없으면 입사일
          ※ 중간정산은 재직 상태를 유지한 채 근속기간만 새로 기산되는 것이므로(퇴사가 아님),
            재직일수/근속연수 계산 기산일에만 반영한다.

  [일반직원] 근로기준법 평균임금 산식:
    1일평균임금 = (월평균급여 × 3 + 상여금(연간) / 4) / 92   (최근 3개월간 임금총액 ÷ 최근 3개월간 총일수 근사)
    재직일수 = (당기 결산기준일 − 기산일) + 1   (재직 상태의 존속기간이므로 초일(기산일)을 산입)
    퇴직금 추계액 = 1일평균임금 × 30일 × (재직일수 / 365)

  [임원] 부서명이 정확히 '임원'인 인원 — 소득세법상 임원 퇴직금 한도 산식(평균급여 × 10% × 근속연수 × 배수):
    근속연수 = 당기 결산기준일의 연도 − 기산일의 연도
    근속월(1년 미만 잔여월 보정) = 12 − 기산일의 월
    퇴직금 추계액 = [(급여기준액 × 근속연수 + 급여기준액 × 근속월 / 10) / 10] × 배수
    배수: '배수(임원)' 컬럼값(임원퇴직금지급규정 등에 따른 배수, 일반 직원은 이 산식 자체가 적용되지 않음)

실행 예:
    python severance_schedule.py kyungnam --fiscal-month 12
    python severance_schedule.py --file severance_kyungnam_information_fy2026.xlsx
"""
import argparse
import calendar
import glob
import os
from datetime import date

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(HERE, "input_data")
OUTPUT_DIR = os.path.join(HERE, "output")

SIG_THRESHOLD_ABS = 1000      # 유의차이 절대금액 기준(원)
SIG_THRESHOLD_PCT = 0.01      # 유의차이 비율 기준(1%)

COST_TYPES = ["제조원가", "판관비"]
BASIS_KEYS = {
    "제조원가": {
        "전기": "전기말 회사계상 퇴직급여충당부채(제조원가분)",
        "당기": "당기말 회사계상 퇴직급여충당부채(제조원가분)",
    },
    "판관비": {
        "전기": "전기말 회사계상 퇴직급여충당부채(판관비분)",
        "당기": "당기말 회사계상 퇴직급여충당부채(판관비분)",
    },
}

EXECUTIVE_DEPT = "임원"  # 부서명이 정확히 이 값일 때만 배수를 적용한다.

CURRENT_SHEET = "당기정보"
PRIOR_SHEET = "전기정보"
LEAVER_SHEET = "당기퇴사자"

DEBIT_BASIS_LABEL = "당기 퇴직급여충당부채 차변(당기지급액, 분개장 기준)"

DISPLAY_COLS = ["사업장", "부서", "사번", "성명", "직급", "원가구분", "입사일", "중간정산일"]

# 요약표 '비고'에 그대로 붙여 수기 검증에 쓰는 계산 공식 텍스트(compute_employee()와 동일 로직).
FORMULA_NOTE = (
    "[일반직원] 1일평균임금=(월평균급여×3+연간상여금/4)÷92 ; "
    "재직일수=(기준일−기산일)+1(초일산입, 기산일=중간정산일 있으면 그 날짜·없으면 입사일) ; "
    "퇴직금추계액=1일평균임금×30×(재직일수/365)  ‖  "
    "[임원, 부서='임원'] 급여기준액=월평균급여+연간상여금×3/12 ; 근속연수=기준일연도−기산일연도 ; "
    "근속월=12−기산일월 ; 퇴직금추계액=((급여기준액×근속연수+급여기준액×근속월/10)/10)×배수(임원)"
)


# ── 공용 헬퍼 ────────────────────────────────────────────────────────────────

def _safe_float(v, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        if pd.isna(v):
            return default
    except (TypeError, ValueError):
        pass
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _safe_date(v):
    if v is None or v == "":
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    ts = pd.Timestamp(v)
    return ts.date() if not pd.isna(ts) else None


def _cost_type(emp: dict) -> str:
    return str(emp.get("원가구분(제조원가/판관비)") or "").strip() or "(미분류)"


def _employee_key(emp: dict):
    """전기/당기 인원 매칭 키. 사번이 있으면 사번, 없으면 성명으로 매칭한다."""
    사번 = emp.get("사번")
    if 사번 not in (None, ""):
        return f"ID::{str(사번).strip()}"
    성명 = emp.get("성명")
    if 성명 not in (None, ""):
        return f"NAME::{str(성명).strip()}"
    return None


def _fy_bounds(target_fy: str, fiscal_month: int) -> tuple:
    """대상 회계연도(target_fy)의 시작월/종료월(YYYY-MM 문자열) 반환.
    depreciation_schedule.py/lease_schedule.py 와 동일 규칙."""
    fy = int(target_fy)
    if fiscal_month == 12:
        return f"{fy}-01", f"{fy}-12"
    return f"{fy - 1}-{fiscal_month + 1:02d}", f"{fy}-{fiscal_month:02d}"


def _apply_interim(fy_end: str, target_fy: str, interim_month: int = None) -> str:
    if not interim_month:
        return fy_end
    interim_ym = f"{target_fy}-{interim_month:02d}"
    return min(fy_end, interim_ym)


def _ym_to_end_date(ym: str) -> date:
    y, m = int(ym[:4]), int(ym[5:7])
    last_day = calendar.monthrange(y, m)[1]
    return date(y, m, last_day)


def _find_input_file(company: str = None, file: str = None) -> str:
    if file:
        path = file if os.path.isabs(file) else os.path.join(INPUT_DIR, file)
        if not os.path.exists(path):
            raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {path}")
        return path

    if company:
        pattern = os.path.join(INPUT_DIR, f"severance_{company}_information_fy*.xlsx")
    else:
        pattern = os.path.join(INPUT_DIR, "severance_*_information_fy*.xlsx")

    matches = [p for p in glob.glob(pattern) if "template" not in os.path.basename(p)]
    if not matches:
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {pattern}")
    if len(matches) > 1 and not company:
        raise ValueError(f"회사를 특정해주세요. 후보 파일 여러 개: {matches}")
    return matches[0]


def _is_significant(diff, base) -> bool:
    if diff is None or pd.isna(diff):
        return False
    if abs(diff) >= SIG_THRESHOLD_ABS and (base in (None, 0) or abs(diff) >= abs(base) * SIG_THRESHOLD_PCT):
        return True
    return False


# ── 입력 로딩 ────────────────────────────────────────────────────────────────

def load_employees(path: str, sheet_name: str) -> list:
    """지정한 인원 시트(1~2행 헤더, 3행부터 데이터)를 읽어 dict 목록으로 반환.
    시트가 없으면 빈 목록을 반환한다(예: '전기정보' 미작성 시 인원변동 명단만 생략됨)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [c.value for c in ws[2]]
    employees = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        rec = dict(zip(headers, row))
        if not rec.get("사번") and not rec.get("성명"):
            continue
        employees.append(rec)
    return employees


def load_basis(path: str) -> dict:
    """'기준정보' 시트를 라벨(A열) 기준으로 읽어 {라벨: 금액} dict 반환. 없으면 빈 dict."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "기준정보" not in wb.sheetnames:
        return {}
    ws = wb["기준정보"]
    basis = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        label, value = row[0], row[1]
        if not isinstance(label, str):
            continue
        label = label.strip()
        if not label:
            continue
        basis[label] = None if value in (None, "") else _safe_float(value)
    return basis


# ── 전기/당기 인원 매칭 (신규입사자/퇴사자 판정) ──────────────────────────────

def match_periods(당기_employees: list, 전기_employees: list) -> dict:
    """당기/전기 인원을 사번(없으면 성명) 기준으로 매칭해 신규입사자/퇴사자를 산출한다."""
    당기_by_key, 전기_by_key = {}, {}
    for e in 당기_employees:
        k = _employee_key(e)
        if k is not None:
            당기_by_key.setdefault(k, e)
    for e in 전기_employees:
        k = _employee_key(e)
        if k is not None:
            전기_by_key.setdefault(k, e)

    신규입사_keys = set(당기_by_key) - set(전기_by_key)
    퇴사_keys = set(전기_by_key) - set(당기_by_key)

    return {
        "신규입사_keys": 신규입사_keys,
        "신규입사자": [당기_by_key[k] for k in 신규입사_keys],
        "퇴사자": [전기_by_key[k] for k in 퇴사_keys],
        "전기_by_key": 전기_by_key,
        "당기_by_key": 당기_by_key,
        "전기인원수": len(전기_by_key),
        "당기인원수": len(당기_by_key),
    }


def compute_prior_balances(전기_by_key: dict, 전기결산일: date) -> dict:
    """'전기정보' 인원 각각의 전기 결산기준일 시점 퇴직금 추계액을 사번(없으면 성명) 키로 반환.
    당기 인원별 표에서 '당기말 - 전기말' 차이를 인별로 계산하는 데 쓰인다."""
    return {k: compute_employee(e, 전기결산일)["퇴직금추계액"] for k, e in 전기_by_key.items()}


def _to_display_df(records: list) -> pd.DataFrame:
    rows = []
    for e in records:
        rows.append({
            "사업장": e.get("사업장") or "",
            "부서": e.get("부서") or "",
            "사번": e.get("사번"),
            "성명": e.get("성명"),
            "직급": e.get("직급"),
            "원가구분": _cost_type(e),
            "입사일": e.get("입사일"),
            "중간정산일": e.get("중간정산일"),
        })
    if not rows:
        return pd.DataFrame(columns=DISPLAY_COLS)
    return pd.DataFrame(rows)[DISPLAY_COLS]


MIN_TENURE_DAYS = 365  # 근로기준법상 퇴직금 지급 요건(계속근로기간 1년 이상) 판정 기준

LEAVER_MATCH_COLS = ["사업장", "부서", "사번", "직급", "원가구분",
                      "전기정보 있으나 당기정보 없음", "실제 퇴사자", "비고"]


def _build_leaver_match_df(전기_by_key: dict, 당기_by_key: dict, 퇴사자_recs: list,
                            leaver_payments: list, 전기결산일: date) -> pd.DataFrame:
    """'전기정보-당기정보 자동 비교로 산출된 퇴사자 명단'과 '당기퇴사자' 시트(사용자 입력) 두 명단을
    사번(없으면 성명) 기준으로 나란히 비교하는 통합 표. 양쪽에 모두 있으면 '이상없음',
    한쪽에만 있으면 그 원인을 비고에 표시한다(전기말 재직 1년 미만·당기정보에 여전히 존재·이중기입 등)."""
    퇴사_keys = {_employee_key(e) for e in 퇴사자_recs if _employee_key(e) is not None}

    당기퇴사자_by_key: dict = {}
    for rec in leaver_payments:
        if rec.get("실제지급액(원)") in (None, ""):
            continue
        k = _employee_key(rec)
        if k is not None:
            당기퇴사자_by_key.setdefault(k, []).append(rec)

    rows = []
    for key in 퇴사_keys | set(당기퇴사자_by_key.keys()):
        전기레코드 = 전기_by_key.get(key)
        당기퇴사자_recs = 당기퇴사자_by_key.get(key, [])
        in_auto = key in 퇴사_keys
        in_actual = len(당기퇴사자_recs) > 0
        source = 전기레코드 or 당기_by_key.get(key)

        # 지급사유가 하나라도 '중간정산'이면 이 인원은 진짜 퇴사자가 아니다(재직 중 중간정산은
        # '당기정보'에 그대로 남아있는 게 정상) — "당기정보에도 존재함" 경고 대신 사유만 표시한다.
        사유들 = {str(rec.get("지급사유(실제퇴사/중간정산)") or "").strip() for rec in 당기퇴사자_recs}
        is_중간정산 = "중간정산" in 사유들

        tags = []
        if in_auto and in_actual:
            if is_중간정산:
                tags.append("중간정산")
            elif key in 당기_by_key:
                tags.append("⚠ '당기정보'에도 그대로 존재함 — 실제 퇴사 여부 확인 필요")
            if len(당기퇴사자_recs) > 1:
                tags.append("⚠ 이중기입의심(동일 인원으로 보이는 항목이 '당기퇴사자' 시트에 여러 번 입력됨 — 동명이인 여부 확인 필요)")
            if not tags:
                tags.append("이상없음")
        elif in_auto and not in_actual:
            tags.append("⚠ '당기퇴사자' 시트에 지급액 입력 없음")
        else:  # in_actual and not in_auto
            if is_중간정산:
                tags.append("중간정산")
            elif 전기레코드 is not None:
                tags.append("⚠ '당기정보'에도 그대로 존재함 — 실제 퇴사 여부 확인 필요")
            else:
                입사일 = _safe_date(당기퇴사자_recs[0].get("입사일(선택)")) if 당기퇴사자_recs else None
                if 입사일 is not None and (전기결산일 - 입사일).days < MIN_TENURE_DAYS:
                    tags.append("전기 결산기준일 기준 근속 1년 미만 — 전기말 퇴충설정대상 아님")
                else:
                    tags.append("⚠ '전기정보'에서 매칭되는 인원을 찾지 못함(확인 필요)")
            if len(당기퇴사자_recs) > 1:
                tags.append("⚠ 이중기입의심(동일 인원으로 보이는 항목이 '당기퇴사자' 시트에 여러 번 입력됨 — 동명이인 여부 확인 필요)")

        성명_auto = ((전기레코드 or {}).get("성명") or (전기레코드 or {}).get("사번") or "") if in_auto else ""
        성명_actual = ((당기퇴사자_recs[0].get("성명") or 당기퇴사자_recs[0].get("사번") or "") if 당기퇴사자_recs else "")
        사번 = (전기레코드 or {}).get("사번") or (당기퇴사자_recs[0].get("사번") if 당기퇴사자_recs else None)

        rows.append({
            "사업장": (source or {}).get("사업장") or "",
            "부서": (source or {}).get("부서") or "",
            "사번": 사번,
            "직급": (source or {}).get("직급") or "",
            "원가구분": _cost_type(source) if source else "(미상)",
            "전기정보 있으나 당기정보 없음": 성명_auto,
            "실제 퇴사자": 성명_actual,
            "비고": " / ".join(tags),
        })

    if not rows:
        return pd.DataFrame(columns=LEAVER_MATCH_COLS)
    df = pd.DataFrame(rows)[LEAVER_MATCH_COLS]
    return df.sort_values(
        ["사업장", "부서", "전기정보 있으나 당기정보 없음", "실제 퇴사자"], na_position="last"
    ).reset_index(drop=True)


def _build_group_summary(d: pd.DataFrame, 전기_calc_df: pd.DataFrame, group_col: str) -> list:
    """사업장별/부서별 요약표. 당기말은 당기_df(d, 당기 재직자)에서, 전기말은 전기_calc_df
    (전기 결산기준일 현재 재직 중이던 전체 인원 — 당기 중 퇴사한 사람 포함)에서 각각 집계한다.
    두 시점의 모집단이 다르므로 반드시 서로 다른 소스에서 집계해야 한다(같은 d에서 전기말까지
    뽑으면 당기 중 퇴사자의 전기말 잔액이 누락됨). 그룹 컬럼에 실제 값이 하나도 없으면 빈 리스트를
    반환해 요약표에서 그 섹션 자체를 생략하도록 한다."""
    당기_by_group: dict = {}
    if not d.empty and group_col in d.columns:
        d2 = d.copy()
        d2[group_col] = d2[group_col].astype(str).str.strip()
        for g, gdf in d2[d2[group_col] != ""].groupby(group_col, sort=False):
            당기_by_group[g] = {
                "인원수": int(len(gdf)),
                "당기말": float(gdf["당기말 퇴직금추계액(계산)"].sum()),
                # by_cost 표와 동일하게, '인원별추계명세'의 당기 퇴직급여(계산) 합계를 그대로 쓴다
                # (퇴사로 인한 잔액 감소가 섞이지 않는 순수 발생액 — 재무제표 퇴직급여 대사용).
                "당기퇴직급여": float(gdf["당기 퇴직급여(계산)"].sum()),
            }

    전기_by_group: dict = {}
    if not 전기_calc_df.empty and group_col in 전기_calc_df.columns:
        p2 = 전기_calc_df.copy()
        p2[group_col] = p2[group_col].astype(str).str.strip()
        for g, gdf in p2[p2[group_col] != ""].groupby(group_col, sort=False):
            전기_by_group[g] = float(gdf["전기말 퇴직금추계액(계산)"].sum())

    all_groups = set(당기_by_group) | set(전기_by_group)
    if not all_groups:
        return []
    rows = []
    for g in sorted(all_groups):
        rows.append({
            group_col: g,
            "당기인원수": 당기_by_group.get(g, {}).get("인원수", 0),
            "전기말(재계산)": 전기_by_group.get(g, 0.0),
            "당기말(재계산)": 당기_by_group.get(g, {}).get("당기말", 0.0),
            "당기 퇴직급여비용(재계산)": 당기_by_group.get(g, {}).get("당기퇴직급여", 0.0),
        })
    return rows


# ── 인원별 퇴직금 추계액 계산 ('당기정보' 시트 기준) ──────────────────────

def compute_employee(emp: dict, 기준일: date) -> dict:
    """지정한 기준일(당기 결산기준일 또는 전기 결산기준일) 현재의 퇴직금 추계액을 계산한다.
    '당기정보'/'전기정보' 두 시트 어느 쪽 인원이든 동일 로직으로 계산할 수 있도록
    기준일을 매개변수로 받는다(전기말 재계산에도 재사용)."""
    입사일 = _safe_date(emp.get("입사일"))
    중간정산일 = _safe_date(emp.get("중간정산일"))
    월평균급여 = _safe_float(emp.get("월평균급여(원)"))
    연간상여금 = _safe_float(emp.get("상여금(연간, 원)"))
    부서 = str(emp.get("부서") or "").strip()
    # 상여금은 연 1회 등 정기 지급을 가정, 근로기준법상 평균임금 산정 관행대로
    # "최근 3개월분 상당액(연간상여금×3/12)"만 월평균급여에 가산한다.
    급여기준액 = 월평균급여 + 연간상여금 * 3 / 12

    result = {
        "재직일수": None, "1일평균임금": None, "퇴직금추계액": 0.0,
        "근속연수": None, "근속월": None,
        "기산일": None, "급여기준액": 급여기준액, "배수": 1.0, "임원배수적용": False,
        "warning": None,
    }

    def _add_warning(w):
        result["warning"] = f"{result['warning']} / {w}".strip(" /") if result["warning"] else w

    # 임원 퇴직금 규정상 배수(예: 정관·임원퇴직금지급규정상 2배, 3배 등)는
    # 부서명이 정확히 '임원'인 인원에만 적용한다(일반 직원은 입력값이 있어도 무시).
    if 부서 == EXECUTIVE_DEPT:
        배수_raw = emp.get("배수(임원)")
        배수 = _safe_float(배수_raw, default=1.0)
        if 배수_raw in (None, ""):
            _add_warning("임원 배수 미입력 — 1배로 계산됨")
        elif 배수 <= 0:
            _add_warning("임원 배수 값 오류(0 이하) — 1배로 계산됨")
            배수 = 1.0
        result["배수"] = 배수
        result["임원배수적용"] = True

    if 입사일 is None:
        _add_warning("입사일 미입력 — 재직일수/추계액 계산 불가")
        return result

    기산일 = 입사일
    if 중간정산일 is not None:
        if 중간정산일 < 입사일:
            _add_warning("중간정산일이 입사일보다 이전 — 확인 필요(입사일 기준으로 계산)")
        elif 중간정산일 > 기준일:
            _add_warning("중간정산일이 결산기준일 이후 — 확인 필요(입사일 기준으로 계산)")
        else:
            기산일 = 중간정산일

    if 입사일 > 기준일:
        _add_warning("입사일이 결산기준일 이후 — 확인 필요(해당 기준일 시점에 아직 미입사)")
        return result

    # 근속기간(재직일수)은 계속된 재직 상태의 존속기간이므로 초일(기산일)을 산입하여 계산한다(+1).
    재직일수 = (기준일 - 기산일).days + 1
    result["재직일수"] = 재직일수
    result["기산일"] = 기산일

    if result["임원배수적용"]:
        # 소득세법상 임원 퇴직금 한도 산식(근로기준법 평균임금 방식과 별개): 평균급여 × 10% × 근속연수 × 배수.
        # 근속연수 = 결산기준연도 - 기산연도(입사연도, 중간정산 이력이 있으면 중간정산연도)
        # 근속월(1년 미만 잔여월 보정항) = 12 - 기산월
        # "10%"는 근속월 나머지항에만 걸리는 게 아니라 (근속연수+근속월/10) 전체에 걸리는 계수이므로
        # 마지막에 전체를 10으로 한 번 더 나눈다(이 나누기가 빠져 있으면 결과가 정확히 10배 커짐).
        근속연수 = 기준일.year - 기산일.year
        근속월 = 12 - 기산일.month
        result["근속연수"] = 근속연수
        result["근속월"] = 근속월
        result["퇴직금추계액"] = ((급여기준액 * 근속연수 + 급여기준액 * 근속월 / 10) / 10) * result["배수"]
    else:
        # 근로기준법 평균임금 산식: 1일평균임금 = 최근 3개월간 임금총액 / 최근 3개월간 총일수(92일 근사).
        # 3개월간 임금총액 = 월평균급여×3 + 상여금 3개월분 상당액(연간상여금/4).
        임금총액_3개월 = 월평균급여 * 3 + 연간상여금 / 4
        일평균임금 = 임금총액_3개월 / 92
        result["1일평균임금"] = 일평균임금
        result["퇴직금추계액"] = 일평균임금 * 30 * (재직일수 / 365) * result["배수"]

    if 급여기준액 <= 0:
        _add_warning("월평균급여(상여금 포함) 미입력 — 추계액 0으로 계산됨")

    return result


# ── 명세서 구성 ('당기정보' 시트 기준) ────────────────────────────────────

def build_schedule_table(employees: list, 당기결산일: date, 신규입사_keys: set, 전기말_balances: dict) -> pd.DataFrame:
    """전기말_balances: {사번(또는 성명) 키: 전기 결산기준일 시점 추계액}. compute_prior_balances()로 생성.
    신규입사자 등 전기 대응값이 없는 인원은 전기말 추계액을 0으로 본다(전기 시점 미재직)."""
    rows = []
    for e in employees:
        r = compute_employee(e, 당기결산일)
        원가구분 = _cost_type(e)
        key = _employee_key(e)
        신규입사 = key in 신규입사_keys
        전기말추계액 = 전기말_balances.get(key, 0.0)
        당기퇴직급여 = r["퇴직금추계액"] - 전기말추계액
        회사계상_raw = e.get("회사계상 퇴직금추계액(원)")
        당기회사계상 = _safe_float(회사계상_raw) if 회사계상_raw not in (None, "") else None
        당기말차이 = None if 당기회사계상 is None else r["퇴직금추계액"] - 당기회사계상
        중간정산일_raw = _safe_date(e.get("중간정산일"))
        중간정산반영 = r["기산일"] is not None and 중간정산일_raw is not None and r["기산일"] == 중간정산일_raw

        비고 = e.get("비고") or ""
        tags = []
        if r["warning"]:
            tags.append(r["warning"])
        if 신규입사:
            tags.append("당기 신규입사(전기정보에는 없음)")
        if 중간정산반영:
            tags.append(f"퇴직금 중간정산 이력 있음(재직일수는 {중간정산일_raw}부터 기산)")
        if r["임원배수적용"]:
            tags.append(
                f"임원 배수 {r['배수']:g}배 적용 — 소득세법상 임원 퇴직금 한도 산식으로 계산"
                f"(근속연수 {r['근속연수']}년, 근속월 보정 {r['근속월']}/10)"
            )
        if tags:
            비고 = f"{비고} / {' / '.join(tags)}".strip(" /")

        rows.append({
            "사업장": e.get("사업장") or "",
            "부서": e.get("부서") or "",
            "사번": e.get("사번"),
            "성명": e.get("성명"),
            "직급": e.get("직급"),
            "원가구분": 원가구분,
            "입사일": e.get("입사일"),
            "중간정산일": e.get("중간정산일"),
            "월평균급여(원)": _safe_float(e.get("월평균급여(원)")),
            "상여금(연간, 원)": _safe_float(e.get("상여금(연간, 원)")),
            "월평균급여(상여금가산후)(계산)": r["급여기준액"],
            "배수(임원)": r["배수"] if r["임원배수적용"] else None,
            "재직일수(당기말기준)": r["재직일수"],
            "1일평균임금(계산)": r["1일평균임금"],
            "근속연수(임원계산용)": r["근속연수"],
            "근속월보정(임원계산용)": r["근속월"],
            "전기말 퇴직금추계액(계산)": 전기말추계액,
            "당기말 퇴직금추계액(계산)": r["퇴직금추계액"],
            "당기말 회사계상 퇴직금추계액(원)": 당기회사계상,
            "당기말 차이(계산-회사계상)": 당기말차이,
            "당기 퇴직급여(계산)": 당기퇴직급여,
            "당기신규입사": 신규입사,
            "중간정산반영": 중간정산반영,
            "비고": 비고,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    return df.sort_values(["사업장", "부서", "원가구분", "사번"], na_position="last").reset_index(drop=True)


def build_prior_schedule_table(전기_employees: list, 전기결산일: date) -> pd.DataFrame:
    """'전기정보' 시트 인원별 전기 결산기준일 시점 추계액과, 입력된 '회사계상 퇴직금추계액(원)'을
    인별로 대사한 명세를 만든다. build_schedule_table()의 전기(당기 대신) 버전 —
    당기 재직 여부·신규입사 태그 등 당기 전용 로직은 없고 순수 전기 시점 계산+대사만 수행한다."""
    rows = []
    for e in 전기_employees:
        r = compute_employee(e, 전기결산일)
        원가구분 = _cost_type(e)
        회사계상_raw = e.get("회사계상 퇴직금추계액(원)")
        회사계상 = _safe_float(회사계상_raw) if 회사계상_raw not in (None, "") else None
        차이 = None if 회사계상 is None else r["퇴직금추계액"] - 회사계상
        중간정산일_raw = _safe_date(e.get("중간정산일"))
        중간정산반영 = r["기산일"] is not None and 중간정산일_raw is not None and r["기산일"] == 중간정산일_raw

        비고 = e.get("비고") or ""
        tags = []
        if r["warning"]:
            tags.append(r["warning"])
        if 중간정산반영:
            tags.append(f"퇴직금 중간정산 이력 있음(재직일수는 {중간정산일_raw}부터 기산)")
        if r["임원배수적용"]:
            tags.append(
                f"임원 배수 {r['배수']:g}배 적용 — 소득세법상 임원 퇴직금 한도 산식으로 계산"
                f"(근속연수 {r['근속연수']}년, 근속월 보정 {r['근속월']}/10)"
            )
        if tags:
            비고 = f"{비고} / {' / '.join(tags)}".strip(" /")

        rows.append({
            "사업장": e.get("사업장") or "",
            "부서": e.get("부서") or "",
            "사번": e.get("사번"),
            "성명": e.get("성명"),
            "직급": e.get("직급"),
            "원가구분": 원가구분,
            "입사일": e.get("입사일"),
            "중간정산일": e.get("중간정산일"),
            "월평균급여(원)": _safe_float(e.get("월평균급여(원)")),
            "상여금(연간, 원)": _safe_float(e.get("상여금(연간, 원)")),
            "월평균급여(상여금가산후)(계산)": r["급여기준액"],
            "배수(임원)": r["배수"] if r["임원배수적용"] else None,
            "재직일수(전기말기준)": r["재직일수"],
            "1일평균임금(계산)": r["1일평균임금"],
            "근속연수(임원계산용)": r["근속연수"],
            "근속월보정(임원계산용)": r["근속월"],
            "전기말 퇴직금추계액(계산)": r["퇴직금추계액"],
            "전기말 회사계상 퇴직금추계액(원)": 회사계상,
            "전기말 차이(계산-회사계상)": 차이,
            "중간정산반영": 중간정산반영,
            "비고": 비고,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    return df.sort_values(["사업장", "부서", "원가구분", "사번"], na_position="last").reset_index(drop=True)


# ── 요약(원가구분별 대사 + 인원변동) ──────────────────────────────────────────

def build_summary(당기_df: pd.DataFrame, 전기_employees: list,
                   신규입사자_recs: list, 퇴사자_recs: list, basis: dict,
                   전기_by_key: dict, 전기말_balances: dict, leaver_payments: list,
                   전기결산일: date, 당기_by_key: dict) -> dict:
    if 당기_df.empty and not 전기_employees and not 신규입사자_recs and not 퇴사자_recs:
        return {}

    d = 당기_df.copy()

    # 전기말(재계산)은 반드시 '전기 결산기준일 현재 재직 중이던 전체 인원'(전기_by_key) 기준으로 집계해야 한다.
    # d(당기_df)는 당기말 재직자만 담고 있어 그 안의 '전기말 퇴직금추계액(계산)' 컬럼만 합산하면
    # 당기 중 퇴사한 사람의 전기말 잔액이 통째로 빠진다 — 이걸 그대로 쓰면 안 됨(전기말 잔액 과소계상 버그).
    전기_calc_rows = [
        {
            "사업장": e.get("사업장") or "",
            "부서": e.get("부서") or "",
            "원가구분": _cost_type(e),
            "전기말 퇴직금추계액(계산)": 전기말_balances.get(k, 0.0),
        }
        for k, e in 전기_by_key.items()
    ]
    전기_calc_df = pd.DataFrame(전기_calc_rows) if 전기_calc_rows else pd.DataFrame(
        columns=["사업장", "부서", "원가구분", "전기말 퇴직금추계액(계산)"]
    )

    by_cost = []
    총_전기재계산 = 0.0
    총_당기재계산 = 0.0
    총_전기입력 = None
    총_당기회사계상 = None
    for cost in COST_TYPES:
        cdf = d[d["원가구분"] == cost] if not d.empty else d
        전기_cdf = 전기_calc_df[전기_calc_df["원가구분"] == cost] if not 전기_calc_df.empty else 전기_calc_df
        전기재계산 = float(전기_cdf["전기말 퇴직금추계액(계산)"].sum()) if not 전기_cdf.empty else 0.0
        당기재계산 = float(cdf["당기말 퇴직금추계액(계산)"].sum()) if not cdf.empty else 0.0
        # 당기 퇴직급여비용은 '인원별추계명세'의 당기 퇴직급여(계산) 합계 그대로 사용한다(당기 재직자
        # 개인별 발생액의 합 — 신규입사자는 전기말 0원 기준, 퇴사자는 당기 표 자체에 없어 제외됨).
        # 손익계산서 퇴직급여 계정과 대사하려는 목적이므로, 퇴사로 인한 충당부채 잔액 감소(지급 효과)까지
        # 섞여버리는 "당기말(재계산)-전기말(재계산)" 순증감과는 값이 다르다(그건 잔액 T계정 검증에만 씀).
        당기퇴직급여 = float(cdf["당기 퇴직급여(계산)"].sum()) if not cdf.empty else 0.0
        전기입력 = basis.get(BASIS_KEYS[cost]["전기"])
        당기회사계상 = basis.get(BASIS_KEYS[cost]["당기"])
        by_cost.append({
            "구분": cost,
            "전기말(재계산)": 전기재계산,
            "전기말(회사계상)": 전기입력,
            "당기말(재계산)": 당기재계산,
            "당기말(회사계상)": 당기회사계상,
            "당기 퇴직급여비용(재계산)": 당기퇴직급여,
            "대사차이(재계산-회사계상)": None if 당기회사계상 is None else 당기재계산 - 당기회사계상,
        })
        총_전기재계산 += 전기재계산
        총_당기재계산 += 당기재계산
        if 전기입력 is not None:
            총_전기입력 = (총_전기입력 or 0.0) + 전기입력
        if 당기회사계상 is not None:
            총_당기회사계상 = (총_당기회사계상 or 0.0) + 당기회사계상

    미분류_당기 = d[~d["원가구분"].isin(COST_TYPES)] if not d.empty else d
    미분류_전기 = 전기_calc_df[~전기_calc_df["원가구분"].isin(COST_TYPES)] if not 전기_calc_df.empty else 전기_calc_df
    if not 미분류_당기.empty or not 미분류_전기.empty:
        미분류_전기재계산 = float(미분류_전기["전기말 퇴직금추계액(계산)"].sum()) if not 미분류_전기.empty else 0.0
        미분류_당기재계산 = float(미분류_당기["당기말 퇴직금추계액(계산)"].sum()) if not 미분류_당기.empty else 0.0
        미분류_당기퇴직급여 = float(미분류_당기["당기 퇴직급여(계산)"].sum()) if not 미분류_당기.empty else 0.0
        by_cost.append({
            "구분": "(미분류)",
            "전기말(재계산)": 미분류_전기재계산,
            "전기말(회사계상)": None,
            "당기말(재계산)": 미분류_당기재계산,
            "당기말(회사계상)": None,
            "당기 퇴직급여비용(재계산)": 미분류_당기퇴직급여,
            "대사차이(재계산-회사계상)": None,
        })
        총_전기재계산 += 미분류_전기재계산
        총_당기재계산 += 미분류_당기재계산

    총_당기퇴직급여 = float(d["당기 퇴직급여(계산)"].sum()) if not d.empty else 0.0
    total_row = {
        "구분": "합계",
        "전기말(재계산)": 총_전기재계산,
        "전기말(회사계상)": 총_전기입력,
        "당기말(재계산)": 총_당기재계산,
        "당기말(회사계상)": 총_당기회사계상,
        "당기 퇴직급여비용(재계산)": 총_당기퇴직급여,
        "대사차이(재계산-회사계상)": None if 총_당기회사계상 is None else 총_당기재계산 - 총_당기회사계상,
    }

    전기_df = _to_display_df(전기_employees)
    신규입사자_df = _to_display_df(신규입사자_recs)

    headcount = {
        "전기말인원수": int(len(전기_df)),
        "신규입사인원수": int(len(신규입사자_df)),
        "퇴사인원수": len(퇴사자_recs),
        "당기말인원수": int(len(d)),
    }
    headcount_by_cost = []
    for cost in COST_TYPES:
        headcount_by_cost.append({
            "원가구분": cost,
            "전기말인원수": int(len(전기_df[전기_df["원가구분"] == cost])) if not 전기_df.empty else 0,
            "신규입사인원수": int(len(신규입사자_df[신규입사자_df["원가구분"] == cost])) if not 신규입사자_df.empty else 0,
            "퇴사인원수": sum(1 for e in 퇴사자_recs if _cost_type(e) == cost),
            "당기말인원수": int(len(d[d["원가구분"] == cost])) if not d.empty else 0,
        })

    # 사업장별/부서별 요약(해당 컬럼에 실제 값이 있을 때만 생성 — 없으면 write_summary_sheet에서 섹션 생략)
    site_summary = _build_group_summary(d, 전기_calc_df, "사업장")
    dept_summary = _build_group_summary(d, 전기_calc_df, "부서")

    # T계정 검증(tie-out): 전기말(회사계상) + 당기 퇴직급여(재계산) - 당기지급액(분개장, 입력) =? 당기말(회사계상)
    # 전액 회사가 실제 보고/기표한 값(전기말·당기말·지급액)에 우리 재계산 전입액만 얹어보는 교차검증.
    당기지급액 = basis.get(DEBIT_BASIS_LABEL)
    tie_out = None
    if 총_전기입력 is not None and 당기지급액 is not None and 총_당기회사계상 is not None:
        계산상기말 = 총_전기입력 + total_row["당기 퇴직급여비용(재계산)"] - 당기지급액
        tie_out = {
            "전기말(회사계상)": 총_전기입력,
            "당기 퇴직급여(재계산)": total_row["당기 퇴직급여비용(재계산)"],
            "당기지급액(분개장)": 당기지급액,
            "계산상 당기말": 계산상기말,
            "당기말(회사계상)": 총_당기회사계상,
            "차이(계산상당기말-회사계상)": 계산상기말 - 총_당기회사계상,
        }

    # 퇴사자 명단 대사 — 자동 산출(전기정보-당기정보 차이) vs '당기퇴사자' 시트(사용자 입력) 통합 비교표
    leaver_match_df = _build_leaver_match_df(전기_by_key, 당기_by_key, 퇴사자_recs, leaver_payments, 전기결산일)

    # 퇴사자 금액차이 분석 — '당기퇴사자' 시트에 입력된 경우만, 전기말 추계액과 실제지급액을 금액으로만 비교
    # (품질 경고성 비고는 위 leaver_match_df 쪽으로 일원화하고, 이 표는 순수 금액 비교로 남긴다)
    leaver_recon_rows = []
    for rec in leaver_payments:
        실제지급액_raw = rec.get("실제지급액(원)")
        if 실제지급액_raw in (None, ""):
            continue
        key = _employee_key(rec)
        실제지급액 = _safe_float(실제지급액_raw)
        전기말추계액 = 전기말_balances.get(key)
        전기레코드 = 전기_by_key.get(key)
        leaver_recon_rows.append({
            "사업장": (전기레코드 or {}).get("사업장") or "",
            "부서": (전기레코드 or {}).get("부서") or "",
            "사번": rec.get("사번"),
            "성명": rec.get("성명"),
            "직급": (전기레코드 or {}).get("직급") or "",
            "원가구분": _cost_type(전기레코드) if 전기레코드 else "(미상)",
            "전기말 추계액(계산)": 전기말추계액,
            "실제지급액(입력)": 실제지급액,
            "차이(추계액-실제지급액)": None if 전기말추계액 is None else 전기말추계액 - 실제지급액,
        })
    leaver_recon_df = pd.DataFrame(leaver_recon_rows) if leaver_recon_rows else pd.DataFrame(
        columns=["사업장", "부서", "사번", "성명", "직급", "원가구분",
                 "전기말 추계액(계산)", "실제지급액(입력)", "차이(추계액-실제지급액)"]
    )

    return {
        "by_cost": by_cost,
        "total_row": total_row,
        "site_summary": site_summary,
        "dept_summary": dept_summary,
        "tie_out": tie_out,
        "leaver_match": leaver_match_df,
        "leaver_recon": leaver_recon_df,
        "신규입사자": 신규입사자_df[["사업장", "부서", "사번", "성명", "직급", "원가구분", "입사일"]],
        "headcount": headcount,
        "headcount_by_cost": headcount_by_cost,
    }


def write_summary_sheet(ws, summary: dict, company: str, target_fy: str,
                         당기결산일: date, 전기결산일: date, interim_month: int = None):
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="203864")
    section_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill("solid", fgColor="9DC3E6")
    sig_fill = PatternFill("solid", fgColor="FFFF00")
    bold = Font(bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    period_note = f", ~{interim_month}월 중간결산(반기 등)" if interim_month else ""
    ws.cell(row=1, column=1,
            value=(f"퇴직급여충당부채 요약표 (회사: {company}, 회계연도: {target_fy}{period_note}, "
                   f"전기말: {전기결산일}, 당기말: {당기결산일})")).font = Font(bold=True, size=13)
    for col in "ABCDEFG":
        ws.column_dimensions[col].width = 18

    r = 3
    if not summary:
        ws.cell(row=r, column=1, value="(인원 데이터 없음)")
        return

    # 1) 원가구분별 전기·당기 비교 + 대사 (전기말/당기말 모두 인별 재계산 기준)
    ws.cell(row=r, column=1, value="■ 퇴직급여충당부채 원가구분별 전기·당기 비교 (대사)").fill = section_fill
    ws.cell(row=r, column=1).font = section_font
    for c in range(2, 9):
        ws.cell(row=r, column=c).fill = section_fill
    r += 2

    cost_headers = ["구분", "전기말(재계산)", "전기말(회사계상)", "당기말(재계산)", "당기말(회사계상)",
                     "당기 퇴직급여비용(재계산)", "대사차이(재계산-회사계상)", "비고(계산 공식)"]
    for i, h in enumerate(cost_headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    r += 1
    ws.column_dimensions["H"].width = 70

    for row_data in summary["by_cost"] + [summary["total_row"]]:
        is_total = row_data["구분"] == "합계"
        cell = ws.cell(row=r, column=1, value=row_data["구분"])
        cell.border = border
        vals = [row_data["전기말(재계산)"], row_data["전기말(회사계상)"], row_data["당기말(재계산)"],
                row_data["당기말(회사계상)"], row_data["당기 퇴직급여비용(재계산)"],
                row_data["대사차이(재계산-회사계상)"]]
        if is_total:
            cell.font = bold
            cell.fill = total_fill
        for i, v in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            if v is not None:
                cell.number_format = "#,##0"
            if is_total:
                cell.font = bold
                cell.fill = total_fill
            elif i == 7 and _is_significant(v, row_data["당기말(회사계상)"]):
                cell.fill = sig_fill
        note_cell = ws.cell(row=r, column=8, value=FORMULA_NOTE if is_total else None)
        note_cell.border = border
        note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if is_total:
            note_cell.fill = total_fill
            ws.row_dimensions[r].height = 60
        r += 1
    r += 2

    # 1-1) T계정 검증 (분개장 당기지급액 입력 시에만 표시)
    tie_out = summary.get("tie_out")
    if tie_out is not None:
        ws.cell(row=r, column=1,
                value="■ 퇴직급여충당부채 T계정 검증 (기초+당기전입-당기지급액 =? 기말, 전액 회사계상·분개장 기준)").fill = section_fill
        ws.cell(row=r, column=1).font = section_font
        for c in range(2, 7):
            ws.cell(row=r, column=c).fill = section_fill
        r += 2

        tie_headers = ["전기말(회사계상)", "당기 퇴직급여(재계산)", "당기지급액(분개장)",
                        "계산상 당기말", "당기말(회사계상)", "차이(계산상당기말-회사계상)"]
        for i, h in enumerate(tie_headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1
        vals = [tie_out["전기말(회사계상)"], tie_out["당기 퇴직급여(재계산)"], tie_out["당기지급액(분개장)"],
                tie_out["계산상 당기말"], tie_out["당기말(회사계상)"], tie_out["차이(계산상당기말-회사계상)"]]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            cell.number_format = "#,##0"
            if i == 6 and _is_significant(v, tie_out["당기말(회사계상)"]):
                cell.fill = sig_fill
        r += 2

    # 1-2)/1-3) 사업장별/부서별 요약 (해당 컬럼에 실제 값이 있는 경우에만 표시)
    for group_col, section_title in (("사업장", "■ 사업장별 요약"), ("부서", "■ 부서별 요약")):
        rows_data = summary.get("site_summary" if group_col == "사업장" else "dept_summary") or []
        if not rows_data:
            continue

        ws.cell(row=r, column=1, value=section_title).fill = section_fill
        ws.cell(row=r, column=1).font = section_font
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = section_fill
        r += 2

        grp_headers = [group_col, "당기인원수", "전기말(재계산)", "당기말(재계산)", "당기 퇴직급여비용(재계산)"]
        for i, h in enumerate(grp_headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1

        총인원수 = 총전기말 = 총당기말 = 총퇴직급여 = 0
        for row_data in rows_data:
            cell = ws.cell(row=r, column=1, value=row_data[group_col])
            cell.border = border
            vals = [row_data["당기인원수"], row_data["전기말(재계산)"], row_data["당기말(재계산)"], row_data["당기 퇴직급여비용(재계산)"]]
            for i, v in enumerate(vals, start=2):
                cell = ws.cell(row=r, column=i, value=v)
                cell.border = border
                if i > 2:
                    cell.number_format = "#,##0"
            총인원수 += row_data["당기인원수"]
            총전기말 += row_data["전기말(재계산)"]
            총당기말 += row_data["당기말(재계산)"]
            총퇴직급여 += row_data["당기 퇴직급여비용(재계산)"]
            r += 1

        cell = ws.cell(row=r, column=1, value="합계")
        cell.font = bold
        cell.fill = total_fill
        cell.border = border
        for i, v in enumerate([총인원수, 총전기말, 총당기말, 총퇴직급여], start=2):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            cell.font = bold
            cell.fill = total_fill
            if i > 2:
                cell.number_format = "#,##0"
        r += 3

    # 2) 인원 변동 요약
    ws.cell(row=r, column=1, value="■ 인원 변동 요약").fill = section_fill
    ws.cell(row=r, column=1).font = section_font
    for c in range(2, 6):
        ws.cell(row=r, column=c).fill = section_fill
    r += 2

    hc_headers = ["구분", "전기말인원수", "신규입사인원수", "퇴사인원수", "당기말인원수"]
    for i, h in enumerate(hc_headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    r += 1

    for hc in summary["headcount_by_cost"]:
        cell = ws.cell(row=r, column=1, value=hc["원가구분"])
        cell.border = border
        for i, k in enumerate(["전기말인원수", "신규입사인원수", "퇴사인원수", "당기말인원수"], start=2):
            cell = ws.cell(row=r, column=i, value=hc[k])
            cell.border = border
        r += 1
    hc = summary["headcount"]
    cell = ws.cell(row=r, column=1, value="합계")
    cell.font = bold
    cell.fill = total_fill
    cell.border = border
    for i, k in enumerate(["전기말인원수", "신규입사인원수", "퇴사인원수", "당기말인원수"], start=2):
        cell = ws.cell(row=r, column=i, value=hc[k])
        cell.border = border
        cell.font = bold
        cell.fill = total_fill
    r += 3

    # 3) 신규입사자 명단
    ws.cell(row=r, column=1,
            value="■ 신규입사자 명단 ('당기정보'에는 있으나 '전기정보'에는 없음)").fill = section_fill
    ws.cell(row=r, column=1).font = section_font
    for c in range(2, 8):
        ws.cell(row=r, column=c).fill = section_fill
    r += 2

    new_headers = ["사업장", "부서", "사번", "성명", "직급", "원가구분", "입사일"]
    for i, h in enumerate(new_headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    r += 1
    신규입사자 = summary["신규입사자"]
    if 신규입사자.empty:
        ws.cell(row=r, column=1, value="(해당 없음)")
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = border
        r += 1
    else:
        for _, row in 신규입사자.iterrows():
            for i, h in enumerate(new_headers, start=1):
                val = row[h]
                cell = ws.cell(row=r, column=i, value=(None if pd.isna(val) else val))
                cell.border = border
                if h == "입사일" and val is not None and not pd.isna(val):
                    cell.number_format = "yyyy-mm-dd"
            r += 1
    r += 2

    # 4) 퇴사자 명단 대사 — 자동 산출('전기정보'에는 있으나 '당기정보'에는 없음) vs '당기퇴사자' 시트(실제 퇴사자) 비교
    ws.cell(row=r, column=1,
            value="■ 퇴사자 명단 대사 (자동 산출 명단 vs '당기퇴사자' 시트 입력 명단 비교)").fill = section_fill
    ws.cell(row=r, column=1).font = section_font
    for c in range(2, 9):
        ws.cell(row=r, column=c).fill = section_fill
    r += 1
    ws.cell(row=r, column=1,
            value="※ 양쪽 명단에 모두 있으면 '이상없음', 한쪽에만 있으면 비고에 원인을 표시합니다. "
                  "사번(없으면 성명) 기준 매칭입니다.").font = Font(italic=True, color="808080", size=9)
    r += 1

    match_headers = ["사업장", "부서", "사번", "직급", "원가구분",
                      "전기정보 있으나 당기정보 없음", "실제 퇴사자", "비고"]
    for i, h in enumerate(match_headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        if h == "비고":
            ws.column_dimensions[get_column_letter(i)].width = 46
    r += 1
    leaver_match = summary.get("leaver_match")
    if leaver_match is None or leaver_match.empty:
        ws.cell(row=r, column=1, value="(해당 없음)")
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = border
        r += 1
    else:
        for _, row in leaver_match.iterrows():
            for i, h in enumerate(match_headers, start=1):
                val = row[h]
                cell = ws.cell(row=r, column=i, value=(None if pd.isna(val) or val == "" else val))
                cell.border = border
                if h == "비고" and val:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if val != "이상없음":
                        cell.fill = sig_fill
            r += 1

        # 마지막 데이터 행 바로 아래에 두 명단(전기정보 있으나 당기정보 없음 / 실제 퇴사자) 각각의 인원수 합계
        전기없음_인원수 = int((leaver_match["전기정보 있으나 당기정보 없음"] != "").sum())
        실제퇴사_인원수 = int((leaver_match["실제 퇴사자"] != "").sum())
        cell = ws.cell(row=r, column=1, value="합계(인원수)")
        cell.font = bold
        cell.fill = total_fill
        cell.border = border
        for c in range(2, 9):
            ws.cell(row=r, column=c).fill = total_fill
            ws.cell(row=r, column=c).border = border
        cell = ws.cell(row=r, column=6, value=전기없음_인원수)
        cell.font = bold
        cell.alignment = center
        cell = ws.cell(row=r, column=7, value=실제퇴사_인원수)
        cell.font = bold
        cell.alignment = center
        r += 1
    r += 2

    # 5) 퇴사자 금액차이 분석 ('당기퇴사자' 시트에 입력된 경우만 표시) — 순수 금액 비교, 품질 경고는 4)번 표 참고
    leaver_recon = summary.get("leaver_recon")
    if leaver_recon is not None and not leaver_recon.empty:
        ws.cell(row=r, column=1,
                value="■ 퇴사자 금액차이 분석 ('당기퇴사자' 시트 입력분 — 전기말 추계액 vs 실제지급액)").fill = section_fill
        ws.cell(row=r, column=1).font = section_font
        for c in range(2, 10):
            ws.cell(row=r, column=c).fill = section_fill
        r += 2

        recon_headers = ["사업장", "부서", "사번", "성명", "직급", "원가구분",
                          "전기말 추계액(계산)", "실제지급액(입력)", "차이(추계액-실제지급액)"]
        for i, h in enumerate(recon_headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1
        for _, row in leaver_recon.iterrows():
            for i, h in enumerate(recon_headers, start=1):
                val = row[h]
                cell = ws.cell(row=r, column=i, value=(None if pd.isna(val) or val == "" else val))
                cell.border = border
                if h in ("전기말 추계액(계산)", "실제지급액(입력)", "차이(추계액-실제지급액)") and val is not None and not pd.isna(val):
                    cell.number_format = "#,##0"
                    if h == "차이(추계액-실제지급액)" and _is_significant(val, row["전기말 추계액(계산)"]):
                        cell.fill = sig_fill
            r += 1

        # 전기말 추계액(계산)/실제지급액(입력)/차이 합계
        총전기말추계액 = float(leaver_recon["전기말 추계액(계산)"].sum(skipna=True))
        총실제지급액 = float(leaver_recon["실제지급액(입력)"].sum(skipna=True))
        cell = ws.cell(row=r, column=1, value="합계")
        cell.font = bold
        cell.fill = total_fill
        cell.border = border
        for c in range(2, 10):
            ws.cell(row=r, column=c).fill = total_fill
            ws.cell(row=r, column=c).border = border
        for col, v in ((7, 총전기말추계액), (8, 총실제지급액), (9, 총전기말추계액 - 총실제지급액)):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = bold
            cell.number_format = "#,##0"


# ── 엑셀 저장 ────────────────────────────────────────────────────────────────

MONEY_COLS = ["월평균급여(원)", "상여금(연간, 원)", "월평균급여(상여금가산후)(계산)", "1일평균임금(계산)",
              "전기말 퇴직금추계액(계산)", "당기말 퇴직금추계액(계산)",
              "당기말 회사계상 퇴직금추계액(원)", "당기말 차이(계산-회사계상)", "당기 퇴직급여(계산)"]
DATE_COLS = ["입사일", "중간정산일"]
# 대사 차이 컬럼 → 유의성 판단 기준(분모)이 되는 회사계상액 컬럼 (depreciation_analyzer와 동일 패턴)
DIFF_BASE_COLS = {"당기말 차이(계산-회사계상)": "당기말 회사계상 퇴직금추계액(원)"}

PRIOR_MONEY_COLS = ["월평균급여(원)", "상여금(연간, 원)", "월평균급여(상여금가산후)(계산)", "1일평균임금(계산)",
                     "전기말 퇴직금추계액(계산)", "전기말 회사계상 퇴직금추계액(원)", "전기말 차이(계산-회사계상)"]
PRIOR_DIFF_BASE_COLS = {"전기말 차이(계산-회사계상)": "전기말 회사계상 퇴직금추계액(원)"}


def save_results(df: pd.DataFrame, output_path: str, company: str, target_fy: str,
                  당기결산일: date, 전기결산일: date, 전기_employees: list,
                  신규입사자_recs: list, 퇴사자_recs: list, basis: dict,
                  전기_by_key: dict, 전기말_balances: dict, leaver_payments: list,
                  당기_by_key: dict, interim_month: int = None):
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약표"
    summary = build_summary(df, 전기_employees, 신규입사자_recs, 퇴사자_recs, basis,
                             전기_by_key, 전기말_balances, leaver_payments, 전기결산일, 당기_by_key)
    write_summary_sheet(ws_summary, summary, company, target_fy, 당기결산일, 전기결산일, interim_month)

    ws = wb.create_sheet("인원별추계명세")

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    subtotal_fill = PatternFill("solid", fgColor="D9E1F2")
    total_fill = PatternFill("solid", fgColor="9DC3E6")
    settle_fill = PatternFill("solid", fgColor="FCE4D6")
    sig_fill = PatternFill("solid", fgColor="FFFF00")
    bold = Font(bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    period_note = f", ~{interim_month}월 중간결산(반기 등)" if interim_month else ""
    ws.cell(row=1, column=1,
            value=(f"인원별 퇴직금 추계 명세서 (회사: {company}, 회계연도: {target_fy}{period_note}, "
                   f"당기말 결산기준일: {당기결산일})")).font = Font(bold=True, size=13)

    headers = list(df.columns) if not df.empty else []
    header_row = 3
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = 16 if h != "비고" else 40
    ws.freeze_panes = f"A{header_row + 1}"

    r = header_row + 1
    if df.empty:
        ws.cell(row=r, column=1, value="(인원 데이터 없음)")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        return

    totals = {c: 0.0 for c in MONEY_COLS}
    grand_totals = {c: 0.0 for c in MONEY_COLS}

    for cost, gdf in df.groupby("원가구분", sort=False):
        for _, row in gdf.iterrows():
            is_settled = bool(row.get("중간정산반영"))
            for i, h in enumerate(headers, start=1):
                val = row[h]
                cell = ws.cell(row=r, column=i, value=(None if pd.isna(val) else val))
                cell.border = border
                if h in DATE_COLS and val is not None and not pd.isna(val):
                    cell.number_format = "yyyy-mm-dd"
                if h in MONEY_COLS:
                    cell.number_format = "#,##0"
                if h in ("당기신규입사", "중간정산반영"):
                    cell.value = "O" if val is True else None
                    cell.alignment = center
                if h in DIFF_BASE_COLS and _is_significant(val, row.get(DIFF_BASE_COLS[h])):
                    cell.fill = sig_fill
                if is_settled and h not in ("비고",):
                    cell.fill = settle_fill
            for c in MONEY_COLS:
                v = row.get(c)
                if v is not None and not pd.isna(v):
                    totals[c] += v
                    grand_totals[c] += v
            r += 1

        ws.cell(row=r, column=1, value=f"[{cost} 소계]").font = bold
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=i)
            cell.fill = subtotal_fill
            cell.border = border
            if h in MONEY_COLS:
                cell.value = totals[h]
                cell.number_format = "#,##0"
                cell.font = bold
        r += 1
        totals = {c: 0.0 for c in MONEY_COLS}

    ws.cell(row=r, column=1, value="총계").font = Font(bold=True, size=11)
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=i)
        cell.fill = total_fill
        cell.border = border
        if h in MONEY_COLS:
            cell.value = grand_totals[h]
            cell.number_format = "#,##0"
            cell.font = bold

    # 전기인원별추계명세 — '전기정보' 인원별 전기말 재계산액과 '회사계상 퇴직금추계액(원)' 대사
    prior_df = build_prior_schedule_table(전기_employees, 전기결산일)
    ws_prior = wb.create_sheet("전기인원별추계명세")
    ws_prior.cell(row=1, column=1,
                  value=(f"전기인원별 퇴직금 추계 명세서 (회사: {company}, 회계연도: {target_fy}, "
                         f"전기말 결산기준일: {전기결산일})")).font = Font(bold=True, size=13)

    prior_headers = list(prior_df.columns) if not prior_df.empty else []
    for i, h in enumerate(prior_headers, start=1):
        c = ws_prior.cell(row=header_row, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
        ws_prior.column_dimensions[get_column_letter(i)].width = 16 if h != "비고" else 40
    ws_prior.freeze_panes = f"A{header_row + 1}"

    pr = header_row + 1
    if prior_df.empty:
        ws_prior.cell(row=pr, column=1, value="(인원 데이터 없음)")
    else:
        p_totals = {c: 0.0 for c in PRIOR_MONEY_COLS}
        p_grand_totals = {c: 0.0 for c in PRIOR_MONEY_COLS}
        for cost, gdf in prior_df.groupby("원가구분", sort=False):
            for _, row in gdf.iterrows():
                is_settled = bool(row.get("중간정산반영"))
                for i, h in enumerate(prior_headers, start=1):
                    val = row[h]
                    cell = ws_prior.cell(row=pr, column=i, value=(None if pd.isna(val) else val))
                    cell.border = border
                    if h in DATE_COLS and val is not None and not pd.isna(val):
                        cell.number_format = "yyyy-mm-dd"
                    if h in PRIOR_MONEY_COLS:
                        cell.number_format = "#,##0"
                    if h == "중간정산반영":
                        cell.value = "O" if val is True else None
                        cell.alignment = center
                    if h in PRIOR_DIFF_BASE_COLS and _is_significant(val, row.get(PRIOR_DIFF_BASE_COLS[h])):
                        cell.fill = sig_fill
                    if is_settled and h not in ("비고",):
                        cell.fill = settle_fill
                for c in PRIOR_MONEY_COLS:
                    v = row.get(c)
                    if v is not None and not pd.isna(v):
                        p_totals[c] += v
                        p_grand_totals[c] += v
                pr += 1

            ws_prior.cell(row=pr, column=1, value=f"[{cost} 소계]").font = bold
            for i, h in enumerate(prior_headers, start=1):
                cell = ws_prior.cell(row=pr, column=i)
                cell.fill = subtotal_fill
                cell.border = border
                if h in PRIOR_MONEY_COLS:
                    cell.value = p_totals[h]
                    cell.number_format = "#,##0"
                    cell.font = bold
            pr += 1
            p_totals = {c: 0.0 for c in PRIOR_MONEY_COLS}

        ws_prior.cell(row=pr, column=1, value="총계").font = Font(bold=True, size=11)
        for i, h in enumerate(prior_headers, start=1):
            cell = ws_prior.cell(row=pr, column=i)
            cell.fill = total_fill
            cell.border = border
            if h in PRIOR_MONEY_COLS:
                cell.value = p_grand_totals[h]
                cell.number_format = "#,##0"
                cell.font = bold

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="퇴직급여충당부채 검증앱(일반기업회계기준·퇴직금 추계액 방식)")
    parser.add_argument("company", nargs="?", default=None, help="처리할 회사명 (생략 시 파일 자동 탐색)")
    parser.add_argument("--file", default=None, help="처리할 특정 입력 파일명 (input_data/ 기준)")
    parser.add_argument("--fiscal-month", type=int, default=12, help="결산월 (기본 12월). 예: 6월 결산법인이면 6")
    parser.add_argument("--fiscal-year", default=None, help="검증 대상 회계연도 (예: 2026). 생략 시 입력파일명의 fy 뒤 숫자 사용")
    parser.add_argument("--interim-month", type=int, default=None,
                         help="반기 등 중간결산 검토월 (예: 6). 당기말 결산기준일만 해당 월말로 앞당기고, "
                              "전기말은 직전 회계연도 결산기준일 그대로 사용")
    args = parser.parse_args()

    input_path = _find_input_file(args.company, args.file)
    company = args.company or os.path.basename(input_path).split("_")[1]

    target_fy = args.fiscal_year
    if not target_fy:
        base = os.path.basename(input_path)
        idx = base.lower().find("fy")
        if idx == -1:
            raise ValueError("--fiscal-year 를 지정하거나 파일명에 'fy<연도>'를 포함하세요.")
        digits = ""
        for ch in base[idx + 2:]:
            if ch.isdigit():
                digits += ch
            else:
                break
        target_fy = f"20{digits}" if len(digits) == 2 else digits

    _, fy_end_ym = _fy_bounds(target_fy, args.fiscal_month)
    fy_end_ym = _apply_interim(fy_end_ym, target_fy, args.interim_month)
    당기결산일 = _ym_to_end_date(fy_end_ym)

    prior_fy = str(int(target_fy) - 1)
    _, prior_fy_end_ym = _fy_bounds(prior_fy, args.fiscal_month)
    전기결산일 = _ym_to_end_date(prior_fy_end_ym)

    print(f"[입력] {input_path}")
    interim_note = f", 중간결산월={args.interim_month}(반기 등)" if args.interim_month else ""
    print(f"[대상] 회사={company}, 회계연도={target_fy}, 결산월={args.fiscal_month}{interim_note}")
    print(f"[기준일] 당기말={당기결산일}, 전기말={전기결산일}")

    당기_employees = load_employees(input_path, CURRENT_SHEET)
    전기_employees = load_employees(input_path, PRIOR_SHEET)
    leaver_payments = load_employees(input_path, LEAVER_SHEET)
    basis = load_basis(input_path)
    print(f"[인원 수] 당기={len(당기_employees)}건, 전기={len(전기_employees)}건, 당기퇴사자(지급액 입력)={len(leaver_payments)}건")

    matched = match_periods(당기_employees, 전기_employees)
    전기말_balances = compute_prior_balances(matched["전기_by_key"], 전기결산일)
    df = build_schedule_table(당기_employees, 당기결산일, matched["신규입사_keys"], 전기말_balances)

    suffix = f"_interim{args.interim_month:02d}" if args.interim_month else ""
    output_path = os.path.join(OUTPUT_DIR, f"severance_schedule_{company}_{target_fy}{suffix}.xlsx")
    save_results(df, output_path, company, target_fy, 당기결산일, 전기결산일,
                 전기_employees, matched["신규입사자"], matched["퇴사자"], basis,
                 matched["전기_by_key"], 전기말_balances, leaver_payments,
                 matched["당기_by_key"], args.interim_month)
    print(f"[완료] {output_path}")


if __name__ == "__main__":
    main()
