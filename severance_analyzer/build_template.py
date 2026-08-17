"""표준 입력 템플릿(severance_template.xlsx) 생성 스크립트.

실행: python build_template.py
input_data/severance_template.xlsx 를 새로 만든다(이미 있으면 덮어씀).
회사별 파일은 이 템플릿을 복사해 severance_<company>_information_fy<year>.xlsx 로 저장해서 사용한다.

설계 원칙: 일반기업회계기준 퇴직급여충당부채(퇴직금 추계액 방식) 검증.
  - 결산기준일(당기말/전기말)은 셀에 직접 입력하지 않고, 파일명(fy<연도>)과 실행 시 --fiscal-month로
    depreciation_analyzer/lease_analyzer와 동일한 규칙으로 앱이 계산한다.
  - depreciation_analyzer와 달리 기초(전기말) 잔액을 신뢰하지 않는다. 당기말 퇴충잔액은 인원별
    재직일수·급여만으로 전액 독립 재계산하며, '기준정보' 시트의 전기말/당기말 회사계상 퇴직급여충당부채
    (제조원가분/판관비분)는 대사(비교)용 참고값으로만 쓰인다.
  - 전기/당기 인원은 회사별로 별도 파일을 만드는 대신, 이 한 파일 안에 '당기정보'/'전기정보'
    두 시트로 나눠 입력받는다(회사가 매년 스냅샷으로 제공하는 인원현황을 그대로 붙여넣는 방식).
    당기 계산(추계액)은 '당기정보'만 사용하고, '전기정보'는 사번(없으면 성명) 기준으로 매칭해
    신규입사자/퇴사자 명단을 산출하는 데에만 쓰인다.
  - '중간정산일'은 재직 상태를 유지한 채 근속기간만 재기산되는 것(퇴사가 아님)이므로,
    재직일수 계산의 기산일(입사일 대신 중간정산일부터 기산)에만 반영된다.
  - '부서' 값이 정확히 '임원'인 행은 일반 직원과 다른 산식(소득세법상 임원 퇴직금 한도 산식)을 쓴다:
    (급여기준액×근속연수 + 급여기준액×근속월/10) × 배수. '배수(임원)'는 이 행에서만 읽어 적용한다.
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(HERE, "input_data", "severance_template.xlsx")

# (그룹헤더, 상세헤더, 열너비) — '당기정보'/'전기정보' 두 시트가 동일한 구조를 공유한다.
COLUMNS = [
    ("인적사항", "사업장", 14),
    ("인적사항", "부서", 14),
    ("인적사항", "사번", 12),
    ("인적사항", "성명", 12),
    ("인적사항", "직급", 10),
    ("재직정보", "원가구분(제조원가/판관비)", 14),
    ("재직정보", "입사일", 12),
    ("재직정보", "중간정산일", 12),
    ("임금정보", "월평균급여(원)", 14),
    ("임금정보", "상여금(연간, 원)", 14),
    ("임금정보", "배수(임원)", 10),
    ("기타", "비고", 30),
]

MONEY_FIELDS = ("월평균급여(원)", "상여금(연간, 원)")
MULTIPLIER_FIELDS = ("배수(임원)",)
DATE_FIELDS = ("입사일", "중간정산일")

COST_TYPE_OPTIONS = ["제조원가", "판관비"]
EXECUTIVE_DEPT = "임원"

BASIS_ROWS = [
    "전기말 회사계상 퇴직급여충당부채(제조원가분)",
    "전기말 회사계상 퇴직급여충당부채(판관비분)",
    "당기말 회사계상 퇴직급여충당부채(제조원가분)",
    "당기말 회사계상 퇴직급여충당부채(판관비분)",
    "당기 퇴직급여충당부채 차변(당기지급액, 분개장 기준)",
]

# '당기퇴사자' 시트 — 선택 입력. 퇴사자 실제 지급액을 알면 전기말 추계액과 인별 대사까지 가능하다.
LEAVER_COLUMNS = [("퇴사자 지급정보", "사번", 14), ("퇴사자 지급정보", "성명", 14),
                   ("퇴사자 지급정보", "실제지급액(원)", 16), ("퇴사자 지급정보", "비고", 30)]
LEAVER_EXAMPLES = [
    ["EMP-0999", "예시)최과장", 38500000,
     "예시) 전기정보에 있는 사번으로 매칭 — 전기말 추계액과 자동 대사됨. 사번 없으면 성명으로 매칭."],
]

# 예시행: 기존재직자(제조/판관, 상여금 있는 경우 포함), 당기 신규입사자, 중간정산 이력자, 임원(배수 적용)
# — '당기정보'/'전기정보' 두 시트에 걸쳐 계속재직자는 양쪽 모두, 신규입사자는 당기만,
#   퇴사자는 전기만 등장시켜 매칭 로직(사번 기준)을 그대로 보여준다.
EXAMPLES_CURRENT = [
    [
        "본사", "생산1팀", "EMP-1001", "예시)홍길동", "과장",
        "제조원가", "2019-03-02", None,
        4200000, 6000000, None, "예시 행 — 실제 인원으로 교체(기존 재직자, 연간상여금 600만원 지급)",
    ],
    [
        "본사", "경영지원팀", "EMP-1002", "예시)김영희", "대리",
        "판관비", "2021-11-15", None,
        3800000, 0, None, "예시 행 — 기존 재직자(상여금 없음)",
    ],
    [
        "본사", "영업팀", "EMP-1003", "예시)이철수", "사원",
        "판관비", "2026-04-01", None,
        3200000, 0, None, "예시) 당기 중 신규입사 — '전기정보'에는 없어 요약표의 '신규입사자 명단'에 자동 표시됨",
    ],
    [
        "제2공장", "생산2팀", "EMP-1004", "예시)박민수", "사원",
        "제조원가", "2018-06-01", "2023-06-30",
        3500000, 4200000, None,
        "예시) 퇴직금 중간정산 이력 있음 — 재직 상태는 계속 유지되며(퇴사 아님), "
        "재직일수만 입사일(2018-06-01)이 아닌 중간정산일(2023-06-30)부터 다시 기산됨",
    ],
    [
        "본사", "임원", "EMP-2001", "예시)정대표", "대표이사",
        "판관비", "2015-01-01", None,
        10000000, 20000000, 2.0,
        "예시) 임원퇴직금지급규정상 배수 2배 적용 — 부서명이 정확히 '임원'인 인원에만 배수가 적용됨",
    ],
]

# 전기 시트: 계속재직자(EMP-1001/1002/1004/2001)는 그대로 유지, 신규입사자(EMP-1003)는 빼고,
# 당기에는 없는 퇴사자(EMP-0999)를 하나 추가해 '퇴사자 명단' 매칭을 보여준다.
EXAMPLES_PRIOR = [
    [
        "본사", "생산1팀", "EMP-1001", "예시)홍길동", "과장",
        "제조원가", "2019-03-02", None,
        4000000, 5500000, None, "예시 행 — 당기와 동일 인원(전기 시점 급여). 이 시트의 급여는 계산에 쓰이지 않음(매칭용)",
    ],
    [
        "본사", "경영지원팀", "EMP-1002", "예시)김영희", "대리",
        "판관비", "2021-11-15", None,
        3800000, 0, None, "예시 행 — 당기와 동일 인원",
    ],
    [
        "제2공장", "생산2팀", "EMP-1004", "예시)박민수", "사원",
        "제조원가", "2018-06-01", "2023-06-30",
        3500000, 4200000, None, "예시 행 — 당기와 동일 인원(중간정산 이력 포함)",
    ],
    [
        "본사", "임원", "EMP-2001", "예시)정대표", "대표이사",
        "판관비", "2015-01-01", None,
        10000000, 18000000, 2.0, "예시 행 — 당기와 동일 인원",
    ],
    [
        "본사", "생산1팀", "EMP-0999", "예시)최과장", "과장",
        "제조원가", "2012-02-01", None,
        4100000, 0, None,
        "예시) 당기 중 퇴사 — '당기정보'에는 이 사번이 없어 요약표의 '퇴사자 명단'에 자동 표시됨",
    ],
]


def build():
    wb = openpyxl.Workbook()
    ws_current = wb.active
    ws_current.title = "당기정보"
    ws_prior = wb.create_sheet("전기정보")

    group_fill = PatternFill("solid", fgColor="D9E1F2")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    group_font = Font(bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _col_letter(header_name: str) -> str:
        idx = next(i for i, (_, name, _) in enumerate(COLUMNS, start=1) if name == header_name)
        return get_column_letter(idx)

    def build_sheet(ws, examples: list):
        # 1행: 그룹 헤더 (병합)
        start = 1
        prev_group = COLUMNS[0][0]
        for i, (group, _, _) in enumerate(COLUMNS, start=1):
            if group != prev_group:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=i - 1)
                start = i
                prev_group = group
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=len(COLUMNS))

        for i, (group, name, width) in enumerate(COLUMNS, start=1):
            c1 = ws.cell(row=1, column=i, value=group if i == 1 or COLUMNS[i - 2][0] != group else None)
            c1.fill = group_fill
            c1.font = group_font
            c1.alignment = center
            c1.border = border

            c2 = ws.cell(row=2, column=i, value=name)
            c2.fill = header_fill
            c2.font = header_font
            c2.alignment = center
            c2.border = border

            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 32
        ws.freeze_panes = "A3"

        for r, row in enumerate(examples, start=3):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = border
                field = COLUMNS[c - 1][1]
                if field in DATE_FIELDS and val:
                    cell.number_format = "yyyy-mm-dd"
                if field in MONEY_FIELDS:
                    cell.number_format = "#,##0"
                if field in MULTIPLIER_FIELDS and val is not None:
                    cell.number_format = "0.0"

        # 데이터 유효성 검사 (300행까지)
        last_row = 300
        dv_cost = DataValidation(
            type="list", formula1=f'"{",".join(COST_TYPE_OPTIONS)}"',
            allow_blank=True, showErrorMessage=True,
        )
        dv_cost.error = "제조원가 또는 판관비 중 선택하세요."
        ws.add_data_validation(dv_cost)
        col = _col_letter("원가구분(제조원가/판관비)")
        dv_cost.add(f"{col}3:{col}{last_row}")

    build_sheet(ws_current, EXAMPLES_CURRENT)
    build_sheet(ws_prior, EXAMPLES_PRIOR)

    # 당기퇴사자 시트 — 선택 입력. 채우면 퇴사자별 전기말 추계액 vs 실제지급액 대사가 요약표에 추가된다.
    ws_leaver = wb.create_sheet("당기퇴사자")
    ws_leaver.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(LEAVER_COLUMNS))
    c1 = ws_leaver.cell(row=1, column=1, value=LEAVER_COLUMNS[0][0])
    c1.fill = group_fill
    c1.font = group_font
    c1.alignment = center
    c1.border = border
    for i, (_, name, width) in enumerate(LEAVER_COLUMNS, start=1):
        c2 = ws_leaver.cell(row=2, column=i, value=name)
        c2.fill = header_fill
        c2.font = header_font
        c2.alignment = center
        c2.border = border
        ws_leaver.column_dimensions[get_column_letter(i)].width = width
    ws_leaver.row_dimensions[1].height = 20
    ws_leaver.row_dimensions[2].height = 32
    ws_leaver.freeze_panes = "A3"
    for r, row in enumerate(LEAVER_EXAMPLES, start=3):
        for c, val in enumerate(row, start=1):
            cell = ws_leaver.cell(row=r, column=c, value=val)
            cell.border = border
            if LEAVER_COLUMNS[c - 1][1] == "실제지급액(원)":
                cell.number_format = "#,##0"

    # 기준정보 시트 — 전기말/당기말 회사계상 충당부채(대사용)
    basis = wb.create_sheet("기준정보")
    basis.column_dimensions["A"].width = 42
    basis.column_dimensions["B"].width = 18
    basis.column_dimensions["C"].width = 46

    basis.cell(row=1, column=1, value="퇴직급여충당부채 기준정보 (회사계상액 — 대사용)").font = Font(bold=True, size=13)
    header_row = 3
    for i, h in enumerate(["항목", "금액(원)", "설명"], start=1):
        cell = basis.cell(row=header_row, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    notes = [
        "전기말 재무제표(또는 결산정산표)상 실제 계상액 — 앱이 '전기정보' 시트로 인별 재계산한 전기말 잔액과 비교해 차이(대사)를 표시함(재계산 자체에는 반영되지 않음)",
        "전기말 재무제표(또는 결산정산표)상 실제 계상액 — 위와 동일",
        "당기말 재무제표상 회사 계상액 — 앱의 인원별 재계산 합계와 비교해 차이(대사)를 표시함",
        "당기말 재무제표상 회사 계상액",
        "분개장분석(journal_analyzer)에서 '퇴직급여충당부채' 계정의 당기 차변(퇴직금 지급으로 인한 감소) 합계를 뽑아 입력. "
        "전기말(회사계상)+당기 퇴직급여(재계산)-이 값이 당기말(회사계상)과 맞는지 요약표에서 자동 검증(T계정 tie-out). 모르면 비워둬도 됨(해당 검증만 생략).",
    ]
    for i, (label, note) in enumerate(zip(BASIS_ROWS, notes), start=header_row + 1):
        c1 = basis.cell(row=i, column=1, value=label)
        c1.border = border
        c2 = basis.cell(row=i, column=2)
        c2.border = border
        c2.number_format = "#,##0"
        c2.fill = PatternFill("solid", fgColor="FFF2CC")
        c3 = basis.cell(row=i, column=3, value=note)
        c3.border = border
        c3.alignment = Alignment(wrap_text=True, vertical="top")

    basis.cell(row=header_row + len(BASIS_ROWS) + 2, column=1,
               value="※ 결산기준일(당기말/전기말)은 이 시트에 입력하지 않습니다 — 실행 시 파일명(fy<연도>)과 --fiscal-month 옵션으로 자동 결정됩니다.").font = Font(italic=True, color="808080")

    # 안내 시트
    guide = wb.create_sheet("작성안내")
    guide.column_dimensions["A"].width = 100
    lines = [
        "퇴직급여충당부채 검증앱(일반기업회계기준·퇴직금 추계액 방식) — 입력 템플릿 작성 안내",
        "",
        "핵심 원칙: 결산기준일 현재 재직 중인 임직원을 대상으로, 근로기준법상 퇴직금 산정 방식(간편식)으로",
        "  개인별 퇴직금 추계액을 계산해 합산한 값을 '재계산된 퇴직급여충당부채'로 봅니다.",
        "  감가상각비 앱과 달리 기초(전기말) 잔액도 회사계상 입력값을 신뢰하지 않고, 전기말 잔액 역시",
        "  인원별로 동일한 방식으로 독립 재계산합니다 — 회사계상 전기말/당기말 잔액은 대사(비교)용으로만 쓰입니다.",
        "",
        "0. 시트 구성 — '당기정보'와 '전기정보' 두 시트로 나뉩니다(회사별 파일을 따로 만들 필요 없음).",
        "   '당기정보': 이번 결산기준일 현재 재직 중인 인원 전체 + 급여정보. 당기말 추계액 계산에 쓰입니다.",
        "   '전기정보': 직전 결산기준일 현재 재직 중이었던 인원 전체 + 그 시점 급여정보(회사가 작년에 준",
        "     인원현황을 그대로 붙여넣으면 됨). 이 시트도 같은 산식으로 전기말 추계액을 계산하는 데 쓰이고,",
        "     사번(없으면 성명) 기준으로 '당기정보'와 매칭해 신규입사자/퇴사자 명단을 만드는 데에도 쓰입니다.",
        "   ※ 사번이 두 시트에서 반드시 동일해야 정확히 매칭됩니다. 사번이 없으면 성명으로 매칭하니 동명이인에 주의하세요.",
        "   ※ '당기 퇴직급여(전입액에 해당)' = 인원별 (당기말 추계액 − 전기말 추계액)의 합계로 요약표에 자동 산출됩니다.",
        "     신규입사자는 전기말 추계액을 0으로 보고, 퇴사자는 '당기 퇴직급여' 계산에 포함되지 않습니다",
        "     (당기말 표에 아예 나타나지 않는 인원이므로 — 이 부분은 별도 지급액 관리가 필요합니다).",
        "",
        "1. 인원 1명 = 1행. 두 시트 모두 같은 컬럼 구조입니다. 계속 재직 중인 사람은 두 시트 모두에,",
        "   당기 신규입사자는 '당기정보'에만, 당기 중 퇴사한 사람은 '전기정보'에만 입력하면 됩니다.",
        "",
        "2. 재직정보",
        "   '원가구분': 급여를 제조원가(생산직 등)로 처리하는지 판관비(관리직 등)로 처리하는지 선택.",
        "   '입사일': 필수. 근속일수 계산의 기준이 됩니다.",
        "   '중간정산일': 퇴직금 중간정산을 받은 이력이 있는 인원만 입력. 중간정산은 재직 상태를 그대로",
        "     유지한 채 근속기간만 다시 기산되는 것(퇴사가 아님)이므로, 재직일수 계산의 기산일이",
        "     입사일 대신 이 날짜로 바뀔 뿐입니다. 정산 이력이 없으면 비워둡니다.",
        "",
        "3. 임금정보 (두 시트 모두 동일 산식 — '당기정보'는 당기말, '전기정보'는 전기말 추계액에 각각 쓰임)",
        "   '월평균급여(원)': 결산기준일 현재(또는 최근 급여 수준)의 월 급여(기본급 등, 상여금 제외).",
        "   '상여금(연간, 원)': 연간 지급된 상여금 총액. 상여금이 없으면 0 또는 비워둡니다.",
        "   급여기준액 = 월평균급여 + 상여금(연간) × 3 / 12  (상여금 중 최근 3개월분 상당액만 가산)",
        "",
        "   [일반 직원] 근로기준법 간편 산식",
        "     1일평균임금 = 급여기준액 × 12 / 365",
        "     재직일수 = 당기 결산기준일 − 기산일(중간정산일이 있으면 그 날짜, 없으면 입사일)",
        "     퇴직금 추계액 = 1일평균임금 × 30일 × (재직일수 / 365)",
        "",
        "   [임원] '부서' 값이 정확히 '임원'인 행 — 소득세법상 임원 퇴직금 한도 산식으로 계산됩니다(위 일반 직원 산식과 다름).",
        "     근속연수 = 당기 결산기준일의 연도 − 기산일의 연도",
        "     근속월(1년 미만 잔여월 보정) = 13 − 기산일의 월",
        "     퇴직금 추계액 = (급여기준액 × 근속연수 + 급여기준액 × 근속월 / 10) × 배수",
        "   '배수(임원)': 임원퇴직금지급규정 등에 따라 배수를 적용하는 회사만 사용합니다.",
        "     '부서' 값이 정확히 '임원'인 행에서만 이 값을 읽어 적용합니다(오탈자·다른 부서명이면 무시되고 일반 직원 산식으로 계산됨).",
        "     일반 직원 행은 이 칸에 값이 있어도 무시됩니다. 임원인데 배수를 비워두면 1배로 계산되고 비고에 안내가 남습니다.",
        "",
        "4. 신규입사자/퇴사자 명단 (요약표에 자동 산출)",
        "   신규입사자 = '당기정보'에는 있으나 '전기정보'에는 없는 사번(또는 성명).",
        "   퇴사자 = '전기정보'에는 있으나 '당기정보'에는 없는 사번(또는 성명).",
        "   '전기정보' 시트를 비워두면(또는 시트 자체를 지우면) 인원변동 명단만 생략되고 당기 계산은 정상 진행됩니다.",
        "   퇴사자 명단의 '비고'에는 다음 두 가지가 자동으로 표시됩니다.",
        "     - 전기 결산기준일 기준 근속기간이 1년 미만이면 '전기 퇴직급여설정대상 아님(근속기간 1년 미만으로 추정)'",
        "       (근로기준법상 계속근로기간 1년 미만은 퇴직금 지급 요건을 충족하지 못하므로, 이 인원은 애초에 전기말",
        "       퇴직급여충당부채 설정 대상이 아니었을 가능성이 있다는 안내입니다).",
        "     - '당기퇴사자' 시트(아래 5번)에 대응하는 지급액 입력이 없으면 '명단 비교 시 차이 있음' 경고.",
        "",
        "5. '당기퇴사자' 시트 (선택) — 퇴사자의 실제 지급액을 알면 채워서 인별 대사를 추가로 받을 수 있습니다.",
        "   '사번'(없으면 성명) + '실제지급액(원)'만 입력하면, 앱이 '전기정보' 시트에서 같은 사번(성명)을 찾아",
        "   그 사람의 전기말 추계액과 자동으로 비교(대사)해 요약표에 표시합니다. 사업장/부서/직급 등은",
        "   '전기정보'에서 그대로 가져오므로 다시 입력할 필요가 없습니다. 모르는 퇴사자는 행을 비워두면 됩니다",
        "   (해당 인원만 대사가 생략되고, 나머지 재계산·집계에는 영향이 없습니다).",
        "   같은 인원(사번/성명)이 두 번 이상 입력되면 '이중기입의심' 경고가 표시됩니다(동명이인일 수도 있으니 확인).",
        "",
        "6. '기준정보' 시트 — 전기말/당기말 회사계상 퇴직급여충당부채(제조원가분/판관비분) 4칸과,",
        "   당기 퇴직급여충당부채 차변(당기지급액) 1칸을 입력합니다. 다섯 값 모두 인별 재계산액과의 대사(비교)용",
        "   참고값일 뿐, 재계산 자체(전기말·당기말·당기 퇴직급여)에는 반영되지 않습니다.",
        "   '당기지급액'은 journal_analyzer(분개장분석) 메뉴에서 '퇴직급여충당부채' 계정의 당기 차변 합계를",
        "   뽑아 입력하면 됩니다 — 전기말(회사계상)+당기 퇴직급여(재계산)-당기지급액이 당기말(회사계상)과",
        "   맞는지 요약표에서 자동으로 T계정 검증(tie-out)합니다. 모르면 비워둬도 됩니다(해당 검증만 생략).",
        "",
        "7. 결산기준일은 이 파일에 입력하지 않습니다. 실행 시 파일명의 'fy<연도>'와 --fiscal-month 옵션(기본 12월)으로",
        "   depreciation_analyzer/lease_analyzer와 동일한 규칙으로 당기말·전기말 결산기준일을 자동 계산합니다.",
        "   예) --fiscal-month 12 --fiscal-year 2026 → 당기말 2026-12-31, 전기말 2025-12-31.",
        "",
        "8. 파일명 규칙: 이 템플릿을 복사해 'severance_<회사명>_information_fy<회계연도>.xlsx' 로 저장하세요.",
        "   예) severance_kyungnam_information_fy2026.xlsx (전기·당기 데이터가 모두 이 한 파일 안에 들어갑니다).",
    ]
    for i, line in enumerate(lines, start=1):
        cell = guide.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    os.makedirs(os.path.join(HERE, "input_data"), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"템플릿 생성 완료: {OUT_PATH}")


if __name__ == "__main__":
    build()
