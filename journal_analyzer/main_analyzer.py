#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
journal_analyzer/main_analyzer.py  v3
analysis.py 20개 분석 메뉴 — DataFrame 반환 방식

폴더 구조:
  journal_analyzer/
  ├── main_analyzer.py
  └── {company}/
      ├── task_list_{company}.xlsx      ← 실행 제어 + 파라미터
      ├── data/
      │   ├── current/                  ← 당기 분개장 (csv/xlsx)
      │   └── previous/                 ← 전기 분개장 (csv/xlsx)
      └── results/                      ← 분석 결과 저장

task_list 파라미터 시트 규격:
  거래처비교   : 계정과목 / 금액열(차변·대변·both) / 실행여부
  벤포드분석   : 계정과목 / 금액열 / 실행여부
  일자차이분석 : 기준일수 / 실행여부
  상대계정분석 : 계정과목 / 금액열 / 실행여부
  키워드검색   : 키워드 / 실행여부
  라운드넘버   : 계정과목 / 금액열 / 최소금액 / 실행여부
  특수관계자   : 거래처명 / 실행여부
  자산부채교차 : 구분(자산·부채) / 계정과목 / 실행여부
  매출비용교차 : 구분(매출·비용) / 계정과목 / 실행여부
  심층분석     : 계정과목 / 개수 / 금액열(차변·대변·both) / 실행여부
  AI계정별분석 : 계정과목 / 실행여부
  AI계정별분석_실행 : 계정과목 / 실행여부  (Gemini 호출 → AI검토결과 + AI검토_확인전표 시트)
  감가상각_평가손익분석 : 계정과목 / 금액열 / 그룹기준열 / 실행여부 / 비고
      (8번 상대계정분석과 동일 로직 — 감가상각비/외화환산손익/평가손익/대손상각비 등
       손익 계정의 상대계정(자산 누계액 등) 금액을 전표 단위로 매칭해 추출. (Phase 1)
       같은 task_list에 '감가상각_유형자산롤포워드' 시트가 있으면 Phase 2도 함께 실행:
  감가상각_유형자산롤포워드 : 유형자산계정명 / 감가상각누계액계정명 / 대체상대계정 /
                              취득원가_수동조정 / 상각누계액_수동조정 / 실행여부
      (유형자산계정별 취득원가·감가상각누계액 기초/당기증가/당기감소/기말 롤포워드.
       당기 감가상각비는 Phase 1의 상대_감가상각* 매칭 결과에서 가져옴.
       기초잔액은 data/previous의 전기 계정별_거래처별명세 파일이 있는 회사만 채워짐.
       대체상대계정: 다른 유형자산으로 계정대체(예: 건물→투자부동산_건물, 건설중인자산→
       본계정)가 있는 경우 그 상대쪽 유형자산계정명을 적으면, 전표 매칭으로 당기증가/
       당기감소를 '대체'분과 '기타'분으로 분리 표시함. 콤마로 여러 개 지정 가능
       (예: 건물 행 = "건설중인자산,투자부동산_건물" — 건설중인자산에서 들어오고
       투자부동산_건물로 나가는 두 방향을 동시에 잡음). 없으면 비워둠.
       수동조정 2종: 전표 매칭으로 못 잡는 잔여 차이(상각누계액_미매칭차이 컬럼 참고)를
       감사인이 직접 확인 후 숫자로 입력하면 기말잔액 계산에 반영됨. 기본 0.)
       같은 task_list에 '유형자산_처분손익' 시트가 있으면 Phase 3도 함께 실행:
  유형자산_처분손익 : 처분이익계정명 / 처분손실계정명 / 실행여부
      (처분이익·처분손실 계정이 등장한 전표(전표그룹키)마다 같은 전표의 유형자산
       취득원가·감가상각누계액(감가상각_유형자산롤포워드 시트의 계정 쌍 기준)을
       찾아 처분 건별로 한 줄씩 유형자산_처분손익명세 시트를 만든다. 장부가액=
       취득원가_처분분-감가상각누계액_처분분, 처분가액(역산)=장부가액+처분손익.
       한계: 개별 자산이 아닌 계정과목×전표 단위로만 구분됨 — 한 전표에 여러
       자산을 묶어 처분하면 그 전표 전체가 한 줄로 잡힘.)
  거래처분석   : 작업명 / 계정과목 / 거래처명 / 금액열 / 실행여부
  벤포드이탈   : 계정과목 / 금액열 / 임계값 / 최대건수 / 실행여부

4번(데이터개요) 부가 기능 — 당기 계정별원장 기초/기말잔액표:
  data/current 폴더에 파일명에 '계정별원장'이 들어간 xlsx가 있으면(파라미터 불필요,
  회사 무관 자동 감지), 4번 실행 시 '계정별원장_잔액표' 시트를 추가로 만든다.
  시트 1개 = 계정 1개(예: '0_당좌예금(10200)') 형식을 전제로 시트명 앞의 순번 접두어와
  뒤의 '(계정코드)' 접미어를 정규식으로 제거해 계정명을 뽑는다.
  각 시트에서 적요란이 '전기이월'/'기초잔액'인 행(공백·대괄호 제거 후 비교)의
  차변/대변 금액을 기초잔액으로, '월계'/'누계'/'합계' 소계 행은 제외하고 나머지
  실제 거래행의 차변합계·대변합계를 구한 뒤, 자산/비용 성격(기초+차변-대변)과
  부채/자본/수익 성격(기초-차변+대변) 두 계산식 중 원장의 마지막 행 잔액과 일치하는
  쪽으로 구분을 자동 판정한다(수기 계정과목표 매핑 불필요). 둘 다 불일치하면
  '검증필요'로 표시.

실행:  python main_analyzer.py sejoong
"""

# =============================================================================
# 0. Imports & 상수
# =============================================================================
import sys, os, glob, re, io, argparse, warnings, unicodedata

# Windows 터미널 한글 깨짐 방지 (cp949 → UTF-8 강제)
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore', category=pd.errors.DtypeWarning)

BASE_DIR          = os.path.dirname(os.path.abspath(__file__))
TASK_MASTER_SHEET = '분석목록'

# .env 로드 (GEMINI_API_KEY 등 — 프로젝트 루트 기준)
from dotenv import load_dotenv
load_dotenv(os.path.join(BASE_DIR, '..', '.env'))

# 컬럼명 상수
COL_DATE       = '전표일자'
COL_JOURNAL_ID = '전표번호'
COL_ACCOUNT    = '계정명'
COL_CLIENT     = '거래처명'
COL_DESC       = '적요'
COL_DEBIT      = '차변'
COL_CREDIT     = '대변'
COL_EMPLOYEE   = '사원명'

# 전표 단위 매칭용 그룹키 (전표일자+전표번호 결합).
# 일부 회사(예: kyungnam)는 전표번호가 전표일자별로 재사용되는 일별 순번이라
# 전표번호 단독으로는 "같은 전표"를 유일하게 특정할 수 없다 (2026-08-12 발견).
# 상대계정/대체 매칭처럼 "같은 전표에 속한 거래" 판정이 필요한 로직은
# 전표번호 대신 이 컬럼을 사용한다. 화면 표시용 전표번호(COL_JOURNAL_ID)는 그대로 둔다.
COL_JOURNAL_KEY = '전표그룹키'

# 분석 상수
BENFORD_MIN_ROWS        = 5
BENFORD_PROBS           = {1:0.301,2:0.176,3:0.125,4:0.097,5:0.079,6:0.067,7:0.058,8:0.051,9:0.046}
DEFAULT_BENFORD_TARGETS = [('복리후생비','차변'),('접대비','차변'),('여비교통비','차변')]
DEFAULT_KEYWORDS        = ['상품권','접대','가수금','선물','회식','결산수정','비자금','가지급','리베이트','현금']
MASK_TARGET_COLS        = ['적요','관리항목4','비고','내용']
GLOBAL_SAFE_MAP: dict   = {}
_CLIENT_COUNTER         = 1
_COMPANY_DIR            = None


# =============================================================================
# 1. 유틸 함수
# =============================================================================

def _nv(val, blank_vals=('(전체)','nan','none','')) -> str:
    s = str(val).strip()
    return '' if s.lower() in blank_vals else s

def _safe_float(val, default: float) -> float:
    try:
        f = float(val)
        return default if pd.isna(f) else f
    except (TypeError, ValueError):
        return default

def get_first_digit(number):
    try:
        s = str(abs(int(number)))
        return int(s[0]) if s[0] != '0' else 0
    except Exception:
        return 0

def _has_expected_columns(df):
    cols = [str(c).strip() for c in df.columns]
    keywords = ('차변','대변','계정','전표일자','거래처','전표번호','적요')
    return any(any(kw in c for kw in keywords) for c in cols)

def _read_excel_with_header_detection(path, dtype_map=None):
    dtype_map = dtype_map or {}
    for header_row in range(3):
        try:
            trial = pd.read_excel(path, engine='openpyxl', header=header_row, dtype=dtype_map)
            if trial.empty or len(trial.columns) < 2:
                continue
            if _has_expected_columns(trial):
                return trial, header_row
        except Exception:
            continue
    df = pd.read_excel(path, engine='openpyxl', dtype=dtype_map)
    return df, 0


def _load_with_parquet_cache(path: str, dtype_map: dict = None) -> tuple:
    """Excel 파일을 읽되, Parquet 캐시가 최신이면 캐시에서 빠르게 로드."""
    dtype_map = dtype_map or {}
    cache_dir = os.path.join(os.path.dirname(os.path.dirname(path)), 'cache')
    cache_name = os.path.splitext(os.path.basename(path))[0] + '.parquet'
    cache_path = os.path.join(cache_dir, cache_name)

    if os.path.isfile(cache_path) and os.path.getmtime(cache_path) >= os.path.getmtime(path):
        print(f'     ⚡ 캐시 로드: {cache_name}', flush=True)
        df = pd.read_parquet(cache_path)
        for col, dt in dtype_map.items():
            if col in df.columns:
                df[col] = df[col].astype(dt)
        return df, 0

    df, header_row = _read_excel_with_header_detection(path, dtype_map)
    try:
        os.makedirs(cache_dir, exist_ok=True)
        df.to_parquet(cache_path, index=False)
        print(f'     💾 캐시 저장: {cache_name}', flush=True)
    except Exception as e:
        print(f'     ⚠️ 캐시 저장 실패: {e}', flush=True)
    return df, header_row

def _normalize_account_for_match(s):
    s = str(s)
    for full, half in [('（','('),('）',')'),('⦅','('),('⦆',')'),
                       ('﹙','('),('﹚',')'),('【','('),('】',')')]:
        s = s.replace(full, half)
    s = s.replace(' ',' ').replace('＆','&')
    s = re.sub(r'\s+','',s)
    s = re.sub(r'[()（）]+','',s)
    s = s.replace('권','').lower()
    return s

def _account_match_flexible(acct_series, acct_str):
    acct_str = str(acct_str).strip()
    if not acct_str:
        return pd.Series(False, index=acct_series.index)
    norm_series = acct_series.fillna('').astype(str).apply(_normalize_account_for_match)
    norm_user   = _normalize_account_for_match(acct_str)
    # 정확 일치를 먼저 시도 — "건물" 검색 시 "건물관리비" 오매칭 방지
    mask = norm_series == norm_user
    if mask.any(): return mask
    mask = norm_series.str.startswith(norm_user)
    if mask.any(): return mask
    mask = norm_series.str.contains(re.escape(norm_user), na=False, regex=True, case=False)
    if mask.any(): return mask
    def _row(val):
        nv = _normalize_account_for_match(str(val) if pd.notna(val) else '')
        if not nv: return False
        return norm_user in nv or nv in norm_user
    return acct_series.fillna('').apply(_row)

def _to_numeric_amount(series):
    s = series.astype(str).str.strip()
    s = s.str.replace(',','',regex=False).str.replace(' ','',regex=False)
    s = s.str.replace(r'^#+$','0',regex=True)
    # (금액) 형식 음수 → -금액 변환 (한국 ERP 음수 표기 방식)
    s = s.str.replace(r'^\((\d+(?:\.\d+)?)\)$', r'-\1', regex=True)
    return pd.to_numeric(s, errors='coerce').fillna(0)

def _normalize_date_journal_columns(df):
    df.columns = [str(c).strip() for c in df.columns]
    if COL_DATE not in df.columns:
        for c in df.columns:
            if c in ('일자','전표일자','전표일','거래일자','적요일자'):
                df = df.rename(columns={c: COL_DATE}); break
    if COL_JOURNAL_ID not in df.columns:
        for c in df.columns:
            if c in ('전표등록번호','전표번호','전표NO','전표 no'):
                df = df.rename(columns={c: COL_JOURNAL_ID}); break
    if COL_CLIENT not in df.columns:
        for c in df.columns:
            if c in ('거래처명','거래처','상대거래처','거래처명칭'):
                df = df.rename(columns={c: COL_CLIENT}); break
    if COL_ACCOUNT not in df.columns:
        for c in df.columns:
            if c in ('계정명','계정과목','계정','과목'):
                df = df.rename(columns={c: COL_ACCOUNT}); break
    return df

def _normalize_debit_credit_columns(df):
    df.columns = [str(c).strip() for c in df.columns]
    # 중복 컬럼명이 있으면 df[col]이 DataFrame을 반환해 오류 → 중복 제거
    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated(keep='first')]
    debit_cols = [c for c in df.columns if '차변' in c]
    if debit_cols:
        for c in debit_cols: df[c] = _to_numeric_amount(df[c])
        primary = COL_DEBIT if COL_DEBIT in debit_cols else debit_cols[0]
        df[COL_DEBIT] = df[primary].copy()
        for c in debit_cols:
            if c != COL_DEBIT:
                df[COL_DEBIT] = df[COL_DEBIT].fillna(df[c])
                df = df.drop(columns=[c], errors='ignore')
    credit_cols = [c for c in df.columns if '대변' in c and c != COL_DEBIT]
    if credit_cols:
        for c in credit_cols: df[c] = _to_numeric_amount(df[c])
        primary = COL_CREDIT if COL_CREDIT in credit_cols else credit_cols[0]
        df[COL_CREDIT] = df[primary].copy()
        for c in credit_cols:
            if c != COL_CREDIT:
                df[COL_CREDIT] = df[COL_CREDIT].fillna(df[c])
                df = df.drop(columns=[c], errors='ignore')
    return df

def _preprocess_df(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    # 공백 제거 후 중복 컬럼 발생 시(예: CSV+xlsx concat) 첫 번째 열만 유지
    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated(keep='first')]
    df = _normalize_date_journal_columns(df)
    df = _normalize_debit_credit_columns(df)
    if COL_DATE in df.columns:
        ser = df[COL_DATE]
        try:
            num = pd.to_numeric(ser, errors='coerce')
            if num.notna().sum() > len(ser) * 0.5:
                ser = num.fillna(0).astype('int64').astype(str)
                df[COL_DATE] = pd.to_datetime(ser, format='%Y%m%d', errors='coerce')
            else:
                df[COL_DATE] = pd.to_datetime(ser, errors='coerce')
        except Exception:
            df[COL_DATE] = pd.to_datetime(ser, errors='coerce')
    if COL_DEBIT  in df.columns: df[COL_DEBIT]  = _to_numeric_amount(df[COL_DEBIT])
    if COL_CREDIT in df.columns: df[COL_CREDIT] = _to_numeric_amount(df[COL_CREDIT])
    reg_names = ['등록일자','등록일','작성일자','작성일','생성일자','생성일','입력일자','입력일']
    for col in df.columns:
        if any(n in str(col).strip() for n in reg_names):
            try:
                ser_r = df[col]
                num_r = pd.to_numeric(ser_r, errors='coerce')
                if num_r.notna().sum() > len(ser_r) * 0.5:
                    df[col] = pd.to_datetime(num_r.fillna(0).astype('int64').astype(str), format='%Y%m%d', errors='coerce')
                else:
                    df[col] = pd.to_datetime(ser_r, errors='coerce', format='mixed')
            except Exception:
                df[col] = pd.to_datetime(df[col], errors='coerce', format='mixed')
            break

    # 전표 단위 매칭용 그룹키: 전표일자+전표번호 결합 (전표번호 단독 재사용 대응)
    # 회사별 preprocess.py 가 이미 전표그룹키를 만들어뒀으면 존중하고 덮어쓰지 않는다
    # (예: sejoong은 전표번호 자체가 손상돼 있어 항번호 리셋 기반으로 별도 재구성함)
    if COL_JOURNAL_KEY not in df.columns and COL_DATE in df.columns and COL_JOURNAL_ID in df.columns:
        date_str = df[COL_DATE].dt.strftime('%Y%m%d').fillna('nodate')
        df[COL_JOURNAL_KEY] = date_str + '_' + df[COL_JOURNAL_ID].fillna('').astype(str).str.strip()

    return df

def _get_gubun_col(df):
    for c in df.columns:
        if str(c).strip() == '구분': return c
    return None

def get_safe_client_name(real_name):
    global _CLIENT_COUNTER
    if pd.isna(real_name) or str(real_name).strip() == '': return '(미기재)'
    real_name = str(real_name).strip()
    if real_name in GLOBAL_SAFE_MAP: return GLOBAL_SAFE_MAP[real_name]
    alias = f'Client_{_CLIENT_COUNTER:03d}'
    GLOBAL_SAFE_MAP[real_name] = alias
    _CLIENT_COUNTER += 1
    return alias

def mask_sensitive_info(text):
    if not isinstance(text, str):
        text = str(text) if text is not None else ''
        if text == '': return text
    text = re.sub(r'(\d{2,})[-](\d{2,})[-](\d{2,})', r'\1-****-\3', text)
    def _mask(m):
        s = m.group()
        return s[:4] + '****' + s[-4:] if len(s) >= 10 else s
    return re.sub(r'\b\d{10,}\b', _mask, text)

def _related_party_pattern(party):
    p = str(party).strip()
    if p.startswith('(주)'): return r'(?:\(주\))?' + re.escape(p[3:])
    return re.escape(p)

def draw_benford_chart(account_name, direction, actual_probs, benford_probs):
    try:
        plt.rc('font', family='Malgun Gothic')
        plt.rcParams['axes.unicode_minus'] = False
        digits = range(1, 10)
        plt.figure(figsize=(10, 6))
        plt.plot(digits, [benford_probs[d]*100 for d in digits],
                 color='red', marker='o', linestyle='--', label='벤포드 법칙')
        plt.bar(digits, [actual_probs.get(d,0.0)*100 for d in digits],
                color='skyblue', alpha=0.7, label=f'실제 ({account_name})')
        plt.title(f'벤포드 분석: {account_name} ({direction})')
        plt.legend(); plt.grid(axis='y', linestyle='--', alpha=0.5)
        buf = io.BytesIO()
        plt.savefig(buf, format='png'); plt.close(); buf.seek(0)
        return buf
    except Exception:
        return None

def _safe_sheet(name, max_len=31):
    s = re.sub(r'[\\/*?:\[\]]', '', str(name).strip())
    return s[:max_len]


# =============================================================================
# 2. 회사별 전용 전처리 모듈 동적 로드
# =============================================================================

def _apply_company_preprocess(df: pd.DataFrame, company_name: str) -> pd.DataFrame:
    """
    journal_analyzer/{company}/preprocess.py 가 존재하면 동적으로 임포트하여
    preprocess(df) 함수를 실행한다. 파일이 없으면 df 를 그대로 반환한다.

    규약
    ----
    각 회사 preprocess.py 는 반드시 아래 시그니처를 가져야 한다.
      def preprocess(df: pd.DataFrame) -> pd.DataFrame
    """
    import importlib.util

    module_path = os.path.join(BASE_DIR, company_name, 'preprocess.py')
    if not os.path.isfile(module_path):
        return df   # 전처리 모듈 없음 → 범용 로직으로 진행

    spec   = importlib.util.spec_from_file_location(f'{company_name}.preprocess', module_path)
    module = importlib.util.module_from_spec(spec)
    try:
        spec.loader.exec_module(module)
    except Exception as e:
        print(f'  [전처리 모듈 로드 오류] {company_name}/preprocess.py — {e}')
        return df

    fn = getattr(module, 'preprocess', None)
    if not callable(fn):
        print(f'  [경고] {company_name}/preprocess.py 에 preprocess() 함수가 없습니다.')
        return df

    print(f'  [전처리] {company_name}/preprocess.py 적용 중...')
    return fn(df)


# =============================================================================
# 3. 데이터 로드 (data/current → 당기, data/previous → 전기)
# =============================================================================

def load_data(company_dir: str) -> pd.DataFrame:
    dtype_map   = {COL_JOURNAL_ID: str}
    current_dir  = os.path.join(company_dir, 'data', 'current')
    previous_dir = os.path.join(company_dir, 'data', 'previous')
    all_dfs = []

    def _read(path, fname):
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == '.csv':
                print(f'     📂 CSV: {fname}', flush=True)
                for enc in ['utf-8','cp949','utf-16','euc-kr','latin-1']:
                    try:
                        out = pd.read_csv(path, encoding=enc, dtype=dtype_map)
                        print(f'     ✅ {len(out)}행', flush=True)
                        return out
                    except UnicodeDecodeError:
                        continue
            elif ext == '.xlsx':
                print(f'     📂 엑셀: {fname}', flush=True)
                out, h = _load_with_parquet_cache(path, dtype_map)
                if h > 0: print(f'     ℹ️ 헤더: {h+1}행', flush=True)
                print(f'     ✅ {len(out)}행', flush=True)
                return out
        except Exception as e:
            print(f'     ⚠️ 로드 실패: {fname} — {e}', flush=True)
        return None

    for label, dir_path in [('당기', current_dir), ('전기', previous_dir)]:
        if not os.path.isdir(dir_path):
            print(f'   ℹ️ 폴더 없음 ({label}): {dir_path}')
            continue
        for f in sorted(os.listdir(dir_path)):
            if not (f.endswith('.csv') or f.endswith('.xlsx')): continue
            if f.startswith('~$'): continue
            if '계정별원장' in f:
                print(f'   ℹ️ 계정별원장 파일은 분개장 로드에서 제외 (4번 메뉴 잔액표 전용): {f}')
                continue
            df = _read(os.path.join(dir_path, f), f)
            if df is not None and not df.empty:
                df['구분'] = label
                all_dfs.append(df)
                print(f'   📂 [{label}] {f} ({len(df)}행)')

    if not all_dfs: return None
    return pd.concat(all_dfs, axis=0, ignore_index=True)


# =============================================================================
# 3. 분석 함수 (2~19번)
#    시그니처: (df: DataFrame, params_list: list[dict]) → DataFrame | dict[str, DataFrame]
#    dict 반환 시 키 = 시트명
# =============================================================================

# ── 2. 거래처 전기/당기 비교 ──────────────────────────────────────────────────
def analyze_client_comparison(df: pd.DataFrame, params_list: list) -> dict:
    # 계정별 방향 매핑 (금액열 → 구분 순으로 fallback)
    acct_dir = {}
    for p in params_list:
        acct = _nv(p.get('계정과목',''))
        if not acct: continue
        vtype = (_nv(p.get('금액열',''), blank_vals=('nan','none',''))
                 or _nv(p.get('구분',''), blank_vals=('nan','none',''))
                 or '차변')
        if vtype not in ('차변','대변','both'): vtype = '차변'
        acct_dir[acct] = vtype
    if not acct_dir: acct_dir = {'접대비': '차변'}

    if COL_ACCOUNT not in df.columns or COL_CLIENT not in df.columns: return {}
    if '구분' not in df.columns: return {}

    df_w = df.copy()
    gc = _get_gubun_col(df_w)
    if gc: df_w[gc] = df_w[gc].astype(str).str.strip()

    # 계정별로 개별 방향 적용 후 합치기
    parts = []
    for acct, vtype in acct_dir.items():
        sub = df_w[_account_match_flexible(df_w[COL_ACCOUNT], acct)].copy()
        if sub.empty: continue
        if vtype == '차변':
            sub['_amt'] = pd.to_numeric(sub[COL_DEBIT], errors='coerce').fillna(0)
        elif vtype == '대변':
            sub['_amt'] = pd.to_numeric(sub[COL_CREDIT], errors='coerce').fillna(0)
        else:
            sub['_amt'] = (pd.to_numeric(sub[COL_DEBIT], errors='coerce').fillna(0)
                          + pd.to_numeric(sub[COL_CREDIT], errors='coerce').fillna(0))
        parts.append(sub)
    if not parts: return {}
    filtered = pd.concat(parts, ignore_index=True)

    pivot = filtered.pivot_table(index=[COL_ACCOUNT, COL_CLIENT], columns='구분',
                                 values='_amt', aggfunc=['sum','count'], fill_value=0)
    if pivot.empty: return {}

    result = pd.DataFrame(index=pivot.index)
    for col in ['전기금액','당기금액','전기전표수','당기전표수']: result[col] = 0
    for (agg_fn, gubun) in pivot.columns:
        g = str(gubun).strip()
        col_key = f'{g}금액' if agg_fn == 'sum' else f'{g}전표수'
        result[col_key] = pivot[(agg_fn, gubun)].reindex(result.index).fillna(0)
    result = result.fillna(0)
    for c in ['전기전표수','당기전표수']:
        if c in result.columns: result[c] = result[c].astype(int)
    result['증감금액'] = result.get('당기금액', 0) - result.get('전기금액', 0)
    result['증감비율(%)'] = result.apply(
        lambda r: (r['증감금액']/r['전기금액']*100) if r.get('전기금액',0) != 0 else 0.0, axis=1)
    result = (result.assign(_abs=result['증감금액'].abs())
                    .sort_values([COL_ACCOUNT,'_abs','당기금액'], ascending=[True,False,False])
                    .drop(columns=['_abs']).reset_index())
    cols = list(result.columns)
    renames = {}
    if cols[0] != '계정명': renames[cols[0]] = '계정명'
    if len(cols)>1 and cols[1] != '거래처명': renames[cols[1]] = '거래처명'
    if renames: result = result.rename(columns=renames)

    out = {}
    if '계정명' in result.columns:
        for acct in result['계정명'].unique():
            sub   = result[result['계정명'] == acct].drop(columns=['계정명'])
            sname = _safe_sheet(f'비교_{re.sub(r"[^가-힣a-zA-Z0-9]","",str(acct))[:20]}')
            out[sname] = sub
    else:
        out['거래처_전기당기비교'] = result
    return out


# ── 3. 벤포드 분석 ────────────────────────────────────────────────────────────
def analyze_benford(df: pd.DataFrame, params_list: list) -> dict:
    targets = []
    for p in params_list:
        acct = _nv(p.get('계정과목',''))
        col  = _nv(p.get('금액열',''), blank_vals=('nan','none','')) or '차변'
        if col not in ('차변','대변'): col = '차변'
        if acct: targets.append((acct, col))
    if not targets: targets = list(DEFAULT_BENFORD_TARGETS)

    out, images = {}, []
    for acct, direction in targets:
        tcol   = COL_DEBIT if direction == '차변' else COL_CREDIT
        mask   = _account_match_flexible(df[COL_ACCOUNT], acct)
        subset = df[mask & (df[tcol] > 0)].copy()
        n      = len(subset)
        sheet_key = f'벤포드_{acct}_{direction}'
        if n < BENFORD_MIN_ROWS:
            out[sheet_key] = pd.DataFrame([{'계정':acct,'방향':direction,'숫자':'-','발생건수':0,
                                            '실제비율(%)':0,'이론비율(%)':0,'차이(%p)':0,
                                            '비고':f'데이터 부족({n}건)'}])
            continue
        subset['Digit'] = subset[tcol].apply(get_first_digit)
        dg     = subset[subset['Digit'] >= 1]['Digit']
        counts = dg.value_counts(normalize=True).sort_index()
        raw    = dg.value_counts().sort_index()
        img    = draw_benford_chart(acct, direction, counts, BENFORD_PROBS)
        if img: images.append((acct, direction, img))
        rows = []
        for d in range(1, 10):
            actual = counts.get(d, 0.0)
            theory = BENFORD_PROBS[d]
            rows.append({'계정':acct,'방향':direction,'숫자':d,'발생건수':int(raw.get(d,0)),
                         '실제비율(%)':round(actual*100,2),'이론비율(%)':round(theory*100,2),
                         '차이(%p)':round((actual-theory)*100,2),
                         '이상여부':'Y' if abs(actual-theory)>0.05 else ''})
        out[sheet_key] = pd.DataFrame(rows)

    if images: out['_benford_images'] = images   # 특수 키: save_results에서 차트 삽입
    return out


# ── 당기 계정별원장 기초/기말잔액표 공용 헬퍼 (4번 데이터개요에서 사용) ─────────
_LEDGER_SHEET_PREFIX_RE = re.compile(r'^\d+_')
_LEDGER_SHEET_CODE_RE   = re.compile(r'\(\d+\)\s*$')
_LEDGER_OPEN_KEYWORDS   = ('전기이월', '기초잔액')
_LEDGER_SKIP_KEYWORDS   = ('월계', '누계', '합계')

# 차감계정 라벨 세분화 (계산식은 손대지 않고 '구분' 표시만 정교화).
# - 자산차감 키워드: 대변성격으로 감지되면 자산차감(예: 감가상각누계액·대손충당금은
#   실제로 대변성격이며 자산을 순액으로 줄이는 역할)
# - 부채차감 키워드: 차변성격으로 감지되면 부채차감(예: 사채할인발행차금, 퇴직급여충당금/
#   퇴직급여충당부채를 상계하는 사외적립자산·퇴직연금운용자산 — 2026-08-12 사용자 확인)
# - 애매(양쪽 다 나올 수 있는 계정): 감지된 방향에 따라 자산차감/부채차감을 자동 판정
#   (예: 현재가치할인차금은 대상이 채권이면 자산차감, 채무면 부채차감)
# 주의: '충당금'이 들어간다고 전부 차감계정은 아님 — 퇴직급여충당금(퇴직급여충당부채)은
# 그 자체가 부채 계정이라 목록에서 제외함 (2026-08-12 사용자 확인).
_LEDGER_CONTRA_ASSET_KEYWORDS = ('감가상각누계액', '대손충당금', '재고평가충당금',
                                  '재고자산평가충당금', '손상차손누계액')
_LEDGER_CONTRA_LIAB_KEYWORDS  = ('사채할인발행차금', '퇴직연금운영자산', '퇴직연금운용자산',
                                  '사외적립자산')
_LEDGER_CONTRA_AMBIGUOUS_KEYWORDS = ('현재가치할인차금',)

def _refine_ledger_gubun(acct_name: str, gubun: str) -> str:
    """'자산/비용'·'부채/자본/수익' 라벨을 차감계정이면 '자산차감'·'부채차감'으로 세분화."""
    if gubun not in ('자산/비용', '부채/자본/수익'):
        return gubun
    if any(k in acct_name for k in _LEDGER_CONTRA_ASSET_KEYWORDS):
        return '자산차감'
    if any(k in acct_name for k in _LEDGER_CONTRA_LIAB_KEYWORDS):
        return '부채차감'
    if any(k in acct_name for k in _LEDGER_CONTRA_AMBIGUOUS_KEYWORDS):
        return '자산차감' if gubun == '부채/자본/수익' else '부채차감'
    return gubun

def _find_current_ledger_files(current_dir: str) -> list:
    """
    data/current 폴더에서 파일명에 '계정별원장'이 들어간 당기 파일 경로를 모두 탐색.
    sejoong처럼 자산/부채/수익비용이 파일 단위로 나뉜 회사는 여러 개가 나올 수 있다.
    """
    if not os.path.isdir(current_dir):
        return []
    return [os.path.join(current_dir, f) for f in sorted(os.listdir(current_dir))
            if '계정별원장' in f and f.endswith('.xlsx') and not f.startswith('~$')]

def _clean_ledger_account_name(sheet_name: str) -> str:
    """시트명 '0_당좌예금(10200)' -> '당좌예금' (순번 접두어 + 계정코드 접미어 제거)."""
    name = _LEDGER_SHEET_PREFIX_RE.sub('', str(sheet_name))
    name = _LEDGER_SHEET_CODE_RE.sub('', name).strip()
    return name

def _clean_ledger_text(s) -> str:
    """적요란 텍스트에서 공백·대괄호 제거 (예: '[ 전 기 이 월 ]' -> '전기이월')."""
    if s is None:
        return ''
    return re.sub(r'[\s\[\]]', '', str(s))

def _compute_ledger_balance(records: list) -> dict:
    """
    ERP 형식과 무관한 공용 계산부. records는 계정 1개 분량의 거래를 시간순으로 나열한
    (종류, 차변, 대변, 잔액) 튜플 리스트 — 종류는 'open'(기초잔액 행) 또는 'txn'(실거래 행)만
    포함하고, 월계/누계/소계 같은 집계 행은 호출부(형식별 추출 함수)에서 이미 제외한 상태여야 한다.

    기초잔액·차변합계·대변합계·원장상 최종 표시잔액을 구한다.
    구분(자산/비용 vs 부채/자본/수익) 판정: 마지막 행 하나만 보면 그 행의 잔액이
    비어 있는 특수 케이스(예: 기말 결산정리 행이 잔액을 갱신하지 않고 남겨두는 경우)에
    취약하므로, 기초잔액 이후 각 거래 행마다 누적(차변-대변)을 계속 추적하면서
    잔액이 채워진 모든 행에 대해 '자산/비용 부호'와 '부채/자본/수익 부호' 중 어느 쪽과
    일치하는지 투표(voting)하여 다수결로 구분을 정한다. records가 비어 있으면 None.
    """
    if not records:
        return None

    open_debit = open_credit = 0.0
    sum_debit  = sum_credit  = 0.0
    last_balance = 0.0
    running_signed = 0.0        # 누적(차변-대변) — 자산/비용 부호 기준
    votes_asset = votes_liab = 0

    for typ, debit, credit, balance in records:
        if typ == 'open':
            open_debit, open_credit = debit, credit
            running_signed = open_debit - open_credit
        else:
            sum_debit  += debit
            sum_credit += credit
            running_signed += debit - credit

        if balance is not None:
            try:
                actual = float(balance)
                last_balance = actual
                TOL = 1
                if abs(running_signed - actual) <= TOL:
                    votes_asset += 1
                elif abs(-running_signed - actual) <= TOL:
                    votes_liab += 1
            except (TypeError, ValueError):
                pass

    calc_asset = running_signed          # 자산/비용: 기초+차변-대변 누적
    calc_liab  = -running_signed         # 부채/자본/수익: 기초-차변+대변 누적

    if open_debit == 0 and open_credit == 0 and sum_debit == 0 and sum_credit == 0:
        gubun, calc = '데이터없음', 0.0        # 기초잔액도 당기 거래도 없는 휴면 계정
    elif votes_asset == 0 and votes_liab == 0:
        gubun, calc = '검증필요', calc_asset
    elif votes_asset >= votes_liab:
        gubun, calc = '자산/비용', calc_asset
    else:
        gubun, calc = '부채/자본/수익', calc_liab

    return {
        '기초잔액':      open_debit if open_debit else (open_credit if open_credit else 0.0),
        '차변합계':      sum_debit,
        '대변합계':      sum_credit,
        '구분':          gubun,
        '기말잔액(계산)': calc,
        '최종표시잔액(원장)': last_balance,
        '차이':          round(calc - last_balance, 2),
    }

def _extract_duzon_style_records(ws) -> list:
    """
    더존 계열(예: graphy, dae_il) 시트 1개(열: 날짜/적요란/코드/거래처명/사업자등록번호/
    차변/대변/잔액)에서 (종류, 차변, 대변, 잔액) 레코드 리스트를 뽑는다.
    """
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) < 8:
            continue
        desc, debit, credit, balance = row[1], row[5] or 0, row[6] or 0, row[7]
        debit, credit = float(debit) if debit else 0.0, float(credit) if credit else 0.0
        text = _clean_ledger_text(desc)
        if not text and not debit and not credit and balance is None:
            continue
        if any(k in text for k in _LEDGER_OPEN_KEYWORDS):
            records.append(('open', debit, credit, balance))
        elif text in _LEDGER_SKIP_KEYWORDS:
            continue
        else:
            records.append(('txn', debit, credit, balance))
    return records

# sejoong 계열: 파일 1개 = 대분류(자산/부채/수익비용 등) 1개, 시트 1개에 여러 계정이
# '전월이월' 행으로 구분되어 연속 나열됨 (2026-08-12 확인).
# 컬럼(0-idx): 0 계정과목 / 1 계정과목명 / 4 일자 / 5 전표번호 / 9 차변금액 / 10 대변금액 / 11 잔액
_SEJOONG_STYLE_OPEN_LABELS = ('전월이월', '전기이월', '기초잔액')
_SEJOONG_STYLE_SKIP_LABELS = ('계정별월별소계', '계정별누계', '총누계', '총계', '합계')

def _iter_sejoong_style_ledger_accounts(ws):
    """
    sejoong 계열 시트에서 계정별로 (계정명, records) 를 순서대로 만들어 낸다.
    파일이 대용량이라 원본 파일 내에 헤더 행이 중간에 재삽입된 경우(계정과목=='계정과목')도
    있어 이를 건너뛴다.
    """
    current_name = None
    current_records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) < 12:
            continue
        code, name, jeonpyo = row[0], row[1], row[5]
        if code == '계정과목':                  # 재삽입된 헤더 행
            continue
        jeonpyo_norm = _clean_ledger_text(jeonpyo)
        debit  = float(row[9])  if row[9]  else 0.0
        credit = float(row[10]) if row[10] else 0.0
        balance = row[11]

        if jeonpyo_norm in _SEJOONG_STYLE_OPEN_LABELS:
            if current_name is not None:
                yield current_name, current_records
            current_name = str(name).strip() if name else ''
            current_records = [('open', debit, credit, balance)]
            continue

        if code is None:
            continue                             # 소계/누계/총계 행 또는 빈 행 — 실거래 아님

        if current_name is None:
            continue                             # 전월이월 행을 만나기 전의 고아 행 방어

        current_records.append(('txn', debit, credit, balance))

    if current_name is not None:
        yield current_name, current_records

def _build_ledger_balance_table(ledger_path: str) -> pd.DataFrame:
    """
    계정별원장 파일 1개를 읽어 계정별 기초/기말잔액표를 만든다. 두 형식을 자동 감지한다:
      - 더존 계열: 시트 1개 = 계정 1개, 시트명 '0_계정명(코드)'
      - sejoong 계열: 시트 1개에 '계정과목'/'계정과목명' 헤더로 여러 계정이 나열됨
    """
    wb = openpyxl.load_workbook(ledger_path, read_only=True, data_only=True)
    try:
        first_ws = wb[wb.sheetnames[0]]
        header = next(first_ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        is_sejoong_style = (len(wb.sheetnames) == 1 and len(header) >= 2
                             and header[0] == '계정과목' and header[1] == '계정과목명')

        rows = []
        if is_sejoong_style:
            print(f'  [계정별원장] sejoong 계열 형식(단일 시트·전월이월 구분) 감지: '
                  f'{os.path.basename(ledger_path)}')
            for acct_name, records in _iter_sejoong_style_ledger_accounts(first_ws):
                result = _compute_ledger_balance(records)
                if result is None:
                    continue
                result['구분'] = _refine_ledger_gubun(acct_name, result['구분'])
                rows.append({'계정명': acct_name, **result})
        else:
            for sn in wb.sheetnames:
                result = _compute_ledger_balance(_extract_duzon_style_records(wb[sn]))
                if result is None:
                    continue
                acct_name = _clean_ledger_account_name(sn)
                result['구분'] = _refine_ledger_gubun(acct_name, result['구분'])
                rows.append({'계정명': acct_name, **result})
    finally:
        wb.close()
    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ── 4. 데이터 개요 ────────────────────────────────────────────────────────────
def analyze_data_overview(df: pd.DataFrame, params_list: list) -> dict:
    summary = pd.DataFrame({
        '항목': ['총 행수','총 차변','총 대변','계정 수','시작일','종료일'],
        '값': [len(df),
               df[COL_DEBIT].sum()  if COL_DEBIT   in df.columns else 0,
               df[COL_CREDIT].sum() if COL_CREDIT  in df.columns else 0,
               df[COL_ACCOUNT].nunique() if COL_ACCOUNT in df.columns else 0,
               df[COL_DATE].min() if COL_DATE in df.columns else '-',
               df[COL_DATE].max() if COL_DATE in df.columns else '-']
    })
    if COL_ACCOUNT in df.columns and COL_DEBIT in df.columns and COL_CREDIT in df.columns:
        d = df[df[COL_DEBIT]!=0].groupby(COL_ACCOUNT)[COL_DEBIT].agg(['count','sum','mean','std'])
        d.columns = ['전표건수(차)','차변합계','평균금액(차)','표준편차(차)']
        c = df[df[COL_CREDIT]!=0].groupby(COL_ACCOUNT)[COL_CREDIT].agg(['count','sum','mean','std'])
        c.columns = ['전표건수(대)','대변합계','평균금액(대)','표준편차(대)']
        stats = pd.concat([d, c], axis=1).fillna(0).sort_values('차변합계',ascending=False).reset_index()
    else:
        stats = pd.DataFrame()

    out = {'데이터개요_요약': summary, '데이터개요_계정별': stats}

    if _COMPANY_DIR:
        ledger_paths = _find_current_ledger_files(os.path.join(_COMPANY_DIR, 'data', 'current'))
        ledger_tables = []
        for path in ledger_paths:
            t = _build_ledger_balance_table(path)
            if not t.empty:
                t.insert(0, '출처파일', os.path.basename(path))
                ledger_tables.append(t)
        if ledger_tables:
            out['계정별원장_잔액표'] = pd.concat(ledger_tables, axis=0, ignore_index=True)

    return out


# ── 5. 계정명 리스트 ──────────────────────────────────────────────────────────
def analyze_account_list(df: pd.DataFrame, params_list: list) -> dict:
    if COL_ACCOUNT not in df.columns:
        return {'계정명리스트': pd.DataFrame({'오류':['계정명 컬럼 없음']})}
    accs   = sorted({str(a).strip() for a in df[COL_ACCOUNT].dropna() if str(a).strip()})
    simple = pd.DataFrame({'계정명': accs})
    jkey_acc = COL_JOURNAL_KEY if COL_JOURNAL_KEY in df.columns else COL_JOURNAL_ID
    agg_map = {COL_DEBIT:['sum','count'], COL_CREDIT:['sum','count']}
    if jkey_acc in df.columns: agg_map[jkey_acc] = 'nunique'
    stats = df.groupby(COL_ACCOUNT).agg(agg_map).reset_index()
    base_cols = ['계정명','차변합계','차변건수','대변합계','대변건수']
    stats.columns = base_cols + (['전표개수'] if jkey_acc in df.columns else [])
    stats['최대금액'] = stats[['차변합계','대변합계']].max(axis=1)
    stats = stats.sort_values('최대금액', ascending=False).drop(columns=['최대금액'])
    return {'계정명_간단': simple, '계정명_통계': stats}


# ── 6. 사원별 집계 ────────────────────────────────────────────────────────────
def analyze_employee_summary(df: pd.DataFrame, params_list: list) -> pd.DataFrame:
    emp_col = next(
        (c for c in df.columns if any(n in str(c) for n in ['사원명','작성자','사용자','User','Employee'])),
        None)
    if emp_col is None:
        return pd.DataFrame({'오류':['사원명 컬럼 없음. 컬럼 목록: ' + str(list(df.columns))]})
    df_emp = df[df[emp_col].notna() & (df[emp_col].astype(str).str.strip() != '')].copy()
    if df_emp.empty: return pd.DataFrame({'오류':['사원명 데이터 없음']})

    jid = (COL_JOURNAL_KEY if COL_JOURNAL_KEY in df.columns
           else COL_JOURNAL_ID if COL_JOURNAL_ID in df.columns else None)
    rows = []
    for emp in sorted(df_emp[emp_col].astype(str).unique()):
        sub   = df_emp[df_emp[emp_col].astype(str) == emp]
        d_sub = sub[sub[COL_DEBIT]  != 0] if COL_DEBIT  in sub.columns else pd.DataFrame()
        c_sub = sub[sub[COL_CREDIT] != 0] if COL_CREDIT in sub.columns else pd.DataFrame()
        dagg  = {COL_DEBIT:'sum'}
        cagg  = {COL_CREDIT:'sum'}
        if jid: dagg[jid]='nunique'; cagg[jid]='nunique'
        dg = d_sub.groupby(COL_ACCOUNT).agg(dagg).reset_index() if not d_sub.empty else pd.DataFrame()
        cg = c_sub.groupby(COL_ACCOUNT).agg(cagg).reset_index() if not c_sub.empty else pd.DataFrame()
        for i in range(max(len(dg), len(cg), 1)):
            row = {'사원명': emp if i == 0 else ''}
            if i < len(dg):
                row['차변계정명'] = dg.iloc[i][COL_ACCOUNT]
                row['차변금액']   = dg.iloc[i][COL_DEBIT]
                row['차변전표수'] = int(dg.iloc[i][jid]) if jid and jid in dg.columns else ''
            else:
                row.update({'차변계정명':'','차변금액':0,'차변전표수':''})
            if i < len(cg):
                row['대변계정명'] = cg.iloc[i][COL_ACCOUNT]
                row['대변금액']   = cg.iloc[i][COL_CREDIT]
                row['대변전표수'] = int(cg.iloc[i][jid]) if jid and jid in cg.columns else ''
            else:
                row.update({'대변계정명':'','대변금액':0,'대변전표수':''})
            rows.append(row)
    return pd.DataFrame(rows)


# ── 7. 일자차이 분석 ──────────────────────────────────────────────────────────
def analyze_date_difference(df: pd.DataFrame, params_list: list) -> dict:
    days_threshold = None
    for p in params_list:
        v = p.get('기준일수', p.get('일수'))
        if v is not None:
            try: days_threshold = int(float(v)); break
            except (TypeError, ValueError): pass
    if not days_threshold or days_threshold <= 0:
        return {'일자차이분석': pd.DataFrame({'안내':['기준일수 파라미터가 없거나 0 이하입니다.']})}

    reg_names = ['등록일자','등록일','작성일자','작성일','생성일자','입력일자','입력일']
    reg_col = next((c for c in df.columns if any(n in str(c) for n in reg_names)), None)
    if reg_col is None:
        return {'일자차이분석': pd.DataFrame({'오류':['등록일자 컬럼 없음',
                                                       f'컬럼: {list(df.columns[:15])}']})}

    df2 = df[(df[COL_DATE].notna()) & (df[reg_col].notna())].copy()
    if df2.empty:
        return {'일자차이분석': pd.DataFrame({'오류':['전표일자 또는 등록일자 없음']})}

    df2['일자차이'] = (df2[reg_col] - df2[COL_DATE]).dt.days
    filtered = df2[df2['일자차이'] >= days_threshold].copy()
    if filtered.empty:
        return {'일자차이분석': pd.DataFrame({'결과':[f'{days_threshold}일 이상 차이 전표 없음']})}

    jkey = COL_JOURNAL_KEY if COL_JOURNAL_KEY in filtered.columns else COL_JOURNAL_ID
    j_sum = filtered.groupby(jkey).agg(
        전표번호=(COL_JOURNAL_ID, 'first'),
        전표일자=(COL_DATE, 'first'),
        등록일자=(reg_col, 'first'),
        일자차이=('일자차이', 'first'),
        계정명=(COL_ACCOUNT, lambda x: ', '.join(x.unique()[:5])),
        차변합계=(COL_DEBIT, 'sum'),
        대변합계=(COL_CREDIT, 'sum'),
    ).reset_index(drop=True).sort_values('일자차이', ascending=False)

    detail = filtered.sort_values(['일자차이', COL_JOURNAL_ID], ascending=[False, True])
    dc = [c for c in ['구분', COL_JOURNAL_ID, COL_DATE, reg_col, '일자차이',
                       COL_ACCOUNT, COL_DEBIT, COL_CREDIT, COL_CLIENT, COL_DESC] if c in detail.columns]
    return {'일자차이_요약': j_sum, '일자차이_상세': detail[dc]}


# ── 8. 상대계정 분석 ──────────────────────────────────────────────────────────
def analyze_counterpart(df: pd.DataFrame, params_list: list) -> dict:
    out = {}
    for p in params_list:
        acct      = _nv(p.get('계정과목',''))
        direction = (_nv(p.get('금액열',''), blank_vals=('nan','none',''))
                     or _nv(p.get('구분',''), blank_vals=('nan','none',''))
                     or '차변')
        if direction not in ('차변','대변'): direction = '차변'
        if not acct: continue

        # 그룹기준열: 파라미터에서 읽고, 없으면 전표그룹키(전표일자+전표번호) 사용
        # (전표번호 단독은 회사에 따라 전표일자별로 재사용되는 순번일 수 있어 신뢰 불가 — 2026-08-12)
        default_group_col = COL_JOURNAL_KEY if COL_JOURNAL_KEY in df.columns else COL_JOURNAL_ID
        group_col_name = _nv(p.get('그룹기준열', ''), blank_vals=('nan', 'none', ''))
        if group_col_name == COL_JOURNAL_ID and COL_JOURNAL_KEY in df.columns:
            # "전표번호로 묶어라"는 지정의 실제 의도는 "같은 전표 단위로 묶어라"이므로
            # 전표그룹키가 있으면 그쪽을 우선한다 (전표번호 단독은 신뢰 불가 — 2026-08-12)
            group_col = COL_JOURNAL_KEY
        elif group_col_name:
            matched_col = next((c for c in df.columns
                                if c.strip() == group_col_name
                                or group_col_name in c), None)
            if matched_col is None:
                print(f'    [경고] 그룹기준열 "{group_col_name}" 컬럼 없음 → 전표그룹키로 대체')
                group_col = default_group_col
            else:
                group_col = matched_col
        else:
            group_col = default_group_col

        tcol  = COL_DEBIT if direction == '차변' else COL_CREDIT
        mask  = _account_match_flexible(df[COL_ACCOUNT], acct) & (df[tcol] != 0)
        target = df[mask]
        if target.empty: continue
        jids    = target[group_col].unique()
        related = df[df[group_col].isin(jids)].copy()
        # 지정 방향의 반대 측(상대계정) 금액만 집계
        counter_col = COL_CREDIT if direction == '차변' else COL_DEBIT
        sum_label   = '대변합계' if direction == '차변' else '차변합계'
        cnt_label   = '대변건수' if direction == '차변' else '차변건수'
        summary = (related[related[counter_col] != 0]
                   .groupby(COL_ACCOUNT)[[counter_col]]
                   .agg(['sum','count']).reset_index())
        summary.columns = ['상대계정명', sum_label, cnt_label]
        summary = summary.sort_values(sum_label, ascending=False)
        sname   = _safe_sheet(f'상대_{re.sub(r"[^가-힣a-zA-Z0-9]","",acct)[:18]}')
        out[sname] = summary
    return out or {'상대계정분석': pd.DataFrame({'안내':['파라미터에 계정과목이 없습니다.']})}


# ── 27. 감가상각/평가손익분석 (Phase 1: 상대계정 매칭) ─────────────────────────
def analyze_depreciation_valuation(df: pd.DataFrame, params_list: list) -> dict:
    """감가상각비·외화환산손익·평가손익·대손상각비 등 손익 계정의 상대계정 금액을
    8번 상대계정분석과 동일한 로직(전표 단위 매칭)으로 추출. (Phase 1)
    이어서 '감가상각_유형자산롤포워드' 파라미터 시트가 있으면 유형자산계정별
    취득원가·감가상각누계액 기초/당기증가/당기감소/기말 표를 추가한다. (Phase 2)
    '유형자산_처분손익' 파라미터 시트가 있으면 처분 건별 유형자산처분손익명세서를
    추가한다. (Phase 3)
    """
    out = analyze_counterpart(df, params_list)
    out.update(_depreciation_rollforward(df, out))
    out.update(_disposal_schedule(df))
    return out


def _depreciation_rollforward(df: pd.DataFrame, phase1_results: dict) -> dict:
    """유형자산계정별 취득원가·감가상각누계액 롤포워드.

    취득원가        : 기초 + 당기증가(대체+기타) - 당기감소(대체+기타) = 기말
    감가상각누계액  : 기초 + 당기감가상각비(Phase1 상대_감가상각* 매칭분) + 당기증가_대체
                      - 당기감소_대체 - 당기감소_기타 + 수동조정 = 기말
    - '대체'분: 파라미터의 대체상대계정과 같은 전표에서 함께 나타나는 금액만 전표 매칭으로
      분리(예: 건물→투자부동산_건물 계정대체, 건설중인자산→본계정 완성대체).
    - 감가상각누계액의 '당기증가'는 감가상각비 상대계정 매칭분(정상적인 상각)과 대체 매칭분만
      코드로 잡고, 그 외(예: 손상차손환입 등)는 '미매칭차이' 컬럼에 남겨 감사인이 확인 후
      수동조정 컬럼에 직접 입력하도록 한다 — 모든 대변 조정을 자동으로 상각비로 잡지 않기
      위한 의도적 설계(8-6차 세션 요청사항).
    기초잔액은 data/previous의 전기 계정별_거래처별명세 파일이 있을 때만 채우고,
    없으면 공란(None)으로 둔다 (분개장에서 역산하지 않음).
    """
    global _COMPANY_DIR
    if _COMPANY_DIR is None:
        return {}

    task_path = os.path.join(_COMPANY_DIR, f'task_list_{os.path.basename(_COMPANY_DIR)}.xlsx')
    if not os.path.isfile(task_path):
        return {}
    sheet_name = '감가상각_유형자산롤포워드'
    try:
        xl = pd.ExcelFile(task_path)
        if sheet_name not in xl.sheet_names:
            return {}
        pairs = pd.read_excel(xl, sheet_name=sheet_name).dropna(how='all')
    except Exception:
        return {}
    if '실행여부' in pairs.columns:
        flag = pairs['실행여부'].astype(str).str.strip().str.upper()
        pairs = pairs[flag.isin(['Y', 'O'])]
    if pairs.empty:
        return {}

    # 전기명세 파일 (있으면 기초잔액용)
    prev_file = _find_prev_detail_file(os.path.join(_COMPANY_DIR, 'data', 'previous'))
    prev_xl = pd.ExcelFile(prev_file) if prev_file else None

    # Phase 1 '상대_감가상각*' 결과에서 {감가상각누계액계정명: 당기감가상각비} 집계
    dep_lookup: dict = {}
    for sname, sdf in phase1_results.items():
        if not sname.startswith('상대_감가상각') or '상대계정명' not in sdf.columns:
            continue
        amt_col = next((c for c in sdf.columns if c.endswith('합계')), None)
        if amt_col is None:
            continue
        for _, r in sdf.iterrows():
            key = str(r['상대계정명']).strip()
            dep_lookup[key] = dep_lookup.get(key, 0) + (r[amt_col] or 0)

    # 유형자산계정명 -> 감가상각누계액계정명 (대체상대계정의 상각누계액계정 역참조용)
    asset_to_dep: dict = {}
    for _, pr in pairs.iterrows():
        an = _nv(pr.get('유형자산계정명', ''))
        if an:
            asset_to_dep[an] = _nv(pr.get('감가상각누계액계정명', ''))

    def _period_sum(acct_name: str):
        mask = _account_match_flexible(df[COL_ACCOUNT], acct_name)
        sub = df[mask]
        return sub[COL_DEBIT].sum(), sub[COL_CREDIT].sum()

    def _split_names(raw: str) -> list:
        if not raw:
            return []
        return [n.strip() for n in re.split(r'[,\n]', str(raw)) if n.strip()]

    def _transfer_amount(acct_name: str, counterpart_names, col: str):
        """acct_name 계정의 col(차변/대변) 중, 같은 전표번호에 counterpart_names(리스트 또는
        문자열, 콤마로 복수 지정 가능) 계정 중 하나라도 함께 나타나는 금액만 전표 단위로
        매칭해 집계 (계정대체 분리용). 예: 건물은 건설중인자산에서 들어오고 투자부동산_건물로
        나가는 두 방향이 동시에 있을 수 있으므로 상대계정을 복수로 받는다."""
        names = counterpart_names if isinstance(counterpart_names, list) else _split_names(counterpart_names)
        names = [n for n in names if n]
        if not names:
            return 0
        mask_this = _account_match_flexible(df[COL_ACCOUNT], acct_name) & (df[col] != 0)
        if not mask_this.any():
            return 0
        counterpart_mask = pd.Series(False, index=df.index)
        for n in names:
            counterpart_mask = counterpart_mask | _account_match_flexible(df[COL_ACCOUNT], n)
        jkey = COL_JOURNAL_KEY if COL_JOURNAL_KEY in df.columns else COL_JOURNAL_ID
        valid_jids = set(df.loc[counterpart_mask, jkey].unique())
        sub = df[mask_this & df[jkey].isin(valid_jids)]
        return sub[col].sum()

    def _manual_num(p, col):
        v = pd.to_numeric(p.get(col, 0), errors='coerce')
        return 0 if pd.isna(v) else v

    rows = []
    for _, p in pairs.iterrows():
        asset_acct = _nv(p.get('유형자산계정명', ''))
        dep_acct   = _nv(p.get('감가상각누계액계정명', ''))
        transfer_partner = _nv(p.get('대체상대계정', ''), blank_vals=('nan', 'none', ''))
        cost_manual = _manual_num(p, '취득원가_수동조정')
        dep_manual  = _manual_num(p, '상각누계액_수동조정')
        if not asset_acct:
            continue

        a_incr_total, a_decr_total = _period_sum(asset_acct)
        a_incr_xfer = _transfer_amount(asset_acct, transfer_partner, COL_DEBIT)
        a_decr_xfer = _transfer_amount(asset_acct, transfer_partner, COL_CREDIT)
        a_incr_etc  = a_incr_total - a_incr_xfer
        a_decr_etc  = a_decr_total - a_decr_xfer
        a_open  = _prev_balance_total(prev_xl, asset_acct)
        a_close = (a_open + a_incr_xfer + a_incr_etc - a_decr_xfer - a_decr_etc + cost_manual
                   if a_open is not None else None)

        row = {
            '유형자산계정': asset_acct,
            '취득원가_기초': a_open,
            '취득원가_당기증가_대체': a_incr_xfer, '취득원가_당기증가_기타': a_incr_etc,
            '취득원가_당기감소_대체': a_decr_xfer, '취득원가_당기감소_기타': a_decr_etc,
            '취득원가_수동조정': cost_manual,
            '취득원가_기말': a_close,
        }

        if dep_acct:
            d_debit_total, d_credit_total = _period_sum(dep_acct)  # 차변=감소(전체), 대변=증가(전체)
            d_dep_expense = dep_lookup.get(dep_acct, 0)             # 당기감가상각비(감가상각비 매칭분)
            # 대체상대계정(복수 가능) 각각의 감가상각누계액계정을 역참조해서 상각누계액 대체 매칭에 사용
            dep_partner_accts = [asset_to_dep.get(n, '') for n in _split_names(transfer_partner)]
            dep_partner_accts = [d for d in dep_partner_accts if d]
            d_incr_xfer = _transfer_amount(dep_acct, dep_partner_accts, COL_CREDIT)
            d_decr_xfer = _transfer_amount(dep_acct, dep_partner_accts, COL_DEBIT)
            d_decr_etc  = d_debit_total - d_decr_xfer
            d_open  = _prev_balance_total(prev_xl, dep_acct)
            d_unexplained = d_credit_total - d_dep_expense - d_incr_xfer
            d_close = (d_open + d_dep_expense + d_incr_xfer - d_decr_xfer - d_decr_etc + dep_manual
                       if d_open is not None else None)
            row.update({
                '감가상각누계액계정': dep_acct,
                '상각누계액_기초': d_open,
                '상각누계액_당기감가상각비': d_dep_expense,
                '상각누계액_당기증가_대체': d_incr_xfer,
                '상각누계액_당기감소_대체': d_decr_xfer,
                '상각누계액_당기감소_기타': d_decr_etc,
                '상각누계액_미매칭차이': d_unexplained,
                '상각누계액_수동조정': dep_manual,
                '상각누계액_기말': d_close,
            })
        rows.append(row)

    if not rows:
        return {}
    return {'유형자산_롤포워드': pd.DataFrame(rows)}


def _disposal_schedule(df: pd.DataFrame) -> dict:
    """유형자산 처분손익명세서 (Phase 3).

    '유형자산_처분손익' 파라미터 시트(처분이익계정명/처분손실계정명)가 있으면,
    그 계정이 등장한 전표(전표그룹키)마다 같은 전표에 있는 유형자산 취득원가·
    감가상각누계액 금액(전표그룹키 매칭, '감가상각_유형자산롤포워드' 시트의
    유형자산계정명/감가상각누계액계정명 쌍 기준)을 찾아 처분 건별로 한 줄씩 뽑는다.

    장부가액 = 취득원가_처분분 - 감가상각누계액_처분분
    처분손익 = 처분손익계정 순액(대변-차변, 이익이면 +, 손실이면 -)
    처분가액(역산) = 장부가액 + 처분손익

    한계: 개별 자산(고정자산대장 항목) 단위가 아니라 계정과목×전표 단위로만
    구분된다 — 한 전표에 여러 자산을 묶어 처분/폐기하면 그 전표 전체가 한 줄로 잡힘.
    """
    global _COMPANY_DIR
    if _COMPANY_DIR is None:
        return {}
    task_path = os.path.join(_COMPANY_DIR, f'task_list_{os.path.basename(_COMPANY_DIR)}.xlsx')
    if not os.path.isfile(task_path):
        return {}
    try:
        xl = pd.ExcelFile(task_path)
        if '유형자산_처분손익' not in xl.sheet_names or '감가상각_유형자산롤포워드' not in xl.sheet_names:
            return {}
        disp_params = pd.read_excel(xl, sheet_name='유형자산_처분손익').dropna(how='all')
        pairs = pd.read_excel(xl, sheet_name='감가상각_유형자산롤포워드').dropna(how='all')
    except Exception:
        return {}
    if '실행여부' in disp_params.columns:
        flag = disp_params['실행여부'].astype(str).str.strip().str.upper()
        disp_params = disp_params[flag.isin(['Y', 'O'])]
    if disp_params.empty:
        return {}

    asset_dep_pairs = []
    for _, pr in pairs.iterrows():
        an = _nv(pr.get('유형자산계정명', ''))
        dn = _nv(pr.get('감가상각누계액계정명', ''))
        if an:
            asset_dep_pairs.append((an, dn))
    if not asset_dep_pairs:
        return {}

    jkey = COL_JOURNAL_KEY if COL_JOURNAL_KEY in df.columns else COL_JOURNAL_ID

    rows = []
    for _, p in disp_params.iterrows():
        gain_acct = _nv(p.get('처분이익계정명', ''), blank_vals=('nan', 'none', ''))
        loss_acct = _nv(p.get('처분손실계정명', ''), blank_vals=('nan', 'none', ''))
        disp_accts = [a for a in (gain_acct, loss_acct) if a]
        if not disp_accts:
            continue
        # 이익계정·손실계정이 같은 전표에 함께 나타나는 경우(혼합 처분) 자산 금액이
        # 중복 집계되지 않도록, 두 계정을 하나로 합쳐서 전표 단위로 순액을 구한다
        mask = pd.Series(False, index=df.index)
        for a in disp_accts:
            mask = mask | _account_match_flexible(df[COL_ACCOUNT], a)
        sub = df[mask]
        if sub.empty:
            continue
        for jid, g in sub.groupby(jkey):
            pl_net = g[COL_CREDIT].sum() - g[COL_DEBIT].sum()
            disposal_date = g[COL_DATE].iloc[0] if COL_DATE in g.columns else None

            for asset_acct, dep_acct in asset_dep_pairs:
                a_mask = _account_match_flexible(df[COL_ACCOUNT], asset_acct)
                a_sub = df[a_mask & (df[jkey] == jid)]
                if a_sub.empty:
                    continue
                cost_decr = a_sub[COL_CREDIT].sum() - a_sub[COL_DEBIT].sum()
                dep_decr = 0
                if dep_acct:
                    d_mask = _account_match_flexible(df[COL_ACCOUNT], dep_acct)
                    d_sub = df[d_mask & (df[jkey] == jid)]
                    dep_decr = d_sub[COL_DEBIT].sum() - d_sub[COL_CREDIT].sum()
                book_value = cost_decr - dep_decr
                rows.append({
                    '전표그룹': jid,
                    '처분일자': disposal_date,
                    '유형자산계정': asset_acct,
                    '취득원가_처분분': cost_decr,
                    '감가상각누계액_처분분': dep_decr,
                    '장부가액': book_value,
                    '처분손익': pl_net,
                    '처분가액(역산)': book_value + pl_net,
                })

    if not rows:
        return {}
    result = pd.DataFrame(rows).sort_values(['처분일자', '유형자산계정']).reset_index(drop=True)
    return {'유형자산_처분손익명세': result}


# ── 9. 키워드 검색 ────────────────────────────────────────────────────────────
def analyze_keyword_search(df: pd.DataFrame, params_list: list) -> pd.DataFrame:
    keywords = [_nv(p.get('키워드','')) for p in params_list if _nv(p.get('키워드',''))]
    if not keywords: keywords = list(DEFAULT_KEYWORDS)
    if COL_DESC not in df.columns:
        return pd.DataFrame({'오류':['적요 컬럼 없음']})
    pattern = '|'.join(re.escape(k) for k in keywords)
    result  = df[df[COL_DESC].str.contains(pattern, na=False, regex=True)].copy()
    if result.empty:
        return pd.DataFrame({'결과':[f'검색어 없음: {", ".join(keywords)}']})
    result['AbsAmt'] = result[[COL_DEBIT, COL_CREDIT]].abs().max(axis=1)
    return result.sort_values('AbsAmt', ascending=False).drop(columns=['AbsAmt'])


# ── 10. 라운드넘버 분석 ───────────────────────────────────────────────────────
def analyze_round_numbers(df: pd.DataFrame, params_list: list) -> pd.DataFrame:
    ALL_UNITS = [100_000, 500_000, 1_000_000, 5_000_000, 10_000_000]
    records = []
    for p in params_list:
        acct     = _nv(p.get('계정과목',''))
        col_f    = _nv(p.get('금액열',''), blank_vals=('nan','none','')) or '차변'
        min_unit = _safe_float(p.get('최소금액'), 100_000.0)
        units    = [u for u in ALL_UNITS if u >= min_unit]
        sub      = df.copy()
        if acct and COL_ACCOUNT in sub.columns:
            sub = sub[sub[COL_ACCOUNT].astype(str).str.contains(acct, na=False)]
        amt_cols = [col_f] if col_f in ('차변','대변') and col_f in sub.columns \
                   else [c for c in ('차변','대변') if c in sub.columns]
        for col in amt_cols:
            work = sub[sub[col] != 0].copy()
            work['라운드단위'] = work[col].apply(
                lambda x: next((f'{u:,}원 배수' for u in sorted(units,reverse=True) if abs(x)%u==0), None))
            records.append(work[work['라운드단위'].notna()].assign(금액열=col))
    if not records: return pd.DataFrame({'결과':['라운드넘버 없음']})
    result  = pd.concat(records, ignore_index=True)
    amt_col = next((c for c in ('차변','대변','금액') if c in result.columns), None)
    out_cols = [c for c in ['구분', COL_DATE, COL_JOURNAL_ID, COL_ACCOUNT, '금액열', amt_col,
                             COL_DESC, COL_CLIENT, '라운드단위'] if c and c in result.columns]
    return result[out_cols].sort_values(amt_col, ascending=False) if amt_col else result[out_cols]


# ── 11. 특수관계자 분석 ───────────────────────────────────────────────────────
def analyze_related_party(df: pd.DataFrame, params_list: list) -> dict:
    parties = [_nv(p.get('거래처명','')) for p in params_list if _nv(p.get('거래처명',''))]
    if not parties:
        return {'특수관계자': pd.DataFrame({'안내':['task_list 특수관계자 시트에 거래처명을 입력하세요.']})}
    if COL_CLIENT not in df.columns:
        return {'특수관계자': pd.DataFrame({'오류':['거래처명 컬럼 없음']})}
    pattern = '|'.join(_related_party_pattern(p) for p in parties)
    related = df[df[COL_CLIENT].str.contains(pattern, na=False, regex=True)].copy()
    if related.empty:
        return {'특수관계자': pd.DataFrame({'결과':['해당 특수관계자 거래 없음']})}
    piv_d = related.pivot_table(index=COL_ACCOUNT, columns=COL_CLIENT, values=COL_DEBIT,
                                 aggfunc='sum', fill_value=0).reset_index()
    piv_c = related.pivot_table(index=COL_ACCOUNT, columns=COL_CLIENT, values=COL_CREDIT,
                                 aggfunc='sum', fill_value=0).reset_index()
    summary = related.groupby([COL_CLIENT, COL_ACCOUNT])[[COL_DEBIT, COL_CREDIT]]\
                     .agg(['sum','count']).reset_index()
    summary.columns = ['거래처명','계정명','차변합계','차변건수','대변합계','대변건수']
    return {'특수관계자_차변피벗': piv_d,
            '특수관계자_대변피벗': piv_c,
            '특수관계자_요약':     summary,
            '특수관계자_상세':     related}


# ── 12. 자산 vs 부채 교차 ────────────────────────────────────────────────────
def analyze_asset_liability_cross(df: pd.DataFrame, params_list: list) -> pd.DataFrame:
    assets = [_nv(p.get('계정과목','')) for p in params_list
              if str(p.get('구분','')).strip() == '자산' and _nv(p.get('계정과목',''))]
    liabs  = [_nv(p.get('계정과목','')) for p in params_list
              if str(p.get('구분','')).strip() == '부채' and _nv(p.get('계정과목',''))]
    if not assets or not liabs:
        return pd.DataFrame({'안내':['task_list 자산부채교차 시트에 구분(자산/부채)·계정과목을 입력하세요.']})
    am = pd.Series(False, index=df.index)
    for a in assets: am |= df[COL_ACCOUNT].str.contains(a, na=False, regex=False)
    lm = pd.Series(False, index=df.index)
    for l in liabs:  lm |= df[COL_ACCOUNT].str.contains(l, na=False, regex=False)
    adf = df[am & (df[COL_DEBIT]  != 0)]
    ldf = df[lm & (df[COL_CREDIT] != 0)]
    if adf.empty or ldf.empty: return pd.DataFrame({'결과':['교차 데이터 없음']})
    ga = adf.groupby(COL_CLIENT).agg({COL_DEBIT:'sum',  COL_ACCOUNT: lambda x: ','.join(set(x))}).reset_index()
    gl = ldf.groupby(COL_CLIENT).agg({COL_CREDIT:'sum', COL_ACCOUNT: lambda x: ','.join(set(x))}).reset_index()
    ga.columns = ['거래처명','자산_금액','자산_계정']
    gl.columns = ['거래처명','부채_금액','부채_계정']
    merged = pd.merge(ga, gl, on='거래처명', how='inner').sort_values('자산_금액', ascending=False)
    return merged if not merged.empty else pd.DataFrame({'결과':['동시 발생 거래처 없음']})


# ── 13. 매출 vs 비용 교차 ────────────────────────────────────────────────────
def analyze_revenue_expense_cross(df: pd.DataFrame, params_list: list) -> pd.DataFrame:
    revs = [_nv(p.get('계정과목','')) for p in params_list
            if str(p.get('구분','')).strip() == '매출' and _nv(p.get('계정과목',''))]
    exps = [_nv(p.get('계정과목','')) for p in params_list
            if str(p.get('구분','')).strip() == '비용' and _nv(p.get('계정과목',''))]
    if not revs or not exps:
        return pd.DataFrame({'안내':['task_list 매출비용교차 시트에 구분(매출/비용)·계정과목을 입력하세요.']})
    rm = pd.Series(False, index=df.index)
    for r in revs: rm |= df[COL_ACCOUNT].str.contains(r, na=False, regex=False)
    em = pd.Series(False, index=df.index)
    for e in exps: em |= df[COL_ACCOUNT].str.contains(e, na=False, regex=False)
    rdf = df[rm & (df[COL_CREDIT] != 0)]
    edf = df[em & (df[COL_DEBIT]  != 0)]
    if rdf.empty or edf.empty: return pd.DataFrame({'결과':['매출 또는 비용 데이터 없음']})
    gr = rdf.groupby(COL_CLIENT).agg({COL_CREDIT:'sum', COL_ACCOUNT: lambda x: ','.join(set(x))}).reset_index()
    ge = edf.groupby(COL_CLIENT).agg({COL_DEBIT:'sum',  COL_ACCOUNT: lambda x: ','.join(set(x))}).reset_index()
    gr.columns = ['거래처명','매출_금액','매출_계정']
    ge.columns = ['거래처명','비용_금액','비용_계정']
    merged = pd.merge(gr, ge, on='거래처명', how='inner').sort_values('매출_금액', ascending=False)
    return merged if not merged.empty else pd.DataFrame({'결과':['동시 발생 거래처 없음']})


# ── 14. 심층분석 (계정별 Top) ─────────────────────────────────────────────────
def analyze_top_accounts(df: pd.DataFrame, params_list: list) -> dict:
    config = []
    for p in params_list:
        acct = _nv(p.get('계정과목',''))
        if not acct: continue
        top_n = int(_safe_float(p.get('개수', p.get('top_n', 10)), 10))
        direction = _nv(p.get('금액열',''), blank_vals=('nan','none','')) or 'both'
        if direction not in ('차변','대변','both'): direction = 'both'
        config.append((acct, top_n, direction))
    if not config: return {'심층분석': pd.DataFrame({'안내':['계정과목 파라미터가 없습니다.']})}

    gc = _get_gubun_col(df)
    has_g = gc is not None
    out = {}

    for idx, (acct_name, top_n, direction) in enumerate(config, 1):
        filtered = df[df[COL_ACCOUNT].str.contains(acct_name, na=False, regex=False)]
        if filtered.empty: continue
        sname     = _safe_sheet(f'Top_{idx}_{re.sub(r"[^가-힣a-zA-Z0-9]","",acct_name)[:16]}')
        grp_cols  = [gc, COL_CLIENT] if has_g else [COL_CLIENT]

        debit_top = pd.DataFrame()
        if direction in ('차변', 'both'):
            d_rows = filtered[filtered[COL_DEBIT] != 0]
            if not d_rows.empty and COL_CLIENT in df.columns:
                debit_top = (d_rows.groupby(grp_cols)[COL_DEBIT].agg(['count','sum'])
                                   .reset_index().sort_values('sum', ascending=False)
                                   .head(top_n*(2 if has_g else 1)))
                rn = {grp_cols[-1]:'거래처명','count':'전표수(차)','sum':'차변금액'}
                if has_g: rn[gc] = '구분'
                debit_top = debit_top.rename(columns=rn)
                debit_top.insert(0, '계정명', acct_name)

        credit_top = pd.DataFrame()
        if direction in ('대변', 'both'):
            c_rows = filtered[filtered[COL_CREDIT] != 0]
            if not c_rows.empty and COL_CLIENT in df.columns:
                credit_top = (c_rows.groupby(grp_cols)[COL_CREDIT].agg(['count','sum'])
                                    .reset_index().sort_values('sum', ascending=False)
                                    .head(top_n*(2 if has_g else 1)))
                rn = {grp_cols[-1]:'거래처명','count':'전표수(대)','sum':'대변금액'}
                if has_g: rn[gc] = '구분'
                credit_top = credit_top.rename(columns=rn)
                credit_top.insert(0, '계정명', acct_name)

        if direction == 'both':
            combined = pd.concat([debit_top, credit_top], axis=1)
        elif direction == '차변':
            combined = debit_top
        else:
            combined = credit_top

        if not combined.empty:
            out[sname] = combined

    return out or {'심층분석': pd.DataFrame({'결과':['데이터 없음']})}


# ── 15. AI 계정별 분석 ────────────────────────────────────────────────────────
def _prepare_ai_material(df: pd.DataFrame, acct: str):
    """계정과목 1건에 대해 (월별집계, 마스킹샘플) 튜플을 반환. 대상 없으면 (None, None).

    메뉴15(analyze_ai_preparation)·메뉴26(analyze_ai_review) 공용 전처리.
    """
    filtered = df[df[COL_ACCOUNT].str.contains(acct, na=False, regex=False)].copy()
    if filtered.empty:
        return None, None

    filtered['YM'] = pd.to_datetime(filtered[COL_DATE], errors='coerce').dt.strftime('%Y-%m')
    monthly = filtered.groupby('YM')[[COL_DEBIT, COL_CREDIT]].agg(['sum','count']).reset_index()
    monthly.columns = ['YM','차변합계','차변건수','대변합계','대변건수']

    filtered['MaxAmt'] = filtered[[COL_DEBIT, COL_CREDIT]].max(axis=1)
    sample = (filtered[filtered['MaxAmt'] >= filtered['MaxAmt'].quantile(0.90)].copy()
              if len(filtered) > 10 else filtered.copy())
    if COL_CLIENT in sample.columns:
        sample[COL_CLIENT] = sample[COL_CLIENT].apply(get_safe_client_name)
    for col in MASK_TARGET_COLS:
        if col in sample.columns: sample[col] = sample[col].apply(mask_sensitive_info)
    sample = sample.drop(columns=['MaxAmt','YM'], errors='ignore')

    return monthly, sample


def analyze_ai_preparation(df: pd.DataFrame, params_list: list) -> dict:
    targets = [_nv(p.get('계정과목','')) for p in params_list if _nv(p.get('계정과목',''))]
    if not targets: return {'AI분석': pd.DataFrame({'안내':['계정과목 파라미터 없음']})}
    out = {}
    for acct in targets:
        monthly, sample = _prepare_ai_material(df, acct)
        if monthly is None: continue
        safe_nm = re.sub(r'[\\/*?:\[\]]', '', acct)[:10]
        out[_safe_sheet(f'AI_{safe_nm}_월별')] = monthly
        out[_safe_sheet(f'AI_{safe_nm}_샘플')] = sample

    if GLOBAL_SAFE_MAP:
        out['_암호해독표'] = pd.DataFrame(list(GLOBAL_SAFE_MAP.items()), columns=['실명','가명'])
    return out or {'AI분석': pd.DataFrame({'결과':['분석 대상 없음']})}


# ── 26. AI 계정별 분석 실행 (Gemini 구조화 응답) ───────────────────────────────
_AI_REVIEW_SCHEMA = {
    'type': 'OBJECT',
    'properties': {
        '위험평가':    {'type': 'STRING', 'enum': ['높음', '중간', '낮음']},
        '주요특이사항': {'type': 'STRING'},
        '결론':        {'type': 'STRING', 'enum': ['적정', '추가확인필요', '부적정']},
        '결론근거':     {'type': 'STRING'},
        '추가확인사항': {'type': 'STRING'},
        '확인필요전표': {
            'type': 'ARRAY',
            'description': (
                '결론근거·추가확인사항의 근거가 되는 개별 전표. 반드시 프롬프트에 '
                '제공된 샘플 거래의 전표번호만 인용할 것(임의 생성 금지). '
                '확인할 특정 전표가 없으면 빈 배열([])로 둘 것.'
            ),
            'items': {
                'type': 'OBJECT',
                'properties': {
                    '전표번호': {'type': 'STRING', 'description': '제공된 샘플 데이터의 전표번호를 그대로 인용'},
                    '확인사유': {'type': 'STRING'},
                },
                'required': ['전표번호', '확인사유'],
            },
        },
    },
    'required': ['위험평가', '주요특이사항', '결론', '결론근거'],
}


def _get_gemini_client_and_config():
    """GEMINI_API_KEY(.env)로 genai.Client 생성 + 구조화 출력용 config 반환.

    google-generativeai(구 SDK)는 지원 종료되어 신규 google-genai SDK 사용.
    """
    from google import genai
    from google.genai import types

    api_key = os.environ.get('GEMINI_API_KEY', '').strip()
    if not api_key:
        raise RuntimeError(
            "GEMINI_API_KEY가 설정되어 있지 않습니다. 프로젝트 루트 .env 파일에 "
            "GEMINI_API_KEY=발급받은키 형식으로 추가한 뒤 다시 실행하세요."
        )
    # 2.5 계열은 신규 API 키에 404(no longer available to new users)로 막히는 사례 확인됨
    # (2026-08 기준) → '-latest' 별칭 기본값으로 향후 모델 교체에도 안전하게 대응
    model_name = os.environ.get('GEMINI_MODEL', 'gemini-flash-lite-latest').strip()
    client = genai.Client(api_key=api_key)
    config = types.GenerateContentConfig(
        response_mime_type='application/json',
        response_schema=_AI_REVIEW_SCHEMA,
    )
    return client, model_name, config


def _build_ai_review_prompt(company_name: str, acct: str, monthly: pd.DataFrame, sample: pd.DataFrame) -> str:
    monthly_txt = monthly.to_csv(index=False)
    sample_txt  = sample.head(30).to_csv(index=False)
    return f"""당신은 외부감사인입니다. 아래는 '{company_name}'의 '{acct}' 계정 분개장 데이터입니다.
거래처명·민감정보는 이미 가명 처리·마스킹되어 있습니다.

[월별 차변/대변 합계·건수]
{monthly_txt}

[금액 상위 10% 샘플 거래 (최대 30건, 전표번호 포함)]
{sample_txt}

위 데이터를 바탕으로 이 계정의 위험평가, 주요 특이사항, 감사 결론, 결론근거, 추가로 확인이 필요한 사항을
JSON 스키마에 맞춰 한국어로 답변하세요. 특히 결론근거·추가확인사항에서 언급한 우려사항과 직접 관련된
개별 전표가 위 샘플 데이터에 있다면, 그 전표번호를 반드시 '확인필요전표'에 그대로 인용해 나열하세요.
샘플에 없는 전표번호를 임의로 만들어내지 마세요. 특정할 전표가 없으면 '확인필요전표'를 빈 배열로 두세요."""


def analyze_ai_review(df: pd.DataFrame, params_list: list) -> dict:
    """메뉴15와 동일한 계정과목 파라미터를 읽어 Gemini에 구조화 분석을 요청한다.

    두 개의 표를 만든다.
      AI검토결과      : 계정과목당 1행 요약(위험평가/결론 등) — mapping_list에서
                        그대로(remarks 비움) 또는 AI_INJECT remarks로 주입.
      AI검토_확인전표 : AI가 결론 근거로 지목한 개별 전표 1건당 1행. 전표번호로
                        원본 샘플 데이터(sample)를 다시 찾아 금액·거래처·일자를
                        채운다(AI가 만든 숫자를 그대로 믿지 않고 원본 대조).
    """
    import json

    targets = [_nv(p.get('계정과목','')) for p in params_list if _nv(p.get('계정과목',''))]
    if not targets:
        return {'AI검토결과': pd.DataFrame({'안내': ['계정과목 파라미터 없음']})}

    company_name = os.path.basename(_COMPANY_DIR) if _COMPANY_DIR else ''
    client, model_name, config = _get_gemini_client_and_config()

    summary_rows, detail_rows = [], []
    for acct in targets:
        monthly, sample = _prepare_ai_material(df, acct)
        if monthly is None:
            continue
        # 전표번호 → 원본 행 조회용 (AI가 지목한 전표의 실제 금액·거래처·일자를 원본에서 재확인)
        jid_lookup = {}
        if COL_JOURNAL_ID in sample.columns:
            for _, r in sample.iterrows():
                jid_lookup.setdefault(str(r[COL_JOURNAL_ID]), r)

        print(f'    [AI검토] {acct} → Gemini({model_name}) 호출 중...', flush=True)
        try:
            prompt = _build_ai_review_prompt(company_name, acct, monthly, sample)
            resp = client.models.generate_content(model=model_name, contents=prompt, config=config)
            data = json.loads(resp.text)
            summary_rows.append({
                '계정과목':     acct,
                '위험평가':     data.get('위험평가', ''),
                '주요특이사항': data.get('주요특이사항', ''),
                '결론':         data.get('결론', ''),
                '결론근거':     data.get('결론근거', ''),
                '추가확인사항': data.get('추가확인사항', ''),
            })
            for item in data.get('확인필요전표') or []:
                jid = str(item.get('전표번호', '')).strip()
                src_row = jid_lookup.get(jid)
                detail_rows.append({
                    '계정과목':   acct,
                    '전표번호':   jid,
                    '전표일자':   src_row[COL_DATE] if src_row is not None and COL_DATE in src_row else '',
                    '거래처명':   src_row[COL_CLIENT] if src_row is not None and COL_CLIENT in src_row else '',
                    '차변':       src_row[COL_DEBIT] if src_row is not None and COL_DEBIT in src_row else '',
                    '대변':       src_row[COL_CREDIT] if src_row is not None and COL_CREDIT in src_row else '',
                    '확인사유':   item.get('확인사유', ''),
                    '비고':       '' if src_row is not None else '원본 샘플에서 전표번호 미확인 — AI 응답 그대로 기재',
                })
        except Exception as e:
            print(f'    [AI검토] {acct} 오류: {e}', flush=True)
            summary_rows.append({
                '계정과목': acct, '위험평가': '', '주요특이사항': '',
                '결론': 'API오류', '결론근거': str(e), '추가확인사항': '',
            })

    out = {
        'AI검토결과': pd.DataFrame(summary_rows) if summary_rows else pd.DataFrame({'결과': ['분석 대상 없음']}),
    }
    if detail_rows:
        out['AI검토_확인전표'] = pd.DataFrame(detail_rows)
    return out


# ── 16. 데이터·헤더 확인 ──────────────────────────────────────────────────────
def analyze_header_check(df: pd.DataFrame, params_list: list) -> pd.DataFrame:
    required = [(COL_DEBIT,'차변'),(COL_CREDIT,'대변'),(COL_ACCOUNT,'계정명'),
                (COL_CLIENT,'거래처명'),(COL_DATE,'전표일자'),(COL_JOURNAL_ID,'전표번호')]
    rows = [{'항목': name, '인식': 'O' if col in df.columns else 'X',
              '실제컬럼': col if col in df.columns else '없음'} for col, name in required]
    gc = _get_gubun_col(df)
    if gc:
        rows.append({'항목':'구분(당기/전기)','인식':'O','실제컬럼':str(df[gc].unique().tolist())})
    rows.append({'항목':'총 행수','인식':'-','실제컬럼':str(len(df))})
    rows.append({'항목':'전체 컬럼','인식':'-','실제컬럼':str(list(df.columns))})
    return pd.DataFrame(rows)


# ── 17. 거래처 분석 ───────────────────────────────────────────────────────────
def analyze_client_detail(df: pd.DataFrame, params_list: list) -> dict:
    out = {}
    for i, p in enumerate(params_list, 1):
        accts    = [a.strip() for a in str(p.get('계정과목','')).split(',')
                    if a.strip() and a.strip().lower() not in ('nan','(전체)','')]
        clients  = [c.strip() for c in str(p.get('거래처명','')).split(',')
                    if c.strip() and c.strip().lower() not in ('nan','')]
        vtype    = _nv(p.get('금액열',''), blank_vals=('nan','none','')) or 'both'
        if vtype not in ('차변','대변','both'): vtype = 'both'
        job_name = _nv(p.get('작업명','')) or f'거래처{i:02d}'
        if not clients: continue

        mask_a = pd.Series(True, index=df.index) if not accts else pd.Series(False, index=df.index)
        for acct in accts: mask_a |= _account_match_flexible(df[COL_ACCOUNT], acct)
        mask_c = pd.Series(False, index=df.index)
        cli_col = df[COL_CLIENT].fillna('').astype(str)
        for cli in clients: mask_c |= cli_col.str.contains(cli, na=False, regex=False)

        filtered = df[mask_a & mask_c].copy()
        if filtered.empty: continue
        if vtype == '차변': filtered = filtered[filtered[COL_DEBIT]  != 0]
        elif vtype == '대변': filtered = filtered[filtered[COL_CREDIT] != 0]
        if filtered.empty: continue

        filtered['YM'] = pd.to_datetime(filtered[COL_DATE], errors='coerce').dt.strftime('%Y-%m')
        monthly = filtered.groupby('YM').agg({COL_DEBIT:'sum',COL_CREDIT:'sum'}).reset_index()
        monthly.columns = ['YM','차변합계','대변합계']
        monthly['합계'] = monthly['차변합계'] + monthly['대변합계']

        dc = [c for c in ['구분', COL_DATE, COL_JOURNAL_ID, COL_ACCOUNT,
                           COL_DEBIT, COL_CREDIT, COL_CLIENT, COL_DESC] if c in filtered.columns]
        sname = _safe_sheet(f'거래처_{re.sub(r"[^가-힣a-zA-Z0-9]","",job_name)[:18]}')
        out[sname]               = filtered[dc]
        out[sname + '_월별합산'] = monthly

    return out or {'거래처분석': pd.DataFrame({'안내':['파라미터에 거래처명이 없습니다.']})}


# ── 18. 벤포드 이탈 상세 추출 ────────────────────────────────────────────────
def analyze_benford_deviation(df: pd.DataFrame, params_list: list) -> dict:
    targets, threshold, max_per = [], 0.03, 500
    for p in params_list:
        acct = _nv(p.get('계정과목',''))
        col  = _nv(p.get('금액열',''), blank_vals=('nan','none','')) or '차변'
        if col not in ('차변','대변'): col = '차변'
        if acct: targets.append((acct, col))
        t = _safe_float(p.get('임계값'), 0.03)
        if 0.001 <= t <= 0.3: threshold = t
        m = p.get('최대건수')
        if m:
            try: max_per = int(m)
            except Exception: pass
    if not targets: targets = list(DEFAULT_BENFORD_TARGETS)

    all_detail, summary_rows = [], []
    for acct, direction in targets:
        tcol   = COL_DEBIT if direction == '차변' else COL_CREDIT
        mask   = _account_match_flexible(df[COL_ACCOUNT], acct)
        subset = df[mask & (df[tcol] > 0)].copy()
        n      = len(subset)
        if n < BENFORD_MIN_ROWS:
            summary_rows.append({'계정':acct,'방향':direction,'선행자릿수':'-','이탈정도':'-',
                                  '추출건수':'-','전체건수':n,'비고':'데이터 부족'})
            continue
        subset['Digit'] = subset[tcol].apply(get_first_digit)
        dg     = subset[subset['Digit'] >= 1]['Digit']
        counts = dg.value_counts(normalize=True).sort_index()
        deviant_set = {d for d in range(1,10) if abs(counts.get(d,0)-BENFORD_PROBS[d]) >= threshold}
        if not deviant_set:
            summary_rows.append({'계정':acct,'방향':direction,'선행자릿수':'-','이탈정도':'-',
                                  '추출건수':'-','전체건수':n,'비고':'임계값 이상 이탈 없음'})
            continue
        deviant_df = subset[subset['Digit'].isin(deviant_set)].copy()
        deviant_df['이탈정도'] = deviant_df['Digit'].map(
            lambda d: round(counts.get(d,0)-BENFORD_PROBS[d],3))
        out_cols = [c for c in ['구분', COL_DATE, COL_JOURNAL_ID, COL_ACCOUNT, COL_CLIENT, COL_DESC,
                                 tcol, 'Digit', '이탈정도'] if c in deviant_df.columns]
        for d in sorted(deviant_set):
            sub = deviant_df[deviant_df['Digit'] == d].head(max_per)
            summary_rows.append({'계정':acct,'방향':direction,'선행자릿수':d,
                                  '이탈정도':round(counts.get(d,0)-BENFORD_PROBS[d],3),
                                  '추출건수':len(sub),
                                  '전체건수':int((deviant_df['Digit']==d).sum()),'비고':''})
            exp = sub[[c for c in out_cols if c in sub.columns]].copy()
            exp['계정분류'] = f'{acct}({direction})'
            all_detail.append(exp)

    out = {}
    if summary_rows: out['벤포드이탈_요약'] = pd.DataFrame(summary_rows)
    if all_detail:   out['벤포드이탈_상세'] = pd.concat(all_detail, ignore_index=True)
    return out or {'벤포드이탈': pd.DataFrame({'결과':['추출 데이터 없음']})}


# ── 19. 월별 전계정 분석 ──────────────────────────────────────────────────────
def analyze_monthly_full_account(df: pd.DataFrame, params_list: list) -> pd.DataFrame:
    work = df.copy()
    work['Month'] = pd.to_datetime(work[COL_DATE], errors='coerce').dt.month
    grp = ['구분', COL_ACCOUNT, 'Month'] if '구분' in work.columns else [COL_ACCOUNT, 'Month']
    return work.groupby(grp)[[COL_DEBIT, COL_CREDIT]].sum().reset_index()


# ── 전기 계정별_거래처별명세 공용 헬퍼 (20번·27번에서 공유) ─────────────────────
def _find_prev_detail_file(prev_dir: str):
    """data/previous 폴더에서 전기 계정별_거래처별명세 파일 경로 탐색. 없으면 None."""
    if not os.path.isdir(prev_dir):
        return None
    for f in sorted(os.listdir(prev_dir)):
        if ('계정별' in f or '거래처별' in f) and f.endswith('.xlsx') and not f.startswith('~$'):
            if '명세' in f or '거래처별' in f:
                return os.path.join(prev_dir, f)
    return None


def _find_prev_sheet(prev_sheets: list, acct_name: str):
    """계정명과 가장 근접한 전기명세 시트명을 찾는다 (공백/특수문자 정규화 후 매칭)."""
    if acct_name in prev_sheets:
        return acct_name
    norm = acct_name.replace(' ', '')
    for s in prev_sheets:
        if s.replace(' ', '') == norm:
            return s
    norm2 = re.sub(r'[\s()（）]', '', acct_name)
    for s in prev_sheets:
        if re.sub(r'[\s()（）]', '', s) == norm2:
            return s
    for s in prev_sheets:
        if acct_name in s or s in acct_name:
            return s
    return None


def _load_prev_balances(prev_xl: pd.ExcelFile, sheet_name: str) -> dict:
    """전기명세 시트에서 {거래처명: 잔액} 딕셔너리 로드."""
    pdf = pd.read_excel(prev_xl, sheet_name=sheet_name, header=0)
    pdf.columns = [str(c).strip() for c in pdf.columns]
    vendor_col = next((c for c in pdf.columns if '거래처' in c), None)
    bal_col = next((c for c in pdf.columns if '잔' in c), None)
    if not vendor_col or not bal_col:
        return {}
    balances = {}
    for _, row in pdf.iterrows():
        vendor = str(row[vendor_col]).strip() if pd.notna(row[vendor_col]) else ''
        if not vendor or vendor.replace(' ', '') in ('합계:', '합계', 'nan', ''):
            continue
        try:
            bal = float(row[bal_col]) if pd.notna(row[bal_col]) else 0
        except (ValueError, TypeError):
            bal = 0
        balances[vendor] = balances.get(vendor, 0) + bal
    return balances


def _prev_balance_total(prev_xl, acct_name: str):
    """전기명세에서 계정명 전체 합산 잔액. 시트가 없으면 None (표시 안 함)."""
    if prev_xl is None:
        return None
    sheet = _find_prev_sheet(prev_xl.sheet_names, acct_name)
    if not sheet:
        return None
    return sum(_load_prev_balances(prev_xl, sheet).values())


# ── 20. 당기증감분석 (계정별 거래처별) ───────────────────────────────────────
def analyze_balance_movement(df: pd.DataFrame, params_list: list) -> dict:
    """전기 계정별_거래처별명세에서 기초잔액, 당기 분개장에서 증감 산출하여 기말잔액 계산.
    자산(차변): 기초잔액 + 당기증가(차변) - 당기감소(대변) = 기말잔액
    부채(대변): 기초잔액 + 당기증가(대변) - 당기감소(차변) = 기말잔액
    """
    global _COMPANY_DIR
    if _COMPANY_DIR is None:
        return {'잔액증감분석': pd.DataFrame({'오류': ['company_dir 미설정']})}

    prev_dir = os.path.join(_COMPANY_DIR, 'data', 'previous')
    prev_file = _find_prev_detail_file(prev_dir)
    if not prev_file:
        return {'잔액증감분석': pd.DataFrame({'오류': ['전기 계정별_거래처별명세 파일 없음']})}

    # ── 거래처 매핑 파일 탐색 ──
    vendor_mapping = {}
    for f in sorted(os.listdir(prev_dir)):
        if '매핑' in f and not f.startswith('~$'):
            map_path = os.path.join(prev_dir, f)
            try:
                if f.endswith('.csv'):
                    for enc in ['utf-8', 'cp949', 'euc-kr']:
                        try:
                            mdf = pd.read_csv(map_path, encoding=enc)
                            break
                        except UnicodeDecodeError:
                            continue
                else:
                    mdf = pd.read_excel(map_path)
                if '분개장거래처' in mdf.columns and '전기명세거래처' in mdf.columns:
                    for _, row in mdf.iterrows():
                        acct = str(row.get('계정명', '')).strip()
                        j_v = str(row.get('분개장거래처', '')).strip()
                        p_v = str(row.get('전기명세거래처', '')).strip()
                        if j_v and p_v and p_v not in ('nan', ''):
                            vendor_mapping[(acct, j_v)] = p_v
                    print(f'    └ 거래처 매핑: {len(vendor_mapping)}건')
            except Exception as e:
                print(f'    ⚠️ 매핑 파일 로드 실패: {e}')
            break

    # ── 파라미터 로드 (params_list → 직접 탐색 fallback) ──
    targets = []
    for p in params_list:
        acct = _nv(p.get('계정명', p.get('계정과목', '')))
        gubun = _nv(p.get('구분', ''))
        if acct and gubun in ('차변', '대변', '자산', '부채'):
            targets.append((acct, gubun))

    if not targets:
        task_path = os.path.join(
            _COMPANY_DIR, f'task_list_{os.path.basename(_COMPANY_DIR)}.xlsx')
        if os.path.isfile(task_path):
            txl = pd.ExcelFile(task_path)
            for s in txl.sheet_names:
                if '증감' in s and ('거래처' in s or '계정' in s or '게정' in s):
                    tdf = pd.read_excel(txl, sheet_name=s, header=None)
                    for i, row in tdf.iterrows():
                        vals = [str(v).strip() for v in row if pd.notna(v)]
                        if '구분' in vals:
                            for j in range(i + 1, len(tdf)):
                                drow = tdf.iloc[j]
                                a = str(drow.iloc[0]).strip() if pd.notna(drow.iloc[0]) else ''
                                g = str(drow.iloc[1]).strip() if len(drow) > 1 and pd.notna(drow.iloc[1]) else ''
                                if a and a not in ('nan', '') and g in ('차변', '대변', '자산', '부채'):
                                    targets.append((a, g))
                            break
                    break

    if not targets:
        return {'잔액증감분석': pd.DataFrame({'안내': ['파라미터에 계정명/구분이 없습니다.']})}

    print(f'    └ 분석 대상: {len(targets)}개 계정')

    # ── 전기 명세 ExcelFile ──
    prev_xl = pd.ExcelFile(prev_file)
    prev_sheets = prev_xl.sheet_names

    # ── 계정별 분석 실행 ──
    all_results = {}

    for acct_name, gubun in targets:
        is_asset = gubun in ('차변', '자산')

        prev_sheet = _find_prev_sheet(prev_sheets, acct_name)
        prev_balances = _load_prev_balances(prev_xl, prev_sheet) if prev_sheet else {}
        if prev_sheet:
            print(f'      {acct_name}: 전기 {len(prev_balances)}건 (시트: {prev_sheet})')
        else:
            print(f'      {acct_name}: 전기 시트 없음')

        mask = _account_match_flexible(df[COL_ACCOUNT], acct_name)
        current = df[mask].copy()

        journal_vendors = {}
        if not current.empty and COL_CLIENT in current.columns:
            for vendor, group in current.groupby(COL_CLIENT):
                v = str(vendor).strip()
                if v and v != 'nan':
                    journal_vendors[v] = {
                        'debit': group[COL_DEBIT].sum() if COL_DEBIT in group.columns else 0,
                        'credit': group[COL_CREDIT].sum() if COL_CREDIT in group.columns else 0,
                    }

        if not prev_balances and not journal_vendors:
            continue

        acct_map_j2p = {k[1]: v for k, v in vendor_mapping.items() if k[0] == acct_name}
        acct_map_p2j = {v: k[1] for k, v in vendor_mapping.items() if k[0] == acct_name}

        all_vendors = set(prev_balances.keys()) | set(journal_vendors.keys())
        rows = []
        matched_prev = set()
        matched_journal = set()

        for vendor in sorted(all_vendors):
            prev_bal = prev_balances.get(vendor, 0)
            prev_key = vendor
            if prev_bal == 0 and vendor in acct_map_j2p:
                prev_key = acct_map_j2p[vendor]
                prev_bal = prev_balances.get(prev_key, 0)

            jv = journal_vendors.get(vendor, None)
            journal_key = vendor
            if jv is None and vendor in acct_map_p2j:
                journal_key = acct_map_p2j[vendor]
                jv = journal_vendors.get(journal_key, None)
            if jv is None:
                jv = {'debit': 0, 'credit': 0}

            matched_prev.add(prev_key)
            matched_journal.add(journal_key)

            debit_sum = jv['debit']
            credit_sum = jv['credit']

            if is_asset:
                increase = debit_sum
                decrease = credit_sum
                ending = prev_bal + increase - decrease
                row = {
                    '계정명': acct_name, '거래처명': vendor,
                    '기초잔액': prev_bal, '당기증가': increase,
                    '당기감소': decrease, '기말잔액': ending,
                }
            else:
                increase = credit_sum
                decrease = debit_sum
                ending = prev_bal + increase - decrease
                row = {
                    '계정명': acct_name, '거래처명': vendor,
                    '기초잔액': prev_bal, '당기감소': decrease,
                    '당기증가': increase, '기말잔액': ending,
                }
            rows.append(row)

        if rows:
            result_df = pd.DataFrame(rows)
            amount_cols = [c for c in result_df.columns if c not in ('계정명', '거래처명')]
            result_df = result_df[result_df[amount_cols].abs().sum(axis=1) > 0]

            if not result_df.empty:
                result_df = result_df.sort_values('기말잔액', key=abs, ascending=False)
                total = {c: result_df[c].sum() for c in amount_cols}
                total['계정명'] = acct_name
                total['거래처명'] = '합  계'
                result_df = pd.concat(
                    [result_df, pd.DataFrame([total])], ignore_index=True)
                sname = _safe_sheet(
                    f'증감_{re.sub(r"[^가-힣a-zA-Z0-9]", "", acct_name)[:20]}')
                all_results[sname] = result_df

    return all_results or {'잔액증감분석': pd.DataFrame({'결과': ['분석 대상 없음']})}


# ── 21. 총계정원장 ────────────────────────────────────────────────────────────
def analyze_general_ledger(df: pd.DataFrame, params_list: list) -> dict:
    """계정별 월별 차변/대변 집계. 연도 2개 이상이면 연도 비교 형식(행=월, 열=연도)으로 출력.
    구분(자산/부채/매출액)은 시트명 접두어로 사용."""
    out = {}
    for p in params_list:
        acct  = _nv(p.get('계정과목', ''))
        gubun = _nv(p.get('구분', ''))
        if not acct: continue

        mask = _account_match_flexible(df[COL_ACCOUNT], acct)

        subset = df[mask].copy()
        if subset.empty:
            continue

        subset['YM']    = pd.to_datetime(subset[COL_DATE], errors='coerce').dt.strftime('%Y-%m')
        subset          = subset[subset['YM'].notna()]
        subset['Year']  = subset['YM'].str[:4]
        subset['Month'] = subset['YM'].str[5:7].astype(int)

        years = sorted(subset['Year'].unique())
        dr    = subset[subset[COL_DEBIT]  != 0]
        cr    = subset[subset[COL_CREDIT] != 0]

        if len(years) >= 2:
            da = dr.groupby(['Year', 'Month']).agg(
                차변합계=(COL_DEBIT, 'sum'), 차변건수=(COL_DEBIT, 'count')).reset_index()
            ca = cr.groupby(['Year', 'Month']).agg(
                대변합계=(COL_CREDIT, 'sum'), 대변건수=(COL_CREDIT, 'count')).reset_index()
            ym = pd.merge(da, ca, on=['Year', 'Month'], how='outer').fillna(0)

            result = pd.DataFrame({'Month': range(1, 13)})
            for year in years:
                yd = ym[ym['Year'] == year][
                    ['Month', '차변합계', '대변합계', '차변건수', '대변건수']
                ].rename(columns={c: f'{year}_{c}' for c in ['차변합계', '대변합계', '차변건수', '대변건수']})
                result = pd.merge(result, yd, on='Month', how='left').fillna(0)
            for year in years:
                result[f'{year}_차변건수'] = result[f'{year}_차변건수'].astype(int)
                result[f'{year}_대변건수'] = result[f'{year}_대변건수'].astype(int)
            if len(years) == 2:
                y0, y1 = years[0], years[1]
                result['증감_차변합계'] = result[f'{y1}_차변합계'] - result[f'{y0}_차변합계']
                result['증감_대변합계'] = result[f'{y1}_대변합계'] - result[f'{y0}_대변합계']
            result.insert(0, '월', result['Month'].apply(lambda x: f'{x}월'))
            result = result.drop(columns=['Month'])
        else:
            da = dr.groupby('YM').agg(차변합계=(COL_DEBIT, 'sum'), 차변건수=(COL_DEBIT, 'count')).reset_index()
            ca = cr.groupby('YM').agg(대변합계=(COL_CREDIT, 'sum'), 대변건수=(COL_CREDIT, 'count')).reset_index()
            merged = pd.merge(da, ca, on='YM', how='outer').fillna(0)
            # 데이터 없는 월도 표시: 해당 연도 12개월 스켈레톤에 left join
            year_val = years[0]
            all_ym = pd.DataFrame({'YM': [f'{year_val}-{m:02d}' for m in range(1, 13)]})
            result = pd.merge(all_ym, merged, on='YM', how='left').fillna(0)
            result['차변건수'] = result['차변건수'].astype(int)
            result['대변건수'] = result['대변건수'].astype(int)
            result = result.rename(columns={'YM': '월'})

        total = {'월': '합  계'}
        for col in result.columns[1:]:
            total[col] = result[col].sum()
        result = pd.concat([result, pd.DataFrame([total])], ignore_index=True)

        prefix = f'{gubun}_' if gubun else ''
        sname  = _safe_sheet(f'총계정원장_{prefix}{re.sub(r"[^가-힣a-zA-Z0-9]", "", acct)[:18]}')
        out[sname] = result

    return out or {'총계정원장': pd.DataFrame({'결과': ['분석 대상 없음']})}


# ── 22. 은행조회서 완전성 ──────────────────────────────────────────────────────
_BANK_DEFAULT_ACCOUNTS = [
    '이자비용', '단기차입금', '장기차입금',
    '유동성장기부채', '유동성장기차입금',
    '전환사채', '유동성전환사채',
    '전환상환우선주', '유동성전환상환우선주',
]

_FI_PATTERN = re.compile(
    r'([\w가-힣()（）]*'
    r'(?:저축은행|저축|은행|금고|신협|조합|증권|캐피탈|카드|보험|파이낸스|크레딧|리스))'
)

# 계좌 별칭에 '은행' 표기 없이 약칭만 쓰이는 경우(예: "우리7396-주계좌", "하나B7404-주계좌",
# "국민6157-ONE KB", "기업(017)") 대응 — 거래처명 맨 앞의 은행 약칭을 정식명으로 매핑
# (2026-08-31 graphy 검토 중 발견: 위 _FI_PATTERN만으로는 이런 계좌명이 전혀 매칭되지 않아
#  은행조회서완전성 결과의 금융기관_조회서목록 시트 헤더에 해당 계정이 통째로 빠지는 문제)
_FI_BANK_ABBR = {
    '국민': '국민은행', '신한': '신한은행', '우리': '우리은행', '하나': '하나은행',
    '기업': '기업은행', '농협': '농협은행', '수협': '수협은행', '산업': '산업은행',
    '수출입': '수출입은행', '대구': '대구은행', '부산': '부산은행', '광주': '광주은행',
    '전북': '전북은행', '경남': '경남은행', '제주': '제주은행', '씨티': '씨티은행',
    'SC제일': 'SC제일은행', '카카오': '카카오뱅크', '케이뱅크': '케이뱅크', '토스': '토스뱅크',
    '새마을금고': '새마을금고',
}
_FI_ABBR_PATTERN = re.compile(
    '^(' + '|'.join(sorted(_FI_BANK_ABBR, key=len, reverse=True)) + ')'
)


def _extract_fi(client_name: str) -> str:
    name = str(client_name).strip()
    if not name or name.lower() in ('nan', 'none', ''):
        return ''
    m = _FI_PATTERN.search(name)
    if m:
        return m.group(1)
    m = _FI_ABBR_PATTERN.match(name)
    return _FI_BANK_ABBR[m.group(1)] if m else ''


def analyze_bank_confirmation(df: pd.DataFrame, params_list: list) -> dict:
    """22. 은행조회서 완전성: 차입금/이자비용 등 관련 계정 상세내역 + 금융기관 요약."""
    # 파라미터 시트에 계정 목록이 있으면 사용, 없으면 기본값
    account_list = [
        _nv(p.get('계정과목', '') or p.get('계정명', '') or p.get('계정', ''))
        for p in params_list
    ]
    account_list = [a for a in account_list if a]
    if not account_list:
        account_list = list(_BANK_DEFAULT_ACCOUNTS)

    all_rows = []
    for acct in account_list:
        mask = df[COL_ACCOUNT].astype(str).str.contains(acct, na=False, regex=False)
        sub = df[mask].copy()
        if sub.empty:
            continue
        sub.insert(0, '조회계정', acct)
        all_rows.append(sub)

    if not all_rows:
        return {'은행조회서완전성': pd.DataFrame({'안내': ['해당 계정의 전표 내역이 없습니다.']})}

    combined = pd.concat(all_rows, ignore_index=True)

    # 금융기관명 추출 (거래처명 바로 뒤에 컬럼 삽입)
    if COL_CLIENT in combined.columns:
        combined['금융기관명'] = combined[COL_CLIENT].apply(_extract_fi)
    else:
        combined['금융기관명'] = ''

    priority = ['조회계정']
    if '구분' in combined.columns: priority.append('구분')
    for c in [COL_DATE, COL_JOURNAL_ID, COL_ACCOUNT, COL_DEBIT, COL_CREDIT, COL_CLIENT]:
        if c in combined.columns: priority.append(c)
    priority.append('금융기관명')
    if COL_DESC in combined.columns: priority.append(COL_DESC)
    other = [c for c in combined.columns if c not in priority]
    combined = combined[priority + other]

    acct_order = {a: i for i, a in enumerate(account_list)}
    combined['_sort'] = combined['조회계정'].map(acct_order).fillna(999)
    sort_cols = ['_sort', COL_DATE] if COL_DATE in combined.columns else ['_sort']
    combined = combined.sort_values(sort_cols).drop(columns=['_sort'])

    results = {'은행조회서완전성': combined}

    # 금융기관별 요약 피벗
    has_fi = combined[combined['금융기관명'].astype(str).str.strip() != '']
    if not has_fi.empty and COL_CLIENT in has_fi.columns:
        pivot = has_fi.groupby(['금융기관명', '조회계정']).size().unstack(fill_value=0)
        acct_cols = [a for a in account_list if a in pivot.columns]
        pivot = pivot.reindex(columns=acct_cols, fill_value=0)
        try:
            pivot_mark = pivot.map(lambda x: '○' if x > 0 else '-')
        except AttributeError:
            pivot_mark = pivot.applymap(lambda x: '○' if x > 0 else '-')

        raw_clients = (
            has_fi.groupby('금융기관명')[COL_CLIENT]
            .apply(lambda s: ', '.join(sorted(s.dropna().astype(str).unique())))
            .rename('거래처명(원본)')
        )
        agg_kw = {'전표건수': ('조회계정', 'count')}
        if COL_DEBIT  in has_fi.columns: agg_kw['차변합계'] = (COL_DEBIT,  'sum')
        if COL_CREDIT in has_fi.columns: agg_kw['대변합계'] = (COL_CREDIT, 'sum')
        totals = has_fi.groupby('금융기관명').agg(**agg_kw)

        summary = pivot_mark.join(raw_clients).join(totals).reset_index()
        summary.insert(1, '조회서발송', 'Y')
        results['금융기관_조회서목록'] = summary

    return results



# ── 23. 계정별 상세거래내역 ──────────────────────────────────────────────────

def analyze_account_transaction_detail(df: pd.DataFrame, params_list: list) -> dict:
    """23. 계정별 상세거래내역: 계정과목별 건별 전표 내역 추출 (차변/대변/전체 선택)."""
    targets = []
    for p in params_list:
        acct = _nv(p.get('계정과목', p.get('계정명', '')))
        col_f = _nv(p.get('금액 유형', p.get('금액유형', p.get('차대구분', ''))),
                    blank_vals=('nan', 'none', ''))
        if acct:
            targets.append((acct, col_f))

    if not targets:
        return {'계정별상세내역': pd.DataFrame({'안내': ['파라미터에 계정과목이 없습니다.']})}

    all_results = {}
    for acct_name, col_f in targets:
        mask = _account_match_flexible(df[COL_ACCOUNT], acct_name)
        sub = df[mask].copy()

        if sub.empty:
            continue

        # 차변/대변 필터
        col_f_norm = col_f.replace(' ', '')
        if col_f_norm in ('차변', '차변만'):
            sub = sub[sub[COL_DEBIT] != 0]
        elif col_f_norm in ('대변', '대변만'):
            sub = sub[sub[COL_CREDIT] != 0]
        # '차변대변모두' 또는 공백이면 전체 유지

        if sub.empty:
            continue

        # 날짜순 정렬
        if COL_DATE in sub.columns:
            sub = sub.sort_values(COL_DATE)

        # 시트명: 상세거래_계정명_차변/대변/차변대변모두
        col_f_label = '차변' if col_f_norm in ('차변', '차변만') \
                      else '대변' if col_f_norm in ('대변', '대변만') \
                      else '차변대변모두'
        acct_short = re.sub(r'[^가-힣a-zA-Z0-9]', '', acct_name)[:12]
        base = _safe_sheet(f'상세거래_{acct_short}_{col_f_label}')
        sheet_name = base
        suffix = 2
        while sheet_name in all_results:
            sheet_name = f'{base[:28]}_{suffix}'
            suffix += 1

        all_results[sheet_name] = sub.reset_index(drop=True)

    if not all_results:
        return {'계정별상세내역': pd.DataFrame({'안내': ['해당 계정의 전표 내역이 없습니다.']})}

    return all_results




# -- 24. 리스완전성검토
_LEASE_FILTER_MOD = None

def _load_lease_filter():
    global _LEASE_FILTER_MOD
    if _LEASE_FILTER_MOD is not None:
        return _LEASE_FILTER_MOD
    try:
        import importlib.util as _ilu
        lf_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               '..', 'account_analyzer', 'lease_analyzer', 'lease_filter.py')
        spec = _ilu.spec_from_file_location('lease_filter', lf_path)
        mod = _ilu.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _LEASE_FILTER_MOD = mod
        return mod
    except Exception as e:
        print(f'    [경고] lease_filter 로드 실패: {e}')
        return None


_DEFAULT_LEASE_ACCOUNTS = ['임차료', '지급수수료', '차량유지비', '건물관리비', '렌탈']


def analyze_lease_completeness(df, params_list):
    """24. 리스완전성검토: K-IFRS 1116 리스 인식 대상 거래 완전성 검토."""
    lf = _load_lease_filter()
    if lf is None:
        return {'리스후보목록': pd.DataFrame({'안내': ['lease_filter 모듈 로드 실패']})}
    target_accounts = [_nv(p.get('계정과목', '')) for p in params_list
                       if _nv(p.get('계정과목', ''))]
    if not target_accounts:
        target_accounts = _DEFAULT_LEASE_ACCOUNTS
    mask = pd.Series(False, index=df.index)
    for acct in target_accounts:
        mask |= _account_match_flexible(df[COL_ACCOUNT], acct)
    sub = df[mask].copy()
    if sub.empty:
        return {'리스후보목록': pd.DataFrame({'안내': ['해당 계정의 데이터가 없습니다.']})}
    col_map = {}
    if COL_ACCOUNT != '계정과목': col_map[COL_ACCOUNT] = '계정과목'
    if COL_CLIENT  != '거래처'  : col_map[COL_CLIENT]  = '거래처'
    if COL_DESC    != '적요'    : col_map[COL_DESC]    = '적요'
    if COL_DEBIT   != '차변'    : col_map[COL_DEBIT]   = '차변'
    if col_map:
        # 거래처(코드)처럼 rename 대상과 같은 이름의 컬럼이 이미 있으면 제거 후 rename
        sub = sub.drop(columns=[v for v in col_map.values() if v in sub.columns], errors='ignore')
        sub = sub.rename(columns=col_map)
    sub = lf.preprocess(sub)
    result_df = lf.aggregate(sub)
    return {'리스후보목록': result_df}


# ── 25. 손익항목 월별 추이 ────────────────────────────────────────────────────
def analyze_pl_comparison(df: pd.DataFrame, params_list: list) -> dict:
    """25. 손익항목 월별 추이 (전기/당기 월별 비교)
    task_list 파라미터: 계정과목 / 구분(차변·대변·both) / 실행여부
    결과: 계정별 월별비교 시트 (거래처별 비교는 2번 거래처비교 메뉴 사용)
    """
    targets = []
    for p in params_list:
        acct = _nv(p.get('계정과목', ''))
        direction = (_nv(p.get('구분', p.get('금액열', '')), blank_vals=('nan', 'none', ''))
                     or '대변')
        if direction not in ('차변', '대변', 'both'):
            direction = '대변'
        if acct:
            targets.append((acct, direction))

    if not targets:
        return {'손익월별분析': pd.DataFrame({'안내': ['계정과목 파라미터가 없습니다.']})}

    if '구분' not in df.columns:
        return {'손익월별분析': pd.DataFrame(
            {'오류': ['구분(전기/당기) 컬럼 없음 — 전기·당기 데이터를 함께 로드하세요.']})}

    out = {}

    for acct_name, direction in targets:
        mask = _account_match_flexible(df[COL_ACCOUNT], acct_name)
        sub = df[mask].copy()
        if sub.empty:
            continue

        sub['Month'] = pd.to_datetime(sub[COL_DATE], errors='coerce').dt.month
        sub['구분_str'] = sub['구분'].astype(str).str.strip()

        if direction == '차변':
            sub['_amt'] = sub[COL_DEBIT].fillna(0)
            label = '차변'
        elif direction == '대변':
            sub['_amt'] = sub[COL_CREDIT].fillna(0)
            label = '대변'
        else:
            sub['_amt'] = sub[COL_DEBIT].fillna(0) + sub[COL_CREDIT].fillna(0)
            label = '합계'

        mon = sub.groupby(['구분_str', 'Month']).agg(
            금액합계=('_amt', 'sum'), 건수=('_amt', 'count')).reset_index()

        prev_m = (mon[mon['구분_str'] == '전기'][['Month', '금액합계', '건수']]
                  .rename(columns={'금액합계': f'전기_{label}', '건수': '전기_건수'}))
        curr_m = (mon[mon['구분_str'] == '당기'][['Month', '금액합계', '건수']]
                  .rename(columns={'금액합계': f'당기_{label}', '건수': '당기_건수'}))

        base = pd.DataFrame({'월': range(1, 13)})
        mr = (base
              .merge(prev_m.rename(columns={'Month': '월'}), on='월', how='left')
              .merge(curr_m.rename(columns={'Month': '월'}), on='월', how='left')
              .fillna(0))
        mr[f'증감_{label}'] = mr[f'당기_{label}'] - mr[f'전기_{label}']
        mr['증감률(%)'] = mr.apply(
            lambda r: round(r[f'증감_{label}'] / r[f'전기_{label}'] * 100, 1)
            if r[f'전기_{label}'] != 0 else 0.0, axis=1)

        total_r = {c: mr[c].sum() if c not in ('월', '증감률(%)') else '' for c in mr.columns}
        total_r['월'] = '합  계'
        if mr[f'전기_{label}'].sum() != 0:
            total_r['증감률(%)'] = round(
                mr[f'증감_{label}'].sum() / mr[f'전기_{label}'].sum() * 100, 1)
        mr = pd.concat([mr, pd.DataFrame([total_r])], ignore_index=True)
        mr['월'] = mr['월'].apply(lambda x: f'{int(x)}월' if isinstance(x, float) else x)
        for c in ['전기_건수', '당기_건수']:
            if c in mr.columns:
                mr[c] = pd.to_numeric(mr[c], errors='coerce').fillna(0).astype(int)

        acct_short = re.sub(r'[^가-힣a-zA-Z0-9]', '', acct_name)[:16]
        out[_safe_sheet(f'손익월별_{acct_short}')] = mr

    return out or {'손익월별분析': pd.DataFrame({'결과': ['분析 대상 없음']})}


# =============================================================================
# 4. 분석 레지스트리  {번호: (이름, 함수)}
# =============================================================================
ANALYSIS_REGISTRY: dict = {
    2:  ('거래처비교',      analyze_client_comparison),
    3:  ('벤포드분석',      analyze_benford),
    4:  ('데이터개요',      analyze_data_overview),
    5:  ('계정명리스트',    analyze_account_list),
    6:  ('사원별집계',      analyze_employee_summary),
    7:  ('일자차이분석',    analyze_date_difference),
    8:  ('상대계정분석',    analyze_counterpart),
    9:  ('키워드검색',      analyze_keyword_search),
    10: ('라운드넘버',      analyze_round_numbers),
    11: ('특수관계자분석',  analyze_related_party),
    12: ('자산부채교차',    analyze_asset_liability_cross),
    13: ('매출비용교차',    analyze_revenue_expense_cross),
    14: ('심층분석',        analyze_top_accounts),
    15: ('AI계정별분석',    analyze_ai_preparation),
    16: ('헤더확인',        analyze_header_check),
    17: ('거래처분석',      analyze_client_detail),
    18: ('벤포드이탈',      analyze_benford_deviation),
    19: ('월별전계정분석',  analyze_monthly_full_account),
    20: ('잔액증감분석',    analyze_balance_movement),
    21: ('총계정원장',      analyze_general_ledger),
    22: ('은행조회서완전성', analyze_bank_confirmation),
    23: ('계정별상세내역',  analyze_account_transaction_detail),
    24: ('리스완전성',       analyze_lease_completeness),
    25: ('손익월별분析',     analyze_pl_comparison),
    26: ('AI계정별분석_실행', analyze_ai_review),
    27: ('감가상각_평가손익분석', analyze_depreciation_valuation),
}

# 결과를 메인 파일이 아닌 별도 파일로 저장하는 task 번호 집합
_SEPARATE_FILE_TASKS = {24}


# =============================================================================
# 5. Task List 읽기
# =============================================================================

def resolve_paths(company_name: str) -> dict:
    company_dir = os.path.join(BASE_DIR, company_name)
    return {
        'company_dir': company_dir,
        'task_list':   os.path.join(company_dir, f'task_list_{company_name}.xlsx'),
        'output':      os.path.join(company_dir, 'results'),
    }

def load_active_tasks(task_list_path: str) -> list:
    if not os.path.isfile(task_list_path):
        raise FileNotFoundError(f'task_list 없음: {task_list_path}')
    xl = pd.ExcelFile(task_list_path)
    df = None
    if TASK_MASTER_SHEET in xl.sheet_names:
        df = pd.read_excel(task_list_path, sheet_name=TASK_MASTER_SHEET)
    else:
        for sh in xl.sheet_names:
            tmp = pd.read_excel(task_list_path, sheet_name=sh)
            if all(any(k in str(c) for c in tmp.columns) for k in ('번호','명','여부')):
                df = tmp; break
    if df is None:
        raise ValueError(f"'분석목록' 시트를 찾을 수 없음: {task_list_path}")

    col_no     = next((c for c in df.columns if '번호' in str(c)), None)
    col_nm     = next((c for c in df.columns if str(c).strip().endswith('명')
                       and '번호' not in str(c) and '설명' not in str(c)), None)
    col_flag   = next((c for c in df.columns if '여부' in str(c)), None)
    col_period = next((c for c in df.columns if '대상' in str(c)), None)
    col_month  = next((c for c in df.columns if '기준월' in str(c) or '종료월' in str(c)), None)
    if not all([col_no, col_nm, col_flag]):
        raise ValueError(f'분석번호/분석명/실행여부 컬럼 없음. 실제 컬럼: {df.columns.tolist()}')

    def _parse_month(val):
        s = str(val).strip().replace('월', '')
        try:
            m = int(float(s))
            return m if 1 <= m <= 12 else None
        except (ValueError, TypeError):
            return None

    flag   = df[col_flag].astype(str).str.strip().str.upper()
    active = df[flag.isin(['Y','O'])].dropna(subset=[col_no])
    tasks  = [
        (int(row[col_no]), str(row[col_nm]).strip(),
         str(row[col_period]).strip() if col_period and str(row[col_period]).strip() not in ('nan', '') else '당기',
         _parse_month(row[col_month]) if col_month else None)
        for _, row in active.iterrows()
    ]
    print(f'  [태스크] {len(tasks)}개: {[f"{n}_{nm}[{p}]" + (f"(~{m}월)" if m else "") for n,nm,p,m in tasks]}')
    return tasks

def load_analysis_params(task_list_path: str, analysis_name: str) -> list:
    xl = pd.ExcelFile(task_list_path)
    # 시트명 공백 normalize 후 매칭 (예: "심층분析 (계정별 Top) " → "심층분析(계정별Top)",
    # "은행조회서 완전성" → "은행조회서완전성" — 중간 공백까지 전부 제거해 비교)
    def _norm_sheet(s: str) -> str:
        return re.sub(r'\s+', '', s)
    norm_map = {_norm_sheet(s): s for s in xl.sheet_names}
    for candidate in [analysis_name, f'{analysis_name}_파라미터']:
        actual = candidate if candidate in xl.sheet_names else norm_map.get(_norm_sheet(candidate))
        if actual is None: continue
        df = pd.read_excel(task_list_path, sheet_name=actual).dropna(how='all')
        if '실행여부' in df.columns:
            flag = df['실행여부'].astype(str).str.strip().str.upper()
            df   = df[flag.isin(['Y','O'])].copy()
        params = df.to_dict('records')
        print(f'    └ 파라미터 시트 [{actual}]: {len(params)}행')
        return params
    return [{}]   # 파라미터 시트 없음 → 함수 내 기본값 사용


# =============================================================================
# 6. 결과 저장
# =============================================================================

def load_settings(task_list_path: str) -> dict:
    """task_list settings 시트에서 회사정보(ClientName/StartDate/EndDate 등) 읽기."""
    try:
        xl = pd.ExcelFile(task_list_path)
        if 'settings' not in xl.sheet_names:
            return {}
        df = pd.read_excel(task_list_path, sheet_name='settings', header=None)
        settings = {}
        for _, row in df.iterrows():
            key = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
            val = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
            if not key or not val or key in ('항목', 'key', 'Key'):
                continue
            # 날짜 형식 정리: '2026-01-01 00:00:00' → '2026-01-01'
            if 'Date' in key or '일자' in key:
                val = val.split(' ')[0].split('T')[0]
            settings[key] = val
        return settings
    except Exception:
        return {}



def _save_lease_completeness_file(results: dict, output_dir: str, company_name: str) -> str:
    """리스완전성검토 결과를 별도 색상 서식 Excel 파일로 저장."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    out_path = os.path.join(output_dir, f'리스완전성_{company_name}.xlsx')
    os.makedirs(output_dir, exist_ok=True)
    _FILL = {
        'O': PatternFill('solid', fgColor='C6EFCE'),
        'X': PatternFill('solid', fgColor='FFCCCC'),
        '?': PatternFill('solid', fgColor='FFEB9C'),
    }
    _FONT = {
        'O': Font(bold=True,  color='276221'),
        'X': Font(bold=False, color='9C0006'),
        '?': Font(bold=False, color='7D4E00'),
    }
    col_labels = {'연간_총발생액': '연간 총 발생액', '거래_발생건수': '거래 발생건수', '대표_적요': '대표 적요'}
    col_widths  = {'거래처': 28, '계정과목': 18, '연간 총 발생액': 16,
                   '거래 발생건수': 12, '대표 적요': 45,
                   '리스인식여부(O/X)': 14, '면제검토': 10, '판단근거': 30, '비고': 20}
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for sheet_name, df in results.items():
            df_out = df.rename(columns=col_labels)
            df_out.to_excel(writer, index=False, sheet_name=sheet_name[:31])
            ws = writer.sheets[sheet_name[:31]]
            hdr_fill = PatternFill('solid', fgColor='D9E1F2')
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for ci, cn in enumerate(df_out.columns, start=1):
                ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(cn, 15)
            if '리스인식여부(O/X)' in df_out.columns:
                jc = df_out.columns.get_loc('리스인식여부(O/X)') + 1
                for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    val = str(row_cells[jc-1].value or '').strip()
                    if val in _FILL:
                        for cell in row_cells: cell.fill = _FILL[val]
                        row_cells[jc-1].font = _FONT[val]
            if '연간 총 발생액' in df_out.columns:
                ac = df_out.columns.get_loc('연간 총 발생액') + 1
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ac, max_col=ac):
                    for cell in row: cell.number_format = '#,##0'
    # -- 판정기준 시트 추가
    _criteria_rows = [
        {'순서': 0, '기준': '적요 - 비리스 키워드',      '판정': 'X', '예시': '유류/주유/보험료/컨설팅/용역비/구독료 등 (임차료 계정 제외)'},
        {'순서': 1, '기준': '계정과목 - 임차료 포함',     '판정': 'O', '예시': '임차료 계정이면 무조건 리스 인식'},
        {'순서': 2, '기준': '적요 - 리스 키워드',         '판정': 'O', '예시': '임차/임대/사무실/창고/공장/호실/렌트카/렌터카 등'},
        {'순서': 3, '기준': '거래처 - 렌탈/리스 업체',    '판정': 'O', '예시': 'SK매직/코웨이/청호나이스/롯데렌터카/AJ렌터카 등'},
        {'순서': 4, '기준': '차량유지비 + 적요 렌트/렌탈','판정': 'O', '예시': '차량유지비 계정이면서 적요에 렌트/렌탈 포함'},
        {'순서': 5, '기준': '수수료 계정 (키워드 없음)',   '판정': 'X', '예시': '지급수수료 등 - 리스 키워드 미발견'},
        {'순서': 6, '기준': '위 어디에도 해당 없음',       '판정': '?', '예시': '추가 검토 필요'},
    ]
    _criteria_df = pd.DataFrame(_criteria_rows)
    _pf2 = {"O": PatternFill("solid", fgColor="C6EFCE"),
            "X": PatternFill("solid", fgColor="FFCCCC"),
            "?": PatternFill("solid", fgColor="FFEB9C")}
    with pd.ExcelWriter(out_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as _w2:
        _criteria_df.to_excel(_w2, index=False, sheet_name="판정기준")
        _wsc = _w2.sheets["판정기준"]
        _hf2 = PatternFill("solid", fgColor="D9E1F2")
        for _cell in _wsc[1]:
            _cell.font = Font(bold=True); _cell.fill = _hf2
            _cell.alignment = Alignment(horizontal="center", vertical="center")
        _jc2 = _criteria_df.columns.get_loc("판정") + 1
        for _row in _wsc.iter_rows(min_row=2, max_row=_wsc.max_row):
            _v = str(_row[_jc2-1].value or "").strip()
            if _v in _pf2:
                for _c in _row: _c.fill = _pf2[_v]
        _cw2 = {"순서": 8, "기준": 30, "판정": 8, "예시": 65}
        for _ci2, _cn2 in enumerate(_criteria_df.columns, 1):
            _wsc.column_dimensions[get_column_letter(_ci2)].width = _cw2.get(_cn2, 15)
        _nr = _wsc.max_row + 2
        _wsc.cell(_nr, 1).value = "※ 참고"
        _wsc.cell(_nr, 1).font = Font(bold=True)
        _wsc.cell(_nr, 2).value = "판정은 거래처+계정과목+적요 패턴 집계 후 대표 적요(상위 2건) 기준으로 자동 판정합니다."
    print(f'  [리스완전성] 별도 저장: {os.path.relpath(out_path)}')
    return out_path


def save_results(results: dict, output_dir: str, company_name: str,
                 settings: dict = None, out_path: str = None) -> str:
    os.makedirs(output_dir, exist_ok=True)
    if out_path is None:
        out_path = os.path.join(output_dir, f'분석결과_{company_name}.xlsx')

    settings      = settings or {}
    client_name   = settings.get('ClientName', company_name)
    start_date    = settings.get('StartDate', '')
    end_date      = settings.get('EndDate', '')
    header_line1  = f'회사명: {client_name}'
    header_line2  = f'분석기간: {start_date} ~ {end_date}' if start_date else ''

    # 특수 키 사전 추출 (ExcelWriter에 넘기지 않음)
    benford_images = results.pop('_benford_images', None)
    decoder        = results.pop('_암호해독표', None)

    # startrow=2: 1~2행을 비워두고 3행부터 컬럼헤더+데이터 기록
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for sheet, df in results.items():
            sname = _safe_sheet(sheet)
            if df is None or (isinstance(df, pd.DataFrame) and df.empty):
                pd.DataFrame([['결과 없음']]).to_excel(
                    writer, sheet_name=sname, index=False, header=False, startrow=2)
            elif isinstance(df, pd.DataFrame):
                df.to_excel(writer, sheet_name=sname, index=False, startrow=2)

    # 회사 정보 상단 삽입 + 벤포드 차트 삽입 (openpyxl로 한 번에 처리)
    wb = openpyxl.load_workbook(out_path)
    from openpyxl.styles import Font
    bold = Font(bold=True)
    for ws in wb.worksheets:
        ws.cell(1, 1).value = header_line1
        ws.cell(1, 1).font  = bold
        if header_line2:
            ws.cell(2, 1).value = header_line2

    if benford_images:
        for _acct, _dir, img_buf in benford_images:
            if not img_buf: continue
            sname = _safe_sheet(f'벤포드_{_acct}_{_dir}')
            if sname in wb.sheetnames:
                wb[sname].add_image(XLImage(img_buf), 'K4')

    wb.save(out_path)

    # 암호해독표 → 별도 파일
    if decoder is not None:
        key_path = os.path.join(output_dir, f'암호해독표_{company_name}.xlsx')
        decoder.to_excel(key_path, index=False)
        print(f'  [암호해독표] {key_path}')

    return out_path


# =============================================================================
# 7. 메인
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='분개장 분석 자동화')
    parser.add_argument('company', nargs='?', help='고객사 이름 (예: sejoong)')
    parser.add_argument('--task', type=int, nargs='+', metavar='N', help='실행할 분석 번호 (예: --task 21  또는  --task 3 8 21)')
    args = parser.parse_args()
    company_name = (args.company or input('고객사 이름: ')).strip()
    if not company_name:
        print('[오류] 고객사 이름이 비어 있습니다.'); sys.exit(1)

    global _COMPANY_DIR
    print(f'\n{"="*60}\n  분개장 분석 자동화 — {company_name}\n{"="*60}')
    paths = resolve_paths(company_name)
    _COMPANY_DIR = paths['company_dir']

    # 1) 태스크 리스트
    try:
        active_tasks = load_active_tasks(paths['task_list'])
    except (FileNotFoundError, ValueError) as e:
        print(f'[오류] {e}'); sys.exit(1)
    if args.task:
        active_tasks = [(n, nm, p, m) for n, nm, p, m in active_tasks if n in args.task]
        print(f'  [필터] --task {args.task} → {len(active_tasks)}개 실행')
    if not active_tasks:
        print('실행할 분석이 없습니다 (Y/O 항목 없음).'); sys.exit(0)

    # 2) 분개장 로드
    print('\n[분개장 로드]')
    df = load_data(paths['company_dir'])
    if df is None:
        print('[오류] data/current 또는 data/previous 폴더에 분개장 파일이 없습니다.')
        sys.exit(1)
    df = _apply_company_preprocess(df, company_name)   # 회사별 전용 전처리
    df = _preprocess_df(df)                            # 공통 표준 전처리
    gc = _get_gubun_col(df)
    print(f'  총 {len(df):,}행'
          + (f' (당기: {(df[gc]=="당기").sum():,}건 / 전기: {(df[gc]=="전기").sum():,}건)' if gc else '')
          + f'\n  컬럼: {df.columns.tolist()}')

    # 3) 분석 순차 실행
    print('\n[분석 실행]')
    all_results: dict = {}
    for task_no, task_name, 분석대상, end_month in active_tasks:
        if task_no not in ANALYSIS_REGISTRY:
            print(f'  [{task_no:>3}] {task_name:<22} → 등록된 함수 없음 (건너뜀)')
            continue
        _, func = ANALYSIS_REGISTRY[task_no]
        params_list = load_analysis_params(paths['task_list'], task_name)
        # 분석대상(당기/전기/전체)에 따라 df 슬라이싱
        if '구분' in df.columns and 분석대상 in ('당기', '전기'):
            task_df = df[df['구분'] == 분석대상].copy()
        else:
            task_df = df
        # 기준월 필터: task_list 분석목록 시트 '기준월' 열에 숫자(1~12) 기재 시 해당 월까지만 사용
        if end_month and COL_DATE in task_df.columns:
            task_df = task_df[task_df[COL_DATE].dt.month <= end_month].copy()
        period_label = 분석대상 + (f' ~{end_month}월' if end_month else '')
        print(f'  [{task_no:>3}] {task_name} [{period_label} {len(task_df):,}행]', flush=True)
        try:
            result = func(task_df, params_list)
            if task_no in _SEPARATE_FILE_TASKS:
                _save_lease_completeness_file(result, paths['output'], company_name)
                print(f'       → 별도 파일 저장')
            elif isinstance(result, dict):
                for sname, sub_df in result.items():
                    all_results[sname] = sub_df
                print(f'       → 시트 {len(result)}개 생성')
            elif isinstance(result, pd.DataFrame):
                all_results[_safe_sheet(task_name)] = result
                print(f'       → 시트 1개 생성')
        except Exception as e:
            import traceback
            print(f'       ⚠️ 오류: {e}')
            traceback.print_exc()

    # 4) 결과 저장
    print('\n[결과 저장]')
    settings = load_settings(paths['task_list'])

    # --task 지정 시 별도 파일로 저장 (덮어쓰기 방지)
    partial_out_path = None
    if args.task and active_tasks:
        import datetime as _dt
        _date_str = _dt.datetime.now().strftime('%Y%m%d')
        _non_sep = [(n, nm) for n, nm, p, m in active_tasks if n not in _SEPARATE_FILE_TASKS]
        if _non_sep:
            _parts = []
            for _n, _nm in _non_sep:
                _safe_nm = re.sub(r'[\\/*?:\[\]<>|]', '', _nm)[:12]
                _parts.append(f'{_n}_{_safe_nm}')
            _fname = '_'.join(_parts) + '_' + _date_str + '.xlsx'
            partial_out_path = os.path.join(paths['output'], _fname)

    if all_results:
        out_path = save_results(all_results, paths['output'], company_name, settings,
                                out_path=partial_out_path)
        print(f'\n  ✅ 완료: {out_path}')
        print(f'  시트 수: {len(all_results)}개')
    else:
        print('\n  ℹ️ 저장할 결과 없음 (별도 파일 태스크만 실행)')


if __name__ == '__main__':
    main()
