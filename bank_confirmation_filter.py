#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
bank_confirmation_filter.py
은행조회서완전성 결과 파일(JS 앱 다운로드)에 금융기관명 컬럼 및 조회서목록 요약 시트를 추가.

실행:
  python bank_confirmation_filter.py --file {path_to_xlsx}
  python bank_confirmation_filter.py --company dae_il
"""

import argparse
import glob
import os
import re
import sys

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))

# 금융기관명 추출 패턴 (저축은행이 은행보다 먼저 매칭되도록 순서 고정)
_FI_PATTERN = re.compile(
    r'([\w가-힣()（）]*'
    r'(?:저축은행|저축|은행|금고|신협|조합|증권|캐피탈|카드|보험|파이낸스|크레딧|리스))'
)


def extract_fi(client_name: str) -> str:
    """거래처명에서 금융기관명 추출. 예: '우리은행 강남지점' → '우리은행'"""
    name = str(client_name).strip()
    if not name or name.lower() in ('nan', 'none', ''):
        return ''
    m = _FI_PATTERN.search(name)
    return m.group(1) if m else ''


def _find_col(df: pd.DataFrame, *keywords) -> str:
    """컬럼명에 키워드가 포함된 첫 번째 컬럼 반환. 없으면 None."""
    for kw in keywords:
        for c in df.columns:
            if kw in str(c):
                return c
    return None


def _pivot_mark(pivot: pd.DataFrame) -> pd.DataFrame:
    """pivot 값 0 → '-', 양수 → '○' 변환 (pandas 버전 호환)."""
    try:
        return pivot.map(lambda x: '○' if x > 0 else '-')
    except AttributeError:
        return pivot.applymap(lambda x: '○' if x > 0 else '-')


def process_file(target: str) -> None:
    print(f"\n[은행조회서완전성] 처리 시작: {os.path.basename(target)}")

    xl = pd.ExcelFile(target)
    sheet_names = [s for s in xl.sheet_names if s != '금융기관_조회서목록']
    if not sheet_names:
        print("  [경고] 처리할 시트가 없습니다.")
        return
    print(f"  시트 목록: {sheet_names}")

    all_frames = {}   # sheet_name → df (금융기관명 컬럼 추가)
    all_combined = [] # 전체 데이터 (summary 생성용)

    for sheet in sheet_names:
        df = pd.read_excel(target, sheet_name=sheet)
        if df.empty:
            all_frames[sheet] = df
            continue

        col_client = _find_col(df, '거래처명', '거래처')

        # 금융기관명 추출 후 거래처명 바로 뒤에 삽입
        if col_client:
            df['금융기관명'] = df[col_client].apply(extract_fi)
            cols = list(df.columns)
            cols.remove('금융기관명')
            cols.insert(cols.index(col_client) + 1, '금융기관명')
            df = df[cols]
        else:
            df['금융기관명'] = ''

        fi_cnt = (df['금융기관명'].astype(str).str.strip() != '').sum()
        print(f"  시트 '{sheet}': {len(df):,}행, 금융기관명 인식 {fi_cnt}건")

        # 요약용 복사본 (시트명 = 조회계정)
        tmp = df.copy()
        tmp.insert(0, '조회계정', sheet)
        all_combined.append(tmp)
        all_frames[sheet] = df

    # ── 요약 시트 생성 ────────────────────────────────────────────────────────
    summary_df = None
    if all_combined:
        combined = pd.concat(all_combined, ignore_index=True)
        col_client = _find_col(combined, '거래처명', '거래처')
        col_debit  = _find_col(combined, '차변')
        col_credit = _find_col(combined, '대변')
        fi_col = '금융기관명'

        has_fi = combined[combined[fi_col].astype(str).str.strip() != ''].copy()

        if not has_fi.empty and col_client:
            # 피벗: 행=금융기관명, 열=조회계정(○/-)
            pivot = has_fi.groupby([fi_col, '조회계정']).size().unstack(fill_value=0)
            acct_cols = [a for a in sheet_names if a in pivot.columns]
            pivot = pivot.reindex(columns=acct_cols, fill_value=0)
            pivot_mark = _pivot_mark(pivot)

            # 원본 거래처명 목록 (금융기관별)
            raw_clients = (
                has_fi.groupby(fi_col)[col_client]
                .apply(lambda s: ', '.join(sorted(s.dropna().astype(str).unique())))
                .rename('거래처명(원본)')
            )

            # 합계
            agg_kwargs = {'전표건수': ('조회계정', 'count')}
            if col_debit:  agg_kwargs['차변합계'] = (col_debit,  'sum')
            if col_credit: agg_kwargs['대변합계'] = (col_credit, 'sum')
            totals = has_fi.groupby(fi_col).agg(**agg_kwargs)

            summary_df = pivot_mark.join(raw_clients).join(totals).reset_index()
            summary_df.rename(columns={fi_col: '금융기관명'}, inplace=True)
            summary_df.insert(1, '조회서발송', 'Y')
            print(f"  금융기관_조회서목록: {len(summary_df)}개 기관 추출")
        else:
            print("  [안내] 거래처명에서 금융기관명을 인식하지 못했습니다.")

    # ── 파일 재작성: 기존 시트 + 요약 시트 ──────────────────────────────────
    with pd.ExcelWriter(target, engine='openpyxl', mode='w') as writer:
        for sheet_name, df in all_frames.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        if summary_df is not None:
            fi_count = len(summary_df)
            # 요약 시트: 2줄 안내 헤더 + 데이터
            ws = writer.book.create_sheet('금융기관_조회서목록')
            ws['A1'] = f'[금융기관 조회서 발송 목록]  총 {fi_count}개 금융기관'
            ws['A2'] = '* 조회서발송: 감사인이 Y/N 직접 표시  |  ○ = 해당 계정 거래 있음  |  - = 없음'
            for r_idx, row in enumerate(
                dataframe_to_rows(summary_df, index=False, header=True), start=3
            ):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

    print(f"  ✅ 완료: {os.path.basename(target)}")


def find_target_file(company: str, base: str = None) -> str:
    """회사 raw_data 폴더에서 은행조회서 결과 파일 자동 탐색."""
    if base:
        raw_dir = os.path.join(PROJECT_ROOT, base, company, 'raw_data')
    else:
        raw_dir = os.path.join(PROJECT_ROOT, company, 'raw_data')

    candidates = []
    for pat in ['*은행조회서*.xlsx', '*은행*조회*.xlsx']:
        candidates += glob.glob(os.path.join(raw_dir, pat))
    candidates = sorted(set(
        f for f in candidates if not os.path.basename(f).startswith('~$')
    ))

    if not candidates:
        raise FileNotFoundError(
            f"[오류] 은행조회서 결과 파일을 찾을 수 없습니다.\n"
            f"  탐색 경로: {raw_dir}\n"
            f"  상세검색_시나리오를 먼저 실행해 주세요."
        )
    return candidates[-1]


def main():
    parser = argparse.ArgumentParser(
        description='은행조회서완전성 금융기관명 추출 + 조회서목록 시트 추가',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            '사용 예:\n'
            '  python bank_confirmation_filter.py --file path/to/은행조회서완전성.xlsx\n'
            '  python bank_confirmation_filter.py --company dae_il\n'
        )
    )
    parser.add_argument('--file', '-f', default=None,
                        help='처리할 xlsx 파일 경로 (JS 앱이 직접 전달)')
    parser.add_argument('--company', '-c', default=None,
                        help='회사 폴더명 (raw_data 에서 자동 탐색)')
    parser.add_argument('--base', default=None,
                        help='회사 폴더의 상위 폴더명 (선택)')
    args = parser.parse_args()

    if args.file:
        target = args.file
    elif args.company:
        try:
            target = find_target_file(args.company, args.base)
        except FileNotFoundError as e:
            print(e)
            sys.exit(1)
    else:
        parser.print_help()
        sys.exit(1)

    if not os.path.isfile(target):
        print(f"[오류] 파일 없음: {target}")
        sys.exit(1)

    process_file(target)


if __name__ == '__main__':
    main()
