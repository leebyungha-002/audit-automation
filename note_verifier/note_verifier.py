#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
note_verifier.py -- 감사주석 검증 도구 (PyQt6 GUI)

[동작]
  감사조서의 주석 시트(숫자 이름)에서
    블록1(왼쪽) : 정산표/DSD 값 자동 채움
    블록2(오른쪽): 감사인이 작성한 주석 값
  두 블록을 비교하여 차이를 색상 표시 + 검증요약 시트 생성.

[색상 기준]
  연녹  : 일치 (차이 1천원 이하)
  연황  : 소차이 (1천원 초과 ~ 1백만원 이하)
  연적  : 큰차이 (1백만원 초과)
  회색  : 소스에 해당 항목 없음
"""

import os
import re
import sys
from datetime import datetime

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import PatternFill, Font
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QFileDialog, QTextEdit, QRadioButton,
    QButtonGroup, QGroupBox, QFrame,
)
from PyQt6.QtCore import QThread, pyqtSignal
from PyQt6.QtGui import QFont, QColor, QTextCursor

# ── 색상 상수 ────────────────────────────────────────────────────────
FILL_OK   = PatternFill("solid", fgColor="C6EFCE")
FILL_DIFF = PatternFill("solid", fgColor="FFEB9C")
FILL_BIG  = PatternFill("solid", fgColor="FFC7CE")
FILL_MISS = PatternFill("solid", fgColor="D9D9D9")
FILL_HDR  = PatternFill("solid", fgColor="BDD7EE")
FONT_BOLD = Font(bold=True)

THRESH_SMALL = 1          # 1천원 이하 → 일치
THRESH_BIG   = 1_000      # 1백만원(천원 단위) 초과 → 큰 차이


# ════════════════════════════════════════════════════════════════════
# 핵심 로직
# ════════════════════════════════════════════════════════════════════

def _norm(s) -> str:
    """공백·괄호·점 제거 + 소문자."""
    return re.sub(r'[\s\(\)\[\]\.,·\-]', '', str(s or '')).lower()

def _norm_sheet(name: str) -> str:
    """'주석5', '5번', '5' → '5'."""
    s = re.sub(r'주석|note|no\.?|번|\s', '', name, flags=re.IGNORECASE)
    return re.sub(r'[^\d]', '', s)

def _is_num(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)

def _note_sheets(wb) -> list:
    """'주석>>>>' 뒤의 시트들, 없으면 숫자로 시작하는 시트들."""
    names = wb.sheetnames
    for i, s in enumerate(names):
        if re.search(r'주석.*>{2,}', s):
            return names[i + 1:]
    return [s for s in names if re.match(r'^\d', s.strip())]

def _hdr_row(ws, max_r=12) -> int:
    """텍스트 셀 3개 이상인 첫 행 (헤더행)."""
    for r in range(1, max_r + 1):
        cnt = sum(
            1 for c in range(1, min((ws.max_column or 1), 30) + 1)
            if isinstance(ws.cell(r, c).value, str)
            and ws.cell(r, c).value.strip()
            and ws.cell(r, c).value.strip() not in ('True', 'False')
        )
        if cnt >= 3:
            return r
    return 1

def _b1_end(ws, hdr: int) -> int:
    """블록1 마지막 열: 연속 빈 열 2개 직전."""
    rows = list(ws.iter_rows(
        min_row=hdr + 1,
        max_row=min(hdr + 8, ws.max_row or 1),
        values_only=True
    ))
    mc = ws.max_column or 1
    for c in range(2, mc):
        c0 = [r[c - 1] if c - 1 < len(r) else None for r in rows]
        c1 = [r[c]     if c     < len(r) else None for r in rows]
        if all(v is None for v in c0) and all(v is None for v in c1):
            return c - 1
    return mc

def _b2_start(ws, b1e: int, hdr: int) -> int:
    """블록2 시작 열: 블록1 끝 이후 첫 비어있지 않은 열."""
    for c in range(b1e + 1, (ws.max_column or 1) + 1):
        for r in range(hdr, min(hdr + 5, ws.max_row or 1) + 1):
            if ws.cell(r, c).value is not None:
                return c
    return -1

def _src_data(ws_src) -> dict:
    """{norm_label: [val_col2, val_col3, ...]} 소스 시트 데이터."""
    hdr = _hdr_row(ws_src)
    out = {}
    for r in range(hdr + 1, (ws_src.max_row or 0) + 1):
        lbl = ws_src.cell(r, 1).value
        if not lbl or not str(lbl).strip():
            continue
        vals = [
            ws_src.cell(r, c).value
            if _is_num(ws_src.cell(r, c).value) else None
            for c in range(2, (ws_src.max_column or 1) + 1)
        ]
        if any(v is not None for v in vals):
            out[_norm(lbl)] = vals
    return out

def _b2_val(ws, row: int, b2s: int) -> float | None:
    """블록2의 해당 행에서 첫 번째 숫자값."""
    for c in range(b2s, (ws.max_column or 1) + 1):
        v = ws.cell(row, c).value
        if _is_num(v):
            return v
    return None

def _b2_hdr_match(ws, row_b1_hdr: int, b1e: int, b2s: int) -> dict:
    """블록1 헤더와 블록2 헤더를 이름으로 매핑 {b1_col: b2_col}."""
    b1_hdrs = {}
    for c in range(2, b1e + 1):
        v = ws.cell(row_b1_hdr, c).value
        if isinstance(v, str) and v.strip():
            b1_hdrs[c] = _norm(v)

    b2_hdrs = {}
    for c in range(b2s, (ws.max_column or 1) + 1):
        v = ws.cell(row_b1_hdr, c).value
        if isinstance(v, str) and v.strip():
            b2_hdrs[c] = _norm(v)

    mapping = {}
    for b1c, b1h in b1_hdrs.items():
        for b2c, b2h in b2_hdrs.items():
            if b1h and b1h == b2h:
                mapping[b1c] = b2c
                break
        if b1c not in mapping:
            # 부분 일치 fallback
            for b2c, b2h in b2_hdrs.items():
                if b1h and (b1h in b2h or b2h in b1h):
                    mapping[b1c] = b2c
                    break
    return mapping


def run_verify(audit_path: str, src_path: str,
               src_label: str, unit_scale: float, log) -> str:
    """
    메인 검증 함수.
    unit_scale: 1.0(천원 그대로) or 0.001(원→천원 변환)
    반환: 저장된 파일 경로
    """
    log(f"감사조서  : {os.path.basename(audit_path)}")
    log(f"소스({src_label}): {os.path.basename(src_path)}")
    if unit_scale != 1.0:
        log("단위 변환 : ÷1,000 (원 → 천원)")

    wb_src   = openpyxl.load_workbook(src_path,   read_only=True, data_only=True)
    wb_audit = openpyxl.load_workbook(audit_path, data_only=True)

    src_map    = {_norm_sheet(s): s for s in wb_src.sheetnames}
    note_list  = _note_sheets(wb_audit)
    log(f"주석 시트 : {len(note_list)}개\n{'─'*48}")

    summary = []

    for sname in note_list:
        ws  = wb_audit[sname]
        key = _norm_sheet(sname)

        src_sname = src_map.get(key)
        if not src_sname:
            log(f"  [{sname:>4}] 소스 시트 없음 — 건너뜀")
            continue

        hdr  = _hdr_row(ws)
        b1e  = _b1_end(ws, hdr)
        b2s  = _b2_start(ws, b1e, hdr)

        if b2s < 0:
            log(f"  [{sname:>4}] 블록2 미발견 — 건너뜀")
            continue

        col_map  = _b2_hdr_match(ws, hdr, b1e, b2s)
        rows_src = _src_data(wb_src[src_sname])
        filled = ok = diff_s = diff_b = miss = 0

        for r in range(hdr + 1, (ws.max_row or 0) + 1):
            lbl_v = ws.cell(r, 1).value
            if not lbl_v or not str(lbl_v).strip():
                continue
            norm_lbl = _norm(lbl_v)
            src_row  = rows_src.get(norm_lbl)

            for ci, b1c in enumerate(range(2, b1e + 1)):
                src_v  = src_row[ci - 1] if src_row and ci - 1 < len(src_row) else None
                cell   = ws.cell(r, b1c)
                b2c    = col_map.get(b1c)
                b2_raw = ws.cell(r, b2c).value if b2c else None

                # 블록2 단위 자동 감지: 값이 5백만 초과이면 원 단위로 판단
                if _is_num(b2_raw) and b2_raw > 5_000_000:
                    audit_v = round(b2_raw / 1000)
                elif _is_num(b2_raw):
                    audit_v = b2_raw
                else:
                    audit_v = None

                if src_v is not None:
                    fill_v = round(src_v * unit_scale)
                    cell.value = fill_v
                    filled += 1

                    if audit_v is not None:
                        diff = abs(fill_v - audit_v)
                        if diff <= THRESH_SMALL:
                            cell.fill = FILL_OK;   ok     += 1
                        elif diff <= THRESH_BIG:
                            cell.fill = FILL_DIFF; diff_s += 1
                        else:
                            cell.fill = FILL_BIG;  diff_b += 1
                            summary.append((
                                sname, str(lbl_v).strip(),
                                fill_v, audit_v, fill_v - audit_v
                            ))
                    else:
                        cell.fill = FILL_OK
                else:
                    if src_row is not None:
                        cell.fill = FILL_MISS
                        miss += 1

        log(f"  [{sname:>4}] 채움 {filled:3}건  ✓{ok}  소차이 {diff_s}  큰차이 {diff_b}  미매칭 {miss}")

    wb_src.close()

    # ── 검증요약 시트 ────────────────────────────────────────────────
    SUM = '검증요약'
    if SUM in wb_audit.sheetnames:
        del wb_audit[SUM]
    ws_s = wb_audit.create_sheet(SUM, 0)
    for cell in ws_s.append(['주석번호', '항목', f'{src_label}값(천원)',
                              '감사조서값(천원)', '차이(천원)']) or ws_s[1]:
        cell.font = FONT_BOLD
        cell.fill = FILL_HDR
    for row in summary:
        ws_s.append(list(row))

    log(f"\n{'─'*48}")
    log(f"큰 차이 항목 합계: {len(summary)}건")

    # ── 저장 ────────────────────────────────────────────────────────
    stem, ext = os.path.splitext(audit_path)
    out = f"{stem}_검증_{datetime.now().strftime('%m%d_%H%M')}{ext}"
    wb_audit.save(out)
    log(f"저장: {os.path.basename(out)}")
    return out


# ════════════════════════════════════════════════════════════════════
# 백그라운드 스레드
# ════════════════════════════════════════════════════════════════════

class Worker(QThread):
    log_sig  = pyqtSignal(str)
    done_sig = pyqtSignal(str)
    err_sig  = pyqtSignal(str)

    def __init__(self, audit, src, label, scale):
        super().__init__()
        self.audit = audit; self.src = src
        self.label = label; self.scale = scale

    def run(self):
        try:
            out = run_verify(self.audit, self.src, self.label,
                             self.scale, lambda m: self.log_sig.emit(m))
            self.done_sig.emit(out)
        except Exception as e:
            import traceback
            self.err_sig.emit(traceback.format_exc())


# ════════════════════════════════════════════════════════════════════
# GUI
# ════════════════════════════════════════════════════════════════════

class NoteVerifier(QMainWindow):
    def __init__(self, default_dir: str = None):
        super().__init__()
        self.setWindowTitle("감사주석 검증 도구")
        self.resize(740, 580)
        self._dir    = default_dir or os.path.expanduser("~")
        self._worker = None
        self._build_ui()

    # ── UI 구성 ──────────────────────────────────────────────────────
    def _build_ui(self):
        cw = QWidget(); self.setCentralWidget(cw)
        vb = QVBoxLayout(cw)
        vb.setContentsMargins(14, 14, 14, 14); vb.setSpacing(10)

        ttl = QLabel("감사주석 검증 도구")
        ttl.setFont(QFont("맑은 고딕", 13, QFont.Weight.Bold))
        vb.addWidget(ttl)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        sep.setFrameShadow(QFrame.Shadow.Sunken); vb.addWidget(sep)

        # 파일 설정 그룹
        grp = QGroupBox("파일 설정")
        gv  = QVBoxLayout(grp); gv.setSpacing(8)

        self._audit_lbl = self._file_row(gv, "감사조서 파일",
                                          lambda: self._pick("감사조서 선택", self._audit_lbl))
        self._src_lbl   = self._file_row(gv, "소스 파일 (정산표/DSD)",
                                          lambda: self._pick("소스 파일 선택", self._src_lbl))

        # 소스 유형 라디오
        h1 = QHBoxLayout()
        h1.addWidget(QLabel("소스 유형:"))
        self._r_tb  = QRadioButton("정산표"); self._r_tb.setChecked(True)
        self._r_dsd = QRadioButton("DSD")
        bg1 = QButtonGroup(self); bg1.addButton(self._r_tb); bg1.addButton(self._r_dsd)
        h1.addWidget(self._r_tb); h1.addWidget(self._r_dsd); h1.addStretch()
        gv.addLayout(h1)

        # 소스 단위 라디오
        h2 = QHBoxLayout()
        h2.addWidget(QLabel("소스 단위:"))
        self._r_1000 = QRadioButton("천원  (그대로 사용)"); self._r_1000.setChecked(True)
        self._r_won  = QRadioButton("원  (÷1,000 자동 변환)")
        bg2 = QButtonGroup(self); bg2.addButton(self._r_1000); bg2.addButton(self._r_won)
        h2.addWidget(self._r_1000); h2.addWidget(self._r_won); h2.addStretch()
        gv.addLayout(h2)

        vb.addWidget(grp)

        # 실행 버튼
        self._btn_run = QPushButton("▶  검증 시작")
        self._btn_run.setFont(QFont("맑은 고딕", 11, QFont.Weight.Bold))
        self._btn_run.setMinimumHeight(40)
        self._btn_run.setStyleSheet(
            "QPushButton{background:#2563EB;color:white;border-radius:6px;}"
            "QPushButton:hover{background:#1D4ED8;}"
            "QPushButton:disabled{background:#9CA3AF;}")
        self._btn_run.clicked.connect(self._run)
        vb.addWidget(self._btn_run)

        # 로그
        grp2 = QGroupBox("실행 로그")
        gv2  = QVBoxLayout(grp2)
        self._log = QTextEdit()
        self._log.setReadOnly(True)
        self._log.setFont(QFont("Consolas", 9))
        self._log.setStyleSheet("background:#1E1E1E;color:#D4D4D4;")
        gv2.addWidget(self._log)
        vb.addWidget(grp2, stretch=1)

    def _file_row(self, parent, label_txt: str, slot) -> QLabel:
        h    = QHBoxLayout()
        lbl  = QLabel(label_txt + ":"); lbl.setFixedWidth(200)
        path = QLabel("(선택 안 됨)")
        path.setStyleSheet(
            "color:#9CA3AF;border:1px solid #E5E7EB;padding:3px 6px;border-radius:4px;")
        btn = QPushButton("파일 선택"); btn.setFixedWidth(90)
        btn.clicked.connect(slot)
        h.addWidget(lbl); h.addWidget(path, 1); h.addWidget(btn)
        parent.addLayout(h)
        return path

    # ── 슬롯 ────────────────────────────────────────────────────────
    def _pick(self, title: str, lbl: QLabel):
        path, _ = QFileDialog.getOpenFileName(
            self, title, self._dir, "Excel 파일 (*.xlsx *.xls)")
        if path:
            lbl.setText(path)
            lbl.setStyleSheet(
                "color:#111827;border:1px solid #D1D5DB;padding:3px 6px;border-radius:4px;")
            self._dir = os.path.dirname(path)

    def _log_ln(self, text: str, color="#D4D4D4"):
        self._log.moveCursor(QTextCursor.MoveOperation.End)
        self._log.setTextColor(QColor(color))
        self._log.insertPlainText(text + "\n")
        self._log.moveCursor(QTextCursor.MoveOperation.End)

    def _run(self):
        audit = self._audit_lbl.text()
        src   = self._src_lbl.text()
        if not os.path.isfile(audit):
            self._log_ln("감사조서 파일을 선택하세요.", "#FBBF24"); return
        if not os.path.isfile(src):
            self._log_ln("소스 파일을 선택하세요.", "#FBBF24"); return

        label = "정산표" if self._r_tb.isChecked() else "DSD"
        scale = 1.0 if self._r_1000.isChecked() else 0.001

        self._log.clear()
        self._log_ln(f"▶ {datetime.now().strftime('%H:%M:%S')} 검증 시작", "#60A5FA")
        self._btn_run.setEnabled(False)

        self._worker = Worker(audit, src, label, scale)
        self._worker.log_sig.connect(lambda m: self._log_ln(m))
        self._worker.done_sig.connect(self._done)
        self._worker.err_sig.connect(self._err)
        self._worker.start()

    def _done(self, out: str):
        self._log_ln(f"\n완료: {out}", "#4ADE80")
        self._btn_run.setEnabled(True)

    def _err(self, msg: str):
        self._log_ln(f"\n오류:\n{msg}", "#F87171")
        self._btn_run.setEnabled(True)


# ════════════════════════════════════════════════════════════════════
# 진입점
# ════════════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument('--company', default=None)
    parser.add_argument('--base',    default=None)
    args, _ = parser.parse_known_args()

    default_dir = None
    if args.company:
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        base = (os.path.join(root, args.base, args.company)
                if args.base else os.path.join(root, args.company))
        if os.path.isdir(base):
            default_dir = base

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = NoteVerifier(default_dir=default_dir)
    win.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
