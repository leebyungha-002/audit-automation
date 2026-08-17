"""유형자산·투자부동산 감가상각비 검증앱 — 엔진.

input_data/depreciation_<company>_information_fy<year>.xlsx 를 읽어
자산별 당기 감가상각비를 재계산하고, 계정과목별 소계가 있는
'고정자산명세서'를 output/depreciation_schedule_<company>_<fy>.xlsx 로 생성한다.

핵심 설계: 취득원가·감가상각누계액·손상차손누계액·정부보조금잔액 전부
"기초잔액은 회사가 보고한 값을 그대로 신뢰하고, 당기 증감만 재계산"하는
동일한 방식을 따른다(다년간 이력을 시뮬레이션하지 않음). 그래서:
  - 기초취득원가/기초 감가상각누계액/기초 손상차손누계액/기초 정부보조금잔액 = 입력값 그대로
  - 당기취득원가(신규취득)/당기처분원가/당기 손상차손인식액/당기 정부보조금수령액 = 입력값(이벤트)
  - 당기감가상각비/당기 정부보조금환입액 = 앱이 계산
  - 기말 각 잔액 = 기초 + 당기증가 - 당기감소 (계산)

당기 회사계상 감가상각비/보조금환입액을 입력해 두면 앱의 재계산 결과와 자동 대사해
차이를 표시한다(유의차이는 노란색 강조).

실행 예:
    python depreciation_schedule.py kyungnam --fiscal-month 6
    python depreciation_schedule.py --file depreciation_kyungnam_information_fy2026.xlsx --fiscal-month 6
"""
import argparse
import glob
import os

import pandas as pd
import openpyxl
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(HERE, "input_data")
OUTPUT_DIR = os.path.join(HERE, "output")

SIG_THRESHOLD_ABS = 1000      # 유의차이 절대금액 기준(원)
SIG_THRESHOLD_PCT = 0.01      # 유의차이 비율 기준(1%)


# ── 공용 헬퍼 ────────────────────────────────────────────────────────────────

def _safe_float(v, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        if pd.isna(v):
            return default
    except (TypeError, ValueError):
        pass
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _safe_date(v):
    if v is None or v == "":
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    ts = pd.Timestamp(v)
    return ts.date() if not pd.isna(ts) else None


def _fiscal_year(ym: str, fiscal_month: int) -> str:
    """결산월 기준 회계연도 문자열 반환 (lease_schedule.py와 동일 규칙).
    결산월=12 → 캘린더 연도 그대로.
    결산월=6  → 7월 이후는 다음 연도로 귀속 (예: 2023-07 → '2024')."""
    if fiscal_month == 12:
        return ym[:4]
    year, month = int(ym[:4]), int(ym[5:7])
    return str(year + 1) if month > fiscal_month else str(year)


def _fy_bounds(target_fy: str, fiscal_month: int) -> tuple:
    """대상 회계연도(target_fy)의 시작월/종료월(YYYY-MM 문자열) 반환."""
    fy = int(target_fy)
    if fiscal_month == 12:
        return f"{fy}-01", f"{fy}-12"
    return f"{fy - 1}-{fiscal_month + 1:02d}", f"{fy}-{fiscal_month:02d}"


def _apply_interim(fy_end: str, target_fy: str, interim_month: int = None) -> str:
    """반기 등 중간결산 검토: 회계연도 종료월(fy_end)을 interim_month까지로 앞당긴다.
    (기초잔액은 회계연도 시작 기준 그대로, 계산 종료월만 축소 — 정상 처분/손상 시점 판정 로직과 동일 방식)"""
    if not interim_month:
        return fy_end
    interim_ym = f"{target_fy}-{interim_month:02d}"
    return min(fy_end, interim_ym)


def _months_between(start_ym: str, end_ym: str) -> int:
    """두 YYYY-MM 사이의 개월 수(양끝 포함)."""
    sy, sm = int(start_ym[:4]), int(start_ym[5:7])
    ey, em = int(end_ym[:4]), int(end_ym[5:7])
    return (ey - sy) * 12 + (em - sm) + 1


def _add_month(ym: str, n: int = 1) -> str:
    y, m = int(ym[:4]), int(ym[5:7])
    m += n
    y += (m - 1) // 12
    m = (m - 1) % 12 + 1
    return f"{y}-{m:02d}"


# ── 입력 로딩 ────────────────────────────────────────────────────────────────

def _find_input_file(company: str = None, file: str = None) -> str:
    if file:
        path = file if os.path.isabs(file) else os.path.join(INPUT_DIR, file)
        if not os.path.exists(path):
            raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {path}")
        return path

    if company:
        pattern = os.path.join(INPUT_DIR, f"depreciation_{company}_information_fy*.xlsx")
    else:
        pattern = os.path.join(INPUT_DIR, "depreciation_*_information_fy*.xlsx")

    matches = [p for p in glob.glob(pattern) if "template" not in os.path.basename(p)]
    if not matches:
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {pattern}")
    if len(matches) > 1 and not company:
        raise ValueError(f"회사를 특정해주세요. 후보 파일 여러 개: {matches}")
    return matches[0]


_SUBTOTAL_EXACT = ("소계", "합계", "총계", "누계", "계", "합 계", "소 계", "총 계")
_SUBTOTAL_SUFFIX = ("소계", "합계", "총계", "누계")  # '계' 단독은 접미어 검사에서 제외 (예: '온도계' 등 실제 자산명 오탐 방지)
_SUBTOTAL_CHECK_COLS = ("자산관리번호", "계정과목", "자산명(세부내역)")


def _looks_like_subtotal_row(rec: dict) -> bool:
    """원장에서 계정별 소계/합계 행이 그대로 복사돼 들어온 경우를 감지해서 걸러낸다.
    (예: 계정과목/자산명 칸에 '기계장치 소계', '합계' 등이 들어있고 취득일 없이 금액만 있는 행)"""
    for col in _SUBTOTAL_CHECK_COLS:
        v = rec.get(col)
        if v is None:
            continue
        text = str(v).strip()
        if not text:
            continue
        if text in _SUBTOTAL_EXACT or text.endswith(_SUBTOTAL_SUFFIX):
            return True
    return False


def load_assets(path: str) -> list:
    """'자산정보' 시트(1~2행 헤더, 3행부터 데이터)를 읽어 dict 목록으로 반환.
    계정별 소계/합계 행이 원장에서 그대로 복사돼 들어온 경우 자동으로 제외한다."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["자산정보"] if "자산정보" in wb.sheetnames else wb.worksheets[0]
    headers = [c.value for c in ws[2]]
    assets = []
    skipped_subtotals = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        rec = dict(zip(headers, row))
        if not rec.get("자산명(세부내역)") and not rec.get("자산관리번호"):
            continue
        if _looks_like_subtotal_row(rec):
            skipped_subtotals += 1
            continue
        assets.append(rec)
    if skipped_subtotals:
        print(f"[안내] 소계/합계로 보이는 행 {skipped_subtotals}건을 자산 목록에서 제외했습니다.")
    return assets


# ── 당기 감가상각/손상차손/정부보조금 계산 ───────────────────────────────────

def compute_asset(a: dict, fiscal_month: int, target_fy: str, interim_month: int = None) -> dict:
    """자산 1건의 당기 활동을 계산한다.
    기초잔액(감가상각누계액/손상차손누계액/정부보조금잔액)은 입력값을 그대로 신뢰하고,
    당기분(상각비/손상차손/보조금환입)만 계산한다. 다년간 이력 시뮬레이션은 하지 않는다.
    interim_month 지정 시 회계연도 전체가 아니라 그 월까지만(반기 검토 등) 계산한다
    — 기초잔액 기준(회계연도 시작월)은 그대로 두고 계산 종료월만 앞당긴다."""
    fy_start, fy_end = _fy_bounds(target_fy, fiscal_month)
    fy_end = _apply_interim(fy_end, target_fy, interim_month)

    기초취득원가 = _safe_float(a.get("기초취득원가"))
    당기취득원가 = _safe_float(a.get("당기취득원가"))
    cost = 기초취득원가 + 당기취득원가
    잔존가치 = _safe_float(a.get("잔존가치"))
    n_months = round(_safe_float(a.get("내용연수(년)")) * 12)
    method = str(a.get("상각방법(정액법/정률법)") or "정액법").strip()
    rate = _safe_float(a.get("상각률(정률법전용)"))
    amort_opt = str(a.get("상각개시(당월/익월)") or "당월").strip()
    offset = 1 if amort_opt == "익월" else 0
    acquire = _safe_date(a.get("취득일"))
    dispose = _safe_date(a.get("처분일"))
    dispose_ym = dispose.strftime("%Y-%m") if dispose else None
    impair_date = _safe_date(a.get("손상차손 인식일"))
    impair_amt = _safe_float(a.get("손상차손 인식액"))
    impair_ym = impair_date.strftime("%Y-%m") if impair_date else None
    grant_new = _safe_float(a.get("당기 정부보조금수령액"))

    result = {
        "기초감가상각누계액": 0.0, "당기감가상각비": 0.0, "처분시감소_누계액": 0.0, "기말감가상각누계액": 0.0,
        "기초손상차손누계액": 0.0, "당기손상차손인식액": 0.0, "처분시감소_손상": 0.0, "기말손상차손누계액": 0.0,
        "기초정부보조금잔액": 0.0, "당기정부보조금수령액": 0.0, "당기정부보조금환입액": 0.0,
        "처분시감소_보조금": 0.0, "기말정부보조금잔액": 0.0,
        "warning": None,
    }

    if cost <= 0 or n_months <= 0:
        result["warning"] = "취득원가(기초+당기)/내용연수 중 필수값 누락 — 상각 계산 불가"
        return result

    if method == "정률법" and rate <= 0:
        method = "정액법"
        result["warning"] = "상각률 미입력 → 정액법으로 대체 계산"

    is_existing = 기초취득원가 > 0
    disposed_in_fy = dispose_ym is not None and fy_start <= dispose_ym <= fy_end
    fixed_monthly_rate = None  # 내용연수 초과 자산: 잔여내용연수 재계산 대신 취득원가/내용연수 고정월상각률 사용

    if is_existing:
        기초누계 = _safe_float(a.get("기초 감가상각누계액"))
        기초손상 = _safe_float(a.get("기초 손상차손누계액"))
        기초보조금 = _safe_float(a.get("기초 정부보조금잔액"))
        # 감가상각은 정부보조금과 무관하게 자산 자체의 장부가(취득원가-감가상각누계액-손상차손누계액) 기준으로 계산한다.
        # 정부보조금은 별도 차감계정으로, 상각비와 같은 비율로 환입만 될 뿐 상각 계산 자체의 기준에는 포함되지 않는다.
        base_book_before_grant = cost - 기초누계 - 기초손상
        held_start = fy_start
        if acquire:
            eff_start_ym = (acquire + relativedelta(months=offset)).strftime("%Y-%m")
            elapsed = max(0, _months_between(eff_start_ym, fy_start) - 1)
            remaining_total_months = max(1, n_months - elapsed)
            if elapsed >= n_months:
                # 경과월수가 이미 내용연수를 초과했지만 장부가액이 남은 경우
                # (잔여내용연수를 1개월로 압축해 잔여장부가액을 한번에 상각해버리는 것을 방지)
                # 취득원가/내용연수 고정월상각률로 잔가에 도달할 때까지 계속 상각한다.
                fixed_monthly_rate = (cost - 잔존가치) / n_months if n_months > 0 else 0.0
                over_warning = (
                    f"⚠ 내용연수초과(경과 {elapsed}개월 > 내용연수 {n_months}개월) "
                    f"— 취득원가/내용연수 고정월상각률로 계속 상각"
                )
                result["warning"] = f"{result['warning']} / {over_warning}".strip(" /") if result["warning"] else over_warning
            else:
                # 경과월수 기준 정액법 예상 누계액과 입력된 기초 감가상각누계액을 비교해
                # 과거 상각 이력이 내용연수와 어긋나 있는지(오상각·내용연수 변경 등) 점검한다.
                expected_accum = elapsed * (cost - 잔존가치) / n_months if n_months > 0 else 0.0
                accum_diff = 기초누계 - expected_accum
                if _is_significant(accum_diff, expected_accum):
                    base_warning = "⚠ 기초잔액 확인필요(경과월수 기준 예상 감가상각누계액과 차이)"
                    result["warning"] = f"{result['warning']} / {base_warning}".strip(" /") if result["warning"] else base_warning
        else:
            remaining_total_months = n_months  # 취득일 미입력 시 근사치(정확한 잔여내용연수 계산 불가)
        grant_ratio = (기초보조금 / base_book_before_grant) if base_book_before_grant > 0 else 0.0
        보조금_당기증가 = 0.0
    else:
        # 당기 신규취득
        if acquire is None:
            result["warning"] = "당기취득원가가 있는 자산은 취득일이 필요합니다"
            return result
        기초누계 = 0.0
        # 건설중인자산 등에서 대체된 당기취득 자산은 전기 이전에 이미 인식된 손상차손누계액을
        # 그대로 넘겨받는 경우가 있으므로(기초취득원가=0이어도) 입력값을 그대로 반영한다.
        기초손상 = _safe_float(a.get("기초 손상차손누계액"))
        기초보조금 = 0.0
        base_book_before_grant = cost - 기초손상
        held_start = (acquire + relativedelta(months=offset)).strftime("%Y-%m")
        if held_start < fy_start:
            held_start = fy_start
        remaining_total_months = n_months
        grant_ratio = (grant_new / 당기취득원가) if 당기취득원가 > 0 else 0.0
        보조금_당기증가 = grant_new

    held_end = dispose_ym if disposed_in_fy else fy_end
    if held_end < held_start:
        result.update({
            "기초감가상각누계액": 기초누계, "기말감가상각누계액": 기초누계,
            "기초손상차손누계액": 기초손상, "기말손상차손누계액": 기초손상,
            "기초정부보조금잔액": 기초보조금, "기말정부보조금잔액": 기초보조금 + 보조금_당기증가,
            "당기정부보조금수령액": 보조금_당기증가,
        })
        return result

    impair_in_period = (
        impair_ym is not None and impair_amt > 0 and held_start <= impair_ym <= held_end
    )

    당기감가상각비 = 0.0
    당기손상 = 0.0

    if method == "정률법":
        book = base_book_before_grant
        m = held_start
        while m <= held_end:
            dep = book * (rate / 12)
            dep = max(0.0, min(dep, book - 잔존가치))
            book -= dep
            당기감가상각비 += dep
            if impair_in_period and m == impair_ym:
                imp = min(impair_amt, book)
                book = max(0.0, book - imp)
                당기손상 += imp
            m = _add_month(m, 1)
    else:  # 정액법
        if fixed_monthly_rate is not None:
            monthly_rate = fixed_monthly_rate
        else:
            monthly_rate = (base_book_before_grant - 잔존가치) / remaining_total_months if remaining_total_months > 0 else 0.0
        if impair_in_period:
            n1 = _months_between(held_start, impair_ym)
            n2 = max(0, _months_between(impair_ym, held_end) - 1)
            dep1 = max(0.0, min(monthly_rate * n1, base_book_before_grant - 잔존가치))
            book_at_impair = base_book_before_grant - dep1
            imp = min(impair_amt, book_at_impair)
            book_after = max(0.0, book_at_impair - imp)
            months_left_in_life = max(1, remaining_total_months - n1)
            monthly_rate2 = (book_after - 잔존가치) / months_left_in_life if months_left_in_life > 0 else 0.0
            dep2 = max(0.0, min(monthly_rate2 * n2, book_after - 잔존가치)) if n2 > 0 else 0.0
            당기감가상각비 = dep1 + dep2
            당기손상 = imp
        else:
            n = _months_between(held_start, held_end)
            당기감가상각비 = max(0.0, min(monthly_rate * n, base_book_before_grant - 잔존가치))

    당기정부보조금환입액 = 당기감가상각비 * grant_ratio

    # 잔존가치를 0으로 입력한 자산이 당기 중 장부가액을 전액(잔존가치까지) 소진하는 경우,
    # 법인세법상 최소 잔존가액(1,000원) 관행과 어긋나므로 회사계상액 입력 여부와 무관하게 안내한다.
    if (
        not impair_in_period and not disposed_in_fy and 잔존가치 == 0
        and 당기감가상각비 > 0
        and abs(당기감가상각비 - (base_book_before_grant - 잔존가치)) < 1
    ):
        tax_residual_warning = "법인세법상 잔존가액(1,000원) 차이로 추정"
        result["warning"] = f"{result['warning']} / {tax_residual_warning}".strip(" /") if result["warning"] else tax_residual_warning

    처분시감소_누계액 = (기초누계 + 당기감가상각비) if disposed_in_fy else 0.0
    기말감가상각누계액 = 기초누계 + 당기감가상각비 - 처분시감소_누계액
    처분시감소_손상 = (기초손상 + 당기손상) if disposed_in_fy else 0.0
    기말손상차손누계액 = 기초손상 + 당기손상 - 처분시감소_손상
    기말정부보조금잔액_처분전 = 기초보조금 + 보조금_당기증가 - 당기정부보조금환입액
    처분시감소_보조금 = 기말정부보조금잔액_처분전 if disposed_in_fy else 0.0
    기말정부보조금잔액 = 기말정부보조금잔액_처분전 - 처분시감소_보조금

    result.update({
        "기초감가상각누계액": 기초누계, "당기감가상각비": 당기감가상각비,
        "처분시감소_누계액": 처분시감소_누계액, "기말감가상각누계액": 기말감가상각누계액,
        "기초손상차손누계액": 기초손상, "당기손상차손인식액": 당기손상,
        "처분시감소_손상": 처분시감소_손상, "기말손상차손누계액": 기말손상차손누계액,
        "기초정부보조금잔액": 기초보조금, "당기정부보조금수령액": 보조금_당기증가,
        "당기정부보조금환입액": 당기정부보조금환입액,
        "처분시감소_보조금": 처분시감소_보조금, "기말정부보조금잔액": 기말정부보조금잔액,
    })
    return result


# ── 명세서 구성 ──────────────────────────────────────────────────────────────

def build_schedule_table(assets: list, fiscal_month: int, target_fy: str, interim_month: int = None) -> pd.DataFrame:
    fy_start, fy_end = _fy_bounds(target_fy, fiscal_month)
    fy_end = _apply_interim(fy_end, target_fy, interim_month)

    rows = []
    for a in assets:
        r = compute_asset(a, fiscal_month, target_fy, interim_month)

        기초취득원가 = _safe_float(a.get("기초취득원가"))
        당기취득원가 = _safe_float(a.get("당기취득원가"))
        당기처분원가 = _safe_float(a.get("당기처분원가"))
        기말취득원가 = 기초취득원가 + 당기취득원가 - 당기처분원가

        순감가상각비 = r["당기감가상각비"] - r["당기정부보조금환입액"]
        기말장부가액 = 기말취득원가 - r["기말감가상각누계액"] - r["기말손상차손누계액"] - r["기말정부보조금잔액"]

        company_dep = a.get("당기 회사계상 감가상각비")
        company_dep_f = _safe_float(company_dep) if company_dep not in (None, "") else None
        company_grant_amort = a.get("당기 회사계상 보조금환입액")
        company_grant_amort_f = _safe_float(company_grant_amort) if company_grant_amort not in (None, "") else None

        dispose = _safe_date(a.get("처분일"))
        dispose_ym = dispose.strftime("%Y-%m") if dispose else None
        disposed_in_fy = dispose_ym is not None and fy_start <= dispose_ym <= fy_end

        당기상각비_차이 = None if company_dep_f is None else r["당기감가상각비"] - company_dep_f

        비고 = a.get("비고") or ""
        if r["warning"]:
            비고 = f"{비고} / {r['warning']}".strip(" /")
        if disposed_in_fy:
            비고 = f"{비고} / 당기 처분(취득원가·상각누계액 전액 제거)".strip(" /")

        rows.append({
            "사업장": a.get("사업장") or "",
            "자산분류": a.get("자산분류(유형자산/투자부동산/무형자산)") or "",
            "계정과목": a.get("계정과목") or "(미분류)",
            "자산관리번호": a.get("자산관리번호"),
            "자산명(세부내역)": a.get("자산명(세부내역)"),
            "취득일": a.get("취득일"),
            "내용연수(년)": a.get("내용연수(년)"),
            "기초취득원가": 기초취득원가,
            "당기증가(신규취득)": 당기취득원가,
            "당기감소(처분)": 당기처분원가,
            "기말취득원가": 기말취득원가,
            "기초감가상각누계액": r["기초감가상각누계액"],
            "당기감가상각비(계산)": r["당기감가상각비"],
            "처분시감소(누계액)": r["처분시감소_누계액"],
            "기말감가상각누계액(계산)": r["기말감가상각누계액"],
            "기초손상차손누계액": r["기초손상차손누계액"],
            "당기 손상차손인식액(계산)": r["당기손상차손인식액"],
            "기말 손상차손누계액(계산)": r["기말손상차손누계액"],
            "정부보조금 계정명": a.get("정부보조금 계정명") or "",
            "기초정부보조금잔액": r["기초정부보조금잔액"],
            "당기 정부보조금수령액": r["당기정부보조금수령액"],
            "당기 정부보조금환입액(계산)": r["당기정부보조금환입액"],
            "기말 정부보조금잔액(계산)": r["기말정부보조금잔액"],
            "순 감가상각비(보조금차감후)(계산)": 순감가상각비,
            "기말장부가액(계산)": 기말장부가액,
            "당기 회사계상 상각비": company_dep_f,
            "당기상각비 차이": 당기상각비_차이,
            "당기 회사계상 보조금환입액": company_grant_amort_f,
            "당기환입액 차이": (None if company_grant_amort_f is None else r["당기정부보조금환입액"] - company_grant_amort_f),
            "원가구분": a.get("원가구분"),
            "비고": 비고,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    return df.sort_values(["사업장", "계정과목", "자산명(세부내역)"], na_position="last").reset_index(drop=True)


# ── 계정분류별 요약표 ─────────────────────────────────────────────────────────

CATEGORY_ORDER = ["유형자산", "투자부동산", "무형자산"]


def build_category_summary(df: pd.DataFrame) -> dict:
    """자산분류(유형자산/투자부동산/무형자산)별 요약 지표 + 사업장×원가구분 감가상각비 피벗 계산."""
    if df.empty:
        return {}

    d = df.copy()
    d["자산분류"] = d["자산분류"].astype(str).str.strip().replace({"": "(미분류)", "nan": "(미분류)"})
    d["사업장"] = d["사업장"].astype(str).str.strip().replace({"": "(미분류)", "nan": "(미분류)"})
    d["원가구분"] = d["원가구분"].astype(str).str.strip().replace({"": "(미분류)", "None": "(미분류)", "nan": "(미분류)"})
    d["기초장부금액"] = (
        d["기초취득원가"] - d["기초감가상각누계액"] - d["기초손상차손누계액"] - d["기초정부보조금잔액"]
    )
    # 회사계상 순감가상각비(회사계상 상각비 - 회사계상 보조금환입액). 회사계상 상각비 미입력 자산은 NaN으로 남겨
    # 합계 시 자동 제외되도록 한다(0으로 채우면 "회사가 0으로 보고"한 것과 구분이 안 됨).
    d["순 감가상각비(회사계상)"] = d["당기 회사계상 상각비"] - d["당기 회사계상 보조금환입액"].fillna(0)

    present = list(dict.fromkeys(d["자산분류"]))
    ordered = [c for c in CATEGORY_ORDER if c in present] + [c for c in present if c not in CATEGORY_ORDER]

    summaries = {}
    for cat in ordered:
        cdf = d[d["자산분류"] == cat]
        # 사업장×원가구분 피벗: 계산값/회사계상값 각각, 그리고 그 차이(계산-회사계상)
        pivot = cdf.pivot_table(
            index="사업장", columns="원가구분", values="순 감가상각비(보조금차감후)(계산)",
            aggfunc="sum", fill_value=0.0,
        )
        pivot["소계"] = pivot.sum(axis=1)
        grand = pivot.sum(axis=0)
        grand.name = "총계"
        pivot = pd.concat([pivot, grand.to_frame().T])

        pivot_company = cdf.pivot_table(
            index="사업장", columns="원가구분", values="순 감가상각비(회사계상)",
            aggfunc="sum", fill_value=0.0,
        )
        pivot_company["소계"] = pivot_company.sum(axis=1)
        grand_c = pivot_company.sum(axis=0)
        grand_c.name = "총계"
        pivot_company = pd.concat([pivot_company, grand_c.to_frame().T])
        pivot_company = pivot_company.reindex(index=pivot.index, columns=pivot.columns, fill_value=0.0)

        pivot_diff = pivot - pivot_company

        # 계정과목별 내역(토지/건물/기계장치 등) — 최초 등장 순서 그대로 유지
        by_account = []
        for acct in dict.fromkeys(cdf["계정과목"]):
            adf = cdf[cdf["계정과목"] == acct]
            calc = {
                "계정과목": acct,
                "자산수": int(len(adf)),
                "기초장부금액": adf["기초장부금액"].sum(),
                "당기감가상각비": adf["당기감가상각비(계산)"].sum(),
                "당기정부보조금환입액": adf["당기 정부보조금환입액(계산)"].sum(),
                "순감가상각비": adf["순 감가상각비(보조금차감후)(계산)"].sum(),
                "당기손상차손": adf["당기 손상차손인식액(계산)"].sum(),
                "기말장부금액": adf["기말장부가액(계산)"].sum(),
            }
            by_account.append(calc)

        # 회사계상분 계정과목별 내역(계산표와 동일한 계정과목 순서·행 구성 유지)
        by_account_company = []
        for acct in dict.fromkeys(cdf["계정과목"]):
            adf = cdf[cdf["계정과목"] == acct]
            by_account_company.append({
                "계정과목": acct,
                "회사계상 입력자산수": int(adf["당기 회사계상 상각비"].notna().sum()),
                "당기감가상각비(회사계상)": adf["당기 회사계상 상각비"].sum(),
                "정부보조금환입액(회사계상)": adf["당기 회사계상 보조금환입액"].sum(),
                "순감가상각비(회사계상)": adf["순 감가상각비(회사계상)"].sum(),
            })

        # 계산 vs 회사계상 차이(계정과목별)
        by_account_diff = []
        for calc_row, comp_row in zip(by_account, by_account_company):
            by_account_diff.append({
                "계정과목": calc_row["계정과목"],
                "당기감가상각비 차이": calc_row["당기감가상각비"] - comp_row["당기감가상각비(회사계상)"],
                "정부보조금환입액 차이": calc_row["당기정부보조금환입액"] - comp_row["정부보조금환입액(회사계상)"],
                "순감가상각비 차이": calc_row["순감가상각비"] - comp_row["순감가상각비(회사계상)"],
            })

        summaries[cat] = {
            "자산수": int(len(cdf)),
            "기초장부금액": cdf["기초장부금액"].sum(),
            "당기감가상각비": cdf["당기감가상각비(계산)"].sum(),
            "당기정부보조금환입액": cdf["당기 정부보조금환입액(계산)"].sum(),
            "순감가상각비": cdf["순 감가상각비(보조금차감후)(계산)"].sum(),
            "당기손상차손": cdf["당기 손상차손인식액(계산)"].sum(),
            "기말장부금액": cdf["기말장부가액(계산)"].sum(),
            "회사계상 입력자산수": int(cdf["당기 회사계상 상각비"].notna().sum()),
            "당기감가상각비(회사계상)": cdf["당기 회사계상 상각비"].sum(),
            "정부보조금환입액(회사계상)": cdf["당기 회사계상 보조금환입액"].sum(),
            "순감가상각비(회사계상)": cdf["순 감가상각비(회사계상)"].sum(),
            "pivot": pivot,
            "pivot_company": pivot_company,
            "pivot_diff": pivot_diff,
            "by_account": by_account,
            "by_account_company": by_account_company,
            "by_account_diff": by_account_diff,
        }
    return summaries


def write_summary_sheet(ws, summaries: dict, company: str, target_fy: str, interim_month: int = None):
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="203864")
    section_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill("solid", fgColor="9DC3E6")
    sig_fill = PatternFill("solid", fgColor="FFFF00")
    bold = Font(bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    period_note = f", ~{interim_month}월 중간결산(반기 등)" if interim_month else ""
    ws.cell(row=1, column=1, value=f"고정자산 계정분류별 요약표 (회사: {company}, 회계연도: {target_fy}{period_note})").font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 16
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 16

    r = 3
    if not summaries:
        ws.cell(row=r, column=1, value="(자산 데이터 없음)")
        return

    METRIC_COLS = ["상각대상 자산수", "기초장부금액", "당기감가상각비(총액)", "정부보조금환입액",
                   "순감가상각비", "당기손상차손", "기말장부금액"]

    for cat, s in summaries.items():
        ws.cell(row=r, column=1, value=f"■ {cat}").fill = section_fill
        ws.cell(row=r, column=1).font = section_font
        for c in range(2, len(METRIC_COLS) + 1):
            ws.cell(row=r, column=c).fill = section_fill
        r += 2

        for i, h in enumerate(METRIC_COLS, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1
        values = [s["자산수"], s["기초장부금액"], s["당기감가상각비"], s["당기정부보조금환입액"],
                  s["순감가상각비"], s["당기손상차손"], s["기말장부금액"]]
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            if i > 1:
                cell.number_format = "#,##0"
        r += 2

        # 계정과목별 내역(토지/건물/기계장치 등 — 소계 대신 계정별 1행씩 + 합계행)
        ws.cell(row=r, column=1, value=f"{cat} 계정과목별 내역").font = bold
        r += 1
        acct_headers = ["계정과목"] + METRIC_COLS
        for i, h in enumerate(acct_headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1
        for acct_row in s["by_account"]:
            cell = ws.cell(row=r, column=1, value=acct_row["계정과목"])
            cell.border = border
            acct_values = [acct_row["자산수"], acct_row["기초장부금액"], acct_row["당기감가상각비"],
                           acct_row["당기정부보조금환입액"], acct_row["순감가상각비"],
                           acct_row["당기손상차손"], acct_row["기말장부금액"]]
            for i, v in enumerate(acct_values, start=2):
                cell = ws.cell(row=r, column=i, value=v)
                cell.border = border
                cell.number_format = "#,##0"
            r += 1
        cell = ws.cell(row=r, column=1, value="합계")
        cell.font = bold
        cell.fill = total_fill
        cell.border = border
        for i, v in enumerate(values, start=2):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            cell.font = bold
            cell.fill = total_fill
            cell.number_format = "#,##0"
        r += 2

        # 계정과목별 내역(회사계상) — 계산표와 동일한 계정과목 순서
        ws.cell(row=r, column=1, value=f"{cat} 계정과목별 내역 (회사계상)").font = bold
        r += 1
        comp_headers = ["계정과목", "회사계상 입력자산수", "당기감가상각비(회사계상)",
                         "정부보조금환입액(회사계상)", "순감가상각비(회사계상)"]
        for i, h in enumerate(comp_headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1
        comp_totals = [0, 0.0, 0.0, 0.0]
        for crow in s["by_account_company"]:
            cell = ws.cell(row=r, column=1, value=crow["계정과목"])
            cell.border = border
            cvals = [crow["회사계상 입력자산수"], crow["당기감가상각비(회사계상)"],
                     crow["정부보조금환입액(회사계상)"], crow["순감가상각비(회사계상)"]]
            for i, v in enumerate(cvals, start=2):
                cell = ws.cell(row=r, column=i, value=v)
                cell.border = border
                cell.number_format = "#,##0"
            for i, v in enumerate(cvals):
                comp_totals[i] += v
            r += 1
        cell = ws.cell(row=r, column=1, value="합계")
        cell.font = bold
        cell.fill = total_fill
        cell.border = border
        for i, v in enumerate(comp_totals, start=2):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            cell.font = bold
            cell.fill = total_fill
            cell.number_format = "#,##0"
        r += 2

        # 계정과목별 차이(계산 - 회사계상) — 유의차이(SIG_THRESHOLD_ABS 이상)는 노란색 강조
        ws.cell(row=r, column=1, value=f"{cat} 계정과목별 차이 (계산 - 회사계상)").font = bold
        r += 1
        diff_headers = ["계정과목", "당기감가상각비 차이", "정부보조금환입액 차이", "순감가상각비 차이"]
        for i, h in enumerate(diff_headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1
        diff_totals = [0.0, 0.0, 0.0]
        for drow in s["by_account_diff"]:
            cell = ws.cell(row=r, column=1, value=drow["계정과목"])
            cell.border = border
            dvals = [drow["당기감가상각비 차이"], drow["정부보조금환입액 차이"], drow["순감가상각비 차이"]]
            for i, v in enumerate(dvals, start=2):
                cell = ws.cell(row=r, column=i, value=v)
                cell.border = border
                cell.number_format = "#,##0"
                if abs(v) >= SIG_THRESHOLD_ABS:
                    cell.fill = sig_fill
            for i, v in enumerate(dvals):
                diff_totals[i] += v
            r += 1
        cell = ws.cell(row=r, column=1, value="합계")
        cell.font = bold
        cell.fill = total_fill
        cell.border = border
        for i, v in enumerate(diff_totals, start=2):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            cell.font = bold
            cell.fill = total_fill
            cell.number_format = "#,##0"
        r += 2

        ws.cell(row=r, column=1, value="사업장별 순 당기감가상각비 (원가구분별, 정부보조금환입액 차감후)").font = bold
        r += 1
        pivot = s["pivot"]
        cost_cols = list(pivot.columns)
        headers = ["사업장"] + cost_cols
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1
        for site, row in pivot.iterrows():
            is_total = site == "총계"
            cell = ws.cell(row=r, column=1, value=site)
            cell.border = border
            if is_total:
                cell.font = bold
                cell.fill = total_fill
            for i, c in enumerate(cost_cols, start=2):
                cell = ws.cell(row=r, column=i, value=row[c])
                cell.number_format = "#,##0"
                cell.border = border
                if is_total:
                    cell.font = bold
                    cell.fill = total_fill
            r += 1
        r += 2

        ws.cell(row=r, column=1, value="사업장별 순 당기감가상각비 (회사계상, 원가구분별)").font = bold
        r += 1
        pivot_company = s["pivot_company"]
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1
        for site, row in pivot_company.iterrows():
            is_total = site == "총계"
            cell = ws.cell(row=r, column=1, value=site)
            cell.border = border
            if is_total:
                cell.font = bold
                cell.fill = total_fill
            for i, c in enumerate(cost_cols, start=2):
                cell = ws.cell(row=r, column=i, value=row[c])
                cell.number_format = "#,##0"
                cell.border = border
                if is_total:
                    cell.font = bold
                    cell.fill = total_fill
            r += 1
        r += 2

        ws.cell(row=r, column=1, value="사업장별 순 당기감가상각비 차이 (계산 - 회사계상, 원가구분별, 유의차이 강조)").font = bold
        r += 1
        pivot_diff = s["pivot_diff"]
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        r += 1
        for site, row in pivot_diff.iterrows():
            is_total = site == "총계"
            cell = ws.cell(row=r, column=1, value=site)
            cell.border = border
            if is_total:
                cell.font = bold
                cell.fill = total_fill
            for i, c in enumerate(cost_cols, start=2):
                v = row[c]
                cell = ws.cell(row=r, column=i, value=v)
                cell.number_format = "#,##0"
                cell.border = border
                if is_total:
                    cell.font = bold
                    cell.fill = total_fill
                elif abs(v) >= SIG_THRESHOLD_ABS:
                    cell.fill = sig_fill
            r += 1
        r += 2


# ── 엑셀 저장 ────────────────────────────────────────────────────────────────

MONEY_COLS = [
    "기초취득원가", "당기증가(신규취득)", "당기감소(처분)", "기말취득원가",
    "기초감가상각누계액", "당기감가상각비(계산)", "처분시감소(누계액)", "기말감가상각누계액(계산)",
    "기초손상차손누계액", "당기 손상차손인식액(계산)", "기말 손상차손누계액(계산)",
    "기초정부보조금잔액", "당기 정부보조금수령액", "당기 정부보조금환입액(계산)", "기말 정부보조금잔액(계산)",
    "순 감가상각비(보조금차감후)(계산)",
    "기말장부가액(계산)",
    "당기 회사계상 상각비", "당기상각비 차이",
    "당기 회사계상 보조금환입액", "당기환입액 차이",
]

# 대사 차이 컬럼 → 유의성 판단 기준(분모)이 되는 회사계상액 컬럼
DIFF_BASE_COLS = {
    "당기상각비 차이": "당기 회사계상 상각비",
    "당기환입액 차이": "당기 회사계상 보조금환입액",
}


def _is_significant(diff, base) -> bool:
    if diff is None or pd.isna(diff):
        return False
    if abs(diff) >= SIG_THRESHOLD_ABS and (base in (None, 0) or abs(diff) >= abs(base) * SIG_THRESHOLD_PCT):
        return True
    return False


def save_results(df: pd.DataFrame, output_path: str, company: str, target_fy: str, interim_month: int = None):
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약표"
    write_summary_sheet(ws_summary, build_category_summary(df), company, target_fy, interim_month)

    ws = wb.create_sheet("고정자산명세서")

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    subtotal_fill = PatternFill("solid", fgColor="D9E1F2")
    total_fill = PatternFill("solid", fgColor="9DC3E6")
    sig_fill = PatternFill("solid", fgColor="FFFF00")
    over_life_fill = PatternFill("solid", fgColor="FFC000")
    bold = Font(bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    period_note = f", ~{interim_month}월 중간결산(반기 등)" if interim_month else ""
    ws.cell(row=1, column=1, value=f"고정자산명세서 (회사: {company}, 회계연도: {target_fy}{period_note})").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns) if not df.empty else 15)

    headers = list(df.columns)
    header_row = 3
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = 16 if h not in ("자산명(세부내역)", "비고") else 26
    ws.freeze_panes = f"A{header_row + 1}"

    r = header_row + 1
    totals = {c: 0.0 for c in MONEY_COLS}
    site_totals = {c: 0.0 for c in MONEY_COLS}
    grand_totals = {c: 0.0 for c in MONEY_COLS}

    if df.empty:
        ws.cell(row=r, column=1, value="(자산 데이터 없음)")
        wb.save(output_path)
        return

    # 사업장 컬럼에 실제 값이 하나라도 있으면 사업장→계정과목 2단 소계, 없으면 기존처럼 계정과목 소계만
    has_site = "사업장" in df.columns and (df["사업장"].astype(str).str.strip() != "").any()
    if has_site:
        df = df.copy()
        df["사업장"] = df["사업장"].astype(str).str.strip().replace("", "(미분류)")
        outer_groups = list(df.groupby("사업장", sort=False))
    else:
        outer_groups = [(None, df)]

    for site, site_df in outer_groups:
        for account, gdf in site_df.groupby("계정과목", sort=False):
            for _, row in gdf.iterrows():
                is_over_life = "내용연수초과" in str(row.get("비고") or "")
                for i, h in enumerate(headers, start=1):
                    val = row[h]
                    cell = ws.cell(row=r, column=i, value=(None if pd.isna(val) else val))
                    cell.border = border
                    if h == "취득일" and val is not None and not pd.isna(val):
                        cell.number_format = "yyyy-mm-dd"
                    if h in MONEY_COLS:
                        cell.number_format = "#,##0"
                    if is_over_life:
                        cell.fill = over_life_fill
                    if h in DIFF_BASE_COLS:
                        if _is_significant(val, row.get(DIFF_BASE_COLS[h])):
                            cell.fill = sig_fill
                for c in MONEY_COLS:
                    v = row.get(c)
                    if v is not None and not pd.isna(v):
                        totals[c] += v
                        site_totals[c] += v
                        grand_totals[c] += v
                r += 1

            # 계정과목 소계 행
            ws.cell(row=r, column=1, value=f"[{account} 소계]").font = bold
            for i, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=i)
                cell.fill = subtotal_fill
                cell.border = border
                if h in MONEY_COLS:
                    cell.value = totals[h]
                    cell.number_format = "#,##0"
                    cell.font = bold
            r += 1
            totals = {c: 0.0 for c in MONEY_COLS}

        if has_site:
            # 사업장 소계 행
            ws.cell(row=r, column=1, value=f"[{site} 사업장 소계]").font = bold
            for i, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=i)
                cell.fill = total_fill
                cell.border = border
                if h in MONEY_COLS:
                    cell.value = site_totals[h]
                    cell.number_format = "#,##0"
                    cell.font = bold
            r += 1
            site_totals = {c: 0.0 for c in MONEY_COLS}

    # 총계 행
    ws.cell(row=r, column=1, value="총계").font = Font(bold=True, size=11)
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=i)
        cell.fill = total_fill
        cell.border = border
        if h in MONEY_COLS:
            cell.value = grand_totals[h]
            cell.number_format = "#,##0"
            cell.font = bold

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="유형자산/투자부동산 감가상각비 검증앱")
    parser.add_argument("company", nargs="?", default=None, help="처리할 회사명 (생략 시 파일 자동 탐색)")
    parser.add_argument("--file", default=None, help="처리할 특정 입력 파일명 (input_data/ 기준)")
    parser.add_argument("--fiscal-month", type=int, default=12, help="결산월 (기본 12월). 예: 6월 결산법인(회계연도 7월~익년6월)이면 6")
    parser.add_argument("--fiscal-year", default=None, help="검증 대상 회계연도 (예: 2026). 생략 시 입력파일명의 fy 뒤 숫자 사용")
    parser.add_argument("--interim-month", type=int, default=None,
                         help="반기 등 중간결산 검토월 (예: 6 → 회계연도 시작월은 그대로 두고 그 해당월까지만 계산). "
                              "12월 결산 법인의 상반기(1~6월) 검토처럼, 기초잔액은 전기말 그대로 쓰고 당기 계산기간만 앞당길 때 사용")
    args = parser.parse_args()

    input_path = _find_input_file(args.company, args.file)
    company = args.company or os.path.basename(input_path).split("_")[1]

    target_fy = args.fiscal_year
    if not target_fy:
        base = os.path.basename(input_path)
        marker = "fy"
        idx = base.lower().find(marker)
        if idx == -1:
            raise ValueError("--fiscal-year 를 지정하거나 파일명에 'fy<연도>'를 포함하세요.")
        digits = ""
        for ch in base[idx + 2:]:
            if ch.isdigit():
                digits += ch
            else:
                break
        target_fy = f"20{digits}" if len(digits) == 2 else digits

    print(f"[입력] {input_path}")
    interim_note = f", 중간결산월={args.interim_month}(반기 등)" if args.interim_month else ""
    print(f"[대상] 회사={company}, 회계연도={target_fy}, 결산월={args.fiscal_month}{interim_note}")

    assets = load_assets(input_path)
    print(f"[자산 수] {len(assets)}건")

    df = build_schedule_table(assets, args.fiscal_month, target_fy, args.interim_month)

    suffix = f"_interim{args.interim_month:02d}" if args.interim_month else ""
    output_path = os.path.join(OUTPUT_DIR, f"depreciation_schedule_{company}_{target_fy}{suffix}.xlsx")
    save_results(df, output_path, company, target_fy, args.interim_month)
    print(f"[완료] {output_path}")


if __name__ == "__main__":
    main()
