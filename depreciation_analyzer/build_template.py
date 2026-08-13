"""표준 입력 템플릿(depreciation_template.xlsx) 생성 스크립트.

실행: python build_template.py
input_data/depreciation_template.xlsx 를 새로 만든다(이미 있으면 덮어씀).
회사별 파일은 이 템플릿을 복사해 depreciation_<company>_information_fy<year>.xlsx 로 저장해서 사용한다.

설계 원칙: 취득원가·감가상각누계액·손상차손누계액·정부보조금잔액 전부
"기초잔액은 회사가 보고한 값을 그대로 입력, 당기 증감만 앱이 계산"하는 동일한 방식.
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(HERE, "input_data", "depreciation_template.xlsx")

# (그룹헤더, 상세헤더, 열너비)
COLUMNS = [
    ("자산기본정보", "사업장", 14),
    ("자산기본정보", "자산분류(유형자산/투자부동산/무형자산)", 14),
    ("자산기본정보", "자산관리번호", 12),
    ("자산기본정보", "계정과목", 16),
    ("자산기본정보", "자산명(세부내역)", 28),
    ("취득정보", "취득일", 12),
    ("취득정보", "기초취득원가", 14),
    ("취득정보", "당기취득원가", 14),
    ("취득정보", "당기처분원가", 14),
    ("취득정보", "잔존가치", 12),
    ("취득정보", "내용연수(년)", 10),
    ("상각정보", "상각방법(정액법/정률법)", 12),
    ("상각정보", "상각률(정률법전용)", 12),
    ("상각정보", "상각개시(당월/익월)", 12),
    ("처분정보", "처분일", 12),
    ("기초잔액(회사계상값 입력)", "기초 감가상각누계액", 16),
    ("기초잔액(회사계상값 입력)", "기초 손상차손누계액", 16),
    ("기초잔액(회사계상값 입력)", "기초 정부보조금잔액", 16),
    ("손상차손(당기·선택)", "손상차손 인식일", 12),
    ("손상차손(당기·선택)", "손상차손 인식액", 14),
    ("정부보조금(선택)", "정부보조금 계정명", 18),
    ("정부보조금(선택)", "당기 정부보조금수령액", 16),
    ("회사계상액(대사용)", "당기 회사계상 감가상각비", 14),
    ("회사계상액(대사용)", "당기 회사계상 보조금환입액", 16),
    ("기타", "원가구분", 10),
    ("기타", "비고", 20),
]

ACCOUNT_SUGGESTIONS = [
    "건물", "건물(투자)", "구축물", "기계장치", "차량운반구",
    "비품", "시설장치", "사용권자산", "투자부동산",
    "소프트웨어", "산업재산권", "회원권", "개발비", "기타의무형자산",
]

ASSET_CATEGORY_OPTIONS = ["유형자산", "투자부동산", "무형자산"]


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "자산정보"

    group_fill = PatternFill("solid", fgColor="D9E1F2")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    group_font = Font(bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1행: 그룹 헤더 (병합)
    start = 1
    prev_group = COLUMNS[0][0]
    for i, (group, _, _) in enumerate(COLUMNS, start=1):
        if group != prev_group:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=i - 1)
            start = i
            prev_group = group
    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=len(COLUMNS))

    for i, (group, name, width) in enumerate(COLUMNS, start=1):
        c1 = ws.cell(row=1, column=i, value=group if i == 1 or COLUMNS[i - 2][0] != group else None)
        c1.fill = group_fill
        c1.font = group_font
        c1.alignment = center
        c1.border = border

        c2 = ws.cell(row=2, column=i, value=name)
        c2.fill = header_fill
        c2.font = header_font
        c2.alignment = center
        c2.border = border

        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "A3"

    MONEY_FIELDS = (
        "기초취득원가", "당기취득원가", "당기처분원가", "잔존가치",
        "기초 감가상각누계액", "기초 손상차손누계액", "기초 정부보조금잔액",
        "손상차손 인식액", "당기 정부보조금수령액",
        "당기 회사계상 감가상각비", "당기 회사계상 보조금환입액",
    )
    DATE_FIELDS = ("취득일", "처분일", "손상차손 인식일")

    # 예시행: 기존자산(정액법/정률법/손상차손 기이력/정부보조금), 신규취득, 투자부동산, 무형자산
    examples = [
        # 사업장, 자산분류, 자산관리번호, 계정과목, 자산명, 취득일,
        # 기초취득원가, 당기취득원가, 당기처분원가, 잔존가치, 내용연수,
        # 상각방법, 상각률, 상각개시,
        # 처분일,
        # 기초감가상각누계액, 기초손상차손누계액, 기초정부보조금잔액,
        # 손상인식일, 손상인식액, 정부보조금계정명, 당기정부보조금수령액,
        # 당기회사계상감가상각비, 당기회사계상보조금환입액, 원가구분, 비고
        [
            "본사", "유형자산", "FA-0001", "기계장치", "예시) 사출성형기 1호(기존보유)", "2023-03-15",
            120000000, 0, 0, 0, 8,
            "정액법", None, "당월",
            None,
            42500000, 0, 0,
            None, None, None, None,
            15000000, None, "제조", "예시 행 — 실제 자산으로 교체 (기초 감가상각누계액=전기말 회사 보고값 그대로 입력)",
        ],
        [
            "본사", "유형자산", "FA-0002", "차량운반구", "예시) 업무용 승용차(기존보유, 정률법)", "2024-07-01",
            45000000, 0, 0, 0, 5,
            "정률법", 0.451, "익월",
            None,
            21536877, 0, 0,
            None, None, None, None,
            10147500, None, "판관", "예시 행 — 실제 자산으로 교체",
        ],
        [
            "제2공장", "유형자산", "FA-0003", "건물", "예시) CGU 손상평가 반영 자산(기존보유)", "2015-01-01",
            500000000, 0, 0, 0, 40,
            "정액법", None, "당월",
            None,
            137500000, 80000000, 0,
            None, None, None, None,
            None, None, "제조",
            "예시) 25년말 CGU 손상평가로 손상차손 인식(전기) → 기초 손상차손누계액에 이미 반영, 당기 신규 손상 없음",
        ],
        [
            "본사", "유형자산", "FA-0006", "기계장치", "예시) 정부보조금 지원 설비(기존보유)", "2024-01-01",
            120000000, 0, 0, 0, 5,
            "정액법", None, "당월",
            None,
            48000000, 0, 18000000,
            None, None, "기계장치(국고보조금)", None,
            None, None, "제조",
            "예시) 국고보조금으로 취득 — 기초 정부보조금잔액을 그대로 입력, 이후 상각과 동일 비율로 자동 환입",
        ],
        [
            "본사", "유형자산", "FA-0007", "비품", "예시) 당기 신규취득 비품", "2026-03-01",
            0, 24000000, 0, 0, 5,
            "정액법", None, "당월",
            None,
            0, 0, 0,
            None, None, None, None,
            None, None, "판관", "예시) 당기 중 신규취득 — 기초취득원가=0, 당기취득원가에만 입력",
        ],
        [
            "본사", "투자부동산", "FA-0004", "건물(투자)", "예시) 임대용 건물(기존보유)", "2019-05-01",
            300000000, 0, 0, 0, 30,
            "정액법", None, "당월",
            None,
            66666667, 0, 0,
            None, None, None, None,
            None, None, "영업외비용", "예시 행 — 실제 자산으로 교체",
        ],
        [
            "제2공장", "무형자산", "FA-0005", "소프트웨어", "예시) ERP 라이선스(기존보유)", "2022-01-01",
            60000000, 0, 0, 0, 5,
            "정액법", None, "당월",
            None,
            48000000, 0, 0,
            None, None, None, None,
            None, None, "판관", "예시 행 — 실제 자산으로 교체",
        ],
    ]
    for r, row in enumerate(examples, start=3):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            field = COLUMNS[c - 1][1]
            if field in DATE_FIELDS and val:
                cell.number_format = "yyyy-mm-dd"
            if field in MONEY_FIELDS:
                cell.number_format = "#,##0"
            if field == "상각률(정률법전용)" and val is not None:
                cell.number_format = "0.000"

    # 데이터 유효성 검사 (100행까지) — 컬럼명으로 위치를 찾아 열 순서가 바뀌어도 안전하게 적용
    last_row = 100

    def _col_letter(header_name: str) -> str:
        idx = next(i for i, (_, name, _) in enumerate(COLUMNS, start=1) if name == header_name)
        return get_column_letter(idx)

    dv_category = DataValidation(
        type="list", formula1=f'"{",".join(ASSET_CATEGORY_OPTIONS)}"',
        allow_blank=True, showErrorMessage=True,
    )
    dv_category.error = "유형자산 / 투자부동산 / 무형자산 중 선택하세요."
    ws.add_data_validation(dv_category)
    col = _col_letter("자산분류(유형자산/투자부동산/무형자산)")
    dv_category.add(f"{col}3:{col}{last_row}")

    dv_method = DataValidation(type="list", formula1='"정액법,정률법"', allow_blank=True, showErrorMessage=True)
    dv_method.error = "정액법 또는 정률법 중 선택하세요."
    ws.add_data_validation(dv_method)
    col = _col_letter("상각방법(정액법/정률법)")
    dv_method.add(f"{col}3:{col}{last_row}")

    dv_start = DataValidation(type="list", formula1='"당월,익월"', allow_blank=True, showErrorMessage=True)
    dv_start.error = "당월 또는 익월 중 선택하세요."
    ws.add_data_validation(dv_start)
    col = _col_letter("상각개시(당월/익월)")
    dv_start.add(f"{col}3:{col}{last_row}")

    dv_account = DataValidation(
        type="list",
        formula1=f'"{",".join(ACCOUNT_SUGGESTIONS)}"',
        allow_blank=True,
        showErrorMessage=False,  # 목록 외 자유 입력 허용 (신규 계정과목 대응)
    )
    ws.add_data_validation(dv_account)
    col = _col_letter("계정과목")
    dv_account.add(f"{col}3:{col}{last_row}")

    # 안내 시트
    guide = wb.create_sheet("작성안내")
    guide.column_dimensions["A"].width = 100
    lines = [
        "감가상각비 검증앱 — 입력 템플릿 작성 안내",
        "",
        "핵심 원칙: 취득원가·감가상각누계액·손상차손누계액·정부보조금잔액 전부",
        "  '기초잔액은 회사가 보고한 값을 그대로 입력, 당기 증감만 앱이 계산'하는 동일한 방식입니다.",
        "  (여러 해 이력을 앱이 다시 시뮬레이션하지 않고, 전기말 회사 보고값을 신뢰합니다)",
        "",
        "1. 자산 1건 = 1행. '자산정보' 시트에 계속 추가하면 됩니다(계정과목별·사업장별로 시트를 나누지 않음).",
        "1-1. 사업장: 사업장이 여러 곳인 회사만 채우면 됩니다(예: 본사/제2공장). 채워두면 출력물이 사업장별 → 계정과목별 2단 소계로 나옵니다.",
        "1-2. 자산분류: 유형자산/투자부동산/무형자산 중 선택(필수 아님). '요약표' 시트가 이 값 기준으로 자동 집계됩니다.",
        "",
        "2. 취득정보 — 기존 보유 자산인지, 당기 신규취득 자산인지에 따라 딱 한 칸만 채웁니다.",
        "   [기존 보유 자산] '기초취득원가'에 취득원가 전액을 입력, '당기취득원가'는 0(또는 공란).",
        "   [당기 신규취득 자산] '당기취득원가'에 취득원가 전액을 입력, '기초취득원가'는 0(또는 공란). 이 경우 '취득일'이 반드시 필요합니다(월할상각 시작월 계산용).",
        "   [당기 중 처분한 자산] '당기처분원가'에 처분된 취득원가를 입력(전액 처분 가정). 처분월까지는 정상 상각 후, 취득원가·감가상각누계액·손상차손누계액·정부보조금잔액을 당기에 전액 제거합니다.",
        "   기말취득원가 = 기초취득원가 + 당기취득원가 - 당기처분원가 (자동 계산)",
        "",
        "3. 기초잔액(회사계상값 입력) — 기존 보유 자산만 해당(신규취득 자산은 전부 0).",
        "   '기초 감가상각누계액'/'기초 손상차손누계액'/'기초 정부보조금잔액'에 전기말 재무제표(또는 고정자산대장)상 실제 금액을 그대로 입력하세요.",
        "   앱은 이 값을 기초로 삼아 당기분만 재계산합니다 — 취득일이 오래되어도 몇 년치를 다시 계산할 필요가 없습니다.",
        "   '취득일'은 그래도 입력해두는 것을 권장합니다(당기 중 손상차손이 새로 발생하는 경우 잔여내용연수를 정확히 계산하는 데 사용되며, 없으면 '내용연수' 전체를 잔여기간으로 근사합니다).",
        "",
        "4. 상각방법: 정액법 선택 시 '상각률'은 비워두면 됩니다. 정률법 선택 시 '상각률'을 반드시 입력하세요(세법상 상각률표 등 참고).",
        "5. 상각개시(당월/익월): 신규취득 자산의 취득월부터 상각을 시작하면 '당월', 취득 다음달부터 시작하면 '익월'.",
        "6. 처분일: 당기 중 처분/폐기된 자산만 입력. 처분월까지만 상각하고 그 이후는 계산에서 제외됩니다.",
        "",
        "7. 손상차손(당기·선택) — 당기 중에 새로 손상차손을 인식한 자산만 입력(해당 없으면 비워둠).",
        "   과거(전기 이전)에 이미 발생한 손상차손은 여기 넣지 말고 '기초 손상차손누계액'에 반영해서 입력하세요.",
        "   '손상차손 인식일'이 속한 월까지는 정상 상각한 뒤, 그 시점 장부금액에서 '손상차손 인식액'을 차감하고 잔여기간 상각액을 재계산합니다.",
        "",
        "8. 정부보조금(선택) — 국고보조금 등을 받아 취득한 자산만 입력(해당 없으면 비워둠).",
        "   '정부보조금 계정명'은 회사마다 다른 실제 차감계정명을 그대로 적으면 됩니다(예: '기계장치(국)', '기계장치(정부보조금)', '기계장치(국고보조금)' 등 — 계산에는 쓰이지 않고 출력물 표시용).",
        "   [기존 보유 자산] '기초 정부보조금잔액'에 전기말 남은 보조금 잔액을 입력.",
        "   [당기 신규취득 자산] '당기 정부보조금수령액'에 당기 중 받은 보조금 총액을 입력.",
        "   보조금은 해당 자산의 감가상각과 동일한 비율로 매월 자동 환입되어 감가상각비를 차감합니다. 출력물에 '순 감가상각비(보조금차감후)'가 별도 표시되고, 요약표의 사업장별 감가상각비도 이 순액 기준입니다.",
        "",
        "9. 회사계상액(대사용) — '당기 회사계상 감가상각비'/'당기 회사계상 보조금환입액'을 채워두면, 앱 재계산 금액과 자동 비교해 차이를 표시합니다.",
        "   모르면 비워둬도 앱 실행에는 문제 없습니다(대사만 생략됨).",
        "10. 파일명 규칙: 이 템플릿을 복사해 'depreciation_<회사명>_information_fy<회계연도>.xlsx' 로 저장하세요.",
        "    예) depreciation_kyungnam_information_fy2026.xlsx",
    ]
    for i, line in enumerate(lines, start=1):
        cell = guide.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    os.makedirs(os.path.join(HERE, "input_data"), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"템플릿 생성 완료: {OUT_PATH}")


if __name__ == "__main__":
    build()
