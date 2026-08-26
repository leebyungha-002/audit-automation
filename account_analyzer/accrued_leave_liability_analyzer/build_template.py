"""표준 입력 템플릿(leave_template.xlsx) 생성 스크립트.

실행: python build_template.py
input_data/leave_template.xlsx 를 새로 만든다(이미 있으면 덮어씀).
회사별 파일은 이 템플릿을 복사해 leave_<company>_information_fy<year>.xlsx 로 저장해서 사용한다.

설계 원칙: 연월차충당부채(잔여연차일수 × 1일통상임금 방식) 검증.
  - 결산기준일(당기말/전기말)은 셀에 직접 입력하지 않고, 파일명(fy<연도>)과 실행 시 --fiscal-month로
    severance_analyzer/depreciation_analyzer와 동일한 규칙으로 앱이 계산한다.
  - depreciation_analyzer와 마찬가지로 "기초(이월) 잔액은 입력값을 신뢰하고, 당기 발생분만 계산"한다
    (연차는 이월이 본질인 값이라 severance_analyzer처럼 전액 독립 재계산할 수 없음). '당기정보'/'전기정보'
    두 시트 각각이 자기 시점의 '기초 이월연차잔여일수'를 독립 입력받고, 그 위에 당기(그 시트가 나타내는
    회계연도) 부여일수만 앱이 계산해 얹는다.
  - 전기/당기 인원은 회사별로 별도 파일을 만드는 대신, 이 한 파일 안에 '당기정보'/'전기정보'
    두 시트로 나눠 입력받는다. 당기 계산은 '당기정보' 시트만 사용하고, '전기정보'는 사번(없으면 성명)
    기준으로 매칭해 신규입사자/퇴사자 명단을 산출하고 전기말 잔액을 독립 재계산하는 데 쓰인다.
  - '기준정보' 시트의 '연차산정기준'(입사기준/회계기준)에 따라 근속연수를 세는 시점만 달라진다.
    입사기준: 개인별 입사기념일마다 근속연수 갱신(+ 직전 기념일부터 결산일까지 진행 중인 다음
    사이클도 발생주의로 월할 안분). 회계기준: 전 직원이 결산기준일(회계연도 종료일)에 일괄 갱신
    (입사연도는 비례연차). 둘 다 발생주의(그 연차를 만든 근로가 제공된 회계기간 말에 부채 인식,
    K-IFRS 1019 누적유급휴가 원칙) 기준이다. 미입력 시 '입사기준'으로 계산된다.
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

ACCRUAL_RATE_MEMO = (
    "연차사용촉진 반영 부채인정비율 — 계산식 메모 (헷갈리기 쉬운 개념이라 예시로 설명)\n\n"
    "결산일 현재 잔여연차 1일은 다음 셋 중 하나로 귀결됩니다:\n"
    "  A. 실제 사용 예정 → 그 유급휴가를 줄 의무 자체가 이미 당기 근무의 대가로 발생\n"
    "     (일반기준 21.5의2 매칭원칙, 현금유출 여부와 무관) → 부채 O\n"
    "  B. 미사용인데 촉진 절차가 적법 이행돼 금전보상의무까지 완전 면제 → 부채 X (완전소멸)\n"
    "  C. 미사용인데 촉진 실패(절차 하자·퇴사자라 시간 없음·대상 제외 등) → 결국 현금 지급 → 부채 O\n\n"
    "부채인정비율 = (A+C) / (A+B+C) = 1 − B의 비율\n"
    "  ※ 자주 하는 실수: '미사용예상비율'(B+C)이나 '지급될 비율'(C)만 곱하는 것 — 틀립니다.\n"
    "    A(사용예정분)도 촉진과 무관하게 그 자체로 부채입니다.\n\n"
    "예시) 잔여연차 10일, 1일 통상임금 100,000원인 직원\n"
    "  A(사용예정) 7일 + C(미사용·지급예정) 2일 + B(완전소멸) 1일 = 10일\n"
    "  부채인정비율 = (7+2)/10 = 90%\n"
    "  연차충당부채 = 10일 × 100,000원 × 90% = 900,000원\n"
    "  (촉진 미적용이면 B=0이므로 100% 그대로 → 1,000,000원)"
)

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(HERE, "input_data", "leave_template.xlsx")

# (그룹헤더, 상세헤더, 열너비) — '당기정보'/'전기정보' 두 시트가 동일한 구조를 공유한다.
COLUMNS = [
    ("인적사항", "사업장", 14),
    ("인적사항", "부서", 14),
    ("인적사항", "사번", 12),
    ("인적사항", "성명", 12),
    ("인적사항", "직급", 10),
    ("재직정보", "원가구분(제조원가/판관비)", 14),
    ("재직정보", "입사일", 12),
    ("연차현황", "기초 이월연차잔여일수(일)", 18),
    ("연차현황", "당기 연차사용일수(일)", 16),
    ("연차현황", "1일 통상임금(원)", 14),
    ("대사용(선택)", "회사계상 기말 연차충당부채(원)", 20),
    ("기타", "비고", 30),
]

MONEY_FIELDS = ("1일 통상임금(원)", "회사계상 기말 연차충당부채(원)")
DAY_FIELDS = ("기초 이월연차잔여일수(일)", "당기 연차사용일수(일)")
DATE_FIELDS = ("입사일",)

COST_TYPE_OPTIONS = ["제조원가", "판관비"]

BASIS_MODE_LABEL = "연차산정기준(입사기준/회계기준)"
BASIS_MODE_OPTIONS = ["입사기준", "회계기준"]
ACCRUAL_RATE_LABEL_CURRENT = "당기 연차사용촉진 반영 부채인정비율(%, 미입력시 100%)"
ACCRUAL_RATE_LABEL_PRIOR = "전기 연차사용촉진 반영 부채인정비율(%, 미입력시 100%)"
PAYROLL_COUNT_LABEL = "기말 급여대장상 총인원수(명부 미확보 시 참고용)"

BASIS_ROWS = [
    "전기말 회사계상 연차충당부채(제조원가분)",
    "전기말 회사계상 연차충당부채(판관비분)",
    "당기말 회사계상 연차충당부채(제조원가분)",
    "당기말 회사계상 연차충당부채(판관비분)",
    "당기 연차충당부채 차변(당기지급액, 분개장 기준)",
]

# '당기퇴사자' 시트 — 선택 입력. '사번'(없으면 성명) + '실제지급액(원)'만 입력하면
# '전기정보'에서 같은 사람을 찾아 전기말 연차충당부채(계산)와 자동 대사한다.
LEAVER_COLUMNS = [("퇴사자 지급정보", "사번", 14), ("퇴사자 지급정보", "성명", 14),
                   ("퇴사자 지급정보", "실제지급액(원)", 16),
                   ("퇴사자 지급정보", "입사일(선택)", 14), ("퇴사자 지급정보", "비고", 30)]
LEAVER_EXAMPLES = [
    ["EMP-0999", "예시)최과장", 1200000, None,
     "예시) 전기정보에 있는 사번으로 매칭 — 전기말 연차충당부채(계산)와 자동 대사됨. 사번 없으면 성명으로 매칭."],
]

# '급여대장인원명부' 시트 — 선택 입력. 기말 급여대장상 실제 인원명부를 사업장/부서/사번/성명/직급만
# 붙여넣으면, 앱이 '당기정보'(연차수당 대상인원)와 사번(없으면 성명) 기준으로 자동 대사한다.
# 명부를 못 받았으면 이 시트는 비워두고, 대신 '기준정보' 시트의 총인원수 참고값만 입력해도 된다.
PAYROLL_COLUMNS = [("급여대장 인원(선택)", "사업장", 14), ("급여대장 인원(선택)", "부서", 14),
                    ("급여대장 인원(선택)", "사번", 12), ("급여대장 인원(선택)", "성명", 12),
                    ("급여대장 인원(선택)", "직급", 10)]
PAYROLL_EXAMPLES = [
    ["본사", "생산1팀", "EMP-1001", "예시)홍길동", "과장"],
    ["본사", "경영지원팀", "EMP-1002", "예시)김영희", "대리"],
    ["본사", "영업팀", "EMP-1003", "예시)이철수", "사원"],
    ["본사", "기타", "EMP-9001", "예시)김철수", "사원"],
]

# 예시행: 기존재직자(제조/판관), 당기 신규입사자(비례연차/월단위 발생 케이스), 근속 3년 이상(가산휴가 케이스)
# — '당기정보'/'전기정보' 두 시트에 걸쳐 계속재직자는 양쪽 모두, 신규입사자는 당기만,
#   퇴사자는 전기만 등장시켜 매칭 로직(사번 기준)을 그대로 보여준다.
EXAMPLES_CURRENT = [
    [
        "본사", "생산1팀", "EMP-1001", "예시)홍길동", "과장",
        "제조원가", "2019-03-02",
        5, 3, 120000, 600000,
        "예시 행 — 실제 인원으로 교체(기존 재직자). 기초 이월연차잔여일수(전기말 이월분) 5일 + 당기 부여일수(계산)"
        " − 당기 사용 3일로 당기말 잔여가 자동 계산됨. '회사계상 기말 연차충당부채'를 채우면 앱 계산값과 자동 대사",
    ],
    [
        "본사", "경영지원팀", "EMP-1002", "예시)김영희", "대리",
        "판관비", "2021-11-15",
        2, 10, 110000, None, "예시 행 — 기존 재직자(회사계상액 모름 — 비워두면 대사만 생략)",
    ],
    [
        "본사", "영업팀", "EMP-1003", "예시)이철수", "사원",
        "판관비", "2026-06-01",
        0, 0, 100000, None,
        "예시) 당기 중 신규입사 — '전기정보'에는 없어 요약표의 '신규입사자 명단'에 자동 표시됨. "
        "근속 1년 미만이므로 월단위 발생(또는 회계기준이면 비례연차)만 반영됨",
    ],
    [
        "제2공장", "생산2팀", "EMP-1004", "예시)박민수", "사원",
        "제조원가", "2018-06-01",
        8, 15, 105000, None,
        "예시) 근속 3년 이상 — 근로기준법상 가산휴가가 자동 반영됨(3년차 16일, 5년차 17일 … 25일 한도)",
    ],
]

# 전기 시트: 계속재직자(EMP-1001/1002/1004)는 그대로 유지, 신규입사자(EMP-1003)는 빼고,
# 당기에는 없는 퇴사자(EMP-0999)를 하나 추가해 '퇴사자 명단' 매칭을 보여준다.
EXAMPLES_PRIOR = [
    [
        "본사", "생산1팀", "EMP-1001", "예시)홍길동", "과장",
        "제조원가", "2019-03-02",
        4, 4, 118000, 580000,
        "예시 행 — 당기와 동일 인원(전기 시점 데이터). 이 시트의 '기초 이월연차잔여일수'는 전전기말 이월분",
    ],
    [
        "본사", "경영지원팀", "EMP-1002", "예시)김영희", "대리",
        "판관비", "2021-11-15",
        5, 13, 108000, None, "예시 행 — 당기와 동일 인원",
    ],
    [
        "제2공장", "생산2팀", "EMP-1004", "예시)박민수", "사원",
        "제조원가", "2018-06-01",
        6, 14, 103000, None, "예시 행 — 당기와 동일 인원",
    ],
    [
        "본사", "생산1팀", "EMP-0999", "예시)최과장", "과장",
        "제조원가", "2012-02-01",
        10, 10, 100000, None,
        "예시) 당기 중 퇴사 — '당기정보'에는 이 사번이 없어 요약표의 '퇴사자 명단'에 자동 표시됨",
    ],
]


def build():
    wb = openpyxl.Workbook()
    ws_current = wb.active
    ws_current.title = "당기정보"
    ws_prior = wb.create_sheet("전기정보")

    group_fill = PatternFill("solid", fgColor="D9E1F2")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    group_font = Font(bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _col_letter(header_name: str) -> str:
        idx = next(i for i, (_, name, _) in enumerate(COLUMNS, start=1) if name == header_name)
        return get_column_letter(idx)

    def build_sheet(ws, examples: list):
        # 1행: 그룹 헤더 (병합)
        start = 1
        prev_group = COLUMNS[0][0]
        for i, (group, _, _) in enumerate(COLUMNS, start=1):
            if group != prev_group:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=i - 1)
                start = i
                prev_group = group
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=len(COLUMNS))

        for i, (group, name, width) in enumerate(COLUMNS, start=1):
            c1 = ws.cell(row=1, column=i, value=group if i == 1 or COLUMNS[i - 2][0] != group else None)
            c1.fill = group_fill
            c1.font = group_font
            c1.alignment = center
            c1.border = border

            c2 = ws.cell(row=2, column=i, value=name)
            c2.fill = header_fill
            c2.font = header_font
            c2.alignment = center
            c2.border = border

            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 32
        ws.freeze_panes = "A3"

        for r, row in enumerate(examples, start=3):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = border
                field = COLUMNS[c - 1][1]
                if field in DATE_FIELDS and val:
                    cell.number_format = "yyyy-mm-dd"
                if field in MONEY_FIELDS:
                    cell.number_format = "#,##0"
                if field in DAY_FIELDS and val is not None:
                    cell.number_format = "0.0"

        # 데이터 유효성 검사 (300행까지)
        last_row = 300
        dv_cost = DataValidation(
            type="list", formula1=f'"{",".join(COST_TYPE_OPTIONS)}"',
            allow_blank=True, showErrorMessage=True,
        )
        dv_cost.error = "제조원가 또는 판관비 중 선택하세요."
        ws.add_data_validation(dv_cost)
        col = _col_letter("원가구분(제조원가/판관비)")
        dv_cost.add(f"{col}3:{col}{last_row}")

    build_sheet(ws_current, EXAMPLES_CURRENT)
    build_sheet(ws_prior, EXAMPLES_PRIOR)

    # 당기퇴사자 시트 — 선택 입력. 채우면 퇴사자별 전기말 연차충당부채 vs 실제지급액 대사가 요약표에 추가된다.
    ws_leaver = wb.create_sheet("당기퇴사자")
    ws_leaver.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(LEAVER_COLUMNS))
    c1 = ws_leaver.cell(row=1, column=1, value=LEAVER_COLUMNS[0][0])
    c1.fill = group_fill
    c1.font = group_font
    c1.alignment = center
    c1.border = border
    for i, (_, name, width) in enumerate(LEAVER_COLUMNS, start=1):
        c2 = ws_leaver.cell(row=2, column=i, value=name)
        c2.fill = header_fill
        c2.font = header_font
        c2.alignment = center
        c2.border = border
        ws_leaver.column_dimensions[get_column_letter(i)].width = width
    ws_leaver.row_dimensions[1].height = 20
    ws_leaver.row_dimensions[2].height = 32
    ws_leaver.freeze_panes = "A3"
    for r, row in enumerate(LEAVER_EXAMPLES, start=3):
        for c, val in enumerate(row, start=1):
            cell = ws_leaver.cell(row=r, column=c, value=val)
            cell.border = border
            field = LEAVER_COLUMNS[c - 1][1]
            if field == "실제지급액(원)":
                cell.number_format = "#,##0"
            if field == "입사일(선택)" and val:
                cell.number_format = "yyyy-mm-dd"

    # 급여대장인원명부 시트 — 선택 입력. 채우면 요약표에 '연차수당 대상인원 대사' 표가 인별로 추가된다.
    ws_payroll = wb.create_sheet("급여대장인원명부")
    ws_payroll.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PAYROLL_COLUMNS))
    c1 = ws_payroll.cell(row=1, column=1, value=PAYROLL_COLUMNS[0][0])
    c1.fill = group_fill
    c1.font = group_font
    c1.alignment = center
    c1.border = border
    for i, (_, name, width) in enumerate(PAYROLL_COLUMNS, start=1):
        c2 = ws_payroll.cell(row=2, column=i, value=name)
        c2.fill = header_fill
        c2.font = header_font
        c2.alignment = center
        c2.border = border
        ws_payroll.column_dimensions[get_column_letter(i)].width = width
    ws_payroll.row_dimensions[1].height = 20
    ws_payroll.row_dimensions[2].height = 32
    ws_payroll.freeze_panes = "A3"
    for r, row in enumerate(PAYROLL_EXAMPLES, start=3):
        for c, val in enumerate(row, start=1):
            cell = ws_payroll.cell(row=r, column=c, value=val)
            cell.border = border

    # 기준정보 시트 — 연차산정기준(전사 공통 설정) + 전기말/당기말 회사계상 충당부채(대사용)
    basis = wb.create_sheet("기준정보")
    basis.column_dimensions["A"].width = 42
    basis.column_dimensions["B"].width = 18
    basis.column_dimensions["C"].width = 52

    basis.cell(row=1, column=1, value="연월차충당부채 기준정보").font = Font(bold=True, size=13)

    header_row = 3
    for i, h in enumerate(["항목", "값", "설명"], start=1):
        cell = basis.cell(row=header_row, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    mode_row = header_row + 1
    c1 = basis.cell(row=mode_row, column=1, value=BASIS_MODE_LABEL)
    c1.border = border
    c2 = basis.cell(row=mode_row, column=2, value="입사기준")
    c2.border = border
    c2.fill = PatternFill("solid", fgColor="FFF2CC")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c3 = basis.cell(
        row=mode_row, column=3,
        value="입사기준: 개인별 입사기념일마다 근속연수가 갱신되어 그 시점에 연차가 전액 부여되고, "
              "직전 기념일부터 결산기준일까지 진행 중인 다음 사이클도 경과월수만큼 안분 가산됩니다. "
              "회계기준: 전 직원이 결산기준일(회계연도 종료일)에 일괄로 근속연수가 갱신되며, 입사연도에는 "
              "비례연차가 적용됩니다. 둘 다 발생주의(그 연차를 만든 근로가 제공된 회계기간 말에 인식) "
              "기준입니다. 미입력 시 '입사기준'으로 계산됩니다.",
    )
    c3.border = border
    c3.alignment = Alignment(wrap_text=True, vertical="top")

    dv_mode = DataValidation(
        type="list", formula1=f'"{",".join(BASIS_MODE_OPTIONS)}"',
        allow_blank=True, showErrorMessage=True,
    )
    dv_mode.error = "입사기준 또는 회계기준 중 선택하세요."
    basis.add_data_validation(dv_mode)
    dv_mode.add(f"B{mode_row}")

    payroll_count_row = mode_row + 1
    c1 = basis.cell(row=payroll_count_row, column=1, value=PAYROLL_COUNT_LABEL)
    c1.border = border
    c2 = basis.cell(row=payroll_count_row, column=2)
    c2.border = border
    c2.number_format = "0"
    c2.fill = PatternFill("solid", fgColor="FFF2CC")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c3 = basis.cell(
        row=payroll_count_row, column=3,
        value="'급여대장인원명부' 시트에 실제 인원명부(사업장/부서/사번/성명/직급)를 붙여넣을 수 있으면 이 값은 "
              "입력하지 않아도 됩니다(그 경우 인별 대사가 자동으로 이루어짐). 명부를 확보하지 못했을 때만, "
              "기말 급여대장상 총인원수를 여기에 입력하면 연차수당 대상인원수(당기정보 인원수)와 총원 차이만 "
              "요약표에 표시됩니다(수기 검증용).",
    )
    c3.border = border
    c3.alignment = Alignment(wrap_text=True, vertical="top")

    accrual_rate_current_row = payroll_count_row + 1
    c1 = basis.cell(row=accrual_rate_current_row, column=1, value=ACCRUAL_RATE_LABEL_CURRENT)
    c1.border = border
    c2 = basis.cell(row=accrual_rate_current_row, column=2)
    c2.border = border
    c2.number_format = "0.0"
    c2.fill = PatternFill("solid", fgColor="FFF2CC")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    memo = Comment(ACCRUAL_RATE_MEMO, "leave_analyzer")
    memo.width, memo.height = 420, 320
    c2.comment = memo
    c3 = basis.cell(
        row=accrual_rate_current_row, column=3,
        value="연차사용촉진제도(근로기준법 제61조)를 적법하게 이행해도, 잔여연차 중 완전히 소멸(미사용+촉진 "
              "이행으로 금전보상의무까지 면제)되는 부분만 부채가 0입니다. 사용될 것으로 예상되는 부분은 그 "
              "유급휴가를 제공할 의무 자체가 이미 당기 근무의 대가로 발생했으므로(일반기준 21.5의2 매칭원칙 — "
              "현금유출 여부와 무관), 미사용인데 촉진 실패로 금전보상해야 하는 부분과 함께 모두 부채로 남습니다. "
              "인원별로 이 구성비를 확인하기 어려우므로, 전사 공통으로 '완전소멸되지 않고 부채로 남을 것으로 "
              "예상하는 비율(%) = 1 − 완전소멸 예상비율'을 입력하면 이 비율만큼만 당기말 충당부채로 인식합니다"
              "(잔여일수 자체는 그대로 표시되고 금액에만 곱해짐). 예) 70 입력 시 당기말 잔여연차 금액의 70%만 "
              "충당부채로 인식(30%는 미사용+촉진 적법이행으로 완전소멸 예상). 촉진제도를 쓰지 않거나 비율을 "
              "모르면 비워두세요(100%로 계산 — 잔여연차 전액 부채).",
    )
    c3.border = border
    c3.alignment = Alignment(wrap_text=True, vertical="top")

    accrual_rate_prior_row = accrual_rate_current_row + 1
    c1 = basis.cell(row=accrual_rate_prior_row, column=1, value=ACCRUAL_RATE_LABEL_PRIOR)
    c1.border = border
    c2 = basis.cell(row=accrual_rate_prior_row, column=2)
    c2.border = border
    c2.number_format = "0.0"
    c2.fill = PatternFill("solid", fgColor="FFF2CC")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    memo_prior = Comment(ACCRUAL_RATE_MEMO, "leave_analyzer")
    memo_prior.width, memo_prior.height = 420, 320
    c2.comment = memo_prior
    c3 = basis.cell(
        row=accrual_rate_prior_row, column=3,
        value="위와 동일하되 전기말 충당부채 재계산에 적용되는 부채인정비율입니다. 연도별로 촉진 이행 여부·실제 "
              "소멸률이 달라질 수 있어 당기와 별도로 입력받습니다. 전기에는 촉진제도를 쓰지 않았거나 비율을 "
              "모르면 비워두세요(100%로 계산됨).",
    )
    c3.border = border
    c3.alignment = Alignment(wrap_text=True, vertical="top")

    money_header_row = accrual_rate_prior_row + 2
    for i, h in enumerate(["항목", "금액(원)", "설명"], start=1):
        cell = basis.cell(row=money_header_row, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    notes = [
        "전기말 재무제표(또는 결산정산표)상 실제 계상액 — 앱이 '전기정보' 시트로 인별 재계산한 전기말 잔액과 비교해 차이(대사)를 표시함(재계산 자체에는 반영되지 않음)",
        "전기말 재무제표(또는 결산정산표)상 실제 계상액 — 위와 동일",
        "당기말 재무제표상 회사 계상액 — 앱의 인원별 재계산 합계와 비교해 차이(대사)를 표시함",
        "당기말 재무제표상 회사 계상액",
        "분개장분석(journal_analyzer)에서 연차충당부채(또는 관련 미지급비용) 계정의 당기 차변(연차수당 지급으로 "
        "인한 감소) 합계를 뽑아 입력. 전기말(회사계상)+당기 연차수당비용(재계산)-이 값이 당기말(회사계상)과 "
        "맞는지 요약표에서 자동 검증(T계정 tie-out). 모르면 비워둬도 됨(해당 검증만 생략).",
    ]
    for i, (label, note) in enumerate(zip(BASIS_ROWS, notes), start=money_header_row + 1):
        c1 = basis.cell(row=i, column=1, value=label)
        c1.border = border
        c2 = basis.cell(row=i, column=2)
        c2.border = border
        c2.number_format = "#,##0"
        c2.fill = PatternFill("solid", fgColor="FFF2CC")
        c3 = basis.cell(row=i, column=3, value=note)
        c3.border = border
        c3.alignment = Alignment(wrap_text=True, vertical="top")

    basis.cell(row=money_header_row + len(BASIS_ROWS) + 2, column=1,
               value="※ 결산기준일(당기말/전기말)은 이 시트에 입력하지 않습니다 — 실행 시 파일명(fy<연도>)과 "
                     "--fiscal-month 옵션으로 자동 결정됩니다.").font = Font(italic=True, color="808080")

    # 안내 시트
    guide = wb.create_sheet("작성안내")
    guide.column_dimensions["A"].width = 100
    lines = [
        "연월차충당부채 검증앱(잔여연차일수 × 1일통상임금 방식) — 입력 템플릿 작성 안내",
        "",
        "핵심 원칙: 연차는 '이월'이 본질인 값이라 감가상각비 앱과 같은 방식으로 기초잔액을 신뢰합니다.",
        "  '당기정보'/'전기정보' 두 시트 각각이 자기 시점의 '기초 이월연차잔여일수'를 독립적으로 입력받고,",
        "  그 위에 당기(그 시트가 나타내는 회계연도) 부여일수만 이 앱이 근로기준법 규정으로 계산해 얹습니다.",
        "  당기말 잔여연차일수(계산) = 기초 이월연차잔여일수(입력) + 당기부여일수(계산) − 당기연차사용일수(입력)",
        "  당기말 연차충당부채(계산) = 당기말 잔여연차일수(계산) × 1일 통상임금(입력)",
        "",
        "0. 시트 구성 — '당기정보'와 '전기정보' 두 시트로 나뉩니다(회사별 파일을 따로 만들 필요 없음).",
        "   '당기정보': 이번 결산기준일 현재 재직 중인 인원 전체 + 이번 회계연도 연차현황. 당기말 계산에 쓰입니다.",
        "   '전기정보': 직전 결산기준일 현재 재직 중이었던 인원 전체 + 그 시점(직전 회계연도) 연차현황",
        "     (회사가 작년에 준 인원현황을 그대로 붙여넣으면 됨). 같은 산식으로 전기말 연차충당부채를",
        "     계산하는 데 쓰이고, 사번(없으면 성명) 기준으로 '당기정보'와 매칭해 신규입사자/퇴사자 명단을",
        "     만드는 데에도 쓰입니다.",
        "   ※ 사번이 두 시트에서 반드시 동일해야 정확히 매칭됩니다. 사번이 없으면 성명으로 매칭하니 동명이인에 주의하세요.",
        "   ※ '당기 연차수당비용(전입액에 해당)' = 인원별 (당기말 연차충당부채 − 전기말 연차충당부채)의 합계로",
        "     요약표에 자동 산출됩니다. 신규입사자는 전기말 충당부채를 0으로 보고, 퇴사자는 당기 표 자체에",
        "     없어 이 합계에 포함되지 않습니다(퇴사자 정산은 별도로 '당기퇴사자' 시트에서 관리).",
        "",
        "1. 인원 1명 = 1행. 두 시트 모두 같은 컬럼 구조입니다. 계속 재직 중인 사람은 두 시트 모두에,",
        "   당기 신규입사자는 '당기정보'에만, 당기 중 퇴사한 사람은 '전기정보'에만 입력하면 됩니다.",
        "",
        "2. 재직정보",
        "   '원가구분': 급여를 제조원가(생산직 등)로 처리하는지 판관비(관리직 등)로 처리하는지 선택.",
        "   '입사일': 필수. 근속연수(연차 부여일수) 계산의 기준이 됩니다.",
        "",
        "3. 연차현황 (두 시트 모두 동일 컬럼 — '당기정보'는 당기말, '전기정보'는 전기말 계산에 각각 쓰임)",
        "   '기초 이월연차잔여일수(일)': 그 시트가 나타내는 회계연도가 시작되는 시점의 이월 잔여연차일수.",
        "     회사의 연차관리대장·인사시스템상 실제 이월잔여를 그대로 입력하세요(이 값은 재계산하지 않고 신뢰함).",
        "   '당기 연차사용일수(일)': 그 회계연도 중 실제 사용한 연차일수(반차는 0.5로 입력 가능).",
        "   '1일 통상임금(원)': 그 시점 기준 1일 통상임금. 충당부채 = 잔여일수 × 이 값으로 계산됩니다.",
        "",
        "   [당기부여일수 계산 — 근로기준법 제60조]",
        "     근속연수 1년 이상: 부여일수 = min(15 + (근속연수−1)//2, 25)  (3년 이상부터 매 2년마다 1일 가산, 25일 한도)",
        "     근속연수 1년 미만(입사연도): 입사 후 매 1개월 경과 시마다 1일씩 발생(최대 11일, 개근 가정)",
        "     근속연수를 세는 시점은 '기준정보' 시트의 '연차산정기준' 설정에 따라 달라집니다(아래 6번 참고).",
        "",
        "   '회사계상 기말 연차충당부채(원)' (선택): 회사가 인원별로 이미 계산해둔 연차충당부채를 알면",
        "     채워 넣으세요. '당기정보'에 채우면 당기말 계산값과, '전기정보'에 채우면 전기말 계산값과",
        "     인별로 자동 대사되어 '인원별추계명세'/'전기인원별추계명세' 시트에 차이가 표시되고",
        "     유의차이는 노란색으로 강조됩니다. 모르면 비워둬도 앱 실행에는 문제 없습니다.",
        "",
        "4. 신규입사자/퇴사자 명단 (요약표에 자동 산출)",
        "   신규입사자 = '당기정보'에는 있으나 '전기정보'에는 없는 사번(또는 성명).",
        "   퇴사자(자동 산출) = '전기정보'에는 있으나 '당기정보'에는 없는 사번(또는 성명).",
        "   '전기정보' 시트를 비워두면(또는 시트 자체를 지우면) 인원변동 명단만 생략되고 당기 계산은 정상 진행됩니다.",
        "   요약표의 '퇴사자 명단 대사' 표는 이 자동 산출 명단과 '당기퇴사자' 시트(아래 5번, 사용자 입력) 명단을",
        "   나란히 놓고 비교합니다 — 양쪽에 모두 있으면 비고에 '이상없음', 한쪽에만 있으면 원인을 표시합니다.",
        "",
        "5. '당기퇴사자' 시트 (선택) — 퇴직 시 연차정산 지급액을 알면 채워서 인별 대사를 추가로 받을 수 있습니다.",
        "   '사번'(없으면 성명) + '실제지급액(원)'만 입력하면, 앱이 '전기정보' 시트에서 같은 사번(성명)을 찾아",
        "   그 사람의 전기말 연차충당부채(계산)와 자동으로 비교(대사)해 요약표에 표시합니다. 사업장/부서/직급",
        "   등은 '전기정보'에서 그대로 가져오므로 다시 입력할 필요가 없습니다. 모르는 퇴사자는 행을 비워두면",
        "   됩니다(해당 인원만 대사가 생략되고, 나머지 재계산·집계에는 영향이 없습니다).",
        "   '입사일(선택)': '전기정보'에서 매칭이 안 되는 인원에 한해 참고용으로 적어둘 수 있습니다.",
        "   같은 인원(사번/성명)이 두 번 이상 입력되면 '이중기입의심' 경고가 표시됩니다.",
        "",
        "6. '급여대장인원명부' 시트 (선택) — 연차수당 대상인원(당기정보)과 기말 급여대장상 실제 인원이",
        "   일치하는지 대사하려는 용도입니다. 회사로부터 기말 급여대장 인원명부를 엑셀로 받으면, 사업장/부서/",
        "   사번/성명/직급만 이 시트에 붙여넣으세요. 앱이 사번(없으면 성명) 기준으로 '당기정보'와 자동 대사해",
        "   요약표에 '연차수당 대상인원 대사' 표를 추가합니다 — 급여대장에만 있으면(연차 대상 인원 누락 가능),",
        "   당기정보에만 있으면(당기 중 퇴사 등) 각각 원인 확인이 필요하다는 경고가 표시됩니다.",
        "   인원명부를 확보하지 못했다면 이 시트는 비워두고, 대신 '기준정보' 시트의 '기말 급여대장상",
        "   총인원수(명부 미확보 시 참고용)'에 총원 숫자만 입력하세요 — 연차수당 대상인원수와 단순 총인원",
        "   비교만 요약표에 표시됩니다(원인 파악은 수기로 확인). 둘 다 비워두면 이 대사 자체가 생략됩니다.",
        "",
        "7. '기준정보' 시트",
        "   '연차산정기준'(전사 공통 설정, 드롭다운): '입사기준' 또는 '회계기준' 중 회사의 실제 운영 방식을",
        "     선택하세요. 미입력 시 '입사기준'으로 계산됩니다.",
        "       입사기준 — 개인별 입사기념일마다 근속연수가 갱신되어 그 시점에 연차가 전액 개별 부여되고,",
        "         직전 확정 기념일부터 결산기준일까지 '만 1개월 개근'이 완성된 횟수만큼(발생주의 월할,",
        "         기념일 당일이 속한 첫 달은 하루만 근무했어도 카운트 안 함) 다음 사이클을 안분해 추가",
        "         가산됩니다(예: 5/31 입사자, 결산일 12/31 → 6개월 뒤인 11/30까지 6개월 + 12/31 완성",
        "         1개월 = 7개월 → 다음 근속연수분 × 7/12 가산).",
        "       회계기준 — 전 직원이 결산기준일(회계연도 종료일)에 일괄로 근속연수가 갱신되어 한 번에",
        "         부여받습니다(당기 근무로 발생하는 연차를 당기말에 인식 — 법적 사용가능일이 익년이라도).",
        "         입사연도(근속연수 0년차)에는 비례연차(재직개월수 비례, 개근 원칙 — 입사월이라도",
        "         하루만 근무했으면 그 달은 미포함)가 적용됩니다.",
        "       ※ 두 기준 모두 발생주의(K-IFRS 1019/일반기준 21장 누적유급휴가 원칙) — 그 연차를 만든",
        "         근로가 제공된 회계기간 말에 부채로 인식하며, 근로기준법상 법적 청구권 발생일(대개 익년",
        "         1/1 또는 입사기념일 다음날)과는 무관합니다.",
        "   '당기/전기 연차사용촉진 반영 부채인정비율(%)': 연차사용촉진제도(근로기준법 제61조)를 적법하게",
        "     이행해도, 잔여연차 중 완전히 소멸(미사용+촉진 이행으로 금전보상의무까지 면제)되는 부분만",
        "     부채가 0입니다. 사용될 것으로 예상되는 부분은 그 유급휴가를 제공할 의무 자체가 이미 당기",
        "     근무의 대가로 발생했으므로(일반기준 21.5의2 매칭원칙, 현금유출 여부와 무관) 부채이고, 미사용",
        "     인데 촉진 실패로 금전보상해야 하는 부분도 당연히 부채입니다. 인원별 구성비를 확인하기 어려우므로,",
        "     전사 공통으로 '완전소멸되지 않고 부채로 남을 것으로 예상하는 비율(%) = 1−완전소멸 예상비율'을",
        "     입력하면 그 비율만큼만 충당부채로 인식합니다(잔여일수 자체는 그대로 표시되고 금액에만 곱해짐).",
        "     연도마다 촉진 이행 여부·실제 소멸률이 달라질 수 있어 당기/전기 비율을 각각 따로 입력받습니다 —",
        "     당기 비율은 당기말 충당부채(재계산)에, 전기 비율은 전기말 충당부채(재계산)에 각각 적용됩니다.",
        "     촉진제도를 쓰지 않거나 비율을 모르면 비워두세요(100%로 계산 — 잔여연차 전액을 충당부채로 인식).",
        "     [예시] 잔여연차 10일, 1일 통상임금 100,000원인 직원이 있다면: 7일은 내년 실사용 예정(부채 O),",
        "     2일은 미사용인데 촉진 실패로 결국 지급 예정(부채 O), 1일은 미사용+촉진 적법이행으로 완전소멸",
        "     (부채 X) → 부채인정비율 = (7+2)/10 = 90% → 연차충당부채 = 10일×100,000원×90% = 900,000원.",
        "     ※ 흔한 실수: '미사용예상비율'(2+1일=30%)이나 '지급될 비율'(2일=20%)만 곱하면 사용예정분",
        "     7일치가 통째로 빠져 과소계상됩니다 — 이 셀(B열)에 마우스를 올리면 같은 예시가 메모로도 뜹니다.",
        "   전기말/당기말 회사계상 연차충당부채(제조원가분/판관비분) 4칸과, 당기 연차충당부채 차변(당기지급액)",
        "   1칸을 입력합니다. 다섯 값 모두 인별 재계산액과의 대사(비교)용 참고값일 뿐, 재계산 자체에는",
        "   반영되지 않습니다. '당기지급액'은 journal_analyzer(분개장분석) 메뉴에서 관련 계정의 당기 차변",
        "   합계를 뽑아 입력하면 됩니다 — 전기말(회사계상)+당기 연차수당비용(재계산)-당기지급액이 당기말",
        "   (회사계상)과 맞는지 요약표에서 자동으로 T계정 검증(tie-out)합니다. 모르면 비워둬도 됩니다.",
        "",
        "8. 결산기준일은 이 파일에 입력하지 않습니다. 실행 시 파일명의 'fy<연도>'와 --fiscal-month 옵션(기본 12월)으로",
        "   severance_analyzer/depreciation_analyzer와 동일한 규칙으로 당기말·전기말 결산기준일을 자동 계산합니다.",
        "   예) --fiscal-month 12 --fiscal-year 2026 → 당기말 2026-12-31, 전기말 2025-12-31.",
        "",
        "9. 파일명 규칙: 이 템플릿을 복사해 'leave_<회사명>_information_fy<회계연도>.xlsx' 로 저장하세요.",
        "   예) leave_kyungnam_information_fy2026.xlsx (전기·당기 데이터가 모두 이 한 파일 안에 들어갑니다).",
    ]
    for i, line in enumerate(lines, start=1):
        cell = guide.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    os.makedirs(os.path.join(HERE, "input_data"), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"템플릿 생성 완료: {OUT_PATH}")


if __name__ == "__main__":
    build()
