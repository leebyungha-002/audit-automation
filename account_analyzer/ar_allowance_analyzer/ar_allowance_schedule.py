"""매출채권 연령분석 및 대손충당금 설정 검증앱 — 엔진 (발생기준 연령분석 + roll rate).

input_data/ar_allowance_<company>_information_fy<year>.xlsx 를 읽어 거래처별 매출채권을
'발생일 기준 경과개월수'로 연령분석하고, 연령구간별 대손율(비상장사는 회사 설정값, 상장사는 앱이
계산한 전이율(roll rate) 기반 누적손실률)을 적용한 대손충당금을 재산출해 회사계상액과 대사하는
output/ar_allowance_schedule_<company>_<fy>.xlsx 를 생성한다.

핵심 설계(2026-08-27, blue sky 요청으로 due-date 기준에서 발생기준으로 전면 재설계)
1. 연령 = 결제기일이 아니라 발생일 기준 경과개월수. 거래처 잔액이 발생일이 다른 여러 채권의 합계일
   수 있어 결제기일 하나로 전체를 판정하기 어렵기 때문. 경과개월수 = (기준일 연월) - (발생일 연월) + 1
   (월 단위 그리드, 일자는 무시 — 예: 11월 발생분은 12월말 기준 '2개월째').
2. 연령 스프레드를 얻는 방법 두 가지를 모두 지원(기준정보 '연령산정 입력방식'):
   (a) 회사연령표 — 회사가 이미 만든 (거래처×기준일×연령구간) 잔액표를 그대로 사용.
   (b) 차변발생내역 — 대변(입금)은 전혀 고려하지 않고, 최근 발생분부터 거슬러 올라가며 그 기준일
       잔액에 도달할 때까지 누적하는 방식으로 연령을 재구성한다("입금은 항상 오래된 채권부터 먼저
       상계된다"는 가정과 수학적으로 동일). 제공된 발생내역이 잔액을 다 못 채우면 부족분은 자동으로
       최고령 구간으로 처리된다(별도의 기초잔액 입력 불필요).
3. 상장사는 위 스프레드를 결산일 포함 여러(8개 분기 이상 권장) 기준일에 만들어, 연속된 두 기준일
   사이 "구간 i 합계 → 다음 기준일 구간 i+1 합계"의 금액가중평균 비율(전이율)을 계산하고, 이를
   최고령구간부터 역순으로 누적곱해 구간별 누적손실률(=최종 대손율)을 산출한다. 비상장사는 이
   전이율 계산 없이 '연령별대손율' 시트의 회사설정 대손율을 그대로 사용한다.
4. 개별평가(부도/회생/소송)·특수관계자채권은 거래처 플래그로 연령분석(집합평가) 모집단에서 분리한다
   (실무에서 가장 흔한 오류가 이 분리 누락이라는 점이 설계 동기 — 이전 버전과 동일 원칙 유지).
5. 담보/보증 차감액은 연령구간 중 가장 오래된 구간부터 순서대로 차감한다(보수적 가정).

실행 예:
    python ar_allowance_schedule.py kyungnam --fiscal-month 12
    python ar_allowance_schedule.py --file ar_allowance_kyungnam_information_fy2026.xlsx
"""
import argparse
import calendar
import glob
import os
from datetime import date

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

HERE = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(HERE, "input_data")
OUTPUT_DIR = os.path.join(HERE, "output")

SIG_THRESHOLD_ABS = 1000
SIG_THRESHOLD_PCT = 0.01
SIG_THRESHOLD_RATE_PP = 0.01
MIN_RECOMMENDED_PERIODS = 9  # 결산일 포함 9개 기준일 = 전이(분기쌍) 8회

CUSTOMER_SHEET = "거래처정보"
AGING_TABLE_SHEET = "연령분석표"
BALANCE_SHEET = "분기말잔액"
TRANSACTION_SHEET = "차변발생내역"
RATE_SHEET = "연령별대손율"

DEFAULT_BUCKET_THRESHOLDS_MONTHS = [3, 6, 9, 12]

CATEGORY_POOLED = "집합평가(연령분석)"
CATEGORY_INDIVIDUAL = "개별평가"
CATEGORY_RELATED = "특수관계자(별도검토)"

BASIS_CURRENT_PERIOD_LABEL = "당기말 기준일(결산기준일, 선택)"
BASIS_LISTED_LABEL = "상장구분(상장/비상장)"
BASIS_METHOD_LABEL = "연령산정 입력방식(회사연령표/차변발생내역)"
BASIS_THRESHOLDS_LABEL = "연령구간 상한(개월, 콤마구분)"
BASIS_TERMINAL_LOSS_LABEL = "최고령구간 최종손실률(%, 상장사, 미입력시 100%)"
BASIS_FORWARD_LOOKING_LABEL = "Forward-looking(미래전망정보) 조정 반영 여부(Y/N, 상장사 참고)"
BASIS_PRIOR_AR_LABEL = "전기말 회사계상 매출채권 총액(원)"
BASIS_PRIOR_ALLOWANCE_LABEL = "전기말 회사계상 대손충당금(원)"
BASIS_CURRENT_ALLOWANCE_LABEL = "당기말 회사계상 대손충당금(원)"
BASIS_TRANSFER_IN_LABEL = "당기 대손충당금 전입액(손익, 선택, 원)"
BASIS_REVERSAL_LABEL = "당기 대손충당금 환입액(선택, 원)"
BASIS_WRITEOFF_LABEL = "당기 대손금 직접상각(제각)액(선택, 분개장 기준, 원)"

FORMULA_NOTE_LINES = [
    ("핵심 계산식", True),
    ("경과개월수 = (기준일이 속한 연월) − (발생일이 속한 연월) + 1  (일자는 무시하는 월 단위 그리드)", False),
    ("[집합평가] 순채권액(구간별) = 연령 스프레드(회사연령표 또는 차변발생내역 재구성) − 담보차감액(최고령구간부터 차감)", False),
    ("[집합평가] 대손충당금(계산) = Σ 구간별 순채권액 × 그 구간의 적용대손율", False),
    ("  비상장사: 적용대손율 = '연령별대손율' 시트의 회사설정 대손율", False),
    ("  상장사: 적용대손율 = 전이율(roll rate) 누적손실률 = (구간i→구간i+1 전이율) × (구간i+1의 누적손실률), "
     "최고령구간의 누적손실률 = 최고령구간 최종손실률(기준정보, 미입력시 100%)", False),
    ("  전이율(구간i→구간i+1) = Σ(다음 기준일 구간i+1 합계) / Σ(이번 기준일 구간i 합계)  (여러 기준일쌍의 금액을 "
     "각각 합산한 뒤 나누는 금액가중평균, 8개 분기 이상의 과거 데이터 권장)", False),
    (f"[{CATEGORY_INDIVIDUAL}] 대손충당금(계산) = 당기말 총채권액 − 담보차감액 − 개별평가 회수가능예상액"
     "(미입력 시 순채권액 전액을 잠정 계상하고 경고)", False),
    (f"[{CATEGORY_RELATED}] 신용위험 성격이 달라 위 계산에서 제외, 별도 표로만 표시(대손충당금 별도 검토 필요)", False),
    ("차변발생내역 재구성: 최근 발생분부터 거슬러 올라가며 그 기준일 잔액을 채울 때까지 누적(대변/입금은 미고려 "
     "— 입금은 오래된 채권부터 먼저 상계된다는 가정과 동일). 발생내역이 잔액을 못 채우면 부족분은 자동으로 "
     "최고령 구간으로 처리.", False),
]


# ── 공용 헬퍼 ────────────────────────────────────────────────────────────────

def _safe_float(v, default: float = 0.0):
    if v is None or v == "":
        return default
    try:
        if pd.isna(v):
            return default
    except (TypeError, ValueError):
        pass
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _safe_date(v):
    if v is None or v == "":
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            return from_excel(v).date()
        except Exception:
            return None
    try:
        ts = pd.Timestamp(v)
    except (ValueError, TypeError):
        # 존재하지 않는 날짜(예: "2025-02-29", 2025년은 윤년이 아님) 등 파싱 자체가 불가능한 값 —
        # 크래시 대신 None으로 처리하고, 호출부(load_*)에서 이런 행을 모아 경고로 보여준다.
        return None
    return ts.date() if not pd.isna(ts) else None


def _fy_bounds(target_fy: str, fiscal_month: int) -> tuple:
    fy = int(target_fy)
    if fiscal_month == 12:
        return f"{fy}-01", f"{fy}-12"
    return f"{fy - 1}-{fiscal_month + 1:02d}", f"{fy}-{fiscal_month:02d}"


def _apply_interim(fy_end: str, target_fy: str, interim_month: int = None) -> str:
    if not interim_month:
        return fy_end
    interim_ym = f"{target_fy}-{interim_month:02d}"
    return min(fy_end, interim_ym)


def _ym_to_end_date(ym: str) -> date:
    y, m = int(ym[:4]), int(ym[5:7])
    last_day = calendar.monthrange(y, m)[1]
    return date(y, m, last_day)


def _find_input_file(company: str = None, file: str = None) -> str:
    if file:
        path = file if os.path.isabs(file) else os.path.join(INPUT_DIR, file)
        if not os.path.exists(path):
            raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {path}")
        return path
    if company:
        pattern = os.path.join(INPUT_DIR, f"ar_allowance_{company}_information_fy*.xlsx")
    else:
        pattern = os.path.join(INPUT_DIR, "ar_allowance_*_information_fy*.xlsx")
    matches = [p for p in glob.glob(pattern) if "template" not in os.path.basename(p)]
    if not matches:
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {pattern}")
    if len(matches) > 1 and not company:
        raise ValueError(f"회사를 특정해주세요. 후보 파일 여러 개: {matches}")
    return matches[0]


def _is_significant(diff, base, abs_threshold: float = SIG_THRESHOLD_ABS) -> bool:
    if diff is None or pd.isna(diff):
        return False
    if abs(diff) >= abs_threshold and (base in (None, 0) or abs(diff) >= abs(base) * SIG_THRESHOLD_PCT):
        return True
    return False


# ── 연령구간(발생기준, 월 단위) ────────────────────────────────────────────

def parse_bucket_thresholds(basis: dict) -> list:
    raw = basis.get(BASIS_THRESHOLDS_LABEL)
    if raw is None or not str(raw).strip():
        return list(DEFAULT_BUCKET_THRESHOLDS_MONTHS)
    try:
        vals = sorted({int(float(str(x).strip())) for x in str(raw).split(",") if str(x).strip()})
        return vals if vals else list(DEFAULT_BUCKET_THRESHOLDS_MONTHS)
    except ValueError:
        return list(DEFAULT_BUCKET_THRESHOLDS_MONTHS)


def bucket_labels(thresholds: list) -> list:
    labels = [f"{thresholds[0]}개월 이내"]
    prev = thresholds[0]
    for t in thresholds[1:]:
        labels.append(f"{prev}개월초과~{t}개월")
        prev = t
    labels.append(f"{thresholds[-1]}개월초과")
    return labels


def age_in_months(발생일: date, 기준일: date) -> int:
    """경과개월수 = (기준일 연월) - (발생일 연월) + 1. 일자는 무시하는 월 단위 그리드(발생월을 1개월째로 카운트)."""
    return (기준일.year * 12 + 기준일.month) - (발생일.year * 12 + 발생일.month) + 1


def bucket_for_age(age_months, thresholds: list, labels: list) -> str:
    if age_months is None:
        return "(발생일 정보 부족)"
    if age_months <= thresholds[0]:
        return labels[0]
    for i, t in enumerate(thresholds[1:], start=1):
        if age_months <= t:
            return labels[i]
    return labels[-1]


def apply_collateral(bucket_amounts: dict, labels: list, collateral: float) -> dict:
    """담보/보증 차감액을 연령구간 중 가장 오래된 구간부터 순서대로 차감(보수적 가정)."""
    result = dict(bucket_amounts)
    remaining = collateral or 0.0
    if remaining <= 0:
        return result
    for label in reversed(labels):
        if remaining <= 1e-9:
            break
        take = min(result.get(label, 0.0), remaining)
        result[label] = result.get(label, 0.0) - take
        remaining -= take
    return result


def allocate_aging_from_transactions(txns: list, balance: float, thresholds: list, labels: list,
                                      기준일: date) -> tuple:
    """차변발생내역 기반 연령 재구성. txns: [(발생일, 발생액), ...]. 대변(입금)은 고려하지 않고,
    최근 발생분부터 거슬러 올라가며 balance에 도달할 때까지 누적한다(입금은 오래된 채권부터 먼저
    상계된다는 가정과 동일). 부족분은 자동으로 최고령 구간으로 처리한다."""
    result = {l: 0.0 for l in labels}
    if balance is None or balance <= 0:
        return result, None
    valid = [(d, a) for d, a in txns if d is not None and d <= 기준일 and a is not None and a > 0]
    valid.sort(key=lambda x: x[0], reverse=True)
    allocated = 0.0
    for 발생일, 금액 in valid:
        if allocated >= balance - 1e-6:
            break
        take = min(금액, balance - allocated)
        label = bucket_for_age(age_in_months(발생일, 기준일), thresholds, labels)
        result[label] += take
        allocated += take
    warning = None
    if allocated < balance - 1e-6:
        shortfall = balance - allocated
        result[labels[-1]] += shortfall
        warning = (f"차변발생내역이 잔액({balance:,.0f}원)을 모두 커버하지 못함 — 부족분 {shortfall:,.0f}원을 "
                   f"최고령 구간('{labels[-1]}')으로 처리")
    return result, warning


# ── 입력 로딩 ────────────────────────────────────────────────────────────────

def _load_rows(path: str, sheet_name: str) -> list:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [c.value for c in ws[2]]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def load_customers(path: str) -> dict:
    rows = _load_rows(path, CUSTOMER_SHEET)
    return {str(r.get("거래처명")).strip(): r for r in rows if r.get("거래처명")}


def load_aging_table(path: str, labels: list) -> tuple:
    """({(거래처명, 기준일): {연령구간: 금액}}, bad_rows) 반환. 연령구간 컬럼명은 '<label>(원)' 형식.
    bad_rows: 기준일을 날짜로 해석할 수 없었던 (거래처명, 원본값) 목록 — 해당 행은 계산에서 제외된다."""
    rows = _load_rows(path, AGING_TABLE_SHEET)
    result = {}
    bad_rows = []
    for r in rows:
        거래처명 = r.get("거래처명")
        if not 거래처명:
            continue
        raw = r.get("기준일")
        기준일 = _safe_date(raw)
        if 기준일 is None:
            if raw not in (None, ""):
                bad_rows.append((str(거래처명).strip(), raw))
            continue
        bucket_amounts = {}
        for label in labels:
            bucket_amounts[label] = _safe_float(r.get(f"{label}(원)"))
        result[(str(거래처명).strip(), 기준일)] = bucket_amounts
    return result, bad_rows


def load_balances(path: str) -> tuple:
    """({(거래처명, 기준일): 채권잔액총액}, bad_rows) 반환."""
    rows = _load_rows(path, BALANCE_SHEET)
    result = {}
    bad_rows = []
    for r in rows:
        거래처명 = r.get("거래처명")
        if not 거래처명:
            continue
        raw = r.get("기준일")
        기준일 = _safe_date(raw)
        if 기준일 is None:
            if raw not in (None, ""):
                bad_rows.append((str(거래처명).strip(), raw))
            continue
        result[(str(거래처명).strip(), 기준일)] = _safe_float(r.get("채권잔액총액(원)"))
    return result, bad_rows


def load_transactions(path: str) -> tuple:
    """({거래처명: [(발생일, 발생액), ...]}, bad_rows) 반환. bad_rows: 발생일자를 날짜로 해석할 수
    없었던(예: 존재하지 않는 날짜) (거래처명, 원본값) 목록 — 해당 행은 연령 재구성에서 제외된다."""
    rows = _load_rows(path, TRANSACTION_SHEET)
    result = {}
    bad_rows = []
    for r in rows:
        거래처명 = r.get("거래처명")
        if not 거래처명:
            continue
        raw = r.get("발생일자")
        발생일 = _safe_date(raw)
        if 발생일 is None and raw not in (None, ""):
            bad_rows.append((str(거래처명).strip(), raw))
        금액 = _safe_float(r.get("발생액(원)"), default=None)
        result.setdefault(str(거래처명).strip(), []).append((발생일, 금액))
    return result, bad_rows


def load_rate_table(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    if RATE_SHEET not in wb.sheetnames:
        return {}
    ws = wb[RATE_SHEET]
    table = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue
        label = str(row[0]).strip()
        rate_raw = row[1] if len(row) > 1 else None
        actual_raw = row[2] if len(row) > 2 else None
        rate = _safe_float(rate_raw) / 100.0 if rate_raw not in (None, "") else None
        actual = _safe_float(actual_raw) / 100.0 if actual_raw not in (None, "") else None
        table[label] = {"rate": rate, "actual_rate": actual}
    return table


def load_basis(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "기준정보" not in wb.sheetnames:
        return {}
    ws = wb["기준정보"]
    basis = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        label, value = row[0], row[1]
        if not isinstance(label, str):
            continue
        label = label.strip()
        if label:
            basis[label] = value
    return basis


def _basis_float(basis: dict, label: str):
    v = basis.get(label)
    return None if v in (None, "") else _safe_float(v)


def basis_current_period_override(basis: dict):
    """'기준정보' 시트의 '당기말 기준일'을 직접 입력한 경우 그 날짜를 반환(없으면 None) —
    입력하면 파일명(fy<연도>)·--fiscal-month/--fiscal-year/--interim-month 옵션보다 우선한다."""
    return _safe_date(basis.get(BASIS_CURRENT_PERIOD_LABEL))


def listed_type(basis: dict) -> str:
    raw = str(basis.get(BASIS_LISTED_LABEL) or "").strip()
    return raw if raw in ("상장", "비상장") else "비상장"


def input_method(basis: dict) -> str:
    raw = str(basis.get(BASIS_METHOD_LABEL) or "").strip()
    return raw if raw in ("회사연령표", "차변발생내역") else "회사연령표"


def forward_looking_applied(basis: dict) -> str:
    return str(basis.get(BASIS_FORWARD_LOOKING_LABEL) or "").strip().upper() or "N"


def terminal_loss_rate(basis: dict) -> float:
    v = _basis_float(basis, BASIS_TERMINAL_LOSS_LABEL)
    if v is None or v < 0 or v > 100:
        return 1.0
    return v / 100.0


# ── 거래처×기준일 연령 스프레드 ─────────────────────────────────────────────

def get_period_buckets(거래처명: str, 기준일: date, method: str, thresholds: list, labels: list,
                        aging_table: dict, balances: dict, transactions: dict) -> tuple:
    """(bucket_amounts, total, warning) 반환. 데이터가 없으면 전부 0에 warning."""
    if method == "회사연령표":
        buckets = aging_table.get((거래처명, 기준일))
        if buckets is None:
            return {l: 0.0 for l in labels}, 0.0, f"{기준일} 기준 연령분석표 데이터 없음"
        return dict(buckets), sum(buckets.values()), None
    balance = balances.get((거래처명, 기준일))
    if balance is None:
        return {l: 0.0 for l in labels}, 0.0, f"{기준일} 기준 분기말잔액 데이터 없음"
    txns = transactions.get(거래처명, [])
    buckets, warning = allocate_aging_from_transactions(txns, balance, thresholds, labels, 기준일)
    return buckets, balance, warning


def collect_periods(method: str, aging_table: dict, balances: dict) -> list:
    if method == "회사연령표":
        return sorted({k[1] for k in aging_table.keys()})
    return sorted({k[1] for k in balances.keys()})


def resolve_current_period(periods: list, 결산일: date) -> tuple:
    if 결산일 in periods:
        return 결산일, None
    candidates = [p for p in periods if p <= 결산일]
    if not candidates:
        return None, f"결산기준일({결산일}) 이하의 연령 데이터가 없음 — 당기말 계산 불가"
    chosen = max(candidates)
    return chosen, (f"⚠ 결산기준일({결산일})과 정확히 일치하는 기준일 데이터가 없어 가장 최근 기준일"
                     f"({chosen})을 당기말로 간주함")


# ── roll rate(전이율) ────────────────────────────────────────────────────────

def compute_period_bucket_totals(pooled_customer_names: list, periods: list, method: str, thresholds: list,
                                  labels: list, aging_table: dict, balances: dict, transactions: dict) -> dict:
    """집합평가 대상 거래처만 합산한 {기준일: {연령구간: 금액}}. roll rate 계산용(담보차감 미반영 — 전이율은
    총액 기준 통계이므로 특정 거래처의 담보 사정과 무관하게 산출)."""
    totals = {}
    for 기준일 in periods:
        agg = {l: 0.0 for l in labels}
        for 거래처명 in pooled_customer_names:
            buckets, _, _ = get_period_buckets(거래처명, 기준일, method, thresholds, labels, aging_table, balances,
                                                transactions)
            for l in labels:
                agg[l] += buckets.get(l, 0.0)
        totals[기준일] = agg
    return totals


def compute_roll_rates(period_bucket_totals: dict, labels: list) -> dict:
    periods_sorted = sorted(period_bucket_totals.keys())
    rr = {}
    for i in range(len(labels) - 1):
        num = den = 0.0
        for t0, t1 in zip(periods_sorted, periods_sorted[1:]):
            den += period_bucket_totals[t0].get(labels[i], 0.0)
            num += period_bucket_totals[t1].get(labels[i + 1], 0.0)
        rr[labels[i]] = (num / den) if den > 1e-9 else None
    return rr


def compute_cumulative_loss_rates(rr: dict, labels: list, terminal_rate: float) -> dict:
    cum = {labels[-1]: terminal_rate}
    for i in range(len(labels) - 2, -1, -1):
        step = rr.get(labels[i])
        nxt = cum.get(labels[i + 1])
        cum[labels[i]] = None if (step is None or nxt is None) else step * nxt
    return cum


# ── 거래처별 계산 ────────────────────────────────────────────────────────────

def classify_customer(rec: dict) -> str:
    if str(rec.get("특수관계자여부(Y/N)") or "").strip().upper() == "Y":
        return CATEGORY_RELATED
    if str(rec.get("개별평가대상여부(Y/N)") or "").strip().upper() == "Y":
        return CATEGORY_INDIVIDUAL
    return CATEGORY_POOLED


def compute_all(customers: dict, 결산일: date, method: str, thresholds: list, labels: list,
                 aging_table: dict, balances: dict, transactions: dict, rate_by_bucket: dict, basis: dict,
                 listed: str) -> dict:
    periods = collect_periods(method, aging_table, balances)
    current_period, current_period_warning = resolve_current_period(periods, 결산일)

    # '거래처정보'에 없는 거래처가 연령표/잔액표에만 존재하면 기본적으로 집합평가로 간주(경고).
    known_names = set(customers.keys())
    aging_names = ({k[0] for k in aging_table.keys()} if method == "회사연령표" else {k[0] for k in balances.keys()})
    unknown_names = aging_names - known_names
    for name in unknown_names:
        customers[name] = {"거래처명": name}

    pooled_names = [n for n, r in customers.items() if classify_customer(r) == CATEGORY_POOLED]
    individual_names = [n for n, r in customers.items() if classify_customer(r) == CATEGORY_INDIVIDUAL]
    related_names = [n for n, r in customers.items() if classify_customer(r) == CATEGORY_RELATED]

    rr = None
    cum_loss = None
    period_bucket_totals = None
    n_transitions = 0
    if listed == "상장" and current_period is not None:
        period_bucket_totals = compute_period_bucket_totals(pooled_names, periods, method, thresholds, labels,
                                                              aging_table, balances, transactions)
        n_transitions = max(0, len(periods) - 1)
        rr = compute_roll_rates(period_bucket_totals, labels)
        cum_loss = compute_cumulative_loss_rates(rr, labels, terminal_loss_rate(basis))
        applied_rate = {l: cum_loss.get(l) for l in labels}
    else:
        applied_rate = {l: (rate_by_bucket.get(l, {}).get("rate")) for l in labels}

    pooled_rows = []
    pooled_customer_rows = []
    if current_period is not None:
        for 거래처명 in pooled_names:
            rec = customers[거래처명]
            raw_buckets, total, warn = get_period_buckets(거래처명, current_period, method, thresholds, labels,
                                                            aging_table, balances, transactions)
            collateral = _safe_float(rec.get("담보/보증 등 차감액(원)"))
            net_buckets = apply_collateral(raw_buckets, labels, collateral)
            customer_total_charge = 0.0
            customer_net_total = 0.0
            for label in labels:
                금액 = net_buckets.get(label, 0.0)
                customer_net_total += 금액
                rate = applied_rate.get(label)
                charge = None if rate is None else 금액 * rate
                if charge:
                    customer_total_charge += charge
                if abs(금액) > 1e-6 or (charge or 0) > 1e-6:
                    pooled_rows.append({
                        "거래처명": 거래처명, "연령구간": label, "구간채권액(원)": 금액,
                        "적용대손율(%)": None if rate is None else rate * 100.0,
                        "대손충당금(계산,원)": charge, "비고": warn or "",
                    })
            회사계상_raw = rec.get("거래처별 회사계상 대손충당금(원)")
            회사계상 = _safe_float(회사계상_raw) if 회사계상_raw not in (None, "") else None
            pooled_customer_rows.append({
                "거래처명": 거래처명, "순채권액(원)": customer_net_total, "대손충당금(계산,원)": customer_total_charge,
                "거래처별 회사계상 대손충당금(원)": 회사계상,
                "차이(계산-회사계상)": None if 회사계상 is None else customer_total_charge - 회사계상,
            })
    pooled_df = pd.DataFrame(pooled_rows) if pooled_rows else pd.DataFrame(
        columns=["거래처명", "연령구간", "구간채권액(원)", "적용대손율(%)", "대손충당금(계산,원)", "비고"])
    pooled_customer_df = pd.DataFrame(pooled_customer_rows) if pooled_customer_rows else pd.DataFrame(
        columns=["거래처명", "순채권액(원)", "대손충당금(계산,원)", "거래처별 회사계상 대손충당금(원)", "차이(계산-회사계상)"])

    individual_rows = []
    for 거래처명 in individual_names:
        rec = customers[거래처명]
        if current_period is None:
            총채권액, balance_warn = 0.0, "당기말 기준일을 확인할 수 없음"
        else:
            _, 총채권액, balance_warn = get_period_buckets(거래처명, current_period, method, thresholds, labels,
                                                          aging_table, balances, transactions)
        담보차감 = _safe_float(rec.get("담보/보증 등 차감액(원)"))
        순채권액 = 총채권액 - 담보차감
        회수가능액_raw = rec.get("개별평가 회수가능예상액(원)")
        회수가능액 = _safe_float(회수가능액_raw) if 회수가능액_raw not in (None, "") else None
        tags = [f"⚠ {balance_warn}"] if balance_warn else []
        if 회수가능액 is None:
            tags.append("⚠ 개별평가 회수가능예상액 미입력 — 순채권액 전액을 잠정 대손충당금으로 계상")
            대손충당금 = max(순채권액, 0.0)
        else:
            대손충당금 = max(순채권액 - 회수가능액, 0.0)
        warn = " / ".join(tags)
        회사계상_raw = rec.get("거래처별 회사계상 대손충당금(원)")
        회사계상 = _safe_float(회사계상_raw) if 회사계상_raw not in (None, "") else None
        individual_rows.append({
            "거래처명": 거래처명, "순채권액(원)": 순채권액, "개별평가사유": rec.get("개별평가사유(선택)"),
            "개별평가 회수가능예상액(원)": 회수가능액, "대손충당금(계산,원)": 대손충당금,
            "거래처별 회사계상 대손충당금(원)": 회사계상,
            "차이(계산-회사계상)": None if 회사계상 is None else 대손충당금 - 회사계상, "비고": warn,
        })
    individual_df = pd.DataFrame(individual_rows) if individual_rows else pd.DataFrame(
        columns=["거래처명", "순채권액(원)", "개별평가사유", "개별평가 회수가능예상액(원)", "대손충당금(계산,원)",
                 "거래처별 회사계상 대손충당금(원)", "차이(계산-회사계상)", "비고"])

    related_rows = []
    for 거래처명 in related_names:
        if current_period is None:
            총채권액, balance_warn = 0.0, "당기말 기준일을 확인할 수 없음"
        else:
            _, 총채권액, balance_warn = get_period_buckets(거래처명, current_period, method, thresholds, labels,
                                                          aging_table, balances, transactions)
        비고 = "별도 검토 필요" + (f" / ⚠ {balance_warn}" if balance_warn else "")
        related_rows.append({"거래처명": 거래처명, "총채권액(원)": 총채권액, "비고": 비고})
    related_df = pd.DataFrame(related_rows) if related_rows else pd.DataFrame(columns=["거래처명", "총채권액(원)", "비고"])

    return {
        "current_period": current_period, "current_period_warning": current_period_warning,
        "periods": periods, "n_transitions": n_transitions,
        "period_bucket_totals": period_bucket_totals, "roll_rates": rr, "cumulative_loss_rates": cum_loss,
        "applied_rate": applied_rate,
        "pooled_df": pooled_df, "pooled_customer_df": pooled_customer_df,
        "individual_df": individual_df, "related_df": related_df,
    }


def build_bucket_verification(labels: list, applied_rate: dict, rate_table: dict, pooled_df: pd.DataFrame,
                               listed: str) -> list:
    rows = []
    prev_rate = None
    for label in labels:
        sub = pooled_df[pooled_df["연령구간"] == label] if not pooled_df.empty else pooled_df
        순채권액 = float(sub["구간채권액(원)"].sum()) if not sub.empty else 0.0
        대손충당금 = float(sub["대손충당금(계산,원)"].sum(skipna=True)) if not sub.empty else 0.0
        적용율 = applied_rate.get(label)
        참고 = rate_table.get(label, {})
        참고율 = 참고.get("rate") if listed == "상장" else 참고.get("actual_rate")
        tags = []
        if 적용율 is None:
            tags.append("⚠ 이 구간 대손율을 계산할 수 없음(데이터 부족)")
        else:
            if prev_rate is not None and 적용율 < prev_rate - 1e-9:
                tags.append("⚠ 이전(더 짧은) 연령구간보다 대손율이 낮음 — 비정상 패턴 의심")
            prev_rate = 적용율
            if 참고율 is not None and 적용율 < 참고율 - SIG_THRESHOLD_RATE_PP:
                tags.append("⚠ 참고값보다 낮음 — 과소설정 가능성" if listed != "상장" else "⚠ 회사 별도 산출값보다 낮음")
        rows.append({
            "연령구간": label, "순채권액(원)": 순채권액,
            "적용대손율(계산,%)": None if 적용율 is None else 적용율 * 100.0,
            ("회사설정 대손율(참고,%)" if listed == "상장" else "회사설정 대손율(%)"): None if 참고율 is None else 참고율 * 100.0,
            "대손충당금(계산,원)": 대손충당금, "비고": " / ".join(tags),
        })
    return rows


def build_category_summary(pooled_customer_df: pd.DataFrame, individual_df: pd.DataFrame,
                            related_df: pd.DataFrame) -> list:
    def _sum(df, col):
        return float(df[col].sum()) if not df.empty else 0.0

    return [
        {"평가구분": CATEGORY_POOLED, "건수": int(len(pooled_customer_df)), "순채권액(원)": _sum(pooled_customer_df, "순채권액(원)"),
         "대손충당금(계산,원)": _sum(pooled_customer_df, "대손충당금(계산,원)"), "비고": ""},
        {"평가구분": CATEGORY_INDIVIDUAL, "건수": int(len(individual_df)), "순채권액(원)": _sum(individual_df, "순채권액(원)"),
         "대손충당금(계산,원)": _sum(individual_df, "대손충당금(계산,원)"), "비고": ""},
        {"평가구분": CATEGORY_RELATED, "건수": int(len(related_df)), "순채권액(원)": _sum(related_df, "총채권액(원)"),
         "대손충당금(계산,원)": None, "비고": "집합평가·개별평가 합계에서 제외 — 신용위험 성격이 달라 별도 검토 필요"},
    ]


def build_overall_summary(pooled_customer_df: pd.DataFrame, individual_df: pd.DataFrame, basis: dict) -> dict:
    총순채권액 = (float(pooled_customer_df["순채권액(원)"].sum()) if not pooled_customer_df.empty else 0.0) + \
             (float(individual_df["순채권액(원)"].sum()) if not individual_df.empty else 0.0)
    총대손충당금 = (float(pooled_customer_df["대손충당금(계산,원)"].sum()) if not pooled_customer_df.empty else 0.0) + \
              (float(individual_df["대손충당금(계산,원)"].sum()) if not individual_df.empty else 0.0)

    당기말_회사계상 = _basis_float(basis, BASIS_CURRENT_ALLOWANCE_LABEL)
    당기말차이 = None if 당기말_회사계상 is None else 총대손충당금 - 당기말_회사계상
    설정률_당기 = (총대손충당금 / 총순채권액) if 총순채권액 else None

    전기말_매출채권 = _basis_float(basis, BASIS_PRIOR_AR_LABEL)
    전기말_충당금 = _basis_float(basis, BASIS_PRIOR_ALLOWANCE_LABEL)
    설정률_전기 = (전기말_충당금 / 전기말_매출채권) if (전기말_매출채권 and 전기말_충당금 is not None) else None
    설정률차이 = None if (설정률_당기 is None or 설정률_전기 is None) else 설정률_당기 - 설정률_전기

    당기전입액 = _basis_float(basis, BASIS_TRANSFER_IN_LABEL)
    당기환입액 = _basis_float(basis, BASIS_REVERSAL_LABEL) or 0.0
    당기직접상각액 = _basis_float(basis, BASIS_WRITEOFF_LABEL) or 0.0
    tie_out = None
    if 전기말_충당금 is not None and 당기전입액 is not None and 당기말_회사계상 is not None:
        계산상기말 = 전기말_충당금 + 당기전입액 - 당기환입액 - 당기직접상각액
        tie_out = {
            "전기말(회사계상)": 전기말_충당금, "당기 전입액(입력)": 당기전입액, "당기 환입액(입력)": 당기환입액,
            "당기 직접상각액(입력)": 당기직접상각액, "계산상 당기말": 계산상기말, "당기말(회사계상)": 당기말_회사계상,
            "차이(계산상당기말-회사계상)": 계산상기말 - 당기말_회사계상,
        }

    return {
        "총순채권액(집합+개별,원)": 총순채권액, "총대손충당금(계산,원)": 총대손충당금,
        "당기말(회사계상,원)": 당기말_회사계상, "당기말차이(계산-회사계상,원)": 당기말차이,
        "설정률(당기,재계산)": 설정률_당기, "설정률(전기,입력값기준)": 설정률_전기,
        "설정률차이(당기-전기,%p)": None if 설정률차이 is None else 설정률차이 * 100.0,
        "설정률_유의변동": (설정률차이 is not None and abs(설정률차이) >= SIG_THRESHOLD_RATE_PP),
        "tie_out": tie_out,
    }


def check_rate_label_mismatch(labels: list, rate_table: dict) -> list:
    warnings = []
    missing = [l for l in labels if l not in rate_table or rate_table[l].get("rate") is None]
    extra = [l for l in rate_table if l not in labels]
    if missing:
        warnings.append(f"⚠ '{RATE_SHEET}' 시트에 대손율이 없는 연령구간: {', '.join(missing)}")
    if extra:
        warnings.append(f"⚠ '{RATE_SHEET}' 시트에 있으나 현재 연령구간 설정과 맞지 않는(사용되지 않는) 구간: {', '.join(extra)}")
    return warnings


# ── 결과 저장 ────────────────────────────────────────────────────────────────

def save_results(result: dict, output_path: str, company: str, target_fy: str, 결산일: date, labels: list,
                  rate_table: dict, basis: dict, listed: str, method: str) -> None:
    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    subtotal_fill = PatternFill("solid", fgColor="D9E1F2")
    total_fill = PatternFill("solid", fgColor="9DC3E6")
    sig_fill = PatternFill("solid", fgColor="FFFF00")
    bold = Font(bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_table(ws, df: pd.DataFrame, start_row: int, money_cols=(), pct_cols=(), col_widths=None):
        headers = list(df.columns)
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=start_row, column=i, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border
            ws.column_dimensions[get_column_letter(i)].width = (col_widths or {}).get(h, 18)
        r = start_row + 1
        totals = {c: 0.0 for c in money_cols}
        for _, row in df.iterrows():
            for i, h in enumerate(headers, start=1):
                val = row[h]
                val = None if (val is None or (isinstance(val, float) and pd.isna(val))) else val
                cell = ws.cell(row=r, column=i, value=val)
                cell.border = border
                if h in money_cols and val is not None:
                    cell.number_format = "#,##0"
                    totals[h] += val
                if h in pct_cols and val is not None:
                    cell.number_format = "0.00"
                if h == "차이(계산-회사계상)" and val is not None:
                    base_col = "거래처별 회사계상 대손충당금(원)"
                    if base_col in row and _is_significant(val, row[base_col]):
                        cell.fill = sig_fill
                if h == "비고" and val:
                    cell.fill = sig_fill
            r += 1
        if not df.empty and money_cols:
            ws.cell(row=r, column=1, value="합계").font = bold
            for i, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=i)
                cell.fill = total_fill
                cell.border = border
                if h in money_cols:
                    cell.value = totals[h]
                    cell.number_format = "#,##0"
                    cell.font = bold
            r += 1
        return r

    # 1. 집합평가상세
    ws1 = wb.active
    ws1.title = "집합평가상세"
    ws1.cell(row=1, column=1,
             value=f"집합평가(연령분석) 상세 (회사: {company}, 회계연도: {target_fy}, 상장구분: {listed}, "
                   f"당기말 기준일: {result['current_period']})").font = Font(bold=True, size=13)
    if result["current_period_warning"]:
        ws1.cell(row=2, column=1, value=result["current_period_warning"]).font = Font(color="C00000")
    r = 4
    ws1.cell(row=r, column=1, value="[거래처×연령구간 상세]").font = bold
    r += 1
    r = _write_table(ws1, result["pooled_df"], r, money_cols=("구간채권액(원)", "대손충당금(계산,원)"),
                      pct_cols=("적용대손율(%)",), col_widths={"비고": 34, "거래처명": 20})
    r += 2
    ws1.cell(row=r, column=1, value="[거래처별 요약(회사계상 대사)]").font = bold
    r += 1
    _write_table(ws1, result["pooled_customer_df"], r,
                 money_cols=("순채권액(원)", "대손충당금(계산,원)", "거래처별 회사계상 대손충당금(원)", "차이(계산-회사계상)"),
                 col_widths={"거래처명": 20})

    # 2. 개별평가_특수관계자상세
    ws2 = wb.create_sheet("개별평가_특수관계자상세")
    ws2.cell(row=1, column=1, value="[개별평가]").font = bold
    r = _write_table(ws2, result["individual_df"], 2,
                      money_cols=("순채권액(원)", "개별평가 회수가능예상액(원)", "대손충당금(계산,원)",
                                  "거래처별 회사계상 대손충당금(원)", "차이(계산-회사계상)"),
                      col_widths={"거래처명": 20, "개별평가사유": 24, "비고": 40})
    r += 2
    ws2.cell(row=r, column=1, value="[특수관계자채권 — 별도 검토 필요]").font = bold
    r += 1
    _write_table(ws2, result["related_df"], r, money_cols=("총채권액(원)",), col_widths={"거래처명": 20})

    # 3. 연령구간별검증 (+ 상장사면 전이율/기간별 구간합계)
    ws3 = wb.create_sheet("연령구간별검증")
    r = 1
    if listed == "상장":
        ws3.cell(row=r, column=1, value="[기준일별 연령구간 합계(집합평가, 원)]").font = bold
        r += 1
        totals = result["period_bucket_totals"] or {}
        period_headers = ["기준일"] + labels
        for i, h in enumerate(period_headers, start=1):
            c = ws3.cell(row=r, column=i, value=h)
            c.fill = header_fill
            c.font = header_font
            c.border = border
            ws3.column_dimensions[get_column_letter(i)].width = 18
        r += 1
        for 기준일 in sorted(totals.keys()):
            ws3.cell(row=r, column=1, value=기준일).border = border
            for i, label in enumerate(labels, start=2):
                cell = ws3.cell(row=r, column=i, value=totals[기준일].get(label, 0.0))
                cell.number_format = "#,##0"
                cell.border = border
            r += 1
        r += 1
        if result["n_transitions"] < MIN_RECOMMENDED_PERIODS - 1:
            ws3.cell(row=r, column=1,
                     value=f"⚠ 과거 데이터가 {MIN_RECOMMENDED_PERIODS}개 기준일(전이 {MIN_RECOMMENDED_PERIODS - 1}회) "
                           f"미만(현재 전이 {result['n_transitions']}회) — 전이율 추정 신뢰도 낮음, 참고치로만 활용 요망"
                     ).font = Font(color="C00000", bold=True)
            r += 1
        r += 1
        ws3.cell(row=r, column=1, value="[구간별 전이율(roll rate)]").font = bold
        r += 1
        for i, h in enumerate(["구간(from)", "→ 구간(to)", "전이율(%)"], start=1):
            c = ws3.cell(row=r, column=i, value=h)
            c.fill = header_fill
            c.font = header_font
            c.border = border
        r += 1
        rr = result["roll_rates"] or {}
        for i in range(len(labels) - 1):
            ws3.cell(row=r, column=1, value=labels[i]).border = border
            ws3.cell(row=r, column=2, value=labels[i + 1]).border = border
            val = rr.get(labels[i])
            cell = ws3.cell(row=r, column=3, value=None if val is None else val * 100.0)
            cell.border = border
            if val is not None:
                cell.number_format = "0.00"
            r += 1
        r += 2

    ws3.cell(row=r, column=1, value="[연령구간별 최종 검증]").font = bold
    r += 1
    bucket_rows = build_bucket_verification(labels, result["applied_rate"], rate_table, result["pooled_df"], listed)
    bucket_df = pd.DataFrame(bucket_rows)
    money_cols = ("순채권액(원)", "대손충당금(계산,원)")
    pct_cols = [c for c in bucket_df.columns if "%" in c]
    _write_table(ws3, bucket_df, r, money_cols=money_cols, pct_cols=pct_cols, col_widths={"비고": 40})

    warn_row = r + len(bucket_df) + 3
    for w in check_rate_label_mismatch(labels, rate_table):
        ws3.cell(row=warn_row, column=1, value=w).font = Font(color="C00000", bold=True)
        warn_row += 1

    # 4. 요약
    ws4 = wb.create_sheet("요약")
    ws4.cell(row=1, column=1,
             value=f"매출채권 대손충당금 검증 요약 (회사: {company}, 회계연도: {target_fy}, 상장구분: {listed}, "
                   f"입력방식: {method})").font = Font(bold=True, size=13)
    ws4.column_dimensions["A"].width = 46
    ws4.column_dimensions["B"].width = 22
    ws4.column_dimensions["C"].width = 50

    r = 3
    ws4.cell(row=r, column=1, value="[평가구분별 요약]").font = bold
    r += 1
    cat_df = pd.DataFrame(build_category_summary(result["pooled_customer_df"], result["individual_df"],
                                                  result["related_df"]))
    r = _write_table(ws4, cat_df, r, money_cols=("순채권액(원)", "대손충당금(계산,원)"), col_widths={"비고": 46})

    r += 1
    ws4.cell(row=r, column=1, value="[연령구간별 대손충당금 요약(집합평가, 거래처 합산)]").font = bold
    r += 1
    bucket_summary_df = bucket_df[["연령구간", "순채권액(원)", "적용대손율(계산,%)", "대손충당금(계산,원)"]].rename(
        columns={"순채권액(원)": "구간채권액(원)", "적용대손율(계산,%)": "적용대손율(%)", "대손충당금(계산,원)": "대손충당금(원)"}
    ) if not bucket_df.empty else pd.DataFrame(columns=["연령구간", "구간채권액(원)", "적용대손율(%)", "대손충당금(원)"])
    r = _write_table(ws4, bucket_summary_df, r, money_cols=("구간채권액(원)", "대손충당금(원)"), pct_cols=("적용대손율(%)",))

    overall = build_overall_summary(result["pooled_customer_df"], result["individual_df"], basis)
    r += 2
    ws4.cell(row=r, column=1, value="[전체 대사]").font = bold
    r += 1
    for label in ["총순채권액(집합+개별,원)", "총대손충당금(계산,원)", "당기말(회사계상,원)", "당기말차이(계산-회사계상,원)"]:
        ws4.cell(row=r, column=1, value=label).border = border
        cell = ws4.cell(row=r, column=2, value=overall[label])
        cell.border = border
        if overall[label] is not None:
            cell.number_format = "#,##0"
        if label == "당기말차이(계산-회사계상,원)" and _is_significant(overall[label], overall["당기말(회사계상,원)"]):
            cell.fill = sig_fill
        r += 1

    r += 1
    ws4.cell(row=r, column=1, value="[설정률 전기 대비 비교]").font = bold
    r += 1
    for label in ["설정률(당기,재계산)", "설정률(전기,입력값기준)", "설정률차이(당기-전기,%p)"]:
        ws4.cell(row=r, column=1, value=label).border = border
        val = overall[label]
        cell = ws4.cell(row=r, column=2,
                         value=(val * 100.0 if val is not None and label != "설정률차이(당기-전기,%p)" else val))
        cell.border = border
        if val is not None:
            cell.number_format = "0.00"
        if label == "설정률차이(당기-전기,%p)" and overall["설정률_유의변동"]:
            cell.fill = sig_fill
            ws4.cell(row=r, column=3, value="⚠ 전기 대비 설정률이 1%p 이상 변동 — 사유 확인 필요").font = Font(color="C00000")
        r += 1

    tie_out = overall["tie_out"]
    if tie_out:
        r += 1
        ws4.cell(row=r, column=1, value="[대손충당금 T계정 검증(tie-out)]").font = bold
        r += 1
        for label in ["전기말(회사계상)", "당기 전입액(입력)", "당기 환입액(입력)", "당기 직접상각액(입력)",
                      "계산상 당기말", "당기말(회사계상)", "차이(계산상당기말-회사계상)"]:
            ws4.cell(row=r, column=1, value=label).border = border
            cell = ws4.cell(row=r, column=2, value=tie_out[label])
            cell.border = border
            cell.number_format = "#,##0"
            if label == "차이(계산상당기말-회사계상)" and _is_significant(tie_out[label], tie_out["당기말(회사계상)"]):
                cell.fill = sig_fill
            r += 1

    r += 2
    ws4.cell(row=r, column=1, value="[상장구분별 안내]").font = bold
    r += 1
    if listed == "상장":
        fl = forward_looking_applied(basis)
        note = ("Forward-looking(미래전망정보) 조정 반영: " +
                ("Y — 근거 문서화 상태 확인 필요" if fl == "Y" else "N 또는 미입력 — K-IFRS9상 반영 필요 여부 및 근거를 회사에 확인 요망"))
        ws4.cell(row=r, column=1, value=note).font = Font(color="C00000" if fl != "Y" else "000000")
        r += 1
        ws4.cell(row=r, column=1,
                 value=f"전이율(roll rate)은 기준일 {len(result['periods'])}개(전이 {result['n_transitions']}회)로 "
                       f"계산됨 — {MIN_RECOMMENDED_PERIODS}개 기준일(전이 8회) 이상 권장.")
    else:
        ws4.cell(row=r, column=1,
                 value="비상장사 — '연령구간별검증' 표의 '최근 실제대손율' 대비 회사설정 대손율 괴리(과소설정 가능성) "
                       "경고를 우선 확인하세요.")
    r += 2

    ws4.cell(row=r, column=1, value="[참고: 계산 공식]").font = bold
    r += 1
    for text, is_bold in FORMULA_NOTE_LINES:
        cell = ws4.cell(row=r, column=1, value=text)
        if is_bold:
            cell.font = bold
        r += 1

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="매출채권 연령분석/대손충당금 설정 검증앱")
    parser.add_argument("company", nargs="?", default=None, help="처리할 회사명 (생략 시 파일 자동 탐색)")
    parser.add_argument("--file", default=None, help="처리할 특정 입력 파일명 (input_data/ 기준)")
    parser.add_argument("--fiscal-month", type=int, default=12, help="결산월 (기본 12월)")
    parser.add_argument("--fiscal-year", default=None, help="검증 대상 회계연도. 생략 시 입력파일명의 fy 뒤 숫자 사용")
    parser.add_argument("--interim-month", type=int, default=None, help="반기 등 중간결산 검토월")
    args = parser.parse_args()

    input_path = _find_input_file(args.company, args.file)
    company = args.company or os.path.basename(input_path).split("_")[2]
    basis = load_basis(input_path)

    target_fy = args.fiscal_year
    if not target_fy:
        base = os.path.basename(input_path)
        idx = base.lower().find("fy")
        if idx == -1:
            raise ValueError("--fiscal-year 를 지정하거나 파일명에 'fy<연도>'를 포함하세요.")
        digits = ""
        for ch in base[idx + 2:]:
            if ch.isdigit():
                digits += ch
            else:
                break
        target_fy = f"20{digits}" if len(digits) == 2 else digits

    print(f"[입력] {input_path}")
    print(f"[대상] 회사={company}, 회계연도={target_fy}, 결산월={args.fiscal_month}")

    결산일_override = basis_current_period_override(basis)
    if 결산일_override is not None:
        결산일 = 결산일_override
        print(f"[결산기준일] {결산일} (기준정보 시트의 '{BASIS_CURRENT_PERIOD_LABEL}' 값 사용 — "
              f"--fiscal-month/--fiscal-year/--interim-month 옵션 무시됨)")
    else:
        _, fy_end_ym = _fy_bounds(target_fy, args.fiscal_month)
        fy_end_ym = _apply_interim(fy_end_ym, target_fy, args.interim_month)
        결산일 = _ym_to_end_date(fy_end_ym)
        print(f"[결산기준일] {결산일}")

    listed = listed_type(basis)
    method = input_method(basis)
    thresholds = parse_bucket_thresholds(basis)
    labels = bucket_labels(thresholds)
    print(f"[상장구분] {listed}, [입력방식] {method}, [연령구간] {labels}")

    customers = load_customers(input_path)
    rate_table = load_rate_table(input_path)
    aging_table, aging_bad = (load_aging_table(input_path, labels) if method == "회사연령표" else ({}, []))
    balances, balance_bad = (load_balances(input_path) if method == "차변발생내역" else ({}, []))
    transactions, txn_bad = (load_transactions(input_path) if method == "차변발생내역" else ({}, []))

    def _print_bad_dates(title: str, bad_rows: list):
        if not bad_rows:
            return
        print(f"[경고] {title} {len(bad_rows)}건 — 날짜를 해석할 수 없어 해당 행이 제외됨(존재하지 않는 "
              f"날짜이거나 형식 오류일 수 있음, 원본 파일에서 확인 필요):")
        for name, raw in bad_rows[:20]:
            print(f"  - {name}: {raw!r}")
        if len(bad_rows) > 20:
            print(f"  ... 외 {len(bad_rows) - 20}건")

    _print_bad_dates(f"'{AGING_TABLE_SHEET}' 시트의 기준일 파싱 실패", aging_bad)
    _print_bad_dates(f"'{BALANCE_SHEET}' 시트의 기준일 파싱 실패", balance_bad)
    _print_bad_dates(f"'{TRANSACTION_SHEET}' 시트의 발생일자 파싱 실패", txn_bad)

    result = compute_all(customers, 결산일, method, thresholds, labels, aging_table, balances, transactions,
                          rate_table, basis, listed)
    print(f"[당기말 기준일] {result['current_period']} (전체 기준일 {len(result['periods'])}개)")
    if result["current_period_warning"]:
        print(f"[경고] {result['current_period_warning']}")

    suffix = f"_interim{args.interim_month:02d}" if args.interim_month else ""
    output_path = os.path.join(OUTPUT_DIR, f"ar_allowance_schedule_{company}_{target_fy}{suffix}.xlsx")
    save_results(result, output_path, company, target_fy, 결산일, labels, rate_table, basis, listed, method)
    print(f"[완료] {output_path}")


if __name__ == "__main__":
    main()
