"""표준 입력 템플릿(ar_allowance_template.xlsx) 생성 스크립트.

실행: python build_template.py
input_data/ar_allowance_template.xlsx 를 새로 만든다(이미 있으면 덮어씀).
회사별 파일은 이 템플릿을 복사해 ar_allowance_<company>_information_fy<year>.xlsx 로 저장해서 사용한다.

설계 원칙(2026-08-27 발생기준 연령분석 방식으로 전면 재설계):
  - 연령은 결제기일(만기일)이 아니라 '발생일 기준 경과개월수'로 판정한다. 거래처 잔액이 발생일자가
    서로 다른 여러 채권의 합계일 수 있어, 결제기일 하나로 전체 잔액의 연령을 판정하는 방식은 무리가
    있기 때문이다. 연령구간은 3/6/9/12개월/12개월초과 5구간이 기본값이다.
  - 연령 스프레드(거래처 잔액을 연령구간별로 나누는 것)를 얻는 방법은 두 가지를 모두 지원한다
    ('기준정보' 시트의 '연령산정 입력방식'으로 선택):
      (1) 회사연령표 — 회사가 이미 만든 (거래처×기준일×연령구간별) 연령분석표를 그대로 입력.
      (2) 차변발생내역 — 결산기간 매출채권 차변(청구/매출인식) 발생내역만 입력받아, 대변(입금 등)은
          전혀 고려하지 않고 '최근 발생분부터 거슬러 올라가며 결산시점 잔액에 도달할 때까지 누적'하는
          방식으로 앱이 직접 연령을 재구성한다. 이는 "입금은 항상 오래된 채권부터 먼저 상계된다"는
          가정과 수학적으로 동일하다 — 즉 결산일 현재 남아있는 잔액은 최근 발생분부터 채워진다.
          예) 12월말 잔액 100원, 11/1 발생 90원, 7/1 발생 100원 → 11/1분 90원(최근분 전액) +
          7/1분 10원(90원을 채우고 남은 잔액을 채우기 위한 부분만, 나머지 90원은 이미 회수된 것으로
          간주)으로 연령이 재구성된다. 제공된 차변내역이 잔액을 다 못 채우면(더 오래된 채권이 있다는
          뜻) 그 부족분은 자동으로 최고령 구간('12개월초과')으로 처리된다 — 별도의 기초잔액 입력이
          필요 없다.
  - 상장사는 이 연령 스프레드를 결산일뿐 아니라 과거 여러 분기말 시점에도 각각 만들어(8개 분기 이상
    권장) '연령구간별 전이율(roll rate)'을 계산한다. 전이율은 연속된 두 분기 사이 "구간 i 합계 → 다음
    분기 구간 i+1 합계"의 금액가중평균 비율(여러 분기쌍의 분자·분모를 각각 합산한 뒤 나누는 방식)로
    계산하고, 이를 최고령구간까지 누적곱해 구간별 최종 대손율(누적손실률)을 산출한다. 이동매트릭스
    추정 자체를 재현하지 않기로 했던 이전 설계에서, blue sky의 요청으로 이 계산을 앱이 직접 수행하는
    방향으로 전환했다(2026-08-27).
  - 비상장사는 결산일 한 시점의 연령 스프레드만 필요하며, 연령구간별 대손율은 회사가 실무관행상
    직접 설정한 값을 그대로 사용한다(전이율 계산 없음).
  - 개별평가(부도/회생/소송 등)·특수관계자채권은 여전히 거래처별 플래그로 연령분석 모집단에서
    분리한다(과거분석과 동일 원칙).
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(HERE, "input_data", "ar_allowance_template.xlsx")

YN_OPTIONS = ["Y", "N"]
LISTED_OPTIONS = ["상장", "비상장"]
METHOD_OPTIONS = ["회사연령표", "차변발생내역"]

DEFAULT_BUCKET_THRESHOLDS_MONTHS = "3,6,9,12"
DEFAULT_BUCKET_LABELS = ["3개월 이내", "3개월초과~6개월", "6개월초과~9개월", "9개월초과~12개월", "12개월초과"]

group_fill = PatternFill("solid", fgColor="D9E1F2")
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF")
group_font = Font(bold=True)
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
input_fill = PatternFill("solid", fgColor="FFF2CC")

TRANSACTION_METHOD_MEMO = (
    "'차변발생내역' 방식 — 대변(입금)은 전혀 입력하지 않습니다.\n\n"
    "결산일 현재 잔액이 남아있는 채권은 '최근 발생분부터' 채워진다고 가정합니다(= 입금은 항상 오래된\n"
    "채권부터 먼저 상계된다는 가정과 동일). 이 시트에는 그 거래처의 청구/매출인식(차변) 내역만\n"
    "발생일자·금액으로 나열하면, 앱이 최근 발생분부터 거슬러 올라가며 '분기말잔액' 시트의 잔액에\n"
    "도달할 때까지 누적해서 연령을 재구성합니다.\n\n"
    "예) 분기말잔액 100원, 11/1 발생 90원, 7/1 발생 100원\n"
    "  → 11/1분 90원(전액) + 7/1분 10원(90원을 채우고 남은 잔액만, 나머지 90원은 이미 회수된 것으로\n"
    "    간주)으로 연령이 재구성됩니다.\n\n"
    "제공한 차변내역 합계가 잔액에 못 미치면(더 오래된 채권이 있다는 뜻) 부족분은 자동으로 최고령\n"
    "구간('12개월초과')으로 처리됩니다 — 그 이전 발생내역까지 모두 입력할 필요는 없습니다."
)

RATE_TABLE_MEMO = (
    "연령구간별 대손율 입력 메모\n\n"
    "'연령구간' 열은 '기준정보' 시트의 '연령구간 상한(개월, 콤마구분)' 설정과 일치해야 합니다.\n\n"
    "비상장사: '회사설정 대손율(%)'이 대손충당금 계산에 그대로 사용되는 주된 값입니다(회사가 실무관행상\n"
    "직접 설정한 연령별 대손율). '최근 실제대손율(참고, 선택)'을 입력하면 back-test 비교가 추가됩니다.\n\n"
    "상장사: 대손충당금 계산에는 앱이 계산한 전이율(roll rate) 기반 누적손실률이 사용되고, 이 시트의\n"
    "'회사설정 대손율(%)'은 회사가 별도로 산출한 값이 있을 때 참고 비교용으로만 쓰입니다(선택 입력)."
)


def _merge_group_header(ws, columns):
    start = 1
    prev_group = columns[0][0]
    for i, (group, _, _) in enumerate(columns, start=1):
        if group != prev_group:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=i - 1)
            start = i
            prev_group = group
    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=len(columns))

    for i, (group, name, width) in enumerate(columns, start=1):
        c1 = ws.cell(row=1, column=i, value=group if i == 1 or columns[i - 2][0] != group else None)
        c1.fill = group_fill
        c1.font = group_font
        c1.alignment = center
        c1.border = border

        c2 = ws.cell(row=2, column=i, value=name)
        c2.fill = header_fill
        c2.font = header_font
        c2.alignment = center
        c2.border = border
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "A3"


def _write_rows(ws, columns, examples, money_fields=(), date_fields=()):
    field_names = [c[1] for c in columns]
    for r, row in enumerate(examples, start=3):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            field = field_names[c - 1]
            if field in date_fields and val:
                cell.number_format = "yyyy-mm-dd"
            if field in money_fields and val is not None:
                cell.number_format = "#,##0"


# ── '거래처정보' 시트 ────────────────────────────────────────────────────
CUSTOMER_COLUMNS = [
    ("거래처정보", "거래처명", 20),
    ("거래처정보", "거래처코드(선택)", 14),
    ("구분", "특수관계자여부(Y/N)", 14),
    ("구분", "개별평가대상여부(Y/N)", 16),
    ("구분", "개별평가사유(선택)", 20),
    ("당기말 조정(선택)", "개별평가 회수가능예상액(원)", 18),
    ("당기말 조정(선택)", "담보/보증 등 차감액(원)", 16),
    ("대사용(선택)", "거래처별 회사계상 대손충당금(원)", 20),
    ("기타", "비고", 30),
]
CUSTOMER_EXAMPLES = [
    ["㈜예시)월별매입처", "CUST-001", "N", "N", None, None, None, None,
     "예시) 집합평가 대상 — '연령분석표'(방식1) 또는 '분기말잔액'+'차변발생내역'(방식2)에서 같은 거래처명으로 연령이 재구성됨"],
    ["㈜예시)관계사", "CUST-002", "Y", "N", None, None, None, None,
     "예시) 특수관계자채권 — 집합평가 모집단에서 제외되고 별도 표로 표시됨"],
    ["㈜예시)회생절차거래처", "CUST-003", "N", "Y", "회생절차 개시(2026-05월)", 20000000, None, 60000000,
     "예시) 개별평가대상 — 연령분석 대신 (당기말 총채권액-회수가능예상액)으로 개별 계산됨"],
    ["㈜예시)담보보유거래처", "CUST-004", "N", "N", None, None, 2000000, None,
     "예시) 담보/보증 차감액이 있으면 연령구간 중 가장 오래된 구간부터 순서대로 차감됨(보수적 가정)"],
]

# ── '연령분석표' 시트 (방식1: 회사연령표) ──────────────────────────────────
AGING_TABLE_COLUMNS = [
    ("거래처", "거래처명", 20), ("거래처", "기준일", 14),
] + [("연령구간별 채권잔액(원)", f"{label}(원)", 16) for label in DEFAULT_BUCKET_LABELS]

AGING_TABLE_EXAMPLES = [
    ["㈜예시)월별매입처", "2026-12-31", 90000000, 10000000, 0, 0, 0],
    ["㈜예시)담보보유거래처", "2026-12-31", 0, 0, 12000000, 0, 0],
    # 개별평가·특수관계자 대상 거래처도 당기말 총채권액 확인용으로 연령분석표에 함께 기재해야 한다
    # (개별평가는 연령분석 계산에는 쓰이지 않지만 '당기말 총채권액'을 이 표에서 그대로 조회한다).
    ["㈜예시)회생절차거래처", "2026-12-31", 0, 0, 0, 80000000, 0],
    ["㈜예시)관계사", "2026-12-31", 0, 0, 0, 0, 30000000],
    # 상장사 roll rate 계산용 과거 분기 예시(㈜예시)월별매입처 동일 거래처의 과거 3개 분기 스냅샷 —
    # 모든 구간에 잔액이 있어야 전이율(구간간 비율)이 전 구간에서 계산된다. 실제로는 8개 분기 이상을
    # 권장한다 — 여기서는 표의 구조와 전이율 계산 결과를 보여주는 예시로 3개만 넣었다.
    ["㈜예시)월별매입처", "2026-03-31", 50000000, 25000000, 15000000, 7000000, 3000000],
    ["㈜예시)월별매입처", "2026-06-30", 55000000, 20000000, 13000000, 8000000, 4000000],
    ["㈜예시)월별매입처", "2026-09-30", 65000000, 18000000, 10000000, 5000000, 2000000],
]

# ── '분기말잔액' 시트 (방식2: 차변발생내역) ────────────────────────────────
BALANCE_COLUMNS = [("거래처", "거래처명", 20), ("거래처", "기준일", 14), ("거래처", "채권잔액총액(원)", 18)]
BALANCE_EXAMPLES = [
    ["㈜예시)월별매입처", "2026-12-31", 100000000,
     ],
]

# ── '차변발생내역' 시트 (방식2) ────────────────────────────────────────────
TRANSACTION_COLUMNS = [("발생내역", "거래처명", 20), ("발생내역", "발생일자", 14), ("발생내역", "발생액(원)", 16)]
TRANSACTION_EXAMPLES = [
    ["㈜예시)월별매입처", "2026-11-01", 90000000],
    ["㈜예시)월별매입처", "2026-07-01", 100000000],
]


def build():
    wb = openpyxl.Workbook()

    # 1. 거래처정보
    ws_cust = wb.active
    ws_cust.title = "거래처정보"
    _merge_group_header(ws_cust, CUSTOMER_COLUMNS)
    _write_rows(ws_cust, CUSTOMER_COLUMNS, CUSTOMER_EXAMPLES,
                money_fields=("개별평가 회수가능예상액(원)", "담보/보증 등 차감액(원)", "거래처별 회사계상 대손충당금(원)"))
    last_row = 1000
    for label, options, error in [
        ("특수관계자여부(Y/N)", YN_OPTIONS, "Y 또는 N 중 선택하세요."),
        ("개별평가대상여부(Y/N)", YN_OPTIONS, "Y 또는 N 중 선택하세요."),
    ]:
        idx = next(i for i, (_, name, _) in enumerate(CUSTOMER_COLUMNS, start=1) if name == label)
        col = get_column_letter(idx)
        dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True, showErrorMessage=True)
        dv.error = error
        ws_cust.add_data_validation(dv)
        dv.add(f"{col}3:{col}{last_row}")

    # 2. 연령분석표 (방식1)
    ws_aging = wb.create_sheet("연령분석표")
    _merge_group_header(ws_aging, AGING_TABLE_COLUMNS)
    _write_rows(ws_aging, AGING_TABLE_COLUMNS, AGING_TABLE_EXAMPLES,
                money_fields=[c[1] for c in AGING_TABLE_COLUMNS if c[0] == "연령구간별 채권잔액(원)"],
                date_fields=("기준일",))
    ws_aging.cell(row=len(AGING_TABLE_EXAMPLES) + 4, column=1,
                  value="※ '연령산정 입력방식'이 '회사연령표'일 때만 이 시트를 사용합니다. '차변발생내역' 방식이면 "
                        "이 시트는 비워두고 '분기말잔액'+'차변발생내역' 시트를 채우세요.").font = Font(italic=True, color="808080")

    # 3. 분기말잔액 (방식2)
    ws_bal = wb.create_sheet("분기말잔액")
    _merge_group_header(ws_bal, BALANCE_COLUMNS)
    _write_rows(ws_bal, BALANCE_COLUMNS, BALANCE_EXAMPLES, money_fields=("채권잔액총액(원)",), date_fields=("기준일",))
    ws_bal.cell(row=len(BALANCE_EXAMPLES) + 4, column=1,
                value="※ '연령산정 입력방식'이 '차변발생내역'일 때만 사용합니다. 상장사는 과거 분기말(8개 분기 이상 "
                      "권장)도 같은 거래처명으로 행을 추가하세요 — roll rate 계산에 쓰입니다.").font = \
        Font(italic=True, color="808080")

    # 4. 차변발생내역 (방식2)
    ws_txn = wb.create_sheet("차변발생내역")
    _merge_group_header(ws_txn, TRANSACTION_COLUMNS)
    _write_rows(ws_txn, TRANSACTION_COLUMNS, TRANSACTION_EXAMPLES, money_fields=("발생액(원)",), date_fields=("발생일자",))
    memo_cell = ws_txn.cell(row=2, column=3)
    memo = Comment(TRANSACTION_METHOD_MEMO, "ar_allowance_analyzer")
    memo.width, memo.height = 440, 340
    memo_cell.comment = memo
    ws_txn.cell(row=len(TRANSACTION_EXAMPLES) + 4, column=1,
                value="※ 대변(입금) 내역은 입력하지 않습니다 — 최근 발생분부터 잔액을 채우는 방식으로 앱이 "
                      "자동 계산합니다. 여러 분기말의 잔액 연령을 재구성하려면, 가장 오래된 분기말보다 더 과거의 "
                      "발생내역까지 충분히 입력해야 그 시점 연령이 정확합니다(부족하면 최고령 구간으로 자동 처리).").font = \
        Font(italic=True, color="808080")

    # 5. 연령별대손율
    ws_rate = wb.create_sheet("연령별대손율")
    RATE_COLUMNS = [("연령구간", 16), ("회사설정 대손율(%)", 18), ("최근 실제대손율(참고,선택,%)", 20), ("비고", 34)]
    for i, (name, width) in enumerate(RATE_COLUMNS, start=1):
        c = ws_rate.cell(row=1, column=i, value=name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
        ws_rate.column_dimensions[get_column_letter(i)].width = width
        if name == "회사설정 대손율(%)":
            memo = Comment(RATE_TABLE_MEMO, "ar_allowance_analyzer")
            memo.width, memo.height = 440, 320
            c.comment = memo
    ws_rate.row_dimensions[1].height = 32
    ws_rate.freeze_panes = "A2"

    RATE_EXAMPLES = [
        [DEFAULT_BUCKET_LABELS[0], 1.0, 0.8, "비상장사: 계산에 직접 사용됨. 상장사: 회사 별도 산출값이 있을 때만 참고 비교용"],
        [DEFAULT_BUCKET_LABELS[1], 5.0, 4.0, None],
        [DEFAULT_BUCKET_LABELS[2], 15.0, 12.0, None],
        [DEFAULT_BUCKET_LABELS[3], 40.0, 30.0, None],
        [DEFAULT_BUCKET_LABELS[4], 100.0, 90.0, "예시) 통상 12개월 초과 연체분은 100% 설정하는 경우가 많음(회사 정책에 따라 다름)"],
    ]
    for r, row in enumerate(RATE_EXAMPLES, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws_rate.cell(row=r, column=c, value=val)
            cell.border = border
            if c in (2, 3) and val is not None:
                cell.number_format = "0.0"
                cell.fill = input_fill

    # 6. 기준정보
    basis = wb.create_sheet("기준정보")
    basis.column_dimensions["A"].width = 46
    basis.column_dimensions["B"].width = 20
    basis.column_dimensions["C"].width = 58
    basis.cell(row=1, column=1, value="매출채권 대손충당금(연령분석) 기준정보").font = Font(bold=True, size=13)

    header_row = 3
    for i, h in enumerate(["항목", "값", "설명"], start=1):
        cell = basis.cell(row=header_row, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    rows = [
        ("상장구분(상장/비상장)", "비상장",
         "상장사는 전이율(roll rate)을 앱이 계산해 누적손실률을 대손율로 사용하고, 비상장사는 '연령별대손율' "
         "시트의 회사설정 대손율을 그대로 사용합니다. 연령 스프레드 방법(아래 '연령산정 입력방식')과 개별평가/"
         "특수관계자 분리 로직은 상장/비상장 동일합니다.", "listbox_listed"),
        ("연령산정 입력방식(회사연령표/차변발생내역)", "회사연령표",
         "'회사연령표': '연령분석표' 시트에 회사가 이미 만든 (거래처×기준일×연령구간) 잔액을 그대로 입력. "
         "'차변발생내역': '분기말잔액'+'차변발생내역' 시트에 결산시점 잔액과 차변(청구) 발생내역만 입력하면 "
         "앱이 최근 발생분부터 거슬러 올라가며 자동으로 연령을 재구성합니다(대변/입금은 입력하지 않음).", "listbox_method"),
        ("연령구간 상한(개월, 콤마구분)", DEFAULT_BUCKET_THRESHOLDS_MONTHS,
         "예: 3,6,9,12 → 3개월 이내/3개월초과~6개월/6개월초과~9개월/9개월초과~12개월/12개월초과 5개 구간이 "
         "생성됩니다. 바꾸면 '연령분석표'의 구간 컬럼명 또는 '연령별대손율' 시트의 구간명도 맞춰 다시 작성하세요.", "text"),
        ("최고령구간 최종손실률(%, 상장사, 미입력시 100%)", None,
         "상장사 roll rate 계산에서 최고령 구간('12개월초과')에 남아있는 채권이 궁극적으로 대손처리될 것으로 "
         "보는 비율입니다. 이 값에서부터 역순으로(최고령→최연소) 구간별 전이율을 누적곱해 각 구간의 최종 "
         "대손율(누적손실률)을 산출합니다. 모르면 비워두세요(100% 가정).", "money_pct"),
        ("Forward-looking(미래전망정보) 조정 반영 여부(Y/N, 상장사 참고)", "N",
         "K-IFRS9 기대신용손실모형은 과거실적 기반 전이율에 거시경제지표 등 미래전망정보를 가산 조정하도록 "
         "요구합니다. 이 앱은 그 조정치를 별도 계산하지 않으므로, 회사가 반영했는지 여부만 감사조서에 기록하는 "
         "용도입니다.", "listbox_yn"),
        ("전기말 회사계상 매출채권 총액(원)", None,
         "전기말 재무제표상 매출채권 총액. 아래 전기말 대손충당금과 함께 전기 설정률을 구해 당기 설정률과 "
         "자동 비교합니다.", "money"),
        ("전기말 회사계상 대손충당금(원)", None, "전기말 재무제표(또는 결산정산표)상 실제 계상액.", "money"),
        ("당기말 회사계상 대손충당금(원)", None,
         "당기말 재무제표상 회사 계상액. 앱의 재계산 합계와 비교해 차이(대사)를 표시합니다.", "money"),
        ("당기 대손충당금 전입액(손익, 선택, 원)", None,
         "당기 손익계산서상 대손상각비 중 충당금 전입액. T계정 검증(tie-out)에 사용됩니다.", "money"),
        ("당기 대손충당금 환입액(선택, 원)", None, "당기 중 대손충당금 환입액.", "money"),
        ("당기 대손금 직접상각(제각)액(선택, 분개장 기준, 원)", None,
         "journal_analyzer에서 대손충당금(또는 매출채권) 계정의 당기 직접상각(제각) 분개 합계를 뽑아 입력. "
         "전기말(회사계상)+당기전입액-당기환입액-이 값이 당기말(회사계상)과 맞는지 자동 검증(T계정 tie-out)합니다. "
         "모르면 비워둬도 됩니다.", "money"),
    ]

    r = header_row + 1
    for label, default, note, kind in rows:
        c1 = basis.cell(row=r, column=1, value=label)
        c1.border = border
        c2 = basis.cell(row=r, column=2, value=default if not kind.startswith("money") else None)
        c2.border = border
        c2.alignment = Alignment(horizontal="center", vertical="center")
        if kind.startswith("money"):
            c2.number_format = "#,##0" if kind == "money" else "0.0"
            c2.fill = input_fill
        elif kind == "text":
            c2.fill = input_fill
        elif kind.startswith("listbox"):
            c2.fill = input_fill
            options = {"listbox_listed": LISTED_OPTIONS, "listbox_method": METHOD_OPTIONS,
                       "listbox_yn": YN_OPTIONS}[kind]
            dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True, showErrorMessage=True)
            dv.error = f"{'/'.join(options)} 중 선택하세요."
            basis.add_data_validation(dv)
            dv.add(f"B{r}")
        c3 = basis.cell(row=r, column=3, value=note)
        c3.border = border
        c3.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    basis.cell(row=r + 1, column=1,
               value="※ 결산기준일(당기말)은 이 시트에 입력하지 않습니다 — 실행 시 파일명(fy<연도>)과 "
                     "--fiscal-month 옵션으로 다른 계정 검증앱과 동일한 규칙으로 결정되고, '연령분석표'/'분기말잔액' "
                     "시트에서 그 날짜와 정확히 일치하는 '기준일' 행을 당기말 데이터로 사용합니다.").font = \
        Font(italic=True, color="808080")

    # 7. 작성안내
    guide = wb.create_sheet("작성안내")
    guide.column_dimensions["A"].width = 100
    lines = [
        "매출채권 연령분석 및 대손충당금 설정 검증앱 — 입력 템플릿 작성 안내 (발생기준 연령분석)",
        "",
        "핵심 개념: 연령은 결제기일이 아니라 '발생일 기준 경과개월수'로 판정합니다(거래처 잔액이 발생일이",
        "다른 여러 채권의 합계일 수 있어 결제기일 하나로는 전체를 판정하기 어렵기 때문). 연령 경과월수는",
        "'(기준일이 속한 연월) − (발생일이 속한 연월) + 1'로 계산합니다 — 예: 11월 발생분은 12월말 기준",
        "'2개월째'(11월, 12월 두 달에 걸침), 10월 발생분은 '3개월째'입니다.",
        "",
        "0. 시트 구성",
        "   '거래처정보': 거래처별 특수관계자/개별평가 여부 등 속성(항상 필요, 당기 기준).",
        "   '연령분석표': 방식1(회사연령표) 전용 — 회사가 이미 만든 연령분석표를 그대로 입력.",
        "   '분기말잔액'+'차변발생내역': 방식2(차변발생내역) 전용.",
        "   ※ 개별평가·특수관계자 거래처도 연령분석에는 쓰이지 않지만 '당기말 총채권액'은 이 표에서 그대로",
        "     조회하므로, 방식1이면 '연령분석표'(아무 구간에나 총액을 몰아서 적어도 무방), 방식2면",
        "     '분기말잔액'에 당기말 기준일 행을 반드시 넣어야 합니다 — 빠지면 총채권액이 0으로 계산됩니다.",
        "   '연령별대손율': 비상장사는 필수(대손율 직접 입력), 상장사는 선택(회사 별도 산출값 참고 비교용).",
        "   '기준정보': 상장구분·입력방식·연령구간 등 전사 공통 설정.",
        "",
        "1. '거래처정보' 시트",
        "   '특수관계자여부(Y/N)': Y면 집합평가(연령분석) 모집단에서 제외되고 별도 표에 표시됩니다.",
        "   '개별평가대상여부(Y/N)': 부도·회생절차·소송 등 손상 징후가 있는 거래처는 반드시 Y로 표시하세요.",
        "     연령대와 무관하게 (당기말 총채권액−담보차감액−개별평가 회수가능예상액)으로 대손충당금이 별도",
        "     계산됩니다. 회수가능예상액을 모르면 비워두되, 이 경우 순채권액 전액이 잠정 대손충당금으로",
        "     계상되고 경고가 표시됩니다.",
        "   '담보/보증 등 차감액': 있으면 연령구간 중 가장 오래된 구간부터 순서대로 차감됩니다(보수적 가정 —",
        "     담보는 위험이 가장 큰 채권에 먼저 대응한다고 봄). 집합평가 대상에만 이렇게 적용되고, 개별평가",
        "     대상은 총채권액에서 그냥 차감됩니다.",
        "   '거래처별 회사계상 대손충당금': 알면 입력하세요 — 그 거래처 전체(모든 연령구간 합산)의 앱",
        "     계산값과 대사됩니다.",
        "",
        "2-A. '연령분석표' 시트 (방식1: 회사연령표) — '기준정보'의 입력방식이 '회사연령표'일 때만 사용",
        "   거래처×기준일 조합 1개 = 1행. 연령구간별 채권잔액(원) 컬럼에 회사 연령분석표의 금액을 그대로",
        "   옮겨 적으세요. 비상장사는 당기말(기준일=결산기준일) 1행만 있으면 되고, 상장사는 과거 분기말",
        "   (8개 분기 이상 권장)도 같은 거래처명으로 행을 추가해야 roll rate가 계산됩니다.",
        "",
        "2-B. '분기말잔액'+'차변발생내역' 시트 (방식2: 차변발생내역) — 입력방식이 '차변발생내역'일 때만 사용",
        "   '분기말잔액': 거래처×기준일별 채권잔액 총액만 입력(연령구간 스프레드는 앱이 계산).",
        "   '차변발생내역': 그 거래처의 청구/매출인식(차변) 내역을 발생일자·금액으로 나열. 대변(입금)은",
        "     입력하지 않습니다 — '최근 발생분부터 잔액을 채운다'는 가정(=입금은 오래된 채권부터 먼저",
        "     상계된다는 가정과 동일)으로 앱이 각 기준일 시점 잔액에 도달할 때까지 최근 발생분부터 거슬러",
        "     올라가며 누적해 연령을 재구성합니다. 예) 분기말잔액 100원, 11/1 발생 90원, 7/1 발생 100원 →",
        "     11/1분 90원(전액)+7/1분 10원(잔액을 채우기 위한 일부만)으로 연령이 재구성됩니다.",
        "     제공한 발생내역 합계가 잔액에 못 미치면 부족분은 자동으로 최고령 구간으로 처리되므로, 가장",
        "     오래된 기준일보다 더 이전 발생내역까지 전부 입력할 필요는 없습니다(다만 너무 부족하면 그만큼",
        "     최고령 구간이 과대 계상되니, 각 기준일 기준 최소 12개월치 이상 발생내역을 입력하는 것을",
        "     권장합니다).",
        "",
        "3. '연령별대손율' 시트 — '기준정보'의 '연령구간 상한' 설정과 일치하는 구간명으로 행을 구성하세요.",
        "   비상장사: '회사설정 대손율(%)'이 대손충당금 계산에 그대로 사용됩니다. '최근 실제대손율(선택)'을",
        "     입력하면 회사설정율과 자동 비교(back-test)되어 과소/과대 설정 여부를 확인할 수 있습니다.",
        "   상장사: 대손충당금 계산에는 앱이 전이율로 계산한 누적손실률이 사용되고, 이 시트는 회사가 별도",
        "     산출한 대손율이 있을 때 참고 비교용으로만 쓰입니다(없으면 비워둬도 됩니다).",
        "",
        "4. '기준정보' 시트 — 상장구분·입력방식·연령구간 등은 위 설명 참고. 전기말/당기말 회사계상 금액과",
        "   당기 전입액/환입액/직접상각액을 입력하면 설정률 전기 대비 비교와 T계정 검증(tie-out)이 요약표에",
        "   자동으로 표시됩니다.",
        "",
        "5. 결산기준일은 이 파일에 입력하지 않습니다. 실행 시 파일명의 'fy<연도>'와 --fiscal-month 옵션(기본",
        "   12월)으로 다른 계정 검증앱과 동일한 규칙으로 자동 계산되고, '연령분석표'/'분기말잔액' 시트에서",
        "   그 날짜와 정확히 일치하는 '기준일' 행을 당기말로 사용합니다(일치하는 행이 없으면 그 이하 날짜 중",
        "   가장 최근 것을 당기말로 간주하고 경고를 표시합니다).",
        "",
        "6. 파일명 규칙: 이 템플릿을 복사해 'ar_allowance_<회사명>_information_fy<회계연도>.xlsx' 로 저장하세요.",
        "   예) ar_allowance_kyungnam_information_fy2026.xlsx",
    ]
    for i, line in enumerate(lines, start=1):
        cell = guide.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    os.makedirs(os.path.join(HERE, "input_data"), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"템플릿 생성 완료: {OUT_PATH}")


if __name__ == "__main__":
    build()
