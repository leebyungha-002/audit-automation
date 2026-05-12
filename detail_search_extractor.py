#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
상세검색 직접 추출기 — 웹 UI 우회, 계정별원장 로컬 파일에서 직접 추출
Usage: python detail_search_extractor.py <회사명>

task_list의 'host1_상세검색_시나리오' 시트를 읽고,
계정별원장 Excel에서 직접 데이터를 추출하여 results/ 폴더에 저장합니다.
해당 시트가 없으면 아무 작업 없이 정상 종료합니다.
"""

import sys
import os
import re
from datetime import datetime, date

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)

SUMMARY_KEYWORDS = ['[전 기 이 월]', '[월 계]', '[누 계]', '[기 초]', '[합 계]']
OUTPUT_COLS = ['계정과목', '날짜', '적요란', '거래처명', '사업자등록번호', '차변', '대변']


# ─── 파일 탐색 ────────────────────────────────────────────────────────────────

def find_task_list(company_dir: str) -> str | None:
    files = os.listdir(company_dir)
    company_lower = os.path.basename(company_dir).lower()
    match = (
        next((f for f in files if f.lower().startswith('task_list_') and
              f.lower().endswith('.xlsx') and company_lower in f.lower()), None)
        or next((f for f in files if f.lower().startswith('task_list_') and
                 f.lower().endswith('.xlsx')), None)
    )
    return os.path.join(company_dir, match) if match else None


def find_ledger_file(company_dir: str) -> str | None:
    raw_dir = os.path.join(company_dir, 'raw_data')
    if not os.path.isdir(raw_dir):
        return None
    for f in os.listdir(raw_dir):
        if f.startswith('~$'):
            continue
        ext = os.path.splitext(f)[1].lower()
        if ext in ('.xlsx', '.xls') and '원장' in f and '분개장' not in f and '전기' not in f:
            return os.path.join(raw_dir, f)
    return None


def get_output_prefix(company_dir: str) -> str:
    """results/ 내 기존 파일에서 날짜 접두사 자동 감지, 없으면 오늘 날짜 사용"""
    results_dir = os.path.join(company_dir, 'results')
    company_name = os.path.basename(company_dir)
    if os.path.isdir(results_dir):
        for f in sorted(os.listdir(results_dir)):
            m = re.match(rf'^{re.escape(company_name)}_(\d{{8}})_', f)
            if m:
                return f'{company_name}_{m.group(1)}_'
    today = datetime.now().strftime('%Y%m%d')
    return f'{company_name}_{today}_'


# ─── 시나리오 파싱 ────────────────────────────────────────────────────────────

def read_scenario_sheet(task_list_path: str) -> list:
    """
    task_list의 host1_상세검색_시나리오 시트를 파싱하여 그룹 리스트 반환.
    반환 형태: [{'group': str, 'tasks': [{'계정과목', '거래처명', '금액 유형', '표시방식'}, ...]}, ...]
    """
    wb = load_workbook(task_list_path, read_only=True, data_only=True)
    target_name = None
    for sname in wb.sheetnames:
        if '상세검색' in sname:
            target_name = sname
            break

    if target_name is None:
        wb.close()
        return []

    rows = list(wb[target_name].iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(c).strip() if c is not None else '' for c in rows[0]]

    def cell(row_dict, *keys):
        for k in keys:
            v = row_dict.get(k)
            if v is not None and str(v).strip() not in ('', 'None', 'nan'):
                return str(v).strip()
        return ''

    # 같은 작업명을 가진 행들을 하나의 그룹으로 집약 (순서 보존)
    groups_dict: dict = {}
    groups_order: list = []
    current_group_name = None

    for row in rows[1:]:
        if not any(c is not None for c in row):
            continue
        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

        작업명 = cell(row_dict, '작업명')
        계정과목 = cell(row_dict, '계정과목')

        if not 계정과목:
            continue

        # 작업명이 있으면 현재 그룹명 갱신, 없으면 이전 그룹명 유지
        if 작업명:
            current_group_name = 작업명

        if current_group_name is None:
            current_group_name = '기타'

        if current_group_name not in groups_dict:
            groups_dict[current_group_name] = []
            groups_order.append(current_group_name)

        groups_dict[current_group_name].append({
            '계정과목': 계정과목,
            '거래처명': cell(row_dict, '거래처명'),
            '금액 유형': cell(row_dict, '금액 유형', '금액유형'),
            '표시방식': cell(row_dict, '표시방식'),
        })

    return [{'group': name, 'tasks': groups_dict[name]} for name in groups_order]


# ─── 계정별원장 처리 ──────────────────────────────────────────────────────────

def normalize_account_name(sheet_title: str) -> str:
    """'N_계정명(코드)' → '계정명' 정규화"""
    s = re.sub(r'^\d+_', '', sheet_title)
    s = re.sub(r'\(\d+\)$', '', s).strip()
    return s


def find_ledger_sheet(wb, account_name: str):
    """계정별원장 워크북에서 account_name에 맞는 시트 탐색 (정확 일치 → 포함 일치)"""
    exact, partial = None, None
    for sheet in wb.worksheets:
        norm = normalize_account_name(sheet.title)
        if norm == account_name:
            exact = sheet
            break
        if exact is None and (account_name in norm or norm in account_name):
            partial = sheet
    return exact or partial


def normalize_date(val) -> str:
    if val is None:
        return ''
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return s


def read_ledger_sheet(sheet) -> pd.DataFrame:
    """계정별원장 시트 → DataFrame (집계행 제거, 날짜 정규화)"""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    headers = [str(c).strip() if c is not None else '' for c in rows[0]]
    data = []
    for row in rows[1:]:
        if not any(c is not None for c in row):
            continue
        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

        # 집계행 제거
        적요 = str(row_dict.get('적요란', '') or '').strip()
        if any(kw in 적요 for kw in SUMMARY_KEYWORDS):
            continue

        # 날짜 없는 행 제거
        if not row_dict.get('날짜'):
            continue

        data.append(row_dict)

    if not data:
        return pd.DataFrame()

    df = pd.DataFrame(data)
    df['날짜'] = df['날짜'].apply(normalize_date)

    for col in ['차변', '대변', '잔액']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    return df


def apply_amount_filter(df: pd.DataFrame, amount_type: str) -> pd.DataFrame:
    t = str(amount_type)
    if '차변만' in t:
        return df[df['차변'] > 0].copy()
    if '대변만' in t:
        return df[df['대변'] > 0].copy()
    return df.copy()


def apply_counterpart_filter(df: pd.DataFrame, counterpart: str) -> pd.DataFrame:
    if not counterpart:
        return df
    col = next((c for c in ['거래처명', '거래처'] if c in df.columns), None)
    if col is None:
        return df
    return df[df[col].astype(str).str.contains(counterpart, na=False)].copy()


def apply_display_mode(df: pd.DataFrame, display_mode: str, account_name: str) -> pd.DataFrame:
    """상세내역(기본) 또는 월합계 처리 후 OUTPUT_COLS 형태로 반환"""
    # 열 명칭 통일: 거래처 컬럼이 회사마다 다를 수 있음
    if '거래처명' not in df.columns and '거래처' in df.columns:
        df = df.rename(columns={'거래처': '거래처명'})
    if '사업자등록번호' not in df.columns and '코드' in df.columns:
        df = df.rename(columns={'코드': '사업자등록번호'})

    df['계정과목'] = account_name

    if '월합계' in str(display_mode):
        df = df.copy()
        df['_월'] = df['날짜'].astype(str).str[:7]
        grouped = df.groupby('_월', as_index=False).agg({'차변': 'sum', '대변': 'sum'})
        grouped.rename(columns={'_월': '날짜'}, inplace=True)
        grouped['계정과목'] = account_name
        grouped['적요란'] = '월 합계'
        grouped['거래처명'] = ''
        grouped['사업자등록번호'] = ''
        return grouped[OUTPUT_COLS]

    for col in OUTPUT_COLS:
        if col not in df.columns:
            df[col] = ''
    return df[OUTPUT_COLS]


# ─── 결과 저장 ────────────────────────────────────────────────────────────────

def save_group_result(group_name: str, all_rows: list, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = group_name[:31]

    ws.append(OUTPUT_COLS)

    # 헤더 스타일
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in all_rows:
        ws.append([row.get(c, '') for c in OUTPUT_COLS])

    # 숫자 서식
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in (6, 7):  # 차변, 대변
                cell.number_format = '#,##0'

    # 열 너비
    widths = {'계정과목': 18, '날짜': 12, '적요란': 30,
               '거래처명': 25, '사업자등록번호': 16, '차변': 14, '대변': 14}
    for i, col in enumerate(OUTPUT_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 15)

    count = ws.max_row - 1
    if count <= 0:
        ws2 = wb.create_sheet('결과없음')
        ws2['A1'] = '검색 결과가 없습니다.'

    wb.save(out_path)
    print(f'  → {os.path.basename(out_path)} ({count}건)')


# ─── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print('사용법: python detail_search_extractor.py <회사명>')
        sys.exit(1)

    company_name = sys.argv[1]
    base_dir = os.path.dirname(os.path.abspath(__file__))
    company_dir = os.path.join(base_dir, company_name)

    if not os.path.isdir(company_dir):
        print(f'[오류] 회사 폴더를 찾을 수 없습니다: {company_dir}')
        sys.exit(1)

    task_list_path = find_task_list(company_dir)
    if not task_list_path:
        print(f'[안내] {company_name}: task_list_*.xlsx 없음 — 상세검색 추출 건너뜀')
        sys.exit(0)

    print(f'[안내] task_list: {os.path.basename(task_list_path)}')

    groups = read_scenario_sheet(task_list_path)
    if not groups:
        print(f'[안내] {company_name}: host1_상세검색_시나리오 시트 없음 — 상세검색 추출 건너뜀')
        sys.exit(0)

    print(f'[안내] 그룹 {len(groups)}개: {[g["group"] for g in groups]}')

    ledger_path = find_ledger_file(company_dir)
    if not ledger_path:
        print(f'[오류] {company_name}/raw_data/ 에서 계정별원장 파일을 찾을 수 없습니다.')
        sys.exit(1)

    print(f'[안내] 계정별원장: {os.path.basename(ledger_path)}')
    print('[안내] 계정별원장 로드 중...')
    ledger_wb = load_workbook(ledger_path, read_only=True, data_only=True)
    print(f'[안내] 시트 수: {len(ledger_wb.sheetnames)}')

    results_dir = os.path.join(company_dir, 'results')
    os.makedirs(results_dir, exist_ok=True)
    prefix = get_output_prefix(company_dir)

    for group in groups:
        group_name = group['group']
        tasks = group['tasks']
        print(f'\n[그룹] {group_name} ({len(tasks)}개 계정)')

        all_rows = []
        for task in tasks:
            account = task['계정과목']
            amount_type = task['금액 유형']
            display_mode = task['표시방식']
            counterpart = task['거래처명']

            sheet = find_ledger_sheet(ledger_wb, account)
            if sheet is None:
                print(f'  [스킵] {account} — 계정별원장 시트 없음')
                continue

            df = read_ledger_sheet(sheet)
            if df.empty:
                print(f'  [스킵] {account} — 데이터 없음')
                continue

            df = apply_amount_filter(df, amount_type)
            df = apply_counterpart_filter(df, counterpart)
            count = len(df)

            label = f'  {account}: {count}건 ({amount_type})'
            if counterpart:
                label += f' / 거래처필터={counterpart}'
            print(label)

            df = apply_display_mode(df, display_mode, account)
            all_rows.extend(df.to_dict('records'))

        out_path = os.path.join(results_dir, f'{prefix}{group_name}.xlsx')
        save_group_result(group_name, all_rows, out_path)

    ledger_wb.close()
    print(f'\n[완료] {len(groups)}개 그룹 추출 완료.')


if __name__ == '__main__':
    main()
